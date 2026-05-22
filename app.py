import os
import ast
from datetime import datetime, timedelta, timezone
import csv
import io
import openpyxl
import click
from flask import Flask, render_template, redirect, url_for, request, flash, abort, send_file, Response, make_response, jsonify,session
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import text, inspect, case
from sqlalchemy.exc import OperationalError
from sqlalchemy.sql.sqltypes import Integer
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
import mistune
import uuid
import requests
from bs4 import BeautifulSoup
import json
import shutil
import time
import zipfile
from flask_migrate import Migrate, stamp as migrate_stamp
from sqlalchemy import MetaData
import re

import logging
import traceback

from queue_service import enqueue_judge_task
from wikibook_config import configure_app

try:
    import jedi
except ImportError:
    jedi = None

app = Flask(__name__)
configure_app(app)


# 💡 配置日志，将所有报错信息写入 flask_error.log 文件中
logging.basicConfig(filename='flask_error.log', level=logging.DEBUG, 
                    format='%(asctime)s - %(levelname)s - %(message)s')

# 💡 捕获全局 500 错误，并强行把堆栈信息(Traceback)打印到终端和日志里
@app.errorhandler(500)
def internal_server_error(e):
    error_tb = traceback.format_exc()
    app.logger.error(f"发生 500 错误:\n{error_tb}")
    print(f"\n{'='*50}\n🚨 崩溃日志来啦:\n{error_tb}\n{'='*50}\n")
    return f"<h1>服务器崩溃了</h1><p>请查看控制台或 flask_error.log 文件提取日志发给 AI 助手。</p><pre>{error_tb}</pre>", 500



# ==========================================
# 💡 核心修复：定义一套命名约定，拯救 SQLite 迁移
# ==========================================
convention = {
    "ix": 'ix_%(column_0_label)s',
    "uq": "uq_%(table_name)s_%(column_0_name)s",
    "ck": "ck_%(table_name)s_%(constraint_name)s",
    "fk": "fk_%(table_name)s_%(column_0_name)s_%(referred_table_name)s",
    "pk": "pk_%(table_name)s"
}
metadata = MetaData(naming_convention=convention)

db = SQLAlchemy(app, metadata=metadata)
migrate = Migrate(app, db, render_as_batch=app.config.get("DB_RENDER_AS_BATCH", False))

login_manager = LoginManager(app)
login_manager.login_view = "login"

markdown = mistune.create_markdown(escape=False, plugins=["strikethrough", "table", "url", "task_lists"])
STATEMENT_SAMPLE_BLOCK_RE = re.compile(r"^(input|output)(\d+)$", re.IGNORECASE)
UTC8 = timezone(timedelta(hours=8))


def _statement_sample_block_meta(node):
    if getattr(node, "name", None) != "pre":
        return None

    code_node = node.find("code", recursive=False)
    if code_node is None:
        return None

    for class_name in code_node.get("class") or []:
        language_name = class_name[9:] if class_name.startswith("language-") else class_name
        match = STATEMENT_SAMPLE_BLOCK_RE.match(language_name)
        if match:
            return {
                "kind": match.group(1).lower(),
                "index": int(match.group(2)),
                "code": code_node,
            }
    return None


def _build_statement_sample_pair(soup, sample_index):
    pair_node = soup.new_tag("div", attrs={"class": "statement-sample-pair"})
    pair_node.attrs["data-sample-index"] = str(sample_index)

    for kind in ("input", "output"):
        card_node = soup.new_tag("div", attrs={"class": "statement-sample-card"})
        header_node = soup.new_tag("div", attrs={"class": "statement-sample-card__header"})
        header_node.string = f"{'输入' if kind == 'input' else '输出'} {sample_index}"
        card_node.append(header_node)
        pair_node.append(card_node)

    return pair_node


def render_problem_statement(statement_md):
    soup = BeautifulSoup(markdown(statement_md or ""), "html.parser")
    transformed = False

    def transform_parent(parent):
        nonlocal transformed
        children = list(parent.children)
        index = 0

        while index < len(children):
            child = children[index]
            sample_meta = _statement_sample_block_meta(child)
            if sample_meta and sample_meta["kind"] == "input":
                next_index = index + 1
                gap_nodes = []
                paired_output = None

                while next_index < len(children):
                    sibling = children[next_index]
                    if getattr(sibling, "name", None) is None and not str(sibling).strip():
                        gap_nodes.append(sibling)
                        next_index += 1
                        continue

                    sibling_meta = _statement_sample_block_meta(sibling)
                    if (
                        sibling_meta
                        and sibling_meta["kind"] == "output"
                        and sibling_meta["index"] == sample_meta["index"]
                    ):
                        paired_output = sibling
                    break

                if paired_output is not None:
                    pair_node = _build_statement_sample_pair(soup, sample_meta["index"])
                    child.insert_before(pair_node)
                    cards = pair_node.find_all("div", class_="statement-sample-card", recursive=False)
                    for block, card_node in zip((child, paired_output), cards):
                        block_classes = list(block.get("class") or [])
                        if "statement-sample-card__code" not in block_classes:
                            block_classes.append("statement-sample-card__code")
                        block["class"] = block_classes
                        card_node.append(block)
                    for gap_node in gap_nodes:
                        gap_node.extract()
                    transformed = True
                    children = list(parent.children)
                    continue

            if getattr(child, "name", None) and child.name not in {"pre", "code"}:
                transform_parent(child)
            index += 1

    transform_parent(soup)
    return str(soup), transformed

# UTC+8 Helper
def now_utc8():
    return datetime.now(UTC8).replace(tzinfo=None)

def to_utc8_naive(dt):
    if not dt:
        return dt
    if dt.tzinfo is not None and dt.utcoffset() is not None:
        return dt.astimezone(UTC8).replace(tzinfo=None)
    return dt

def datetime_diff_seconds(later, earlier):
    later = to_utc8_naive(later)
    earlier = to_utc8_naive(earlier)
    if not later or not earlier:
        return 0
    return (later - earlier).total_seconds()

def study_session_seconds(session):
    return max(datetime_diff_seconds(session.end_time, session.start_time), 0)

def is_image_icon(s):
    if not s:
        return False
    v = s.strip().lower()
    if v.startswith("http://") or v.startswith("https://") or v.startswith("/"):
        return v.endswith((".png", ".jpg", ".jpeg", ".gif", ".svg", ".webp"))
    if "." in v:
        ext = v.rsplit(".", 1)[-1]
        return ext in ["png", "jpg", "jpeg", "gif", "svg", "webp"]
    return False

def badge_icon_url(s):
    if not s:
        return ""
    v = s.strip()
    if v.startswith("http://") or v.startswith("https://") or v.startswith("/"):
        return v
    if is_image_icon(v):
        if v.startswith("uploads/"):
             return url_for("static", filename=v)
        return url_for("static", filename=f"uploads/{v}")
    return ""

ALLOWED_IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".png", ".gif", ".webp", ".svg"}
ALLOWED_PROBLEM_FILE_EXTENSIONS = {
    ".jpg", ".jpeg", ".png", ".gif", ".webp", ".svg",
    ".yaml", ".yml", ".json", ".txt", ".md", ".pdf", ".csv", ".zip",
    ".in", ".out", ".ans", ".log"
}
TEXT_EDITABLE_PROBLEM_FILE_EXTENSIONS = {
    ".yaml", ".yml", ".json", ".txt", ".md", ".csv", ".in", ".out", ".ans", ".log"
}
TESTDATA_FILE_EXTENSIONS = {".in", ".out", ".ans", ".yaml", ".yml", ".json", ".txt", ".csv", ".log"}

def is_allowed_image_file(filename):
    ext = os.path.splitext(filename or "")[1].lower()
    return ext in ALLOWED_IMAGE_EXTENSIONS

def save_uploaded_image(file_storage, upload_subdir, prefix):
    if not file_storage or not file_storage.filename:
        return ""
    if not is_allowed_image_file(file_storage.filename):
        raise ValueError("不支持的图片格式，请上传 jpg、jpeg、png、gif、webp 或 svg")

    ext = os.path.splitext(secure_filename(file_storage.filename))[1].lower()
    filename = secure_filename(f"{prefix}_{uuid.uuid4().hex}{ext}")
    save_dir = os.path.join(app.root_path, "static", "uploads", upload_subdir)
    os.makedirs(save_dir, exist_ok=True)
    file_storage.save(os.path.join(save_dir, filename))
    return f"/static/uploads/{upload_subdir}/{filename}"

def remove_uploaded_static_file(file_url, upload_subdir):
    prefix = f"/static/uploads/{upload_subdir}/"
    if not file_url or not file_url.startswith(prefix):
        return
    filename = file_url.split("/")[-1]
    path = os.path.join(app.root_path, "static", "uploads", upload_subdir, filename)
    if os.path.exists(path):
        try:
            os.remove(path)
        except OSError:
            pass


def is_allowed_problem_file(filename):
    ext = os.path.splitext(filename or "")[1].lower()
    return ext in ALLOWED_PROBLEM_FILE_EXTENSIONS


def save_problem_asset(file_storage, problem_id):
    if not file_storage or not file_storage.filename:
        raise ValueError("未选择文件")

    original_filename = secure_filename(file_storage.filename)
    if not original_filename:
        raise ValueError("文件名无效")

    if not is_allowed_problem_file(original_filename):
        raise ValueError("不支持的文件类型")

    ext = os.path.splitext(original_filename)[1].lower()
    stored_filename = secure_filename(f"{uuid.uuid4().hex}{ext}")
    relative_subdir = f"problem_assets/problem_{problem_id}"
    absolute_dir = os.path.join(app.root_path, "static", "uploads", relative_subdir)
    os.makedirs(absolute_dir, exist_ok=True)

    absolute_path = os.path.join(absolute_dir, stored_filename)
    file_storage.save(absolute_path)

    return {
        "filename": original_filename,
        "stored_filename": stored_filename,
        "file_path": f"/static/uploads/{relative_subdir}/{stored_filename}",
        "file_size": os.path.getsize(absolute_path),
        "content_type": file_storage.content_type or "",
    }


def save_oj_assignment_asset(file_storage, assignment_id):
    if not file_storage or not file_storage.filename:
        raise ValueError("未选择文件")

    original_filename = secure_filename(file_storage.filename)
    if not original_filename:
        raise ValueError("文件名无效")

    if not is_allowed_problem_file(original_filename):
        raise ValueError("不支持的文件类型")

    ext = os.path.splitext(original_filename)[1].lower()
    stored_filename = secure_filename(f"{uuid.uuid4().hex}{ext}")
    relative_subdir = f"oj_assignment_assets/assignment_{assignment_id}"
    absolute_dir = os.path.join(app.root_path, "static", "uploads", relative_subdir)
    os.makedirs(absolute_dir, exist_ok=True)

    absolute_path = os.path.join(absolute_dir, stored_filename)
    file_storage.save(absolute_path)

    return {
        "filename": original_filename,
        "stored_filename": stored_filename,
        "file_path": f"/static/uploads/{relative_subdir}/{stored_filename}",
        "file_size": os.path.getsize(absolute_path),
        "content_type": file_storage.content_type or "",
    }


def remove_problem_asset_file(file_path):
    prefix = "/static/uploads/problem_assets/"
    if not file_path or not file_path.startswith(prefix):
        return
    relative_path = file_path[len(prefix):]
    absolute_path = os.path.join(app.root_path, "static", "uploads", "problem_assets", relative_path)
    if os.path.exists(absolute_path):
        try:
            os.remove(absolute_path)
        except OSError:
            pass


def remove_oj_assignment_asset_file(file_path):
    prefix = "/static/uploads/oj_assignment_assets/"
    if not file_path or not file_path.startswith(prefix):
        return
    relative_path = file_path[len(prefix):]
    absolute_path = os.path.join(app.root_path, "static", "uploads", "oj_assignment_assets", relative_path)
    if os.path.exists(absolute_path):
        try:
            os.remove(absolute_path)
        except OSError:
            pass


def is_text_editable_problem_file(filename):
    ext = os.path.splitext(filename or "")[1].lower()
    return ext in TEXT_EDITABLE_PROBLEM_FILE_EXTENSIONS


def problem_asset_absolute_path(file_path):
    if not file_path or not file_path.startswith("/static/uploads/"):
        raise ValueError("文件路径不合法")
    return os.path.join(app.root_path, file_path.lstrip("/"))


def create_text_problem_asset(problem_id, filename, content, uploaded_by_id):
    sanitized_name = secure_filename((filename or "").strip())
    if not sanitized_name:
        raise ValueError("文件名不能为空")
    if not is_allowed_problem_file(sanitized_name):
        raise ValueError("不支持的文件类型")
    if not is_text_editable_problem_file(sanitized_name):
        raise ValueError("该文件类型不支持在线编辑")

    ext = os.path.splitext(sanitized_name)[1].lower()
    stored_filename = secure_filename(f"{uuid.uuid4().hex}{ext}")
    relative_subdir = f"problem_assets/problem_{problem_id}"
    absolute_dir = os.path.join(app.root_path, "static", "uploads", relative_subdir)
    os.makedirs(absolute_dir, exist_ok=True)

    absolute_path = os.path.join(absolute_dir, stored_filename)
    with open(absolute_path, "w", encoding="utf-8") as file_handle:
        file_handle.write(content or "")

    return ProblemFile(
        problem_id=problem_id,
        filename=sanitized_name,
        stored_filename=stored_filename,
        file_path=f"/static/uploads/{relative_subdir}/{stored_filename}",
        content_type="text/plain",
        file_size=os.path.getsize(absolute_path),
        uploaded_by_id=uploaded_by_id,
    )


def update_text_problem_asset(problem_file, filename, content):
    sanitized_name = secure_filename((filename or "").strip())
    if not sanitized_name:
        raise ValueError("文件名不能为空")
    if not is_allowed_problem_file(sanitized_name):
        raise ValueError("不支持的文件类型")
    if not is_text_editable_problem_file(sanitized_name):
        raise ValueError("该文件类型不支持在线编辑")

    absolute_path = problem_asset_absolute_path(problem_file.file_path)
    old_ext = os.path.splitext(problem_file.stored_filename or "")[1].lower()
    new_ext = os.path.splitext(sanitized_name)[1].lower()

    if new_ext != old_ext:
        new_stored_filename = secure_filename(f"{uuid.uuid4().hex}{new_ext}")
        new_absolute_path = os.path.join(os.path.dirname(absolute_path), new_stored_filename)
        os.replace(absolute_path, new_absolute_path)
        problem_file.stored_filename = new_stored_filename
        problem_file.file_path = "/".join(problem_file.file_path.split("/")[:-1] + [new_stored_filename])
        absolute_path = new_absolute_path

    with open(absolute_path, "w", encoding="utf-8") as file_handle:
        file_handle.write(content or "")

    problem_file.filename = sanitized_name
    problem_file.content_type = "text/plain"
    problem_file.file_size = os.path.getsize(absolute_path)


def read_problem_asset_text(problem_file):
    if not is_text_editable_problem_file(problem_file.filename):
        return ""
    absolute_path = problem_asset_absolute_path(problem_file.file_path)
    if not os.path.exists(absolute_path):
        return ""
    with open(absolute_path, "r", encoding="utf-8") as file_handle:
        return file_handle.read()


def make_unique_problem_filename(problem_id, filename, reserved=None):
    reserved = reserved or set()
    base_name = secure_filename(os.path.basename((filename or "").strip()))
    if not base_name:
        raise ValueError("文件名无效")

    stem, ext = os.path.splitext(base_name)
    candidate = base_name
    suffix = 2
    while candidate in reserved or ProblemFile.query.filter_by(problem_id=problem_id, filename=candidate).first():
        candidate = f"{stem}_{suffix}{ext}"
        suffix += 1
    reserved.add(candidate)
    return candidate


def create_problem_asset_from_bytes(problem_id, filename, content_bytes, content_type, uploaded_by_id):
    sanitized_name = secure_filename(os.path.basename((filename or "").strip()))
    if not sanitized_name:
        raise ValueError("文件名无效")
    if not is_allowed_problem_file(sanitized_name):
        raise ValueError("不支持的文件类型")

    ext = os.path.splitext(sanitized_name)[1].lower()
    stored_filename = secure_filename(f"{uuid.uuid4().hex}{ext}")
    relative_subdir = f"problem_assets/problem_{problem_id}"
    absolute_dir = os.path.join(app.root_path, "static", "uploads", relative_subdir)
    os.makedirs(absolute_dir, exist_ok=True)

    absolute_path = os.path.join(absolute_dir, stored_filename)
    with open(absolute_path, "wb") as file_handle:
        file_handle.write(content_bytes or b"")

    return ProblemFile(
        problem_id=problem_id,
        filename=sanitized_name,
        stored_filename=stored_filename,
        file_path=f"/static/uploads/{relative_subdir}/{stored_filename}",
        content_type=content_type or "",
        file_size=os.path.getsize(absolute_path),
        uploaded_by_id=uploaded_by_id,
    )


def decode_zip_text(content_bytes):
    try:
        return (content_bytes or b"").decode("utf-8-sig")
    except UnicodeDecodeError:
        return (content_bytes or b"").decode("utf-8", errors="replace")


def is_ajax_request():
    if request.headers.get("X-Requested-With") == "XMLHttpRequest":
        return True
    accepts = request.accept_mimetypes
    return accepts["application/json"] > accepts["text/html"]


def next_problem_case_number(problem_or_files):
    if isinstance(problem_or_files, Problem):
        problem_files = ProblemFile.query.filter_by(problem_id=problem_or_files.id).all()
    else:
        problem_files = problem_or_files
    max_index = 0
    for problem_file in problem_files:
        stem, ext = os.path.splitext(problem_file.filename or "")
        if ext.lower() not in {".in", ".out", ".ans"}:
            continue
        if stem.isdigit():
            max_index = max(max_index, int(stem))
    return max_index + 1 if max_index else 1


def rebuild_hidden_problem_testcases_from_files(problem):
    files = ProblemFile.query.filter_by(problem_id=problem.id).order_by(ProblemFile.filename.asc()).all()
    grouped_files = {}
    for problem_file in files:
        stem, ext = os.path.splitext(problem_file.filename or "")
        ext = ext.lower()
        if ext not in {".in", ".out"}:
            continue
        grouped_files.setdefault(stem, {})[ext] = problem_file

    paired_entries = []
    for stem, pair in grouped_files.items():
        if ".in" in pair and ".out" in pair:
            paired_entries.append((stem, pair[".in"], pair[".out"]))
    paired_entries.sort(key=lambda item: item[0])

    hidden_cases = [case for case in problem.test_cases if case.case_type == "hidden"]
    for case in hidden_cases:
        db.session.delete(case)
    db.session.flush()

    created_count = 0
    for sort_order, (_, input_file, output_file) in enumerate(paired_entries):
        input_absolute_path = os.path.join(app.root_path, input_file.file_path.lstrip("/"))
        output_absolute_path = os.path.join(app.root_path, output_file.file_path.lstrip("/"))
        if not os.path.exists(input_absolute_path) or not os.path.exists(output_absolute_path):
            continue

        with open(input_absolute_path, "r", encoding="utf-8") as input_handle:
            input_data = input_handle.read()
        with open(output_absolute_path, "r", encoding="utf-8") as output_handle:
            expected_output = output_handle.read()

        problem.test_cases.append(
            ProblemTestCase(
                case_type="hidden",
                input_data=input_data,
                expected_output=expected_output,
                score=1,
                sort_order=sort_order,
            )
        )
        created_count += 1

    return created_count


def build_problem_file_workspace_context(problem):
    problem_files = ProblemFile.query.filter_by(problem_id=problem.id).order_by(ProblemFile.uploaded_at.desc(), ProblemFile.id.desc()).all()
    data_files = [problem_file for problem_file in problem_files if problem_file.is_testdata_file]
    asset_files = [problem_file for problem_file in problem_files if not problem_file.is_testdata_file]
    file_contents = {
        problem_file.id: read_problem_asset_text(problem_file)
        for problem_file in problem_files
        if problem_file.is_text_editable
    }
    return {
        "problem_files": problem_files,
        "data_files": data_files,
        "asset_files": asset_files,
        "file_contents": file_contents,
        "next_case_number": next_problem_case_number(problem_files),
    }


def render_problem_file_workspace(problem):
    return render_template(
        "admin/_oj_problem_files_workspace.html",
        problem=problem,
        **build_problem_file_workspace_context(problem),
    )


def problem_file_workspace_response(problem, message, category="success", status_code=200):
    if is_ajax_request():
        return jsonify({
            "ok": category != "error",
            "message": message,
            "category": category,
            "workspace_html": render_problem_file_workspace(problem),
            "workspace": serialize_problem_file_workspace(problem),
        }), status_code
    flash(message, category)
    return redirect(url_for("manage_oj_problem_files", problem_id=problem.id))


def vite_asset(entry_name):
    manifest_path = os.path.join(app.static_folder, "oj-vue", ".vite", "manifest.json")
    try:
        with open(manifest_path, "r", encoding="utf-8") as manifest_file:
            manifest = json.load(manifest_file)
    except (FileNotFoundError, json.JSONDecodeError):
        return None

    entry = manifest.get(entry_name)
    if not entry:
        entry = next(
            (
                item
                for item in manifest.values()
                if item.get("name") == entry_name or item.get("src") == entry_name or item.get("src", "").endswith(entry_name)
            ),
            None,
        )
    if not entry:
        return None
    return url_for("static", filename=f"oj-vue/{entry['file']}")


def vite_css_assets(entry_name):
    manifest_path = os.path.join(app.static_folder, "oj-vue", ".vite", "manifest.json")
    try:
        with open(manifest_path, "r", encoding="utf-8") as manifest_file:
            manifest = json.load(manifest_file)
    except (FileNotFoundError, json.JSONDecodeError):
        return []

    entry = manifest.get(entry_name) or {}
    if not entry:
        entry = next(
            (
                item
                for item in manifest.values()
                if item.get("name") == entry_name or item.get("src") == entry_name or item.get("src", "").endswith(entry_name)
            ),
            {},
        )
    return [
        url_for("static", filename=f"oj-vue/{css_file}")
        for css_file in entry.get("css", [])
    ]


def serialize_admin_problem(problem):
    return {
        "id": problem.id,
        "uid": problem.uid_text,
        "code": problem.problem_code,
        "title": problem.title,
        "slug": problem.slug,
        "difficulty": problem.difficulty,
        "testcaseCount": problem.testcase_count,
        "fileCount": problem.files.count(),
        "allowedLanguages": problem.allowed_languages_text,
        "visible": bool(problem.is_visible),
        "updatedAt": problem.updated_at.strftime("%Y-%m-%d %H:%M") if problem.updated_at else "-",
        "urls": {
            "edit": url_for("edit_oj_problem", problem_id=problem.id),
            "files": url_for("manage_oj_problem_files", problem_id=problem.id),
            "delete": url_for("delete_oj_problem", problem_id=problem.id),
        },
    }


def serialize_problem_edit_workspace(problem):
    context = build_problem_file_workspace_context(problem)
    return {
        "id": problem.id,
        "uid": problem.uid_text,
        "code": problem.problem_code,
        "title": problem.title,
        "slug": problem.slug,
        "difficulty": problem.difficulty,
        "visible": bool(problem.is_visible),
        "timeLimitMs": problem.time_limit_ms,
        "memoryLimitMb": problem.memory_limit_mb,
        "source": problem.source or "",
        "allowedLanguages": problem.allowed_languages_text or "python, cpp, c",
        "statementMd": problem.statement_md or "",
        "updatedAt": problem.updated_at.strftime("%Y-%m-%d %H:%M") if problem.updated_at else "-",
        "stats": {
            "testcaseCount": problem.testcase_count,
            "dataFileCount": len(context["data_files"]),
            "assetFileCount": len(context["asset_files"]),
            "nextCaseNumber": context["next_case_number"],
        },
        "urls": {
            "edit": url_for("edit_oj_problem", problem_id=problem.id),
            "files": url_for("manage_oj_problem_files", problem_id=problem.id),
            "json": url_for("admin_oj_problem_json", problem_id=problem.id),
        },
    }


def serialize_problem_file(problem_file, content=None):
    problem_id = problem_file.problem_id
    return {
        "id": problem_file.id,
        "filename": problem_file.filename,
        "filePath": problem_file.file_path,
        "fileSizeKb": round((problem_file.file_size or 0) / 1024, 1),
        "uploadedAt": problem_file.uploaded_at.strftime("%Y-%m-%d %H:%M") if problem_file.uploaded_at else "",
        "isTextEditable": bool(problem_file.is_text_editable),
        "isTestdataFile": bool(problem_file.is_testdata_file),
        "markdownEmbed": problem_file.markdown_embed,
        "content": content if content is not None else "",
        "urls": {
            "update": url_for("update_text_oj_problem_file", problem_id=problem_id, file_id=problem_file.id),
            "delete": url_for("delete_oj_problem_file", problem_id=problem_id, file_id=problem_file.id),
        },
    }


def serialize_problem_file_workspace(problem):
    context = build_problem_file_workspace_context(problem)
    file_contents = context["file_contents"]
    data_files = [
        serialize_problem_file(problem_file, file_contents.get(problem_file.id, ""))
        for problem_file in context["data_files"]
    ]
    asset_files = [
        serialize_problem_file(problem_file, file_contents.get(problem_file.id, ""))
        for problem_file in context["asset_files"]
    ]
    return {
        "problem": {
            "id": problem.id,
            "uid": problem.uid_text,
            "code": problem.problem_code,
            "title": problem.title,
            "editUrl": url_for("edit_oj_problem", problem_id=problem.id),
            "filesUrl": url_for("manage_oj_problem_files", problem_id=problem.id),
            "listUrl": url_for("manage_oj_problems"),
        },
        "stats": {
            "testcaseCount": problem.testcase_count,
            "dataFileCount": len(data_files),
            "assetFileCount": len(asset_files),
            "nextCaseNumber": context["next_case_number"],
        },
        "dataFiles": data_files,
        "assetFiles": asset_files,
        "urls": {
            "createPair": url_for("create_text_oj_problem_file_pair", problem_id=problem.id),
            "createText": url_for("create_text_oj_problem_file", problem_id=problem.id),
            "upload": url_for("upload_oj_problem_files", problem_id=problem.id),
        },
    }


def oj_problem_source_tags(problem):
    if not problem.source:
        return []
    return [
        tag.strip()
        for tag in problem.source.replace("，", ",").split(",")
        if tag.strip()
    ]


def serialize_oj_problem_row(problem, submission_stat=None, my_status=None):
    submission_stat = submission_stat or {"accepted": 0, "attempts": 0}
    my_status = my_status or {"accepted": False, "attempted": False}
    author_name = "系统"
    if problem.created_by:
        author_name = problem.created_by.real_name or problem.created_by.username
    return {
        "id": problem.id,
        "uid": problem.uid_text,
        "code": problem.problem_code,
        "title": problem.title,
        "slug": problem.slug,
        "difficulty": problem.difficulty,
        "source": problem.source or "",
        "tags": oj_problem_source_tags(problem),
        "visible": bool(problem.is_visible),
        "author": author_name,
        "submissionStat": submission_stat,
        "myStatus": my_status,
        "urls": {
            "detail": url_for("oj_problem_detail", slug=problem.slug),
            "submissions": url_for("oj_submission_list", problem_id=problem.id),
        },
    }


def build_oj_problem_list_payload(q="", difficulty="", visibility="visible"):
    query = Problem.query
    if not current_user.is_admin:
        query = query.filter_by(is_visible=True)
    elif visibility == "hidden":
        query = query.filter_by(is_visible=False)
    elif visibility == "all":
        pass
    else:
        query = query.filter_by(is_visible=True)

    if q:
        like = f"%{q}%"
        query = query.filter(db.or_(Problem.problem_code.like(like), Problem.title.like(like), Problem.slug.like(like), Problem.source.like(like)))

    if difficulty in {"easy", "medium", "hard"}:
        query = query.filter_by(difficulty=difficulty)

    problems = query.order_by(Problem.problem_code.asc(), Problem.id.asc()).all()
    problem_ids = [problem.id for problem in problems]
    problem_submission_stats = {problem_id: {"accepted": 0, "attempts": 0} for problem_id in problem_ids}
    if problem_ids:
        rows = (
            db.session.query(
                JudgeTask.problem_id,
                db.func.count(JudgeTask.id).label("attempts"),
                db.func.sum(case((JudgeTask.status == "accepted", 1), else_=0)).label("accepted"),
            )
            .filter(JudgeTask.problem_id.in_(problem_ids), JudgeTask.assignment_id.is_(None))
            .group_by(JudgeTask.problem_id)
            .all()
        )
        for problem_id, attempts, accepted in rows:
            problem_submission_stats[problem_id] = {"accepted": int(accepted or 0), "attempts": int(attempts or 0)}

    my_problem_status = {problem_id: {"accepted": False, "attempted": False} for problem_id in problem_ids}
    if problem_ids:
        my_rows = (
            db.session.query(
                JudgeTask.problem_id,
                db.func.count(JudgeTask.id).label("attempts"),
                db.func.sum(case((JudgeTask.status == "accepted", 1), else_=0)).label("accepted"),
            )
            .filter(JudgeTask.problem_id.in_(problem_ids), JudgeTask.user_id == current_user.id, JudgeTask.assignment_id.is_(None))
            .group_by(JudgeTask.problem_id)
            .all()
        )
        for problem_id, attempts, accepted in my_rows:
            my_problem_status[problem_id] = {"accepted": int(accepted or 0) > 0, "attempted": int(attempts or 0) > 0}

    stats_query = Problem.query if current_user.is_admin else Problem.query.filter_by(is_visible=True)
    problem_stats = {
        "total": stats_query.count(),
        "easy": stats_query.filter_by(difficulty="easy").count(),
        "medium": stats_query.filter_by(difficulty="medium").count(),
        "hard": stats_query.filter_by(difficulty="hard").count(),
    }
    return {
        "filters": {"q": q, "difficulty": difficulty, "visibility": visibility},
        "stats": problem_stats,
        "problems": [
            serialize_oj_problem_row(problem, problem_submission_stats.get(problem.id), my_problem_status.get(problem.id))
            for problem in problems
        ],
        "count": len(problems),
    }


def serialize_oj_file(problem_file):
    return {
        "id": problem_file.id,
        "filename": problem_file.filename,
        "filePath": problem_file.file_path,
        "fileSizeKb": round((problem_file.file_size or 0) / 1024, 1),
    }


def serialize_oj_task_summary(task):
    if not task:
        return None
    meta = oj_submission_status_meta(task)
    return {
        "id": task.id,
        "status": task.status,
        "statusLabel": meta["label"],
        "statusTone": meta["tone"],
        "passedCount": task.passed_count,
        "totalCount": task.total_count,
        "totalScore": task.total_score,
        "createdAt": task.created_at.strftime("%Y-%m-%d %H:%M") if task.created_at else "-",
        "url": url_for("oj_submission_detail", task_id=task.id),
    }


def serialize_oj_problem_detail(problem, active_assignment=None):
    sample_cases = [
        {"input": case.input_data, "expectedOutput": case.expected_output, "score": case.score}
        for case in problem.sample_cases
    ]
    hidden_case_count = len(problem.hidden_cases)
    all_files = ProblemFile.query.filter_by(problem_id=problem.id).order_by(ProblemFile.filename.asc()).all()
    visible_files = all_files if current_user.is_admin else [problem_file for problem_file in all_files if not problem_file.is_testdata_file]
    assignment_id = active_assignment.id if active_assignment else None
    task_query = JudgeTask.query.filter_by(problem_id=problem.id, user_id=current_user.id, assignment_id=assignment_id)
    my_latest_task = task_query.order_by(JudgeTask.created_at.desc(), JudgeTask.id.desc()).first()
    submission_history = task_query.order_by(JudgeTask.created_at.desc(), JudgeTask.id.desc()).limit(5).all()
    statement_html, statement_has_sample_pairs = render_problem_statement(problem.statement_md or "")
    return {
        "id": problem.id,
        "uid": problem.uid_text,
        "code": problem.problem_code,
        "title": problem.title,
        "slug": problem.slug,
        "difficulty": problem.difficulty,
        "visible": bool(problem.is_visible),
        "timeLimitMs": problem.time_limit_ms,
        "memoryLimitMb": problem.memory_limit_mb,
        "source": problem.source or "",
        "allowedLanguages": problem.allowed_languages_text,
        "statementHtml": statement_html,
        "statementHasSamplePairs": statement_has_sample_pairs,
        "sampleCases": sample_cases,
        "hiddenCaseCount": hidden_case_count,
        "visibleFiles": [serialize_oj_file(problem_file) for problem_file in visible_files],
        "latestTask": serialize_oj_task_summary(my_latest_task),
        "submissionHistory": [serialize_oj_task_summary(task) for task in submission_history],
        "canManage": bool(current_user.is_admin),
        "activeAssignment": {"id": active_assignment.id, "title": active_assignment.title, "url": url_for("oj_assignment_detail", assignment_id=active_assignment.id)} if active_assignment else None,
        "urls": {
            "list": url_for("oj_problem_list"),
            "detail": url_for("oj_problem_detail", slug=problem.slug, assignment_id=assignment_id),
            "code": url_for("oj_problem_code", slug=problem.slug, assignment_id=assignment_id),
            "submit": url_for("submit_oj_problem", slug=problem.slug, assignment_id=assignment_id),
            "submissions": url_for("oj_submission_list", assignment_id=assignment_id),
            "adminEdit": url_for("edit_oj_problem", problem_id=problem.id),
        },
    }


def serialize_submission_row(task):
    meta = oj_submission_status_meta(task)
    return {
        "id": task.id,
        "language": task.language,
        "status": task.status,
        "statusLabel": meta["label"],
        "statusTone": meta["tone"],
        "totalScore": task.total_score,
        "createdAt": task.created_at.strftime("%Y-%m-%d %H:%M") if task.created_at else "-",
        "user": task.user.username if task.user else "未知用户",
        "problem": {
            "id": task.problem.id,
            "title": task.problem.title,
            "slug": task.problem.slug,
            "code": task.problem.problem_code,
            "url": url_for("oj_problem_detail", slug=task.problem.slug),
        } if task.problem else None,
        "url": url_for("oj_submission_detail", task_id=task.id),
    }


def build_oj_submission_list_payload(status_filter="", language_filter="", user_filter="", problem_filter="", problem_id_filter=None, assignment_id_filter=None):
    query = JudgeTask.query.filter(JudgeTask.problem_id.isnot(None))
    language_query = db.session.query(JudgeTask.language).filter(JudgeTask.problem_id.isnot(None))
    if not current_user.is_admin:
        query = query.filter_by(user_id=current_user.id)
        language_query = language_query.filter(JudgeTask.user_id == current_user.id)

    selected_assignment = None
    if assignment_id_filter:
        selected_assignment = OJAssignment.query.filter_by(id=assignment_id_filter).first()
        if selected_assignment and selected_assignment.is_visible_to(current_user):
            query = query.filter(JudgeTask.assignment_id == selected_assignment.id)
            language_query = language_query.filter(JudgeTask.assignment_id == selected_assignment.id)
        else:
            selected_assignment = None
            assignment_id_filter = None
    else:
        query = query.filter(JudgeTask.assignment_id.is_(None))
        language_query = language_query.filter(JudgeTask.assignment_id.is_(None))

    if status_filter in OJ_STATUS_META:
        query = query.filter_by(status=status_filter)
    if language_filter:
        query = query.filter_by(language=language_filter)
    if user_filter and current_user.is_admin:
        like = f"%{user_filter}%"
        user_clauses = [User.username.like(like), User.real_name.like(like)]
        if user_filter.isdigit():
            user_clauses.append(User.id == int(user_filter))
        query = query.filter(JudgeTask.user.has(db.or_(*user_clauses)))

    selected_problem = None
    if problem_id_filter:
        selected_problem = Problem.query.filter_by(id=problem_id_filter).first()
        problem_visible_in_assignment = bool(selected_assignment and any(link.problem_id == problem_id_filter for link in selected_assignment.problem_links))
        if selected_problem and (can_view_problem(selected_problem, current_user) or problem_visible_in_assignment):
            query = query.filter_by(problem_id=problem_id_filter)
        else:
            selected_problem = None
            problem_id_filter = None
    elif problem_filter:
        like = f"%{problem_filter}%"
        query = query.filter(JudgeTask.problem.has(db.or_(Problem.problem_code.like(like), Problem.title.like(like), Problem.slug.like(like))))

    available_languages = [
        language
        for (language,) in language_query.distinct().order_by(JudgeTask.language.asc()).all()
        if language
    ]
    submissions = query.order_by(JudgeTask.created_at.desc(), JudgeTask.id.desc()).limit(100).all()
    return {
        "submissions": [serialize_submission_row(task) for task in submissions],
        "availableLanguages": available_languages,
        "filters": {
            "status": status_filter,
            "language": language_filter,
            "user": user_filter if current_user.is_admin else "",
            "problem": problem_filter,
            "problemId": problem_id_filter,
            "assignmentId": assignment_id_filter,
        },
        "selectedProblem": {"id": selected_problem.id, "code": selected_problem.problem_code, "title": selected_problem.title} if selected_problem else None,
        "selectedAssignment": {"id": selected_assignment.id, "title": selected_assignment.title, "url": url_for("oj_assignment_detail", assignment_id=selected_assignment.id)} if selected_assignment else None,
        "isAdmin": bool(current_user.is_admin),
        "count": len(submissions),
    }


def serialize_submission_detail(task, results=None):
    results = results if results is not None else task.results.order_by(JudgeTaskResult.case_index.asc()).all()
    task_meta = oj_submission_status_meta(task, results)
    failure_feedback = build_oj_failure_feedback(task, results)
    can_view_details = bool(current_user.is_admin)
    return {
        "id": task.id,
        "status": task.status,
        "statusLabel": task_meta["label"],
        "statusTone": task_meta["tone"],
        "passedCount": task.passed_count,
        "totalCount": task.total_count,
        "totalScore": task.total_score,
        "language": task.language,
        "timeLimitMs": task.time_limit_ms,
        "memoryLimitMb": task.memory_limit_mb,
        "errorMessage": task.error_message,
        "sourceCode": task.source_code,
        "problem": {
            "title": task.problem.title,
            "slug": task.problem.slug,
            "url": url_for("oj_problem_detail", slug=task.problem.slug),
        } if task.problem else None,
        "results": [
            {
                "caseIndex": result.case_index,
                "status": result.status,
                "meta": oj_case_status_meta(result.status),
                "score": result.score,
                "timeMs": result.time_ms,
                "inputData": result.input_data if can_view_details else "",
                "expectedOutput": result.expected_output if can_view_details else "",
                "actualOutput": result.actual_output if can_view_details else "",
                "stderrText": result.stderr_text if can_view_details else "",
            }
            for result in results
        ],
        "failureFeedback": failure_feedback,
        "canViewCaseDetails": can_view_details,
        "urls": {
            "list": url_for("oj_submission_list"),
            "detail": url_for("oj_submission_detail", task_id=task.id),
            "detailJson": url_for("oj_submission_detail_json", task_id=task.id),
            "status": url_for("oj_submission_status_api", task_id=task.id),
        },
    }


def serialize_oj_assignment_task(task):
    if not task:
        return None
    meta = oj_task_display_meta(task)
    return {
        "id": task.id,
        "status": task.status,
        "statusLabel": meta["label"],
        "statusTone": meta.get("tone", oj_status_meta(task.status).get("tone", "neutral")),
        "statusBadge": meta.get("badge", "badge-ghost"),
        "totalScore": task.total_score,
        "createdAt": task.created_at.strftime("%Y-%m-%d %H:%M") if task.created_at else "-",
        "createdShort": task.created_at.strftime("%m-%d %H:%M") if task.created_at else "-",
        "problemCode": task.problem.problem_code if task.problem else "-",
        "url": url_for("oj_submission_detail", task_id=task.id),
    }


def serialize_oj_assignment_summary(assignment, latest_by_problem=None, accepted_problem_ids=None):
    latest_by_problem = latest_by_problem or {}
    accepted_problem_ids = accepted_problem_ids or set()
    last_task = None
    if latest_by_problem:
        last_task = max(
            latest_by_problem.values(),
            key=lambda item: ((item.created_at or datetime.min), item.id),
        )
    return {
        "id": assignment.id,
        "title": assignment.title,
        "targetText": assignment.target_text,
        "isActive": bool(assignment.is_active),
        "extensionDays": assignment.extension_days,
        "problemCount": assignment.problem_count,
        "acceptedCount": len(accepted_problem_ids),
        "attemptedCount": len(latest_by_problem),
        "startAt": assignment.start_at.strftime("%Y-%m-%d %H:%M") if assignment.start_at else "-",
        "endAt": assignment.end_at.strftime("%Y-%m-%d %H:%M") if assignment.end_at else "-",
        "endShort": assignment.end_at.strftime("%m-%d %H:%M") if assignment.end_at else "",
        "lastTask": serialize_oj_assignment_task(last_task),
        "urls": {
            "detail": url_for("oj_assignment_detail", assignment_id=assignment.id),
            "detailJson": url_for("oj_assignment_detail_json", assignment_id=assignment.id),
            "scoreboard": url_for("oj_assignment_scoreboard", assignment_id=assignment.id),
            "scoreboardJson": url_for("oj_assignment_scoreboard_json", assignment_id=assignment.id),
            "submissions": url_for("oj_submission_list", assignment_id=assignment.id),
            "edit": url_for("edit_oj_assignment", assignment_id=assignment.id) if current_user.is_admin else None,
        },
    }


def build_oj_assignment_list_payload():
    assignments = get_visible_oj_assignments_for_user(current_user)
    rows = []
    for assignment in assignments:
        latest_by_problem, accepted_problem_ids = assignment_problem_latest_status(assignment, current_user)
        rows.append(serialize_oj_assignment_summary(assignment, latest_by_problem, accepted_problem_ids))
    return {
        "assignments": rows,
        "count": len(rows),
        "isAdmin": bool(current_user.is_admin),
    }


def build_oj_assignment_detail_payload(assignment):
    latest_by_problem, accepted_problem_ids = assignment_problem_latest_status(assignment, current_user)
    detail_rows = []
    problem_ids = [link.problem_id for link in assignment.problem_links]
    total_submissions = (
        JudgeTask.query
        .filter(
            JudgeTask.user_id == current_user.id,
            JudgeTask.problem_id.in_(problem_ids),
        )
        .count()
    ) if problem_ids else 0
    last_submission = None

    for link in assignment.problem_links:
        task = latest_by_problem.get(link.problem_id)
        if task and (not last_submission or ((task.created_at or datetime.min), task.id) > ((last_submission.created_at or datetime.min), last_submission.id)):
            last_submission = task
        detail_rows.append({
            "problem": {
                "id": link.problem.id,
                "code": link.problem.problem_code,
                "title": link.problem.title,
                "slug": link.problem.slug,
                "url": url_for("oj_problem_detail", slug=link.problem.slug, assignment_id=assignment.id),
            },
            "task": serialize_oj_assignment_task(task),
            "accepted": link.problem_id in accepted_problem_ids,
        })

    assignment_submissions = (
        JudgeTask.query
        .filter(
            JudgeTask.user_id == current_user.id,
            JudgeTask.assignment_id == assignment.id,
            JudgeTask.problem_id.in_(problem_ids),
        )
        .order_by(JudgeTask.created_at.desc(), JudgeTask.id.desc())
        .limit(8)
        .all()
    ) if problem_ids else []

    description_html = markdown(assignment.description_md or "暂无作业介绍。")
    score_summary = {
        "acceptedCount": len(accepted_problem_ids),
        "attemptedCount": len(latest_by_problem),
        "problemCount": assignment.problem_count,
        "completionPercent": round((len(accepted_problem_ids) / assignment.problem_count) * 100, 1) if assignment.problem_count else 0,
        "totalSubmissions": total_submissions,
    }
    return {
        "assignment": serialize_oj_assignment_summary(assignment, latest_by_problem, accepted_problem_ids),
        "descriptionHtml": description_html,
        "rows": detail_rows,
        "recentSubmissions": [serialize_oj_assignment_task(task) for task in assignment_submissions],
        "scoreSummary": score_summary,
        "lastSubmission": serialize_oj_assignment_task(last_submission),
        "isAdmin": bool(current_user.is_admin),
    }


def build_oj_assignment_scoreboard_payload(assignment):
    participants = get_oj_assignment_participants(assignment)
    problem_links = list(assignment.problem_links)
    problem_ids = [link.problem_id for link in problem_links]
    participant_ids = [user.id for user in participants]

    tasks = []
    if problem_ids and participant_ids:
        tasks = (
            JudgeTask.query
            .filter(
                JudgeTask.assignment_id == assignment.id,
                JudgeTask.user_id.in_(participant_ids),
                JudgeTask.problem_id.in_(problem_ids),
            )
            .order_by(JudgeTask.user_id.asc(), JudgeTask.problem_id.asc(), JudgeTask.created_at.desc(), JudgeTask.id.desc())
            .all()
        )

    latest_by_pair = {}
    accepted_pairs = set()
    for task in tasks:
        if task.status == 'accepted':
            accepted_pairs.add((task.user_id, task.problem_id))
        latest_by_pair.setdefault((task.user_id, task.problem_id), task)

    rows = []
    for user in participants:
        accepted_count = 0
        submission_count = 0
        cells = []
        for link in problem_links:
            pair_task = latest_by_pair.get((user.id, link.problem_id))
            if pair_task:
                submission_count += 1
            accepted = (user.id, link.problem_id) in accepted_pairs
            if accepted:
                accepted_count += 1
            cells.append({
                "problem": {
                    "id": link.problem.id,
                    "code": link.problem.problem_code,
                    "title": link.problem.title,
                    "slug": link.problem.slug,
                    "url": url_for("oj_problem_detail", slug=link.problem.slug, assignment_id=assignment.id),
                },
                "task": serialize_oj_assignment_task(pair_task),
                "accepted": accepted,
            })

        rows.append({
            "user": {
                "id": user.id,
                "name": user.real_name or user.username,
                "username": user.username,
                "className": user.student_class.name if user.student_class else "未分班",
            },
            "acceptedCount": accepted_count,
            "submissionCount": submission_count,
            "score": accepted_count * 100,
            "cells": cells,
        })

    rows.sort(
        key=lambda item: (
            -item["acceptedCount"],
            -item["score"],
            -item["submissionCount"],
            item["user"]["name"],
        )
    )

    last_score = None
    current_rank = 0
    for index, row in enumerate(rows, start=1):
        score_key = (row["acceptedCount"], row["score"], row["submissionCount"])
        if score_key != last_score:
            current_rank = index
            last_score = score_key
        row["rank"] = current_rank

    return {
        "assignment": serialize_oj_assignment_summary(assignment),
        "problemLinks": [
            {
                "id": link.id,
                "problem": {
                    "id": link.problem.id,
                    "code": link.problem.problem_code,
                    "title": link.problem.title,
                    "slug": link.problem.slug,
                    "url": url_for("oj_problem_detail", slug=link.problem.slug, assignment_id=assignment.id),
                },
            }
            for link in problem_links
        ],
        "rows": rows,
        "stats": {
            "participantCount": len(rows),
            "problemCount": len(problem_links),
            "startAt": assignment.start_at.strftime("%m-%d %H:%M") if assignment.start_at else "-",
            "endAt": assignment.end_at.strftime("%m-%d %H:%M") if assignment.end_at else "-",
        },
        "isAdmin": bool(current_user.is_admin),
    }


app.jinja_env.globals["vite_asset"] = vite_asset
app.jinja_env.globals["vite_css_assets"] = vite_css_assets

app.jinja_env.filters["is_image_icon"] = is_image_icon
app.jinja_env.filters["badge_icon_url"] = badge_icon_url

followers = db.Table('followers',
    db.Column('follower_id', db.Integer, db.ForeignKey('user.id')),
    db.Column('followed_id', db.Integer, db.ForeignKey('user.id'))
)

user_groups = db.Table('user_groups',
    db.Column('user_id', db.Integer, db.ForeignKey('user.id'), primary_key=True),
    db.Column('group_id', db.Integer, db.ForeignKey('group.id'), primary_key=True)
)

class Class(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), unique=True, nullable=False)
    users = db.relationship('User', backref='student_class', lazy='dynamic')

class Group(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), unique=True, nullable=False)
    description = db.Column(db.String(255))
    users = db.relationship('User', secondary=user_groups, lazy='subquery',
        backref=db.backref('groups', lazy=True))

class User(db.Model, UserMixin):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    email = db.Column(db.String(120), unique=True, nullable=False)
    password_hash = db.Column(db.String(256), nullable=False)
    is_admin = db.Column(db.Boolean, default=False)
    created_at = db.Column(db.DateTime, default=now_utc8)
    
    # New fields for student information
    real_name = db.Column(db.String(80))
    student_id = db.Column(db.String(20), unique=True)
    department = db.Column(db.String(100))
    class_name = db.Column(db.String(100))
    class_id = db.Column(db.Integer, db.ForeignKey('class.id'))
    
    # Badge System
    selected_badge_id = db.Column(db.Integer, db.ForeignKey("badge.id"), nullable=True)
    selected_badge = db.relationship("Badge", foreign_keys=[selected_badge_id])
    
    subscriptions = db.relationship("Subscription", backref="user", cascade="all, delete-orphan")
    editor_roles = db.relationship("WikiEditor", backref="user", cascade="all, delete-orphan")
    notes = db.relationship("Note", backref="user", cascade="all, delete-orphan")
    shared_notes = db.relationship("NoteShare", backref="user", cascade="all, delete-orphan")
    earned_badges = db.relationship("UserBadge", backref="user", cascade="all, delete-orphan")

    # bookmarks 书签
    bookmarks = db.relationship('UserBookmark', backref='user', lazy='dynamic', cascade="all, delete-orphan")


    followed = db.relationship(
        'User', secondary=followers,
        primaryjoin=(followers.c.follower_id == id),
        secondaryjoin=(followers.c.followed_id == id),
        backref=db.backref('followers', lazy='dynamic'),
        lazy='dynamic'
    )

    def follow(self, user):
        if not self.is_following(user):
            self.followed.append(user)

    def unfollow(self, user):
        if self.is_following(user):
            self.followed.remove(user)

    def is_following(self, user):
        return self.followed.filter(followers.c.followed_id == user.id).count() > 0

    def is_followed_by(self, user):
        return self.followers.filter(followers.c.follower_id == user.id).count() > 0

    @property
    def is_online(self):
        if not self.active_status or not self.active_status.last_active_at:
            return False
        return to_utc8_naive(self.active_status.last_active_at) > (now_utc8() - timedelta(minutes=5))

    @property
    def last_active_human(self):
        if not self.active_status or not self.active_status.last_active_at:
            return "从未上线"
        seconds = datetime_diff_seconds(now_utc8(), self.active_status.last_active_at)
        if seconds < 60:
            return "刚刚"
        minutes = seconds / 60
        if minutes < 60:
            return f"{int(minutes)}分钟前"
        hours = minutes / 60
        if hours < 24:
            return f"{int(hours)}小时前"
        days = hours / 24
        if days < 7:
            return f"{int(days)}天前"
        if days < 30:
            return f"{int(days/7)}周前"
        if days < 365:
            return f"{int(days/30)}个月前"
        return f"{int(days/365)}年前"

    @property
    def study_time_today_human(self):
        now = now_utc8()
        start_of_day = now.replace(hour=0, minute=0, second=0, microsecond=0)
        sessions = StudySession.query.filter(
            StudySession.user_id == self.id,
            StudySession.start_time >= start_of_day
        ).all()
        total_seconds = sum(study_session_seconds(s) for s in sessions)
        
        if total_seconds < 60:
            return "少于1分钟"
        m, s = divmod(total_seconds, 60)
        h, m = divmod(m, 60)
        if h > 0:
            return f"{int(h)}小时 {int(m)}分钟"
        else:
            return f"{int(m)}分钟"

    @property
    def latest_earned_badge(self):
        return UserBadge.query.filter_by(user_id=self.id).order_by(UserBadge.earned_at.desc()).first()

    @property
    def badges(self):
        return self.earned_badges

    @property
    def wikis(self):
        return Wiki.query.filter_by(created_by_id=self.id).all()

    @property
    def study_partners(self):
        """Return a query for users that are mutually following each other."""
        # Users I follow who also follow me
        # Subquery: IDs of people who follow me
        subquery = db.session.query(followers.c.follower_id).filter(followers.c.followed_id == self.id)
        # Filter my followed list by that subquery
        return self.followed.filter(User.id.in_(subquery))

    def set_password(self, password):
        self.password_hash = generate_password_hash(password)

    def check_password(self, password):
        return check_password_hash(self.password_hash, password)

    def get_selected_user_badge(self):
        if not self.selected_badge_id:
            return None
        return UserBadge.query.filter_by(user_id=self.id, badge_id=self.selected_badge_id).first()

    def calculate_stats(self):
        """Calculate all detailed statistics for the user."""
        stats = {
            "wiki_count": Wiki.query.filter_by(created_by_id=self.id).count(),
            "note_count": Note.query.filter_by(user_id=self.id).count(),
            "badge_count": len(self.earned_badges),
            "study_partner_count": self.study_partners.count(),
            "following_count": self.followed.count(),
            "follower_count": self.followers.count(),
            "comment_count": Comment.query.filter_by(user_id=self.id).count(),
            "wiki_edit_count": WikiPageHistory.query.filter_by(user_id=self.id).count(),
            "streak_days": 0,
            "study_hours": 0.0,
            "night_owl_days": 0,
            "early_bird_days": 0,
            "weekend_study_hours": 0.0,
            "long_sessions": 0,
            "total_views_received": NoteViewLog.query.join(Note).filter(Note.user_id == self.id).count(),
            "share_count": NoteShare.query.join(Note).filter(Note.user_id == self.id).count()
        }

        # Study Session Analysis
        sessions = StudySession.query.filter_by(user_id=self.id).all()
        
        # 1. Study Hours
        total_seconds = sum(study_session_seconds(s) for s in sessions)
        stats["study_hours"] = round(total_seconds / 3600, 1)

        # 2. Streak
        sorted_sessions = sorted(sessions, key=lambda x: to_utc8_naive(x.start_time) or datetime.min, reverse=True)
        dates = sorted(list(set([to_utc8_naive(s.start_time).date() for s in sorted_sessions if to_utc8_naive(s.start_time)])), reverse=True)
        streak = 0
        if dates:
            today = now_utc8().date()
            if dates[0] == today:
                streak = 1
                current = today
            elif dates[0] == today - timedelta(days=1):
                streak = 1
                current = today - timedelta(days=1)
            else:
                streak = 0
            if streak > 0:
                for i in range(1, len(dates)):
                    if dates[i] == current - timedelta(days=1):
                        streak += 1
                        current = dates[i]
                    else:
                        break
        stats["streak_days"] = streak

        # 3. Time Patterns
        night_days = set()
        early_days = set()
        weekend_seconds = 0
        long_count = 0

        for s in sessions:
            start_time = to_utc8_naive(s.start_time)
            if not start_time:
                continue

            # Night Owl
            h = start_time.hour
            if h >= 23 or h < 4:
                logical_date = (start_time - timedelta(hours=4)).date()
                night_days.add(logical_date)
            
            # Early Bird
            if 5 <= h < 8:
                early_days.add(start_time.date())
            
            # Weekend
            if start_time.weekday() in [5, 6]:
                duration = study_session_seconds(s)
                if duration > 0:
                    weekend_seconds += duration
            
            # Long Session
            duration = study_session_seconds(s)
            if duration >= 7200:
                long_count += 1
        
        stats["night_owl_days"] = len(night_days)
        stats["early_bird_days"] = len(early_days)
        stats["weekend_study_hours"] = round(weekend_seconds / 3600, 1)
        stats["long_sessions"] = long_count

        return stats

    @property
    def total_exp(self):
        """计算用户总源力值 (EXP)"""
        stats = self.calculate_stats()
        # 基于你之前的逻辑：每分钟1分 + 每番茄5分 + 每笔记15分 + 每次Wiki贡献10分 + 每次评论2分
        # 我们这里直接复用计算逻辑
        duration_mins = int(stats.get("study_hours", 0) * 60)
        pomos = PomodoroRecord.query.filter_by(user_id=self.id).count()
        notes = stats.get("note_count", 0)
        wiki_edits = stats.get("wiki_edit_count", 0)
        comments = stats.get("comment_count", 0)
        return duration_mins + (pomos * 5) + (notes * 15) + (wiki_edits * 10) + (comments * 2)

    @property
    def level(self):
        """根据 EXP 计算等级：等级 = sqrt(EXP / 10) 的向下取整"""
        import math
        return int(math.sqrt(self.total_exp / 10)) if self.total_exp > 0 else 0

class Notification(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    message = db.Column(db.String(500), nullable=False)
    is_read = db.Column(db.Boolean, default=False)
    created_at = db.Column(db.DateTime, default=now_utc8)
    type = db.Column(db.String(50), default='system') # system, friend_login, friend_achievement
    link = db.Column(db.String(200), nullable=True)

    user = db.relationship("User", backref=db.backref("notifications", lazy="dynamic", cascade="all, delete-orphan"))

class SystemAnnouncement(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    title = db.Column(db.String(200), nullable=False)
    content = db.Column(db.Text, nullable=False) # Markdown
    created_by_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    created_at = db.Column(db.DateTime, default=now_utc8)
    updated_at = db.Column(db.DateTime, default=now_utc8, onupdate=now_utc8)
    is_active = db.Column(db.Boolean, default=True)

    creator = db.relationship("User", backref="created_announcements")

class UserAnnouncementConfirmation(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    announcement_id = db.Column(db.Integer, db.ForeignKey('system_announcement.id'), nullable=False)
    confirmed_at = db.Column(db.DateTime, default=now_utc8)
    
    user = db.relationship("User", backref="announcement_confirmations")
    announcement = db.relationship("SystemAnnouncement", backref="confirmations")

class AnnouncementFile(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    filename = db.Column(db.String(255), nullable=False)
    file_path = db.Column(db.String(500), nullable=False)
    uploaded_by_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=False)
    uploaded_at = db.Column(db.DateTime, default=now_utc8)
    
    uploader = db.relationship("User", backref="uploaded_announcement_files")

class Badge(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    description = db.Column(db.String(255), nullable=False)
    icon = db.Column(db.String(255), nullable=False) # Emoji or FontAwesome class
    condition_type = db.Column(db.String(50), nullable=False) # streak_days, study_hours, featured_count, note_count
    condition_value = db.Column(db.Integer, nullable=False)
    is_hidden = db.Column(db.Boolean, default=False) # Hidden badges won't show conditions until earned
    sticker_count = db.Column(db.Integer, default=1) # Number of stickers this badge provides
    category = db.Column(db.String(50), default="一般") # Achievement category
    rarity = db.Column(db.String(20), default="common") # common, rare, epic, legendary
    is_secret = db.Column(db.Boolean, default=False) # Secret achievements (??? when locked)
    custom_condition_text = db.Column(db.String(255), nullable=True) # Custom description for condition
    total_limit = db.Column(db.Integer, nullable=True) # Max number of users who can earn this badge (Null = infinite)
    issued_count = db.Column(db.Integer, default=0) # Current number of users who earned this
    start_time = db.Column(db.DateTime, nullable=True) # For time-limited badges
    end_time = db.Column(db.DateTime, nullable=True) # For time-limited badges
    created_at = db.Column(db.DateTime, default=now_utc8)
    issuer = db.Column(db.String(100), default="WikiBook") # Issuer of the badge
    serial_prefix = db.Column(db.String(20), default="SF")
    
    users = db.relationship("UserBadge", backref="badge_info", lazy="dynamic")
    target_id = db.Column(db.Integer, nullable=True) # 💡 新增：用于存储具体的 Wiki_id 或 Assignment_id

class UserBadge(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=False)
    badge_id = db.Column(db.Integer, db.ForeignKey("badge.id"), nullable=False)
    earned_at = db.Column(db.DateTime, default=now_utc8)
    serial_number = db.Column(db.String(50), unique=True, nullable=True) # Unique Serial Number
    is_notified = db.Column(db.Boolean, default=False)
    badge = db.relationship("Badge", overlaps="badge_info,users")

    def __init__(self, **kwargs):
        super(UserBadge, self).__init__(**kwargs)


OJ_BADGE_CONDITION_LABELS = {
    'oj_ac_count': 'OJ AC题目数',
    'oj_submission_count': 'OJ提交次数',
    'oj_error_count': 'OJ错误次数',
    'oj_assignment_complete_count': 'OJ作业全完成次数',
    'oj_consecutive_ac_count': 'OJ连续AC次数',
    'oj_consecutive_error_count': 'OJ连续错误次数',
    'oj_first_try_ac_count': 'OJ一发AC题目数',
    'oj_daily_ac_peak': 'OJ单日AC峰值',
    'oj_specific_assignment_complete': '完成指定OJ作业',
}

OJ_FINAL_STATUSES = {'accepted', 'failed', 'system_error'}
OJ_ERROR_STATUSES = {'failed', 'system_error'}
OJ_SPECIFIC_ASSIGNMENT_CONDITIONS = {'oj_specific_assignment_complete'}


def _base_oj_task_query(user_id, final_only=False):
    query = JudgeTask.query.filter(
        JudgeTask.user_id == user_id,
        JudgeTask.problem_id.isnot(None),
    )
    if final_only:
        query = query.filter(JudgeTask.status.in_(OJ_FINAL_STATUSES))
    return query


def _count_current_oj_streak(user_id, accepted=True):
    streak = 0
    tasks = (
        _base_oj_task_query(user_id, final_only=True)
        .order_by(JudgeTask.created_at.desc(), JudgeTask.id.desc())
        .all()
    )
    for task in tasks:
        if accepted and task.status == 'accepted':
            streak += 1
            continue
        if not accepted and task.status in OJ_ERROR_STATUSES:
            streak += 1
            continue
        break
    return streak


def _count_oj_completed_assignments(user_id):
    completed_count = 0
    for assignment in OJAssignment.query.all():
        problem_ids = [link.problem_id for link in assignment.problem_links]
        if not problem_ids:
            continue
        accepted_problem_ids = {
            row[0] for row in (
                JudgeTask.query
                .with_entities(JudgeTask.problem_id)
                .filter(
                    JudgeTask.user_id == user_id,
                    JudgeTask.assignment_id == assignment.id,
                    JudgeTask.problem_id.in_(problem_ids),
                    JudgeTask.status == 'accepted',
                )
                .distinct()
                .all()
            )
        }
        if set(problem_ids).issubset(accepted_problem_ids):
            completed_count += 1
    return completed_count


def _is_oj_assignment_completed(user_id, assignment_id):
    assignment = OJAssignment.query.get(assignment_id)
    if not assignment:
        return False
    problem_ids = [link.problem_id for link in assignment.problem_links]
    if not problem_ids:
        return False
    accepted_problem_ids = {
        row[0] for row in (
            JudgeTask.query
            .with_entities(JudgeTask.problem_id)
            .filter(
                JudgeTask.user_id == user_id,
                JudgeTask.assignment_id == assignment.id,
                JudgeTask.problem_id.in_(problem_ids),
                JudgeTask.status == 'accepted',
            )
            .distinct()
            .all()
        )
    }
    return set(problem_ids).issubset(accepted_problem_ids)


def _count_first_try_oj_accepts(user_id):
    first_tasks_by_problem = {}
    tasks = (
        _base_oj_task_query(user_id, final_only=True)
        .order_by(JudgeTask.problem_id.asc(), JudgeTask.created_at.asc(), JudgeTask.id.asc())
        .all()
    )
    for task in tasks:
        first_tasks_by_problem.setdefault(task.problem_id, task)
    return sum(1 for task in first_tasks_by_problem.values() if task.status == 'accepted')


def _count_daily_oj_ac_peak(user_id):
    accepted_tasks = (
        _base_oj_task_query(user_id)
        .filter(JudgeTask.status == 'accepted')
        .order_by(JudgeTask.created_at.asc(), JudgeTask.id.asc())
        .all()
    )
    ac_by_day = {}
    for task in accepted_tasks:
        if not task.created_at:
            continue
        day = task.created_at.date()
        ac_by_day.setdefault(day, set()).add(task.problem_id)
    if not ac_by_day:
        return 0
    return max(len(problem_ids) for problem_ids in ac_by_day.values())


def get_oj_badge_progress(user, badge):
    if badge.condition_type == 'oj_ac_count':
        return (
            _base_oj_task_query(user.id)
            .filter(JudgeTask.status == 'accepted')
            .with_entities(JudgeTask.problem_id)
            .distinct()
            .count()
        )
    if badge.condition_type == 'oj_submission_count':
        return _base_oj_task_query(user.id).count()
    if badge.condition_type == 'oj_error_count':
        return _base_oj_task_query(user.id).filter(JudgeTask.status.in_(OJ_ERROR_STATUSES)).count()
    if badge.condition_type == 'oj_assignment_complete_count':
        return _count_oj_completed_assignments(user.id)
    if badge.condition_type == 'oj_consecutive_ac_count':
        return _count_current_oj_streak(user.id, accepted=True)
    if badge.condition_type == 'oj_consecutive_error_count':
        return _count_current_oj_streak(user.id, accepted=False)
    if badge.condition_type == 'oj_first_try_ac_count':
        return _count_first_try_oj_accepts(user.id)
    if badge.condition_type == 'oj_daily_ac_peak':
        return _count_daily_oj_ac_peak(user.id)
    if badge.condition_type == 'oj_specific_assignment_complete':
        return 1 if badge.target_id and _is_oj_assignment_completed(user.id, badge.target_id) else 0
    return None


def meets_oj_badge_condition(user, badge):
    progress = get_oj_badge_progress(user, badge)
    if progress is None:
        return False
    target = 1 if badge.condition_type in OJ_SPECIFIC_ASSIGNMENT_CONDITIONS else badge.condition_value
    return progress >= target

# Wiki Permissions Associations
wiki_classes = db.Table('wiki_classes',
    db.Column('wiki_id', db.Integer, db.ForeignKey('wiki.id'), primary_key=True),
    db.Column('class_id', db.Integer, db.ForeignKey('class.id'), primary_key=True)
)

wiki_groups = db.Table('wiki_groups',
    db.Column('wiki_id', db.Integer, db.ForeignKey('wiki.id'), primary_key=True),
    db.Column('group_id', db.Integer, db.ForeignKey('group.id'), primary_key=True)
)

wiki_users = db.Table('wiki_users',
    db.Column('wiki_id', db.Integer, db.ForeignKey('wiki.id'), primary_key=True),
    db.Column('user_id', db.Integer, db.ForeignKey('user.id'), primary_key=True)
)

class Wiki(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    title = db.Column(db.String(200), nullable=False)
    description = db.Column(db.Text, default="")
    created_by_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=False)
    created_at = db.Column(db.DateTime, default=now_utc8)
    
    # 用于首页3D显示
    is_featured = db.Column(db.Boolean, default=False)
    featured_order = db.Column(db.Integer, default=0)


    # Appearance Settings
    bg_color = db.Column(db.String(20), default="#ffffff")
    fg_color = db.Column(db.String(20), default="#1f2937")
    pattern = db.Column(db.String(50), default="pattern-none")
    icon_url = db.Column(db.String(255), nullable=True)
    icon_scale = db.Column(db.Float, default=1.0)
    icon_position_x = db.Column(db.Float, default=50.0)
    icon_position_y = db.Column(db.Float, default=50.0)
    
    # Advanced Appearance (Arc Style)
    blur_level = db.Column(db.Integer, default=60) # px
    gradient_bg = db.Column(db.String(50), default="radial") # radial, linear, conical
    gradient_color = db.Column(db.String(20), default="") # Second color for gradient
    gradient_direction = db.Column(db.String(50), default="to bottom right") # Direction of gradient
    
    # Permissions
    visibility = db.Column(db.String(20), default="public") # public, private, restricted
    
    allowed_classes = db.relationship('Class', secondary=wiki_classes, lazy='subquery',
        backref=db.backref('wikis', lazy=True))
    allowed_groups = db.relationship('Group', secondary=wiki_groups, lazy='subquery',
        backref=db.backref('wikis', lazy=True))
    allowed_users = db.relationship('User', secondary=wiki_users, lazy='subquery',
        backref=db.backref('accessible_wikis', lazy=True))
    
    created_by = db.relationship("User", foreign_keys=[created_by_id])
    pages = db.relationship("WikiPage", backref="wiki", cascade="all, delete-orphan")
    subscriptions = db.relationship("Subscription", backref="wiki", cascade="all, delete-orphan")
    editors = db.relationship("WikiEditor", backref="wiki", cascade="all, delete-orphan")

    def is_visible_to(self, user):
        if self.visibility == 'public':
            return True
        if not user.is_authenticated:
            return False
        if user.is_admin:
            return True
        if user.id == self.created_by_id:
            return True
        
        # Check editors
        if WikiEditor.query.filter_by(wiki_id=self.id, user_id=user.id).first():
            return True
            
        if self.visibility == 'private':
            return False
            
        # restricted
        # Check users
        if user in self.allowed_users:
            return True
        # Check classes
        if user.student_class and user.student_class in self.allowed_classes:
            return True
        # Check groups
        # Intersection of user groups and allowed groups
        user_groups_set = set(user.groups)
        allowed_groups_set = set(self.allowed_groups)
        if not user_groups_set.isdisjoint(allowed_groups_set):
            return True
            
        return False

class WikiPage(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    wiki_id = db.Column(db.Integer, db.ForeignKey("wiki.id"), nullable=False)
    title = db.Column(db.String(200), nullable=False)
    slug = db.Column(db.String(200), nullable=False) # 实际上现在存储的是整数
    content_md = db.Column(db.Text, default="")
    view_count = db.Column(db.Integer, default=0)
    updated_at = db.Column(db.DateTime, default=now_utc8, onupdate=now_utc8)
    
    comment_enabled = db.Column(db.Boolean, default=True)

    tags = db.relationship('Tag', secondary='wiki_page_tags', backref=db.backref('pages', lazy='dynamic'))
    comments = db.relationship('Comment', backref='wiki_page', cascade="all, delete-orphan", order_by="desc(Comment.created_at)")

    folder_id = db.Column(db.Integer, db.ForeignKey('wiki_folder.id'), nullable=True)
    order_weight = db.Column(db.Integer, default=0) # 用于排序

class UserActiveStatus(db.Model):
    user_id = db.Column(db.Integer, db.ForeignKey("user.id"), primary_key=True)
    last_active_at = db.Column(db.DateTime, default=now_utc8)
    current_path = db.Column(db.String(500), default="/")
    current_action = db.Column(db.String(200), default="")
    # 💡 必须在这里定义字段，业务逻辑才能读写它们
    pomo_state = db.Column(db.String(20), default="IDLE") # IDLE, WORK, REST
    pomo_end_time = db.Column(db.Float, default=0.0)      # 存储毫秒级时间戳

    # 💡 新增：用于记录学习报告的最后查阅日期
    last_daily_report = db.Column(db.Date, nullable=True)
    last_weekly_report = db.Column(db.Date, nullable=True)
    last_monthly_report = db.Column(db.Date, nullable=True)
    
    user = db.relationship("User", backref=db.backref("active_status", uselist=False))

class StudySession(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=False)
    start_time = db.Column(db.DateTime, default=now_utc8)
    end_time = db.Column(db.DateTime, default=now_utc8)
    
    user = db.relationship("User", backref="study_sessions")

class WikiViewLog(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    wiki_id = db.Column(db.Integer, db.ForeignKey("wiki.id"), nullable=False)
    user_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=True) # 可以记录匿名访问，如果需要
    timestamp = db.Column(db.DateTime, default=now_utc8)

class WikiPageHistory(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    page_id = db.Column(db.Integer, db.ForeignKey("wiki_page.id"), nullable=False)
    user_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=False)
    title = db.Column(db.String(200), nullable=False)
    slug = db.Column(db.String(200), nullable=False)
    content_md = db.Column(db.Text, default="")
    created_at = db.Column(db.DateTime, default=now_utc8)
    
    user = db.relationship("User", backref="page_history")
    page = db.relationship("WikiPage", backref=db.backref("history", cascade="all, delete-orphan", order_by="desc(WikiPageHistory.created_at)"))

class WikiEditor(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    wiki_id = db.Column(db.Integer, db.ForeignKey("wiki.id"), nullable=False)
    user_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=False)

class WikiFile(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    wiki_id = db.Column(db.Integer, db.ForeignKey("wiki.id"), nullable=False)
    filename = db.Column(db.String(255), nullable=False)
    file_path = db.Column(db.String(500), nullable=False)
    uploaded_by_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=False)
    uploaded_at = db.Column(db.DateTime, default=now_utc8)
    
    wiki = db.relationship("Wiki", backref=db.backref("files", cascade="all, delete-orphan"))
    uploader = db.relationship("User", backref="uploaded_files")

class Subscription(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    wiki_id = db.Column(db.Integer, db.ForeignKey("wiki.id"), nullable=False)
    user_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=False)
    created_at = db.Column(db.DateTime, default=now_utc8)

class MarkdownStyle(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    css_text = db.Column(db.Text, default="")
    updated_at = db.Column(db.DateTime, default=now_utc8, onupdate=now_utc8)

class Note(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    title = db.Column(db.String(200), nullable=False)
    content_md = db.Column(db.Text, default="")
    
    # 👇 新增这两个字段
    is_scrapbook = db.Column(db.Boolean, default=False) # 区分是传统 MD 笔记还是手账
    scrapbook_data = db.Column(db.JSON, nullable=True)  # 存储手账画布的所有坐标与内容
    
    user_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=False)
    is_featured = db.Column(db.Boolean, default=False)
    created_at = db.Column(db.DateTime, default=now_utc8)
    updated_at = db.Column(db.DateTime, default=now_utc8, onupdate=now_utc8)
    icon = db.Column(db.String(255), nullable=True) # Emoji or Image URL
    
    comment_enabled = db.Column(db.Boolean, default=True)

    shares = db.relationship("NoteShare", backref="note", cascade="all, delete-orphan")
    
    tags = db.relationship('Tag', secondary='note_tags', backref=db.backref('notes', lazy='dynamic'))
    comments = db.relationship('Comment', backref='note', cascade="all, delete-orphan", order_by="desc(Comment.created_at)")

    @property
    def view_count(self):
        return NoteViewLog.query.filter_by(note_id=self.id).count()

class NoteShare(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    note_id = db.Column(db.Integer, db.ForeignKey("note.id"), nullable=False)
    user_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=False) # Receiver
    created_at = db.Column(db.DateTime, default=now_utc8)

class Tag(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(50), unique=True, nullable=False)

# Association tables for Tags
note_tags = db.Table('note_tags',
    db.Column('note_id', db.Integer, db.ForeignKey('note.id'), primary_key=True),
    db.Column('tag_id', db.Integer, db.ForeignKey('tag.id'), primary_key=True)
)

wiki_page_tags = db.Table('wiki_page_tags',
    db.Column('page_id', db.Integer, db.ForeignKey('wiki_page.id'), primary_key=True),
    db.Column('tag_id', db.Integer, db.ForeignKey('tag.id'), primary_key=True)
)

class Comment(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=False)
    content = db.Column(db.Text, nullable=False)
    created_at = db.Column(db.DateTime, default=now_utc8)
    is_approved = db.Column(db.Boolean, default=False, nullable=False)
    approved_at = db.Column(db.DateTime, nullable=True)
    approved_by_id = db.Column(db.Integer, nullable=True)
    
    # Polymorphic-like association (nullable FKs)
    note_id = db.Column(db.Integer, db.ForeignKey("note.id"), nullable=True)
    wiki_page_id = db.Column(db.Integer, db.ForeignKey("wiki_page.id"), nullable=True)
    
    user = db.relationship("User", backref="comments")
    # Relationships will be defined in Note and WikiPage models for backrefs

class NoteViewLog(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    note_id = db.Column(db.Integer, db.ForeignKey("note.id"), nullable=False)
    user_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=False)
    timestamp = db.Column(db.DateTime, default=now_utc8)

class UserSticker(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=False)
    badge_id = db.Column(db.Integer, db.ForeignKey("badge.id"), nullable=False)
    page_type = db.Column(db.String(20), nullable=False) # 'wiki' or 'book'
    x = db.Column(db.Float, default=50.0) # Percentage
    y = db.Column(db.Float, default=50.0) # Percentage
    rotation = db.Column(db.Float, default=0.0) # Degrees
    scale = db.Column(db.Float, default=1.0)
    z_index = db.Column(db.Integer, default=1)
    
    badge = db.relationship("Badge")
    user = db.relationship("User", backref="stickers")

class StickerBoardSnapshot(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=False)
    page_type = db.Column(db.String(20), nullable=False)
    title = db.Column(db.String(100), nullable=True)
    caption = db.Column(db.Text, nullable=True)
    visibility = db.Column(db.String(20), default='private') # private, followers, public
    data_json = db.Column(db.JSON, nullable=False)
    created_at = db.Column(db.DateTime, default=now_utc8)
    likes_count = db.Column(db.Integer, default=0)
    
    user = db.relationship("User", backref="sticker_snapshots")


# ==========================================
# 书签系统 Models
# ==========================================

class BookmarkType(db.Model):
    __tablename__ = 'bookmark_types'
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)       # 书签名称，如 "银杏落叶"、"纯金火漆"
    description = db.Column(db.Text, nullable=True)        # 书签背后的寓意描述
    icon = db.Column(db.String(255), nullable=False)       # 可以是 Emoji，也可以是图片的 URL
    rarity = db.Column(db.String(20), default='common')    # 稀有度: common, rare, epic, legendary
    
    # 💡 核心机制：定义如何获得这枚书签
    condition_type = db.Column(db.String(50), nullable=False) # 条件类型：例如 'wiki_read_time', 'pomo_count', 'note_created' 等
    condition_value = db.Column(db.Integer, default=0)        # 触发阈值：例如 60 (表示60分钟)，或 10 (表示10次)
    
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    
    # 关系绑定
    user_bookmarks = db.relationship('UserBookmark', backref='bookmark_type', lazy='dynamic', cascade="all, delete-orphan")


class UserBookmark(db.Model):
    __tablename__ = 'user_bookmarks'
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    bookmark_type_id = db.Column(db.Integer, db.ForeignKey('bookmark_types.id'), nullable=False)
    
    # 获取记录
    earned_at = db.Column(db.DateTime, default=datetime.utcnow)
    is_notified = db.Column(db.Boolean, default=False)     # 用于全局弹窗：False 表示这是刚获得的，还需要弹窗提示
    
    # 💡 使用状态与三维坐标 (物理夹页映射)
    is_placed = db.Column(db.Boolean, default=False)       # False = 闲置(显示+号)，True = 已使用(显示i号)
    
    # 当 is_placed 为 True 时，以下字段才有值：
    target_type = db.Column(db.String(20), nullable=True)  # 夹在哪里：'wiki' 还是 'book'
    target_id = db.Column(db.Integer, nullable=True)       # 具体的 WikiPage.id 或 BookNote.id
    target_url = db.Column(db.String(255), nullable=True)  # 完整的访问路径，方便直接生成跳转链接
    target_block_id = db.Column(db.String(100), nullable=True) # 💡 精确定位：HTML中的段落锚点 ID (如 heading-2, block-15)
    
    # 冗余存储，用于在右上角抽屉里直接展示，不用跨表联查，极大提高渲染速度
    target_title = db.Column(db.String(255), nullable=True) # 书签所在页面的标题
    target_snippet = db.Column(db.Text, nullable=True)      # 书签所夹段落的大致内容（前30个字）摘要
    placed_at = db.Column(db.DateTime, nullable=True)      # 夹入书签的时间


class VinylRecord(db.Model):
    __tablename__ = 'vinyl_record'
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    cover_url = db.Column(db.String(255), nullable=False) # 圆形封面图
    audio_url = db.Column(db.String(255), nullable=False) # 音频文件
    condition_type = db.Column(db.String(50), nullable=False) # 获取条件类型
    condition_value = db.Column(db.Integer, default=0)        # 获取条件阈值
    created_at = db.Column(db.DateTime, default=now_utc8)
    
    users = db.relationship('UserVinyl', backref='vinyl', cascade="all, delete-orphan")

class UserVinyl(db.Model):
    __tablename__ = 'user_vinyl'
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    vinyl_id = db.Column(db.Integer, db.ForeignKey('vinyl_record.id'), nullable=False)
    earned_at = db.Column(db.DateTime, default=now_utc8)
    is_notified = db.Column(db.Boolean, default=False)


class PhoneApp(db.Model):
    __tablename__ = 'phone_app'
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(40), nullable=False)
    description = db.Column(db.Text, default="")
    icon_url = db.Column(db.String(255), default="")
    icon_class = db.Column(db.String(120), default="")
    bg_color = db.Column(db.String(20), default="#82d5bb")
    action_type = db.Column(db.String(20), default="link")  # link, internal
    target_endpoint = db.Column(db.String(80), default="")
    target_url = db.Column(db.String(255), default="")
    internal_key = db.Column(db.String(80), default="")
    sort_order = db.Column(db.Integer, default=0)
    enabled = db.Column(db.Boolean, default=True)
    admin_only = db.Column(db.Boolean, default=False)
    created_at = db.Column(db.DateTime, default=now_utc8)
    updated_at = db.Column(db.DateTime, default=now_utc8, onupdate=now_utc8)

    @property
    def is_builtin_icon(self):
        return bool(self.icon_class and self.icon_class.startswith("animal-phone__app-icon--"))


@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))

def can_edit_wiki(wiki_id):
    if not current_user.is_authenticated:
        return False
    if current_user.is_admin:
        return True
    return WikiEditor.query.filter_by(wiki_id=wiki_id, user_id=current_user.id).first() is not None

def can_view_wiki(wiki_id):
    w = Wiki.query.get(wiki_id)
    if not w:
        return False
    
    if not current_user.is_authenticated:
        # If public, maybe allow? But currently all wikis require login.
        # But if visibility is public, maybe we should allow anonymous?
        # The user said "viewing permission range", implying restrictions.
        # Let's stick to login required for now as base policy.
        return False
        
    return w.is_visible_to(current_user)

DEFAULT_PHONE_APPS = [
    {"name": "主页", "description": "查看个人主页与成长记录。", "icon_class": "animal-phone__app-icon--camera", "bg_color": "#b77dee", "action_type": "link", "target_endpoint": "user_profile", "sort_order": 10},
    {"name": "设置", "description": "修改账号资料、主题与偏好。", "icon_class": "animal-phone__app-icon--miles", "bg_color": "#889df0", "action_type": "link", "target_endpoint": "settings", "sort_order": 20},
    {"name": "广场", "description": "浏览精选 Wiki 与公开知识库。", "icon_class": "animal-phone__app-icon--critterpedia", "bg_color": "#f7cd67", "action_type": "link", "target_endpoint": "index", "sort_order": 30},
    {"name": "Book", "description": "进入笔记与个人知识本。", "icon_class": "animal-phone__app-icon--diy", "bg_color": "#e59266", "action_type": "link", "target_endpoint": "book_index", "sort_order": 40},
    {"name": "贴纸墙", "description": "打开大家的学习贴纸墙。", "icon_class": "animal-phone__app-icon--design", "bg_color": "#f8a6b2", "action_type": "link", "target_endpoint": "sticker_walls_view", "sort_order": 50},
    {"name": "贴纸库", "description": "翻阅你的成就贴纸与系列进度。", "icon_class": "fas fa-book-open", "bg_color": "#82d5bb", "action_type": "link", "target_endpoint": "my_achievement_book", "internal_key": "achievement_book", "sort_order": 55},
    {"name": "排行榜", "description": "查看学习、番茄与贡献榜。", "icon_class": "animal-phone__app-icon--map", "bg_color": "#82d5bb", "action_type": "link", "target_endpoint": "leaderboard", "sort_order": 60},
    {"name": "OJ题库", "description": "浏览编程题目与样例数据。", "icon_class": "fas fa-code", "bg_color": "#5aa7d8", "action_type": "link", "target_endpoint": "oj_problem_list", "internal_key": "oj_problem_list", "sort_order": 65},
    {"name": "公告栏", "description": "阅读系统公告与活动信息。", "icon_class": "animal-phone__app-icon--variant", "bg_color": "#8ac68a", "action_type": "link", "target_endpoint": "announcements_view", "sort_order": 70},
    {"name": "任务栏", "description": "查看委托任务和悬赏。", "icon_class": "animal-phone__app-icon--helicopter", "bg_color": "#fc736d", "action_type": "link", "target_endpoint": "bulletin_view", "sort_order": 80},
    {"name": "退出", "description": "退出当前账号。", "icon_class": "animal-phone__app-icon--chat", "bg_color": "#d1da49", "action_type": "link", "target_endpoint": "logout", "sort_order": 90},
    {"name": "便签", "description": "在手机里查看和编辑你的随手便签。", "icon_class": "fas fa-sticky-note", "bg_color": "#f6c85f", "action_type": "internal", "internal_key": "sticky_notes", "sort_order": 100},
    {"name": "番茄钟", "description": "在手机里开启专注计时、查看今日番茄。", "icon_class": "fas fa-hourglass-half", "bg_color": "#ef7b73", "action_type": "internal", "internal_key": "pomodoro", "sort_order": 110},
]

PHONE_INTERNAL_APP_META = {
    "app_store": {"title": "App Store", "subtitle": "安装在这台小手机里的功能目录"},
    "profile_card": {"title": "我的名片", "subtitle": "在手机内查看个人摘要"},
    "sticky_notes": {"title": "便签", "subtitle": "随手记下灵感、待办和碎片想法"},
    "pomodoro": {"title": "番茄钟", "subtitle": "把专注节奏装进这台小手机"},
}

def seed_default_phone_apps():
    if PhoneApp.query.first() is not None:
        return
    for data in DEFAULT_PHONE_APPS:
        db.session.add(PhoneApp(**data))
    db.session.commit()

def ensure_builtin_phone_apps():
    builtin_keys = [data.get("internal_key") for data in DEFAULT_PHONE_APPS if data.get("internal_key")]
    if not builtin_keys:
        return

    existing_keys = {
        key for (key,) in db.session.query(PhoneApp.internal_key)
        .filter(PhoneApp.internal_key.in_(builtin_keys))
        .all()
    }

    created = False
    for data in DEFAULT_PHONE_APPS:
        internal_key = data.get("internal_key")
        if not internal_key or internal_key in existing_keys:
            continue
        db.session.add(PhoneApp(**data))
        created = True

    if created:
        db.session.commit()

def resolve_phone_app_url(phone_app):
    if not phone_app:
        return "#"
    if phone_app.target_endpoint:
        try:
            return url_for(phone_app.target_endpoint)
        except Exception:
            return phone_app.target_url or "#"
    return phone_app.target_url or "#"

def get_visible_phone_apps(limit=9):
    if not current_user.is_authenticated:
        return []
    query = PhoneApp.query.filter_by(enabled=True).order_by(PhoneApp.sort_order.asc(), PhoneApp.id.asc())
    apps = []
    for phone_app in query.all():
        if phone_app.admin_only and not current_user.is_admin:
            continue
        apps.append(phone_app)
        if limit and len(apps) >= limit:
            break
    return apps

def get_visible_comments(note_id=None, wiki_page_id=None):
    query = Comment.query
    if note_id is not None:
        query = query.filter(Comment.note_id == note_id)
    if wiki_page_id is not None:
        query = query.filter(Comment.wiki_page_id == wiki_page_id)

    if not current_user.is_admin:
        query = query.filter(
            db.or_(Comment.is_approved.is_(True), Comment.user_id == current_user.id)
        )

    return query.order_by(Comment.created_at.desc()).all()

def format_badge_condition(badge):
    """Format badge condition into human-readable Chinese string"""
    if badge.custom_condition_text:
        return badge.custom_condition_text
        
    ctype = badge.condition_type
    val = badge.condition_value
    target_id = getattr(badge, "target_id", None)
    
    if ctype == 'manual':
        return '管理员手动发放'
    elif ctype == 'all_users':
        return '注册即可获得'
    elif ctype == 'login_days_in_range':
        return f'活动期间累计登录 {val} 天'
    elif ctype == 'streak_days':
        return f'连续学习打卡 {val} 天'
    elif ctype == 'study_hours':
        return f'累计学习时长 {val} 小时'
    elif ctype == 'note_count':
        return f'发布公开笔记 {val} 篇'
    elif ctype == 'featured_count':
        return f'获得精选笔记 {val} 篇'
    elif ctype == 'wiki_edit_count':
        return f'参与Wiki编辑 {val} 次'
    elif ctype == 'wiki_create_count':
        return f'创建Wiki知识库 {val} 个'
    elif ctype == 'comment_count':
        return f'发表评论 {val} 条'
    elif ctype == 'night_owl_sessions':
        return f'深夜(23:00-4:00)学习 {val} 天'
    elif ctype == 'early_bird':
        return f'早起(5:00-8:00)学习 {val} 天'
    elif ctype == 'weekend_warrior':
        return f'周末累计学习 {val} 小时'
    elif ctype == 'long_session_count':
        return f'单次专注超过2小时 {val} 次'
    elif ctype == 'share_count':
        return f'分享笔记 {val} 次'
    elif ctype == 'total_views_received':
        return f'笔记累计被阅读 {val} 次'
    elif ctype == 'pomo_count':
        return f'累计完成番茄钟 {val} 个'
    elif ctype == 'user_level':
        return f'用户等级达到 Lv.{val}'
    elif ctype == 'wiki_specific_duration':
        wiki_title = None
        if target_id:
            wiki = Wiki.query.get(target_id)
            wiki_title = wiki.title if wiki else None
        if wiki_title:
            return f'在《{wiki_title}》中累计学习 {val} 分钟'
        return f'在指定 Wiki 中累计学习 {val} 分钟'
    elif ctype == 'assignment_complete':
        assignment_title = None
        if target_id:
            assignment = Assignment.query.get(target_id)
            assignment_title = assignment.title if assignment else None
        if assignment_title:
            return f'完成任务《{assignment_title}》'
        return '完成指定任务'
    elif ctype == 'assignment_stars':
        assignment_title = None
        if target_id:
            assignment = Assignment.query.get(target_id)
            assignment_title = assignment.title if assignment else None
        if assignment_title:
            return f'在任务《{assignment_title}》中获得至少 {val} 颗星'
        return f'在指定任务中获得至少 {val} 颗星'
    elif ctype == 'assignment_grade':
        assignment_title = None
        grade_map = {6: 'S+', 5: 'S', 4: 'A', 3: 'B', 2: 'C', 1: 'D'}
        grade_text = grade_map.get(val, f'等级 {val}')
        if target_id:
            assignment = Assignment.query.get(target_id)
            assignment_title = assignment.title if assignment else None
        if assignment_title:
            return f'在任务《{assignment_title}》中达到 {grade_text} 及以上'
        return f'在指定任务中达到 {grade_text} 及以上'
    elif ctype == 'oj_ac_count':
        return f'OJ 累计 AC {val} 道题'
    elif ctype == 'oj_submission_count':
        return f'OJ 累计提交 {val} 次'
    elif ctype == 'oj_error_count':
        return f'OJ 累计调试错误 {val} 次'
    elif ctype == 'oj_assignment_complete_count':
        return f'完整完成 {val} 次 OJ 作业'
    elif ctype == 'oj_consecutive_ac_count':
        return f'OJ 当前连续 AC {val} 次'
    elif ctype == 'oj_consecutive_error_count':
        return f'OJ 当前连续调试未通过 {val} 次'
    elif ctype == 'oj_first_try_ac_count':
        return f'OJ 一发 AC {val} 道题'
    elif ctype == 'oj_daily_ac_peak':
        return f'OJ 单日 AC 达到 {val} 道题'
    elif ctype == 'oj_specific_assignment_complete':
        assignment_title = None
        if target_id:
            assignment = OJAssignment.query.get(target_id)
            assignment_title = assignment.title if assignment else None
        if assignment_title:
            return f'完成 OJ 作业《{assignment_title}》的全部题目'
        return '完成指定 OJ 作业的全部题目'
    else:
        return f'{ctype}: {val}'

@app.template_filter('badge_condition_text')
def badge_condition_text_filter(badge):
    return format_badge_condition(badge)

def process_tags(tag_string):
    if not tag_string:
        return []
    # Support comma or space separation? Let's stick to comma for multi-word tags support
    # But usually space is easier. Let's support comma.
    # Replace Chinese comma
    tag_string = tag_string.replace('，', ',')
    tag_names = set([t.strip() for t in tag_string.split(',') if t.strip()])
    tags = []
    for name in tag_names:
        tag = Tag.query.filter_by(name=name).first()
        if not tag:
            tag = Tag(name=name)
            db.session.add(tag)
        tags.append(tag)
    return tags

def backup_wiki(wiki_id):
    """
    Backup wiki data to a JSON file and copy associated files.
    Returns the backup directory path.
    """
    w = Wiki.query.get(wiki_id)
    if not w:
        return None

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_root = os.path.join(app.root_path, "backups")
    backup_dir = os.path.join(backup_root, f"wiki_{wiki_id}_{timestamp}")
    
    if not os.path.exists(backup_dir):
        os.makedirs(backup_dir)

    # 1. Backup Metadata and Structure
    data = {
        "wiki": {
            "id": w.id,
            "title": w.title,
            "description": w.description,
            "created_by_id": w.created_by_id,
            "created_at": w.created_at.isoformat() if w.created_at else None
        },
        "pages": [],
        "files": [],
        "editors": [e.user_id for e in w.editors],
        "subscribers": [s.user_id for s in w.subscriptions]
    }

    # Pages
    for p in w.pages:
        page_data = {
            "id": p.id,
            "title": p.title,
            "slug": p.slug,
            "content_md": p.content_md,
            "view_count": p.view_count,
            "updated_at": p.updated_at.isoformat() if p.updated_at else None,
            "tags": [t.name for t in p.tags],
            "comments": [
                {
                    "user_id": c.user_id,
                    "content": c.content,
                    "created_at": c.created_at.isoformat() if c.created_at else None,
                    "is_approved": c.is_approved,
                    "approved_at": c.approved_at.isoformat() if c.approved_at else None,
                    "approved_by_id": c.approved_by_id
                } for c in p.comments
            ],
            "history": [
                {
                    "user_id": h.user_id,
                    "title": h.title,
                    "slug": h.slug,
                    "content_md": h.content_md,
                    "created_at": h.created_at.isoformat() if h.created_at else None
                } for h in p.history
            ]
        }
        data["pages"].append(page_data)

    # Files
    files_dir = os.path.join(backup_dir, "files")
    if not os.path.exists(files_dir):
        os.makedirs(files_dir)

    for f in w.files:
        file_data = {
            "id": f.id,
            "filename": f.filename,
            "file_path": f.file_path,
            "uploaded_by_id": f.uploaded_by_id,
            "uploaded_at": f.uploaded_at.isoformat() if f.uploaded_at else None
        }
        data["files"].append(file_data)
        
        # Copy physical file
        # Assuming f.filename matches the file in UPLOAD_FOLDER
        # However, upload logic uses unique_filename. 
        # In upload_wiki_file: 
        # unique_filename = str(uuid.uuid4()) + ext
        # file.save(os.path.join(upload_dir, unique_filename))
        # wf = WikiFile(..., filename=filename, ...) <- This stores original filename!
        # Wait, the physical filename is NOT stored in WikiFile directly?
        # file_path stores url_for("static", filename=f"uploads/{unique_filename}")
        # So we need to extract unique_filename from file_path.
        
        try:
            # f.file_path is like /static/uploads/UUID.ext
            if "/static/uploads/" in f.file_path:
                unique_filename = f.file_path.split("/static/uploads/")[-1]
                src_path = os.path.join(app.config["UPLOAD_FOLDER"], unique_filename)
                if os.path.exists(src_path):
                    shutil.copy2(src_path, os.path.join(files_dir, unique_filename))
        except Exception as e:
            print(f"Error copying file {f.id}: {e}")

    # Write JSON
    with open(os.path.join(backup_dir, "data.json"), "w", encoding="utf-8") as json_file:
        json.dump(data, json_file, indent=4, ensure_ascii=False)

    return backup_dir

@app.route("/")
def index():
    q = request.args.get("q", "").strip()
    user_id = request.args.get("user_id")
    target_user = None
    
    if user_id:
        if current_user.is_authenticated:
            try:
                potential_target = User.query.get(int(user_id))
                if potential_target:
                    if potential_target.id == current_user.id:
                        target_user = potential_target
                    else:
                        if current_user.is_following(potential_target) and potential_target.is_following(current_user):
                            target_user = potential_target
            except:
                pass
                
    if q:
        # 💡 搜索模式：展示普通网格列表
        wikis = Wiki.query.filter(Wiki.title.contains(q)).order_by(Wiki.created_at.desc()).all()
        pages = WikiPage.query.filter((WikiPage.title.contains(q)) | (WikiPage.content_md.contains(q))).all()
    else:
        # 💡 默认模式/筛选模式：展示 3D 滚动精选卡片
        visibility_filter = request.args.get("visibility", "all")
        
        # 核心修复 1：将 is_featured=True 作为基础查询起点，而不是直接查询 .all()
        query = Wiki.query.filter_by(is_featured=True)
        
        if visibility_filter == "my_class":
            if current_user.is_authenticated and current_user.student_class:
                query = query.outerjoin(wiki_classes).filter(
                    db.or_(Wiki.visibility == 'public', wiki_classes.c.class_id == current_user.student_class.id)
                )
            else:
                query = query.filter(Wiki.visibility == 'public')
                
        elif visibility_filter == "my_groups":
            if current_user.is_authenticated and current_user.groups:
                group_ids = [g.id for g in current_user.groups]
                query = query.outerjoin(wiki_groups).filter(
                    db.or_(Wiki.visibility == 'public', wiki_groups.c.group_id.in_(group_ids))
                )
            else:
                query = query.filter(Wiki.visibility == 'public')
                
        elif visibility_filter.startswith("class_") and current_user.is_authenticated and current_user.is_admin:
            try:
                class_id = int(visibility_filter.split("_")[1])
                query = query.outerjoin(wiki_classes).filter(
                    db.or_(Wiki.visibility == 'public', wiki_classes.c.class_id == class_id)
                )
            except:
                query = query.filter(Wiki.visibility == 'public')
                
        elif visibility_filter.startswith("group_") and current_user.is_authenticated and current_user.is_admin:
            try:
                group_id = int(visibility_filter.split("_")[1])
                query = query.outerjoin(wiki_groups).filter(
                    db.or_(Wiki.visibility == 'public', wiki_groups.c.group_id == group_id)
                )
            except:
                query = query.filter(Wiki.visibility == 'public')
        
        # 核心修复 2：在查询末尾统一加上精选排序规则，再执行 .all()
        fetched_wikis = query.order_by(Wiki.featured_order.asc()).all()
        
        visible_wikis = []
        for w in fetched_wikis:
            # 这里的权限检验作为最后的兜底
            if w.is_visible_to(current_user if current_user.is_authenticated else User(is_admin=False, id=-1)): 
                visible_wikis.append(w)
        
        wikis = visible_wikis
        pages = []
        
    # Context data for filters
    all_classes = []
    all_groups = []
    if current_user.is_authenticated and current_user.is_admin:
        all_classes = Class.query.order_by(Class.name).all()
        all_groups = Group.query.order_by(Group.name).all()

    my_home_stickers = []
    if current_user.is_authenticated:
        stickers = UserSticker.query.filter_by(user_id=current_user.id, page_type='profile').all()
        for s in stickers:
            if not s.badge:
                continue
                
            icon_val = s.badge.icon
            final_icon = badge_icon_url(icon_val) if is_image_icon(icon_val) else icon_val
            
            my_home_stickers.append({
                "id": s.id, 
                "badge_id": s.badge_id,
                "badge_icon": final_icon,
                "badge_name": s.badge.name,
                "x": s.x, 
                "y": s.y, 
                "rotation": s.rotation, 
                "scale": s.scale, 
                "z_index": s.z_index
            })

    return render_template("index.html", 
                           wikis=wikis, 
                           pages=pages, 
                           q=q, 
                           target_user=target_user, 
                           current_filter=request.args.get("visibility", "all"),
                           all_classes=all_classes, 
                           all_groups=all_groups,
                           my_home_stickers=my_home_stickers)

@app.route("/register", methods=["GET", "POST"])
def register():
    # 关闭注册功能
    flash("注册功能已关闭，请联系管理员")
    return redirect(url_for("login"))

@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")
        remember = request.form.get("remember") == "on"  # Get remember me checkbox
        
        u = User.query.filter_by(username=username).first()
        if not u or not u.check_password(password):
            flash("用户名或密码错误")
            return redirect(url_for("login"))
            
        login_user(u, remember=remember) # Use remember me
        return redirect(url_for("index"))
    return render_template("login.html")

@app.route("/logout")
@login_required
def logout():
    logout_user()
    return redirect(url_for("index"))

@app.route("/settings", methods=["GET", "POST"])
@login_required
def settings():
    if request.method == "POST":
        current_password = request.form.get("current_password", "")
        new_password = request.form.get("new_password", "")
        confirm_password = request.form.get("confirm_password", "")
        
        # Update password if requested
        if new_password:
            if not current_password:
                flash("修改密码需要输入当前密码进行验证")
                return redirect(url_for("settings"))
                
            if not current_user.check_password(current_password):
                flash("当前密码错误")
                return redirect(url_for("settings"))
                
            if new_password != confirm_password:
                flash("两次输入的新密码不一致")
                return redirect(url_for("settings"))
                
            current_user.set_password(new_password)
            db.session.commit()
            flash("密码已修改")
        else:
            flash("未做任何修改")
            
        return redirect(url_for("settings"))
        
    return render_template("settings.html")

@app.route("/admin/wikis/new", methods=["GET", "POST"])
@login_required
def create_wiki():
    if not current_user.is_admin:
        abort(403)
    if request.method == "POST":
        title = request.form.get("title", "").strip()
        description = request.form.get("description", "").strip()
        bg_color = request.form.get("bg_color", "#ffffff").strip()
        fg_color = request.form.get("fg_color", "#1f2937").strip()
        pattern = request.form.get("pattern", "pattern-none").strip()
        blur_level = request.form.get("blur_level", 60, type=int)
        gradient_color = request.form.get("gradient_color", "").strip()
        gradient_direction = request.form.get("gradient_direction", "to bottom right").strip()
        
        if not title:
            flash("标题不能为空")
            return redirect(url_for("create_wiki"))
        w = Wiki(
            title=title, 
            description=description, 
            created_by_id=current_user.id,
            bg_color=bg_color,
            fg_color=fg_color,
            pattern=pattern,
            blur_level=blur_level,
            gradient_color=gradient_color,
            gradient_direction=gradient_direction,
            visibility=request.form.get("visibility", "public")
        )
        
        # Permissions
        class_ids = request.form.getlist("allowed_classes")
        if class_ids:
            w.allowed_classes = Class.query.filter(Class.id.in_(class_ids)).all()
            
        group_ids = request.form.getlist("allowed_groups")
        if group_ids:
            w.allowed_groups = Group.query.filter(Group.id.in_(group_ids)).all()
            
        # User permissions (handled as list of IDs)
        # Note: Frontend should provide a way to select users
        user_ids = request.form.getlist("allowed_users")
        if user_ids:
            w.allowed_users = User.query.filter(User.id.in_(user_ids)).all()
            
        db.session.add(w)
        db.session.commit()
        
        # Check initialization options
        init_home = request.form.get("init_home") == "on"
        if init_home:
            home_page = WikiPage(
                wiki_id=w.id,
                title="欢迎",
                slug="home",
                content_md=f"# 欢迎来到 {w.title}\n\n{w.description or '这是一个新的 Wiki 知识库。'}\n\n## 开始使用\n\n- 点击右上角的编辑按钮修改此页面\n- 点击侧边栏的“新建页面”添加更多内容\n"
            )
            db.session.add(home_page)
            db.session.commit()
            
            # Create a view log for the creator
            log = WikiViewLog(wiki_id=w.id, user_id=current_user.id)
            db.session.add(log)
            db.session.commit()
            
            return redirect(url_for("view_page", wiki_id=w.id, slug="home"))
            
        return redirect(url_for("wiki_detail", wiki_id=w.id))
        
    classes = Class.query.order_by(Class.name).all()
    groups = Group.query.order_by(Group.name).all()
    # Users for selection - might be too many, but let's provide all for now or implement search
    # For now, let's just pass all users sorted by name
    all_users = User.query.order_by(User.class_name, User.student_id).all()
    
    return render_template("admin/create_wiki.html", classes=classes, groups=groups, all_users=all_users)

@app.route("/wikis/<int:wiki_id>")
@login_required
def wiki_detail(wiki_id):
    if not can_view_wiki(wiki_id):
        abort(403)
    w = Wiki.query.get_or_404(wiki_id)
    
    # 1. 获取所有文件夹和页面
    all_folders = WikiFolder.query.filter_by(wiki_id=wiki_id).all()
    all_pages = WikiPage.query.filter_by(wiki_id=wiki_id).all()
    
    # 2. 构建混合目录树 (支持文件夹和独立页面混排)
    root_items = []
    
    # A. 组装文件夹及其内部的页面
    for f in all_folders:
        f.item_type = 'folder'
        # 筛选出属于该文件夹的页面，并按 order_weight 升序
        f.children = sorted([p for p in all_pages if p.folder_id == f.id], key=lambda x: x.order_weight)
        root_items.append(f)
        
    # B. 组装根目录下的独立页面
    for p in all_pages:
        if p.folder_id is None:
            p.item_type = 'page'
            root_items.append(p)
            
    # C. 对所有根节点元素统一按 order_weight 进行全局排序
    root_items = sorted(root_items, key=lambda x: x.order_weight)
    
    # 决定默认显示的第一页
    # 为了保证按真实显示顺序打开第一页，我们将树形结构展平来寻找第一个页面
    pages_sorted_flat = []
    for item in root_items:
        if getattr(item, 'item_type', '') == 'folder':
            pages_sorted_flat.extend(getattr(item, 'children', []))
        elif getattr(item, 'item_type', '') == 'page':
            pages_sorted_flat.append(item)
            
    first_page = pages_sorted_flat[0] if pages_sorted_flat else None
    
    is_editor = can_edit_wiki(wiki_id)
    is_subscribed = Subscription.query.filter_by(wiki_id=wiki_id, user_id=current_user.id).first() is not None
    html = markdown(first_page.content_md or "") if first_page else ""
    
    if first_page:
        log = WikiViewLog(wiki_id=wiki_id, user_id=current_user.id)
        db.session.add(log)
        first_page.view_count += 1
        db.session.commit()
        html = parse_assignment_tags(html, first_page.id)
    
    # Fetch contributors
    contributors = []
    try:
        if all_pages:
            page_ids_list = [pg.id for pg in all_pages]
            contributors = db.session.query(
                User, 
                db.func.count(WikiPageHistory.id).label('edit_count')
            ).join(
                WikiPageHistory, WikiPageHistory.user_id == User.id
            ).filter(
                WikiPageHistory.page_id.in_(page_ids_list)
            ).group_by(
                User.id
            ).order_by(
                db.text('edit_count DESC')
            ).all()
    except Exception as e:
        print(f"Error fetching contributors in wiki_detail: {e}")
        contributors = []
        
    my_recent_notes = []
    if current_user.is_authenticated:
        my_recent_notes = Note.query.filter_by(user_id=current_user.id).order_by(Note.updated_at.desc()).limit(15).all()
    visible_comments = get_visible_comments(wiki_page_id=first_page.id) if first_page else []

    return render_template("wiki/view_page.html", 
                           wiki=w, 
                           page=first_page, 
                           root_items=root_items, # 💡 传入树形结构
                           pages=all_pages,       # 兼容旧逻辑保留，但不用于目录渲染
                           html=html, 
                           can_edit=is_editor, 
                           is_subscribed=is_subscribed, 
                           is_home=True, 
                           contributors=contributors,
                           my_recent_notes=my_recent_notes,
                           visible_comments=visible_comments)

import json

@app.route("/admin/wikis/<int:wiki_id>/edit", methods=["GET", "POST"])
@login_required
def edit_wiki(wiki_id):
    w = Wiki.query.get_or_404(wiki_id)
    if not current_user.is_admin and current_user.id != w.created_by_id:
        abort(403)
        
    if request.method == "POST":
        # 1. 基础信息与权限保存
        w.title = request.form.get("title", "").strip()
        w.description = request.form.get("description", "").strip()
        w.bg_color = request.form.get("bg_color", "#ffffff").strip()
        w.fg_color = request.form.get("fg_color", "#1f2937").strip()
        w.pattern = request.form.get("pattern", "pattern-none").strip()
        w.blur_level = request.form.get("blur_level", 60, type=int)
        w.gradient_color = request.form.get("gradient_color", "").strip()
        w.gradient_direction = request.form.get("gradient_direction", "to bottom right").strip()
        w.visibility = request.form.get("visibility", "public")
        w.icon_scale = request.form.get("icon_scale", 1.0, type=float)
        w.icon_position_x = request.form.get("icon_position_x", 50.0, type=float)
        w.icon_position_y = request.form.get("icon_position_y", 50.0, type=float)

        w.icon_scale = max(0.5, min(w.icon_scale, 3.0))
        w.icon_position_x = max(0.0, min(w.icon_position_x, 100.0))
        w.icon_position_y = max(0.0, min(w.icon_position_y, 100.0))

        remove_icon = request.form.get("remove_icon") == "1"
        icon_file = request.files.get("icon_file")

        if remove_icon and w.icon_url:
            remove_uploaded_static_file(w.icon_url, "wiki_icons")
            w.icon_url = None

        if icon_file and icon_file.filename:
            try:
                new_icon_url = save_uploaded_image(icon_file, "wiki_icons", "wiki_icon")
            except ValueError as exc:
                flash(str(exc), "error")
                return redirect(url_for("edit_wiki", wiki_id=w.id))

            if w.icon_url and w.icon_url != new_icon_url:
                remove_uploaded_static_file(w.icon_url, "wiki_icons")
            w.icon_url = new_icon_url
        
        # 权限更新逻辑（保持你原有的逻辑）
        class_ids = request.form.getlist("allowed_classes")
        w.allowed_classes = Class.query.filter(Class.id.in_(class_ids)).all() if class_ids else []
        group_ids = request.form.getlist("allowed_groups")
        w.allowed_groups = Group.query.filter(Group.id.in_(group_ids)).all() if group_ids else []
        user_ids = request.form.getlist("allowed_users")
        w.allowed_users = User.query.filter(User.id.in_(user_ids)).all() if user_ids else []

        db.session.commit()
        flash("基本信息已更新", "success")
        return redirect(url_for("wiki_detail", wiki_id=w.id))
        
    # ==========================================
    # GET 请求：构建用于渲染的层级数据字典
    # ==========================================
    classes = Class.query.order_by(Class.name).all()
    groups = Group.query.order_by(Group.name).all()
    all_users = User.query.order_by(User.class_name, User.student_id).all()
    
    all_folders = WikiFolder.query.filter_by(wiki_id=w.id).all()
    all_pages = WikiPage.query.filter_by(wiki_id=w.id).all()
    
    root_items = []
    # A. 处理文件夹
    for f in all_folders:
        # 获取属于该文件夹的文章，并排序
        f_pages = [p for p in all_pages if p.folder_id == f.id]
        f_pages.sort(key=lambda x: x.order_weight or 0)
        
        root_items.append({
            'type': 'folder',
            'id': f.id,
            'name': f.name,
            'order_weight': f.order_weight or 0,
            'children': f_pages # 这里存放的是 WikiPage 对象，方便读取标题等属性
        })
        
    # B. 处理根目录独立文章
    for p in all_pages:
        if p.folder_id is None:
            root_items.append({
                'type': 'page',
                'id': p.id,
                'title': p.title,
                'order_weight': p.order_weight or 0
            })
            
    # C. 全局按照 order_weight 排序
    root_items.sort(key=lambda x: x['order_weight'])
    
    return render_template(
        "admin/edit_wiki.html", 
        wiki=w, 
        classes=classes, 
        groups=groups, 
        all_users=all_users,
        root_items=root_items
    )

# ==========================================
# 💡 实时保存接口：由前端拖拽完成后自动调用
# ==========================================
@app.route("/admin/wikis/<int:wiki_id>/update_structure", methods=["POST"])
@login_required
def update_wiki_structure(wiki_id):
    w = Wiki.query.get_or_404(wiki_id)
    if not current_user.is_admin and current_user.id != w.created_by_id:
        return {"success": False, "error": "无权限"}, 403

    data = request.json.get("structure_data", [])
    if not data:
        return {"success": True}

    id_mapping = {} # 用于把临时ID映射回真实ID
    received_folder_ids = []

    try:
        # 1. 第一轮：处理文件夹创建与更新
        for item in data:
            if item.get('type') == 'folder':
                f_id = str(item.get('id'))
                if f_id.startswith('new_'):
                    new_f = WikiFolder(wiki_id=w.id, name=item.get('name'), order_weight=item.get('order'))
                    db.session.add(new_f)
                    db.session.flush()
                    id_mapping[f_id] = new_f.id
                    db_folder_id = new_f.id
                else:
                    db_folder_id = int(f_id)
                    folder = WikiFolder.query.get(db_folder_id)
                    if folder:
                        folder.name = item.get('name', folder.name)
                        folder.order_weight = item.get('order')
                received_folder_ids.append(db_folder_id)

        # 2. 清理删除的文件夹
        existing_folders = WikiFolder.query.filter_by(wiki_id=w.id).all()
        for ef in existing_folders:
            if ef.id not in received_folder_ids:
                pages = WikiPage.query.filter_by(folder_id=ef.id).all()
                for pf in pages: pf.folder_id = None # 释放文章
                db.session.delete(ef)
        db.session.flush()

        # 3. 第二轮：处理文章归属和最终排序
        for item in data:
            if item.get('type') == 'folder':
                db_f_id = id_mapping.get(str(item['id'])) or int(item['id'])
                for p_idx, p_item in enumerate(item.get('pages', [])):
                    page = WikiPage.query.get(int(p_item['id']))
                    if page:
                        page.folder_id = db_f_id
                        page.order_weight = p_item['order']
            elif item.get('type') == 'page':
                page = WikiPage.query.get(int(item['id']))
                if page:
                    page.folder_id = None
                    page.order_weight = item['order']

        db.session.commit()
        return {"success": True, "id_mapping": id_mapping}
    except Exception as e:
        db.session.rollback()
        return {"success": False, "error": str(e)}, 500
@app.route("/admin/wikis/<int:wiki_id>/delete", methods=["POST"])
@login_required
def delete_wiki(wiki_id):
    if not current_user.is_admin:
        abort(403)
        
    w = Wiki.query.get_or_404(wiki_id)
    
    # Backup first
    try:
        backup_path = backup_wiki(wiki_id)
        if not backup_path:
             flash("备份失败，操作取消")
             return redirect(url_for("edit_wiki", wiki_id=wiki_id))
    except Exception as e:
        flash(f"备份过程中出错: {str(e)}")
        return redirect(url_for("edit_wiki", wiki_id=wiki_id))
        
    # Delete Wiki
    try:
        db.session.delete(w)
        db.session.commit()
        flash(f"Wiki 已删除。备份已保存至: {backup_path}")
        return redirect(url_for("index"))
    except Exception as e:
        db.session.rollback()
        flash(f"删除失败: {str(e)}")
        return redirect(url_for("edit_wiki", wiki_id=wiki_id))


@app.route("/wikis/<int:wiki_id>/subscribe", methods=["POST"])
@login_required
def subscribe_wiki(wiki_id):
    w = Wiki.query.get_or_404(wiki_id)
    sub = Subscription.query.filter_by(wiki_id=wiki_id, user_id=current_user.id).first()
    if sub:
        # 取消订阅
        db.session.delete(sub)
    else:
        # 订阅
        s = Subscription(wiki_id=wiki_id, user_id=current_user.id)
        db.session.add(s)
    db.session.commit()
    # 获取来源页面，如果未指定则重定向到详情页
    next_page = request.args.get("next")
    if not next_page:
        next_page = url_for("wiki_detail", wiki_id=wiki_id)
    return redirect(next_page)

@app.route("/wikis/<int:wiki_id>/pages/new", methods=["GET", "POST"])
@login_required
def new_page(wiki_id):
    w = Wiki.query.get_or_404(wiki_id)
    if not can_edit_wiki(wiki_id):
        abort(403)
        
    if request.method == "POST":
        title = request.form.get("title", "").strip()
        slug = request.form.get("slug", "").strip()
        content_md = request.form.get("content_md", "")
        tags_str = request.form.get("tags", "")
        
        if not title or not slug:
            flash("标题与标识不能为空")
            return redirect(url_for("new_page", wiki_id=wiki_id))
            
        # 💡 核心修复：计算当前最大的排序权重，确保新页面排在最后面
        max_page = WikiPage.query.filter_by(wiki_id=wiki_id).order_by(WikiPage.order_weight.desc()).first()
        max_folder = WikiFolder.query.filter_by(wiki_id=wiki_id).order_by(WikiFolder.order_weight.desc()).first()
        
        max_p_weight = max_page.order_weight if max_page and max_page.order_weight else 0
        max_f_weight = max_folder.order_weight if max_folder and max_folder.order_weight else 0
        next_order_weight = max(max_p_weight, max_f_weight) + 1
        
        # 保存时带上自动计算好的权重
        p = WikiPage(
            wiki_id=wiki_id, 
            title=title, 
            slug=slug, 
            content_md=content_md,
            order_weight=next_order_weight # <--- 就是这个决定了它会在最后面！
        )
        
        # Process tags
        p.tags = process_tags(tags_str)
        
        db.session.add(p)
        db.session.commit()
        
        # Check badges (Wiki Create Count, Wiki Edit Count)
        check_and_award_badges(current_user)
        
        return redirect(url_for("view_page", wiki_id=wiki_id, slug=slug))
        
    return render_template("wiki/new_page.html", wiki=w)




@app.route("/wikis/<int:wiki_id>/pages/<slug>/edit", methods=["GET", "POST"])
@login_required
def edit_page(wiki_id, slug):
    w = Wiki.query.get_or_404(wiki_id)
    if not can_edit_wiki(wiki_id):
        abort(403)
    p = WikiPage.query.filter_by(wiki_id=wiki_id, slug=slug).first_or_404()
    
    if request.method == "POST":
        # 1. 保存历史版本
        history = WikiPageHistory(
            page_id=p.id, user_id=current_user.id,
            title=p.title, slug=p.slug, content_md=p.content_md
        )
        db.session.add(history)
        
        # 2. 接收基础数据
        p.title = request.form.get("title", "").strip()
        p.slug = request.form.get("slug", "").strip()
        new_content = request.form.get("content_md", "")
        
        # ==========================================
        # 💡 3. 核心升级：作业标签“自动打钢印”系统
        # ==========================================
        import re
        
        # A. 扫描带有 ID 钢印的现有作业 (例如 [作业#15:新名字])，更新名称
        existing_tags = re.findall(r'\[作业#(\d+)[:：]\s*(.*?)\]', new_content)
        for assign_id_str, new_title in existing_tags:
            assign = Assignment.query.get(int(assign_id_str))
            if assign and assign.wiki_page_id == p.id:
                assign.title = new_title.strip()
                
        # B. 扫描没有 ID 的新作业 (例如 [作业:第一次提交])
        # 注意正则：紧跟在“作业”后面的必须是冒号，排除了带有 # 的情况
        new_tags = re.findall(r'\[作业[:：]\s*([^\]]+)\]', new_content)
        for t_title in new_tags:
            t_title = t_title.strip()
            if t_title:
                # 检查是否已经存在同名的老作业 (完美兼容你之前创建但还没打钢印的数据)
                exists = Assignment.query.filter_by(wiki_page_id=p.id, title=t_title).first()
                if not exists:
                    new_assign = Assignment(wiki_page_id=p.id, title=t_title)
                    db.session.add(new_assign)
                    db.session.flush() # 立即获取生成的 ID
                    target_id = new_assign.id
                else:
                    target_id = exists.id
                
                # C. 自动修改 Markdown 内容：注入 ID 钢印
                escaped_title = re.escape(t_title)
                # 精确匹配旧文本，将其替换为带 #ID 的新文本
                replace_pattern = r'\[作业[:：]\s*' + escaped_title + r'\]'
                new_content = re.sub(replace_pattern, f'[作业#{target_id}:{t_title}]', new_content)

        # 4. 保存最终的文本和设置
        p.content_md = new_content
        tags_str = request.form.get("tags", "")
        p.tags = process_tags(tags_str)
        p.comment_enabled = request.form.get("comment_enabled") == "on"
        
        try:
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            flash(f"保存失败，数据库错误: {str(e)}", "error")
            return redirect(url_for("edit_page", wiki_id=wiki_id, slug=slug))
            
        check_and_award_badges(current_user)
        flash("页面已更新", "success")
        return redirect(url_for("view_page", wiki_id=wiki_id, slug=p.slug))
        
    return render_template("wiki/edit_page.html", wiki=w, page=p)

@app.route("/wikis/<int:wiki_id>/pages/<slug>/comment", methods=["POST"])
@login_required
def comment_wiki_page(wiki_id, slug):
    w = Wiki.query.get_or_404(wiki_id)
    # Check view permission
    if not can_view_wiki(wiki_id):
        abort(403)
        
    p = WikiPage.query.filter_by(wiki_id=wiki_id, slug=slug).first_or_404()
    
    content = request.form.get("content", "").strip()
    if content:
        c = Comment(
            content=content,
            user_id=current_user.id,
            wiki_page_id=p.id,
            is_approved=False
        )
        db.session.add(c)
        db.session.commit()
        
        # Check badges
        check_and_award_badges(current_user)
        
        flash("评论已提交，等待管理员审核通过后会展示给其他用户。", "info")
    else:
        flash("评论内容不能为空", "warning")
        
    return redirect(url_for("view_page", wiki_id=wiki_id, slug=slug))

@app.route("/comments/<int:comment_id>/delete", methods=["POST"])
@login_required
def delete_comment(comment_id):
    c = Comment.query.get_or_404(comment_id)
    
    # Check permission:
    # 1. Comment author
    # 2. Admin
    # 3. Content owner (Note or Wiki Page owner) - Logic below
    
    is_allowed = False
    
    if current_user.id == c.user_id or current_user.is_admin:
        is_allowed = True
    elif c.note_id:
        note = Note.query.get(c.note_id)
        if note and note.user_id == current_user.id:
            is_allowed = True
    elif c.wiki_page_id:
        page = WikiPage.query.get(c.wiki_page_id)
        # For wiki pages, "owner" is tricky. Let's say wiki admins/editors can delete?
        # Or specifically the wiki creator? 
        # For now let's stick to Wiki Editor permission
        if page:
             if can_edit_wiki(page.wiki_id):
                 is_allowed = True
                 
    if not is_allowed:
        abort(403)
        
    # Get redirect URL before deletion
    next_url = request.referrer
    if not next_url:
        if c.note_id:
            next_url = url_for("view_note", note_id=c.note_id)
        elif c.wiki_page_id:
            # Need to fetch page/wiki info to construct URL if referrer is missing
            page = WikiPage.query.get(c.wiki_page_id)
            if page:
                next_url = url_for("view_page", wiki_id=page.wiki_id, slug=page.slug)
            else:
                next_url = url_for("index")
        else:
            next_url = url_for("index")
            
    db.session.delete(c)
    db.session.commit()
    flash("评论已删除")
    return redirect(next_url)

@app.route("/wikis/<int:wiki_id>/pages/<slug>/history")
@login_required
def page_history(wiki_id, slug):
    w = Wiki.query.get_or_404(wiki_id)
    p = WikiPage.query.filter_by(wiki_id=wiki_id, slug=slug).first_or_404()
    if not can_edit_wiki(wiki_id):
        abort(403)
    return render_template("wiki/page_history.html", wiki=w, page=p)

@app.route("/wikis/<int:wiki_id>/pages/<slug>/history/<int:history_id>/restore", methods=["POST"])
@login_required
def restore_page(wiki_id, slug, history_id):
    if not can_edit_wiki(wiki_id):
        abort(403)
    p = WikiPage.query.filter_by(wiki_id=wiki_id, slug=slug).first_or_404()
    h = WikiPageHistory.query.get_or_404(history_id)
    
    if h.page_id != p.id:
        abort(404)
        
    current_history = WikiPageHistory(
        page_id=p.id,
        user_id=current_user.id,
        title=p.title,
        slug=p.slug,
        content_md=p.content_md
    )
    db.session.add(current_history)
    
    p.title = h.title
    p.slug = h.slug
    p.content_md = h.content_md
    db.session.commit()
    
    flash(f"已恢复到 {h.created_at.strftime('%Y-%m-%d %H:%M')} 的版本")
    return redirect(url_for("view_page", wiki_id=wiki_id, slug=p.slug))

@app.route("/wikis/<int:wiki_id>/pages/<slug>/delete", methods=["POST"])
@login_required
def delete_page(wiki_id, slug):
    w = Wiki.query.get_or_404(wiki_id)
    if not can_edit_wiki(wiki_id):
        abort(403)
    p = WikiPage.query.filter_by(wiki_id=wiki_id, slug=slug).first_or_404()
    db.session.delete(p)
    db.session.commit()
    flash("页面已删除")
    return redirect(url_for("wiki_detail", wiki_id=wiki_id))

@app.route("/admin/wikis/<int:wiki_id>/stats")
@login_required
def wiki_stats(wiki_id):
    if not current_user.is_admin:
        abort(403)
    w = Wiki.query.get_or_404(wiki_id)
    subscribers = User.query.join(Subscription).filter(Subscription.wiki_id == wiki_id).all()
    now = datetime.utcnow() # 注意：这里尽量统一用 now_utc8()，如果之前的代码混用了请自行统一
    
    # 统计概览
    periods = {
        "1h": timedelta(hours=1),
        "12h": timedelta(hours=12),
        "1d": timedelta(days=1),
        "1w": timedelta(weeks=1),
        "1m": timedelta(days=30),
        "3m": timedelta(days=90),
        "6m": timedelta(days=180),
        "1y": timedelta(days=365)
    }
    view_stats = {}
    sub_stats = {}
    
    for label, delta in periods.items():
        start_time = now - delta
        view_stats[label] = WikiViewLog.query.filter(WikiViewLog.wiki_id == wiki_id, WikiViewLog.timestamp >= start_time).count()
        sub_stats[label] = Subscription.query.filter(Subscription.wiki_id == wiki_id, Subscription.created_at >= start_time).count()

    # --- 新增：生成最近 14 天的每日趋势数据供图表使用 ---
    trend_days = 14
    trend_start = now - timedelta(days=trend_days)
    
    # 获取最近14天的所有日志
    raw_logs = WikiViewLog.query.filter(
        WikiViewLog.wiki_id == wiki_id, 
        WikiViewLog.timestamp >= trend_start
    ).all()
    
    # 按日期聚合
    date_counts = {}
    # 初始化过去14天每一天为0，防止断层
    for i in range(trend_days):
        d = (now - timedelta(days=i)).strftime("%m-%d")
        date_counts[d] = 0
        
    for log in raw_logs:
        # 假设 timestamp 是 UTC，简单转为字符串日期统计
        # 如果需要精确的 +8区，建议 log.timestamp + timedelta(hours=8)
        d_str = (log.timestamp + timedelta(hours=8)).strftime("%m-%d")
        if d_str in date_counts:
            date_counts[d_str] += 1
            
    # 整理成 Chart.js 需要的列表 (按日期升序)
    chart_labels = sorted(date_counts.keys())
    chart_data = [date_counts[d] for d in chart_labels]

    return render_template(
        "admin/wiki_stats.html", 
        wiki=w, 
        subscribers=subscribers, 
        view_stats=view_stats, 
        sub_stats=sub_stats,
        chart_labels=chart_labels, # 传递给模板
        chart_data=chart_data      # 传递给模板
    )
    
@app.route("/admin/wikis/<int:wiki_id>/editors", methods=["GET", "POST"])
@login_required
def manage_editors(wiki_id):
    if not current_user.is_admin:
        abort(403)
    w = Wiki.query.get_or_404(wiki_id)
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        u = User.query.filter_by(username=username).first()
        if not u:
            flash("用户不存在")
            return redirect(url_for("manage_editors", wiki_id=wiki_id))
        if WikiEditor.query.filter_by(wiki_id=wiki_id, user_id=u.id).first() is None:
            e = WikiEditor(wiki_id=wiki_id, user_id=u.id)
            db.session.add(e)
            db.session.commit()
        return redirect(url_for("manage_editors", wiki_id=wiki_id))
    
    # Get all users who are editors for this wiki
    editors = User.query.join(WikiEditor, WikiEditor.user_id == User.id).filter(WikiEditor.wiki_id == wiki_id).all()
    return render_template("admin/manage_editors.html", wiki=w, editors=editors)

@app.route("/admin/wikis/<int:wiki_id>/editors/<int:user_id>/remove", methods=["POST"])
@login_required
def remove_editor(wiki_id, user_id):
    if not current_user.is_admin:
        abort(403)
        
    w = Wiki.query.get_or_404(wiki_id)
    editor = WikiEditor.query.filter_by(wiki_id=wiki_id, user_id=user_id).first_or_404()
    
    # Prevent removing the creator (optional safety, but good practice)
    if user_id == w.created_by_id:
        flash("无法移除 Wiki 创建者的编辑权限")
        return redirect(url_for("manage_editors", wiki_id=wiki_id))
        
    db.session.delete(editor)
    db.session.commit()
    flash("编辑权限已移除")
    return redirect(url_for("manage_editors", wiki_id=wiki_id))

@app.route("/admin/wikis/<int:wiki_id>/files", methods=["GET", "POST"])
@login_required
def manage_files(wiki_id):
    if not can_edit_wiki(wiki_id):
        abort(403)
        
    w = Wiki.query.get_or_404(wiki_id)
    files = WikiFile.query.filter_by(wiki_id=wiki_id).order_by(WikiFile.uploaded_at.desc()).all()
    
    return render_template("admin/wiki_files.html", wiki=w, files=files)

@app.route("/admin/wikis/<int:wiki_id>/files/upload", methods=["POST"])
@login_required
def upload_wiki_file(wiki_id):
    if not can_edit_wiki(wiki_id):
        return {"error": "Permission denied"}, 403
        
    if "image" not in request.files:
        return {"error": "No file part"}, 400
    file = request.files["image"]
    if file.filename == "":
        return {"error": "No selected file"}, 400
    if file:
        filename = secure_filename(file.filename)
        ext = os.path.splitext(filename)[1].lower()
        if ext not in [".jpg", ".jpeg", ".png", ".gif", ".webp"]:
             app.logger.error(f"Upload failed: Invalid extension '{ext}' for file '{filename}'")
             return {"error": f"File type '{ext}' not allowed. Allowed: jpg, png, gif, webp"}, 400
        unique_filename = str(uuid.uuid4()) + ext
        
        # Ensure uploads directory exists
        upload_dir = app.config["UPLOAD_FOLDER"]
        if not os.path.exists(upload_dir):
            os.makedirs(upload_dir)
            
        file.save(os.path.join(upload_dir, unique_filename))
        
        file_path = url_for("static", filename=f"uploads/{unique_filename}")
        
        # Record in DB
        wf = WikiFile(
            wiki_id=wiki_id, 
            filename=filename, 
            file_path=file_path, 
            uploaded_by_id=current_user.id
        )
        db.session.add(wf)
        db.session.commit()
        
        return {"data": {"filePath": file_path, "filename": filename}}

@app.route("/admin/wikis/<int:wiki_id>/files/<int:file_id>/delete", methods=["POST"])
@login_required
def delete_wiki_file(wiki_id, file_id):
    if not can_edit_wiki(wiki_id):
        abort(403)
        
    wf = WikiFile.query.get_or_404(file_id)
    if wf.wiki_id != wiki_id:
        abort(404)
        
    # Optional: Delete actual file from disk? 
    # For now, just remove DB record to keep it simple and safe
    # If deleting from disk, need to map URL back to file path
    
    db.session.delete(wf)
    db.session.commit()
    flash("文件已删除")
    return redirect(url_for("manage_files", wiki_id=wiki_id))

@app.route("/assets/markdown.css")
def markdown_css():
    s = MarkdownStyle.query.first()
    css = s.css_text if s else ""
    return Response(css, mimetype="text/css")

@app.route("/admin/markdown-css", methods=["GET", "POST"])
@login_required
def edit_markdown_css():
    if not current_user.is_admin:
        abort(403)
    s = MarkdownStyle.query.first()
    if request.method == "POST":
        css_text = request.form.get("css_text", "")
        s.css_text = css_text
        db.session.commit()
        return redirect(url_for("edit_markdown_css"))
    return render_template("admin/markdown_css.html", css_text=s.css_text if s else "")

@app.route("/upload", methods=["POST"])
@login_required
def upload_file():
    if "image" not in request.files:
        return {"error": "No file part"}, 400
    file = request.files["image"]
    if file.filename == "":
        return {"error": "No selected file"}, 400
    if file:
        filename = secure_filename(file.filename)
        ext = os.path.splitext(filename)[1].lower()
        if ext not in [".jpg", ".jpeg", ".png", ".gif", ".webp"]:
             app.logger.error(f"Upload failed: Invalid extension '{ext}' for file '{filename}'")
             return {"error": f"File type '{ext}' not allowed. Allowed: jpg, png, gif, webp"}, 400
        unique_filename = str(uuid.uuid4()) + ext
        file.save(os.path.join(app.config["UPLOAD_FOLDER"], unique_filename))
        # EasyMDE expects: {"success": 1, "url": "..."}
        return {"success": 1, "url": url_for("static", filename=f"uploads/{unique_filename}")}

@app.route("/admin/users")
@login_required
def manage_users():
    if not current_user.is_admin:
        abort(403)
    users = User.query.order_by(User.created_at.desc()).all()
    classes = Class.query.order_by(Class.name).all()
    return render_template("admin/manage_users.html", users=users, classes=classes)

@app.route("/admin/comments")
@login_required
def manage_comments():
    if not current_user.is_admin:
        abort(403)

    status = request.args.get("status", "pending")
    query = Comment.query.order_by(Comment.created_at.desc())

    if status == "approved":
        query = query.filter(Comment.is_approved.is_(True))
    elif status == "pending":
        query = query.filter(Comment.is_approved.is_(False))
    elif status != "all":
        status = "pending"
        query = query.filter(Comment.is_approved.is_(False))

    comments = query.all()
    counts = {
        "all": Comment.query.count(),
        "approved": Comment.query.filter(Comment.is_approved.is_(True)).count(),
        "pending": Comment.query.filter(Comment.is_approved.is_(False)).count(),
    }
    return render_template("admin/manage_comments.html", comments=comments, status=status, counts=counts)

@app.route("/admin/comments/<int:comment_id>/moderate", methods=["POST"])
@login_required
def moderate_comment(comment_id):
    if not current_user.is_admin:
        abort(403)

    c = Comment.query.get_or_404(comment_id)
    action = request.form.get("action", "").strip()

    if action == "approve":
        c.is_approved = True
        c.approved_at = now_utc8()
        c.approved_by_id = current_user.id
        flash("评论已通过审核", "success")
    elif action == "pending":
        c.is_approved = False
        c.approved_at = None
        c.approved_by_id = None
        flash("评论已打回待审核状态", "warning")
    else:
        flash("无效的审核操作", "error")
        return redirect(request.referrer or url_for("manage_comments"))

    db.session.commit()
    next_url = request.form.get("next", "").strip()
    if not next_url.startswith("/"):
        next_url = request.referrer or url_for("manage_comments")
    return redirect(next_url)

def get_phone_app_endpoint_choices():
    endpoints = []
    skip_prefixes = ("static", "api_", "service_")
    for rule in app.url_map.iter_rules():
        if "GET" not in rule.methods or "<" in rule.rule:
            continue
        if rule.endpoint.startswith(skip_prefixes):
            continue
        endpoints.append((rule.endpoint, rule.rule))
    return sorted(set(endpoints), key=lambda item: item[0])

def apply_phone_app_form(phone_app):
    phone_app.name = request.form.get("name", "").strip() or "未命名 App"
    phone_app.description = request.form.get("description", "").strip()
    phone_app.icon_class = request.form.get("icon_class", "").strip()
    phone_app.bg_color = request.form.get("bg_color", "").strip() or "#82d5bb"
    phone_app.action_type = request.form.get("action_type", "link").strip()
    phone_app.target_endpoint = request.form.get("target_endpoint", "").strip()
    phone_app.target_url = request.form.get("target_url", "").strip()
    phone_app.internal_key = request.form.get("internal_key", "").strip()
    phone_app.enabled = request.form.get("enabled") == "on"
    phone_app.admin_only = request.form.get("admin_only") == "on"

    if phone_app.action_type not in ["link", "internal"]:
        phone_app.action_type = "link"

    try:
        phone_app.sort_order = int(request.form.get("sort_order", 0))
    except ValueError:
        phone_app.sort_order = 0

    if request.form.get("remove_icon") == "on":
        remove_uploaded_static_file(phone_app.icon_url, "phone_apps")
        phone_app.icon_url = ""

    icon_file = request.files.get("icon_file")
    if icon_file and icon_file.filename:
        remove_uploaded_static_file(phone_app.icon_url, "phone_apps")
        phone_app.icon_url = save_uploaded_image(icon_file, "phone_apps", "phone_app")

    return phone_app

@app.route("/admin/phone-apps", methods=["GET", "POST"])
@login_required
def manage_phone_apps():
    if not current_user.is_admin:
        abort(403)

    if request.method == "POST":
        try:
            phone_app = apply_phone_app_form(PhoneApp())
            db.session.add(phone_app)
            db.session.commit()
            flash("手机 App 已创建", "success")
        except ValueError as error:
            flash(str(error), "error")
        return redirect(url_for("manage_phone_apps"))

    phone_apps = PhoneApp.query.order_by(PhoneApp.sort_order.asc(), PhoneApp.id.asc()).all()
    return render_template(
        "admin/manage_phone_apps.html",
        phone_apps=phone_apps,
        endpoint_choices=get_phone_app_endpoint_choices(),
        internal_app_meta=PHONE_INTERNAL_APP_META,
    )

@app.route("/admin/phone-apps/<int:app_id>", methods=["POST"])
@login_required
def update_phone_app(app_id):
    if not current_user.is_admin:
        abort(403)

    phone_app = PhoneApp.query.get_or_404(app_id)
    try:
        apply_phone_app_form(phone_app)
        db.session.commit()
        flash("手机 App 已更新", "success")
    except ValueError as error:
        flash(str(error), "error")
    return redirect(url_for("manage_phone_apps"))

@app.route("/admin/phone-apps/<int:app_id>/delete", methods=["POST"])
@login_required
def delete_phone_app(app_id):
    if not current_user.is_admin:
        abort(403)

    phone_app = PhoneApp.query.get_or_404(app_id)
    remove_uploaded_static_file(phone_app.icon_url, "phone_apps")
    db.session.delete(phone_app)
    db.session.commit()
    flash("手机 App 已删除", "success")
    return redirect(url_for("manage_phone_apps"))

@app.route("/admin/phone-apps/reset-defaults", methods=["POST"])
@login_required
def reset_phone_apps_defaults():
    if not current_user.is_admin:
        abort(403)

    PhoneApp.query.delete()
    db.session.commit()
    seed_default_phone_apps()
    flash("已恢复默认 3x3 手机 App", "success")
    return redirect(url_for("manage_phone_apps"))

@app.route("/admin/classes", methods=["GET", "POST"])
@login_required
def manage_classes():
    if not current_user.is_admin:
        abort(403)
    
    if request.method == "POST":
        name = request.form.get("name", "").strip()
        if name:
            if Class.query.filter_by(name=name).first():
                flash("班级已存在")
            else:
                c = Class(name=name)
                db.session.add(c)
                db.session.commit()
                flash("班级创建成功")
        return redirect(url_for("manage_classes"))
        
    classes = Class.query.order_by(Class.name).all()
    return render_template("admin/manage_classes.html", classes=classes)

@app.route("/admin/classes/<int:class_id>/edit", methods=["POST"])
@login_required
def edit_class(class_id):
    if not current_user.is_admin:
        abort(403)
    c = Class.query.get_or_404(class_id)
    name = request.form.get("name", "").strip()
    if name:
        if Class.query.filter(Class.name == name, Class.id != class_id).first():
            flash("班级名已存在")
        else:
            c.name = name
            # Update class_name in User table for consistency (legacy field)
            # User.query.filter_by(class_id=c.id).update({"class_name": name})
            # Actually, let's keep them in sync if we are still using class_name somewhere
            for u in c.users:
                u.class_name = name
            db.session.commit()
            flash("班级更新成功")
    return redirect(url_for("manage_classes"))

@app.route("/admin/classes/<int:class_id>/delete", methods=["POST"])
@login_required
def delete_class(class_id):
    if not current_user.is_admin:
        abort(403)
    c = Class.query.get_or_404(class_id)
    # Check if has users
    if c.users.count() > 0:
        flash("无法删除：该班级下还有学生")
    else:
        db.session.delete(c)
        db.session.commit()
        flash("班级已删除")
    return redirect(url_for("manage_classes"))


# ==========================================
# 班级成员管理 (编辑班级内的学生)
# ==========================================
@app.route("/admin/classes/<int:class_id>/members", methods=["GET", "POST"])
@login_required
def edit_class_members(class_id):
    if not current_user.is_admin:
        abort(403)
        
    c = Class.query.get_or_404(class_id)
    
    if request.method == "POST":
        # 1. 获取前端勾选的所有 user_id 列表
        selected_user_ids = request.form.getlist("members")
        
        # 2. 先将原来属于这个班级的学生移出（设置为无班级）
        # 这一步是为了处理那些被管理员取消勾选的学生
        for u in c.users:
            if str(u.id) not in selected_user_ids:
                u.class_id = None
                u.class_name = None
                
        # 3. 将新勾选的学生加入这个班级
        if selected_user_ids:
            users_to_add = User.query.filter(User.id.in_(selected_user_ids)).all()
            for u in users_to_add:
                u.class_id = c.id
                u.class_name = c.name # 保持冗余字段同步
                
        db.session.commit()
        flash(f"【{c.name}】班级成员更新成功！", "success")
        return redirect(url_for("manage_classes"))
        
    # GET 请求：渲染页面
    # 获取所有学生，按班级和学号排序，方便管理员查找
    all_users = User.query.order_by(User.class_id, User.student_id).all()
    
    # 获取当前班级里的学生 ID 列表，传给前端用于默认勾选
    class_user_ids = [u.id for u in c.users]
    
    return render_template("admin/edit_classes.html", current_class=c, all_users=all_users, class_user_ids=class_user_ids)


@app.route("/admin/groups", methods=["GET", "POST"])
@login_required
def manage_groups():
    if not current_user.is_admin:
        abort(403)
        
    if request.method == "POST":
        name = request.form.get("name", "").strip()
        description = request.form.get("description", "").strip()
        if name:
            if Group.query.filter_by(name=name).first():
                flash("组名已存在")
            else:
                g = Group(name=name, description=description)
                db.session.add(g)
                db.session.commit()
                flash("创建成功")
        return redirect(url_for("manage_groups"))
        
    groups = Group.query.order_by(Group.id.desc()).all()
    return render_template("admin/manage_groups.html", groups=groups)

@app.route("/admin/groups/<int:group_id>/edit", methods=["GET", "POST"])
@login_required
def edit_group(group_id):
    if not current_user.is_admin:
        abort(403)
    g = Group.query.get_or_404(group_id)
    
    if request.method == "POST":
        g.name = request.form.get("name", "").strip()
        g.description = request.form.get("description", "").strip()
        
        # Update members
        # We expect a list of user_ids
        member_ids = request.form.getlist("members")
        # Clear existing
        g.users = []
        if member_ids:
            users = User.query.filter(User.id.in_(member_ids)).all()
            g.users.extend(users)
            
        db.session.commit()
        flash("更新成功")
        return redirect(url_for("manage_groups"))
    
    # GET: Show edit form with member selection
    # We need all users to select from
    all_users = User.query.order_by(User.class_name, User.student_id).all()
    return render_template("admin/edit_group.html", group=g, all_users=all_users)

@app.route("/admin/groups/<int:group_id>/delete", methods=["POST"])
@login_required
def delete_group(group_id):
    if not current_user.is_admin:
        abort(403)
    g = Group.query.get_or_404(group_id)
    db.session.delete(g)
    db.session.commit()
    flash("删除成功")
    return redirect(url_for("manage_groups"))

@app.route("/admin/users/template")
@login_required
def download_user_template():
    if not current_user.is_admin:
        abort(403)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "User Import Template"
    
    headers = ['username', 'password', 'permission', 'email', 'real_name', 'student_id', 'department', 'class_name']
    ws.append(headers)
    ws.append(['test_user', '123456', 'user', 'test@example.com', '张三', '20230001', '计算机系', '计科1班'])
    ws.append(['admin_user', '123456', 'admin', 'admin@example.com', '李四', '20230002', '教师组', ''])
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = make_response(output.read())
    response.headers["Content-Disposition"] = "attachment; filename=user_import_template.xlsx"
    response.headers["Content-type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    return response

@app.route("/admin/users/import", methods=["POST"])
@login_required
def import_users():
    if not current_user.is_admin:
        abort(403)
    if "file" not in request.files:
        flash("未选择文件")
        return redirect(url_for("manage_users"))
    file = request.files["file"]
    if file.filename == "":
        flash("未选择文件")
        return redirect(url_for("manage_users"))
        
    filename = file.filename.lower()
    if not (filename.endswith('.csv') or filename.endswith('.xlsx') or filename.endswith('.xls')):
        flash("请上传CSV或Excel文件")
        return redirect(url_for("manage_users"))
        
    try:
        success_count = 0
        error_count = 0
        rows = []
        
        if filename.endswith('.csv'):
            stream = io.StringIO(file.stream.read().decode("UTF8"), newline=None)
            csv_input = csv.DictReader(stream)
            rows = list(csv_input)
        else:
            # Excel file
            wb = openpyxl.load_workbook(file)
            ws = wb.active
            # Assume first row is header
            headers = [cell.value for cell in ws[1]]
            # Map headers to expected keys (handle potential whitespace or case issues)
            # But for simplicity, assume strict matching first, or lowercase matching
            header_map = {str(h).strip().lower(): i for i, h in enumerate(headers) if h}
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                # Construct a dict similar to csv.DictReader
                row_data = {}
                for key, idx in header_map.items():
                    if idx < len(row):
                        row_data[key] = str(row[idx]) if row[idx] is not None else ''
                if row_data:
                    rows.append(row_data)

        for row in rows:
            # Map keys to standard keys if necessary, but template uses standard keys
            # CSV/Excel headers: username, password, permission, email, real_name, student_id, department, class_name
            
            # Helper to get value case-insensitively if needed, but template enforces keys.
            # Let's try to get values robustly
            def get_val(keys):
                for k in keys:
                    if k in row: return str(row[k]).strip()
                return ''
            
            username = get_val(['username'])
            password = get_val(['password'])
            permission = get_val(['permission']).lower()
            email = get_val(['email']).lower()
            real_name = get_val(['real_name', '姓名'])
            student_id = get_val(['student_id', '学号'])
            department = get_val(['department', '系部'])
            class_name = get_val(['class_name', '班级'])
            
            if not username or not password:
                error_count += 1
                continue
                
            # Check for duplicates (username, email, or student_id)
            query_filter = (User.username == username) | (User.email == email)
            if student_id:
                query_filter = query_filter | (User.student_id == student_id)
                
            if User.query.filter(query_filter).first():
                error_count += 1
                continue
                
            u = User(username=username, email=email)
            u.set_password(password)
            u.is_admin = (permission == 'admin')
            u.real_name = real_name
            u.student_id = student_id if student_id else None
            u.department = department
            u.class_name = class_name
            
            if class_name:
                cls = Class.query.filter_by(name=class_name).first()
                if not cls:
                    cls = Class(name=class_name)
                    db.session.add(cls)
                    db.session.flush()
                u.class_id = cls.id
            
            db.session.add(u)
            success_count += 1
        db.session.commit()
        flash(f"导入完成: 成功 {success_count} 个, 失败/重复 {error_count} 个")
    except Exception as e:
        flash(f"导入出错: {str(e)}")
    return redirect(url_for("manage_users"))

@app.route("/admin/users/<int:user_id>/role", methods=["POST"])
@login_required
def update_user_role(user_id):
    if not current_user.is_admin:
        abort(403)
    u = User.query.get_or_404(user_id)
    if u.id == current_user.id:
        flash("不能修改自己的权限")
        return redirect(url_for("manage_users"))
    role = request.form.get("role")
    if role == "admin":
        u.is_admin = True
    else:
        u.is_admin = False
    db.session.commit()
    flash(f"用户 {u.username} 权限已更新")
    return redirect(url_for("manage_users"))

@app.route("/admin/users/<int:user_id>/class", methods=["POST"])
@login_required
def update_user_class(user_id):
    if not current_user.is_admin:
        abort(403)
    u = User.query.get_or_404(user_id)
    class_id = request.form.get("class_id")
    
    if class_id:
        c = Class.query.get(class_id)
        if c:
            u.class_id = c.id
            u.class_name = c.name # Keep legacy field in sync
        else:
            u.class_id = None
            u.class_name = None
    else:
        u.class_id = None
        u.class_name = None
        
    db.session.commit()
    flash(f"用户 {u.username} 班级已更新")
    return redirect(url_for("manage_users"))

@app.route("/admin/users/<int:user_id>/reset_password", methods=["POST"])
@login_required
def reset_user_password(user_id):
    if not current_user.is_admin:
        abort(403)
        
    u = User.query.get_or_404(user_id)
    u.set_password("123456")
    db.session.commit()
    
    flash(f"用户 {u.username} 的密码已重置为 123456")
    return redirect(url_for("manage_users"))


@app.route("/admin/users/<int:user_id>/edit_info", methods=["POST"])
@login_required
def edit_user_info(user_id):
    """编辑用户基础信息"""
    if not current_user.is_admin:
        abort(403)
        
    u = User.query.get_or_404(user_id)
    
    new_username = request.form.get("username", "").strip()
    new_email = request.form.get("email", "").strip()
    
    # 防止用户名或邮箱与其他人重复
    if User.query.filter(User.username == new_username, User.id != u.id).first():
        flash("更新失败：用户名已存在", "error")
        return redirect(url_for("manage_users"))
    if User.query.filter(User.email == new_email, User.id != u.id).first():
        flash("更新失败：邮箱已被注册", "error")
        return redirect(url_for("manage_users"))
        
    u.username = new_username
    u.email = new_email
    u.real_name = request.form.get("real_name", "").strip()
    u.student_id = request.form.get("student_id", "").strip()
    u.department = request.form.get("department", "").strip()
    
    db.session.commit()
    flash(f"用户 {u.username} 的基础信息已更新", "success")
    return redirect(url_for("manage_users"))

@app.route("/admin/users/<int:user_id>/delete", methods=["POST"])
@login_required
def delete_user(user_id):
    """彻底删除用户及其关联的大部分日志数据"""
    if not current_user.is_admin:
        abort(403)
        
    if user_id == current_user.id:
        flash("安全拦截：不能在后台删除当前登录的自己！", "error")
        return redirect(url_for("manage_users"))
        
    u = User.query.get_or_404(user_id)
    
    try:
        # 手动清理一些没有设置级联删除的关联数据，防止外键约束报错
        Comment.query.filter_by(user_id=u.id).delete()
        WikiViewLog.query.filter_by(user_id=u.id).delete()
        NoteViewLog.query.filter_by(user_id=u.id).delete()
        UserActiveStatus.query.filter_by(user_id=u.id).delete()
        StudySession.query.filter_by(user_id=u.id).delete()
        PomodoroRecord.query.filter_by(user_id=u.id).delete()
        
        db.session.delete(u)
        db.session.commit()
        flash(f"用户 {u.username} 已被彻底删除", "success")
    except Exception as e:
        db.session.rollback()
        flash(f"删除失败，存在强关联数据: {str(e)}", "error")
        
    return redirect(url_for("manage_users"))

# ==========================================
# 1. 徽章管理页路由 (新建与列表展示)
# ==========================================
@app.route("/admin/badges", methods=["GET", "POST"])
@login_required
def manage_badges():
    if not current_user.is_admin:
        abort(403)
        
    if request.method == "POST":
        name = request.form.get("name", "").strip()
        description = request.form.get("description", "").strip()
        
        # Handle file upload
        icon = ""
        if "icon_file" in request.files:
            file = request.files["icon_file"]
            if file and file.filename:
                # Validate extension
                if not file.filename.lower().endswith(('.png', '.jpg', '.jpeg', '.gif', '.webp')):
                    flash("不支持的文件格式")
                    return redirect(url_for("manage_badges"))
                
                # Save file
                filename = str(uuid.uuid4()) + os.path.splitext(file.filename)[1]
                save_dir = os.path.join(app.root_path, "static/uploads/badges")
                if not os.path.exists(save_dir):
                    os.makedirs(save_dir)
                
                file.save(os.path.join(save_dir, filename))
                icon = f"/static/uploads/badges/{filename}"
        
        condition_type = request.form.get("condition_type", "").strip()
        try:
            condition_value = int(request.form.get("condition_value", 0))
        except ValueError:
            condition_value = 0
            
        try:
            sticker_count = int(request.form.get("sticker_count", 1))
        except ValueError:
            sticker_count = 1
            
        total_limit_str = request.form.get("total_limit", "").strip()
        total_limit = None
        if total_limit_str:
            try:
                total_limit = int(total_limit_str)
            except ValueError:
                pass
            
        is_hidden = request.form.get("is_hidden") == "on"
        is_secret = request.form.get("is_secret") == "on"
        category = request.form.get("category", "一般").strip() or "一般"
        rarity = request.form.get("rarity", "common").strip()
        issuer = request.form.get("issuer", "WikiBook").strip() or "WikiBook"
        
        # 💡 新增：获取自定义序号前缀
        serial_prefix = request.form.get("serial_prefix", "SF").strip().upper() or "SF"
        custom_condition_text = request.form.get("custom_condition_text", "").strip()

        # 💡 新增：获取 target_id (来自下拉菜单的选择)
        target_id_str = request.form.get("target_id", "").strip()
        target_id = int(target_id_str) if target_id_str.isdigit() else None
        
        # Parse dates
        start_time = None
        end_time = None
        if request.form.get("start_time"):
            try:
                start_time = datetime.strptime(request.form.get("start_time"), "%Y-%m-%d")
            except ValueError:
                pass
        if request.form.get("end_time"):
            try:
                end_time = datetime.strptime(request.form.get("end_time"), "%Y-%m-%d")
            except ValueError:
                pass
            
        if not name or not icon or not condition_type:
            flash("请填写完整信息（包括上传图标）")
            return redirect(url_for("manage_badges"))
            
        b = Badge(
            name=name,
            description=description,
            icon=icon,
            condition_type=condition_type,
            condition_value=condition_value,
            sticker_count=sticker_count,
            category=category,
            rarity=rarity,
            issuer=issuer,
            serial_prefix=serial_prefix,
            custom_condition_text=custom_condition_text,
            total_limit=total_limit,
            is_hidden=is_hidden,
            is_secret=is_secret,
            start_time=start_time,
            end_time=end_time,
            target_id=target_id  # 💡 保存 target_id
        )
        db.session.add(b)
        db.session.commit()
        
        # 💡 修改：全员普发时的序列号生成
        if b.condition_type == 'all_users':
            users = User.query.all()
            for user in users:
                if b.total_limit is not None and b.issued_count >= b.total_limit:
                    break
                    
                if not UserBadge.query.filter_by(user_id=user.id, badge_id=b.id).first():
                    ub = UserBadge(user_id=user.id, badge_id=b.id)
                    ub.serial_number = f"{b.serial_prefix}-{str(b.issued_count + 1).zfill(3)}"
                    db.session.add(ub)
                    b.issued_count += 1
                    
            db.session.commit()
            
        flash("徽章创建成功")
        return redirect(url_for("manage_badges"))
        
    # 自动校准发放数量
    all_badges_for_sync = Badge.query.all()
    needs_commit = False
    for b in all_badges_for_sync:
        actual_count = b.users.count()
        if b.issued_count != actual_count:
            b.issued_count = actual_count
            needs_commit = True
    if needs_commit:
        db.session.commit()

    # 筛选逻辑
    condition_filter = request.args.get('condition_type', 'all')
    query = Badge.query

    if condition_filter != 'all':
        query = query.filter_by(condition_type=condition_filter)

    badges = query.order_by(Badge.created_at.desc()).all()
    users = User.query.all()

    # 💡 1. 后端直接算出总发放量，避免前端 Jinja2 循环
    total_issued = sum(b.issued_count for b in badges)
    
    # 💡 2. 预先处理好每个徽章的已获得用户 ID，避免前端发生 N+1 数据库查询导致卡死
    # 建立一个字典，key 是 badge.id，value 是 'id1,id2,id3' 格式的字符串
    badge_earned_map = {}
    for b in badges:
        if b.condition_type == 'manual':
            # 只为手动发放的徽章查询用户列表，因为其他徽章前端用不到这个数据
            earned_ids = [str(ub.user_id) for ub in b.users.all()]
            badge_earned_map[b.id] = ",".join(earned_ids)
        else:
            badge_earned_map[b.id] = ""

    existing_conditions = [c[0] for c in db.session.query(Badge.condition_type).distinct().all() if c[0]]

    condition_map = {
        'manual': '手动发放',
        'all_users': '全员获取',
        'login_days_in_range': '限时登录天数',
        'streak_days': '连续打卡',
        'study_hours': '累计学习时长',
        'note_count': '累计笔记',
        'featured_count': '精选笔记',
        'wiki_edit_count': 'Wiki编辑',
        'wiki_create_count': 'Wiki创建',
        'comment_count': '评论数量',
        'night_owl_sessions': '深夜学习',
        'early_bird': '早起鸟',
        'weekend_warrior': '周末卷王',
        'long_session_count': '深度专注',
        'share_count': '分享笔记',
        'total_views_received': '笔记被阅读',
        'pomo_count': '番茄达人',
        # 💡 新增映射映射
        'user_level': '用户等级达标',
        'wiki_specific_duration': '特定Wiki学习时长',
        'assignment_complete': '完成指定任务',
        'assignment_stars': '任务实得星星',
        'assignment_grade': '任务等第要求',
        **OJ_BADGE_CONDITION_LABELS,
    }

    # 💡 提取全部任务和 Wiki 供下拉菜单选择
    assignments = Assignment.query.order_by(Assignment.created_at.desc()).all()
    wikis = Wiki.query.order_by(Wiki.created_at.desc()).all()
    oj_assignments = OJAssignment.query.order_by(OJAssignment.created_at.desc()).all()

    return render_template("admin/manage_badges.html", 
                           badges=badges, 
                           users=users,
                           total_issued=total_issued,            # 传给前端
                           badge_earned_map=badge_earned_map,    # 传给前端
                           current_filter=condition_filter,
                           existing_conditions=existing_conditions,
                           condition_map=condition_map,
                           assignments=assignments,
                           wikis=wikis,
                           oj_assignments=oj_assignments)

@app.route("/admin/badges/<int:badge_id>/edit", methods=["GET", "POST"])
@login_required
def edit_badge(badge_id):
    if not current_user.is_admin:
        abort(403)
        
    b = Badge.query.get_or_404(badge_id)
    
    if request.method == "POST":
        b.name = request.form.get("name", "").strip()
        b.description = request.form.get("description", "").strip()
        
        # Handle file upload
        if "icon_file" in request.files:
            file = request.files["icon_file"]
            if file and file.filename:
                if not file.filename.lower().endswith(('.png', '.jpg', '.jpeg', '.gif', '.webp')):
                    flash("不支持的文件格式")
                    return render_template(
                        "admin/edit_badge.html",
                        badge=b,
                        users=User.query.order_by(User.class_name, User.student_id).all(),
                        earned_user_ids=[ub.user_id for ub in b.users.all()],
                        assignments=Assignment.query.order_by(Assignment.created_at.desc()).all(),
                        wikis=Wiki.query.order_by(Wiki.created_at.desc()).all(),
                        oj_assignments=OJAssignment.query.order_by(OJAssignment.created_at.desc()).all(),
                    )
                
                filename = str(uuid.uuid4()) + os.path.splitext(file.filename)[1]
                save_dir = os.path.join(app.root_path, "static/uploads/badges")
                if not os.path.exists(save_dir):
                    os.makedirs(save_dir)
                
                new_path = os.path.join(save_dir, filename)
                file.save(new_path)
                
                if b.icon and b.icon.startswith("/static/uploads/badges/"):
                    old_filename = b.icon.split("/")[-1]
                    old_path = os.path.join(save_dir, old_filename)
                    if os.path.exists(old_path):
                        try:
                            os.remove(old_path)
                        except OSError:
                            pass 
                
                b.icon = f"/static/uploads/badges/{filename}"

        b.condition_type = request.form.get("condition_type", "").strip()
        try:
            b.condition_value = int(request.form.get("condition_value", 0))
        except ValueError:
            b.condition_value = 0
            
        try:
            b.sticker_count = int(request.form.get("sticker_count", 1))
        except ValueError:
            b.sticker_count = 1
            
        total_limit_str = request.form.get("total_limit", "").strip()
        if total_limit_str:
            try:
                b.total_limit = int(total_limit_str)
            except ValueError:
                pass
        else:
            b.total_limit = None
            
        b.is_hidden = request.form.get("is_hidden") == "on"
        b.is_secret = request.form.get("is_secret") == "on"
        b.category = request.form.get("category", "一般").strip() or "一般"
        b.rarity = request.form.get("rarity", "common").strip()
        b.issuer = request.form.get("issuer", "WikiBook").strip() or "WikiBook"
        
        new_prefix = request.form.get("serial_prefix", "SF").strip().upper() or "SF"
        b.serial_prefix = new_prefix
        b.custom_condition_text = request.form.get("custom_condition_text", "").strip()
        
        # 💡 新增：获取 target_id
        target_id_str = request.form.get("target_id", "").strip()
        b.target_id = int(target_id_str) if target_id_str.isdigit() else None

        if request.form.get("start_time"):
            try: b.start_time = datetime.strptime(request.form.get("start_time"), "%Y-%m-%d")
            except ValueError: pass
        else: b.start_time = None
            
        if request.form.get("end_time"):
            try: b.end_time = datetime.strptime(request.form.get("end_time"), "%Y-%m-%d")
            except ValueError: pass
        else: b.end_time = None
        
        db.session.flush() # 先推送基本信息
        
        # 💡 洗号防撞车机制：先全员打上临时随机 UUID，再赋予真实序号
        existing_ubs = UserBadge.query.filter_by(badge_id=b.id).order_by(UserBadge.earned_at.asc()).all()
        
        for ub in existing_ubs:
            ub.serial_number = f"TMP-{uuid.uuid4().hex}"
        db.session.flush() # 彻底释放原来占用的 001, 002
        
        for idx, ub in enumerate(existing_ubs):
            ub.serial_number = f"{b.serial_prefix}-{str(idx + 1).zfill(3)}"
            
        b.issued_count = len(existing_ubs)
        db.session.flush()
        
        if b.condition_type == 'all_users':
            users = User.query.all()
            for user in users:
                if b.total_limit is not None and b.issued_count >= b.total_limit: break
                if not UserBadge.query.filter_by(user_id=user.id, badge_id=b.id).first():
                    ub = UserBadge(user_id=user.id, badge_id=b.id)
                    ub.serial_number = f"{b.serial_prefix}-{str(b.issued_count + 1).zfill(3)}"
                    db.session.add(ub)
                    b.issued_count += 1
            db.session.commit()
            flash("徽章基本信息更新完毕，已为所有用户发放。")
        
        elif b.condition_type != 'manual':
            users = User.query.all()
            for user in users:
                meets_condition = False
                val = b.condition_value
                tid = b.target_id
                
                # 各类条件判定逻辑...
                if b.condition_type == 'login_days_in_range':
                    if b.start_time and b.end_time:
                        sessions = StudySession.query.filter(StudySession.user_id == user.id, StudySession.start_time >= b.start_time, StudySession.start_time <= b.end_time).all()
                        meets_condition = len(set([s.start_time.date() for s in sessions])) >= val
                elif b.condition_type == 'note_count':
                    meets_condition = Note.query.filter_by(user_id=user.id).count() >= val
                elif b.condition_type == 'featured_count':
                    meets_condition = Note.query.filter_by(user_id=user.id, is_featured=True).count() >= val
                elif b.condition_type == 'wiki_edit_count':
                    meets_condition = WikiPageHistory.query.filter_by(user_id=user.id).count() >= val
                elif b.condition_type == 'comment_count':
                    meets_condition = Comment.query.filter_by(user_id=user.id).count() >= val
                elif b.condition_type == 'night_owl_sessions':
                    sessions = StudySession.query.filter_by(user_id=user.id).all()
                    night_days = set()
                    for s in sessions:
                        start_time = to_utc8_naive(s.start_time)
                        if start_time and (start_time.hour >= 23 or start_time.hour < 4):
                            night_days.add((start_time - timedelta(hours=4)).date())
                    meets_condition = len(night_days) >= val
                elif b.condition_type == 'early_bird':
                    sessions = StudySession.query.filter_by(user_id=user.id).all()
                    early_days = set()
                    for s in sessions:
                        start_time = to_utc8_naive(s.start_time)
                        if start_time and 5 <= start_time.hour < 8:
                            early_days.add(start_time.date())
                    meets_condition = len(early_days) >= val
                elif b.condition_type == 'weekend_warrior':
                    sessions = StudySession.query.filter_by(user_id=user.id).all()
                    total_seconds = sum(
                        study_session_seconds(s)
                        for s in sessions
                        if to_utc8_naive(s.start_time) and to_utc8_naive(s.start_time).weekday() in [5, 6]
                    )
                    meets_condition = (total_seconds / 3600) >= val
                elif b.condition_type == 'long_session_count':
                    sessions = StudySession.query.filter_by(user_id=user.id).all()
                    long_count = sum(1 for s in sessions if study_session_seconds(s) >= 7200)
                    meets_condition = long_count >= val
                elif b.condition_type == 'share_count':
                    meets_condition = NoteShare.query.join(Note).filter(Note.user_id == user.id).count() >= val
                elif b.condition_type == 'total_views_received':
                    meets_condition = NoteViewLog.query.join(Note).filter(Note.user_id == user.id).count() >= val
                elif b.condition_type == 'wiki_create_count':
                    from sqlalchemy import func
                    subquery = db.session.query(func.min(WikiPageHistory.id)).group_by(WikiPageHistory.page_id)
                    meets_condition = WikiPageHistory.query.filter(WikiPageHistory.id.in_(subquery), WikiPageHistory.user_id == user.id).count() >= val
                elif b.condition_type == 'study_hours':
                    sessions = StudySession.query.filter_by(user_id=user.id).all()
                    total_seconds = sum(study_session_seconds(s) for s in sessions)
                    meets_condition = (total_seconds / 3600) >= val
                elif b.condition_type == 'pomo_count':
                    meets_condition = PomodoroRecord.query.filter_by(user_id=user.id).count() >= val
                elif b.condition_type == 'streak_days':
                    sessions = StudySession.query.filter_by(user_id=user.id).order_by(StudySession.start_time.desc()).all()
                    dates = sorted(list(set([s.start_time.date() for s in sessions])), reverse=True)
                    streak = 0
                    if dates:
                        today = now_utc8().date()
                        if dates[0] == today: streak = 1; current = today
                        elif dates[0] == today - timedelta(days=1): streak = 1; current = today - timedelta(days=1)
                        if streak > 0:
                            for i in range(1, len(dates)):
                                if dates[i] == current - timedelta(days=1): streak += 1; current = dates[i]
                                else: break
                    meets_condition = streak >= val
                    
                # 💡 新的重算规则补充
                elif b.condition_type == 'user_level':
                    meets_condition = user.level >= val
                elif b.condition_type == 'wiki_specific_duration':
                    if tid:
                        logs_count = WikiViewLog.query.filter_by(user_id=user.id, wiki_id=tid).count()
                        meets_condition = (logs_count * 10) >= val
                elif b.condition_type == 'assignment_complete':
                    if tid:
                        sub = Submission.query.filter_by(user_id=user.id, assignment_id=tid, status='graded').first()
                        meets_condition = sub is not None
                elif b.condition_type == 'assignment_stars':
                    if tid:
                        sub = Submission.query.filter_by(user_id=user.id, assignment_id=tid, status='graded').first()
                        meets_condition = sub and (sub.stars_earned or 0) >= val
                elif b.condition_type == 'assignment_grade':
                    if tid:
                        sub = Submission.query.filter_by(user_id=user.id, assignment_id=tid, status='graded').first()
                        if sub and sub.grade:
                            grade_map = {'S+': 6, 'S': 5, 'A': 4, 'B': 3, 'C': 2, 'D': 1}
                            meets_condition = grade_map.get(sub.grade.upper(), 0) >= val
                elif b.condition_type in OJ_BADGE_CONDITION_LABELS:
                    meets_condition = meets_oj_badge_condition(user, b)

                existing_ub = UserBadge.query.filter_by(user_id=user.id, badge_id=b.id).first()
                
                if meets_condition:
                    if not existing_ub:
                        if b.total_limit is not None and b.issued_count >= b.total_limit: continue 
                        ub = UserBadge(user_id=user.id, badge_id=b.id)
                        ub.serial_number = f"{b.serial_prefix}-{str(b.issued_count + 1).zfill(3)}"
                        db.session.add(ub)
                        b.issued_count += 1
                else:
                    if existing_ub:
                        if user.selected_badge_id == b.id: user.selected_badge_id = None
                        db.session.delete(existing_ub)
                        if b.issued_count > 0: b.issued_count -= 1
            
            db.session.commit()
            flash("徽章已更新，所有序号已重排，并完成了全员进度校准！")
        else:
            db.session.commit()
            flash("徽章基本信息更新完毕，已重排历史序列号。")
            
        return redirect(url_for("manage_badges"))
    
    users = User.query.order_by(User.class_name, User.student_id).all()
    earned_user_ids = [ub.user_id for ub in b.users.all()]
    
    # 💡 提取供下拉菜单用的数据
    assignments = Assignment.query.order_by(Assignment.created_at.desc()).all()
    wikis = Wiki.query.order_by(Wiki.created_at.desc()).all()
    oj_assignments = OJAssignment.query.order_by(OJAssignment.created_at.desc()).all()

    return render_template("admin/edit_badge.html", 
                           badge=b, 
                           users=users, 
                           earned_user_ids=earned_user_ids,
                           assignments=assignments,
                           wikis=wikis,
                           oj_assignments=oj_assignments)

@app.route("/admin/badges/<int:badge_id>/award", methods=["POST"])
@login_required
def award_badge_manually(badge_id):
    if not current_user.is_admin:
        abort(403)
        
    b = Badge.query.get_or_404(badge_id)
    if b.condition_type != 'manual':
        return jsonify({"error": "Only manual badges can be manually awarded"}), 400
        
    data = request.get_json()
    if not data or "user_ids" not in data:
        return jsonify({"error": "Missing user_ids in request"}), 400

    # 安全提取整数ID
    target_user_ids = [int(uid) for uid in data.get("user_ids", []) if str(uid).strip().isdigit()]
    
    current_ubs = UserBadge.query.filter_by(badge_id=b.id).all()
    current_user_ids = [ub.user_id for ub in current_ubs]
    
    to_add_ids = set(target_user_ids) - set(current_user_ids)
    to_remove_ids = set(current_user_ids) - set(target_user_ids)
    
    prefix = b.serial_prefix or "SF"
    
    try:
        # 1. ORM 级安全撤销
        if to_remove_ids:
            remove_list = list(to_remove_ids)
            ubs_to_remove = UserBadge.query.filter(UserBadge.badge_id == b.id, UserBadge.user_id.in_(remove_list)).all()
            for ub in ubs_to_remove:
                db.session.delete(ub)
                
            users_to_update = User.query.filter(User.selected_badge_id == b.id, User.id.in_(remove_list)).all()
            for u in users_to_update:
                u.selected_badge_id = None
                
            db.session.flush() # 同步到数据库
            
        # 2. 新增发放
        added_count = 0
        for uid in to_add_ids:
            current_total = b.issued_count - len(to_remove_ids) + added_count
            if b.total_limit is not None and current_total >= b.total_limit:
                break 
                
            ub = UserBadge(user_id=uid, badge_id=b.id)
            ub.serial_number = f"TMP-{uuid.uuid4().hex[:8]}" # 防撞车占位符
            db.session.add(ub)
            added_count += 1
            
        db.session.flush()

        b.issued_count = b.issued_count - len(to_remove_ids) + added_count
        
        # 3. 绝对安全：重新连号重排
        if to_remove_ids or to_add_ids:
            all_left = UserBadge.query.filter_by(badge_id=b.id).order_by(UserBadge.earned_at.asc()).all()
            
            # 第一阶段：全量变临时占位符
            for item in all_left:
                item.serial_number = f"TMP-{uuid.uuid4().hex}"
            db.session.flush()
            
            # 第二阶段：重新发放 001, 002...
            for idx, item in enumerate(all_left):
                item.serial_number = f"{prefix}-{str(idx + 1).zfill(3)}"
                
        db.session.commit()
        return jsonify({"success": True, "count": b.issued_count})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({"success": False, "error": str(e)}), 500

@app.route("/admin/badges/<int:badge_id>/delete", methods=["POST"])
@login_required
def delete_badge(badge_id):
    if not current_user.is_admin:
        abort(403)
    b = Badge.query.get_or_404(badge_id)
    
    # Cascade delete: Delete all UserBadges associated with this badge
    UserBadge.query.filter_by(badge_id=b.id).delete()
    
    # Also clear selected_badge_id from users who equipped it
    User.query.filter_by(selected_badge_id=b.id).update({"selected_badge_id": None})
    
    # Delete icon file if it's a local upload
    if b.icon and b.icon.startswith("/static/uploads/badges/"):
        filename = b.icon.split("/")[-1]
        file_path = os.path.join(app.root_path, "static/uploads/badges", filename)
        if os.path.exists(file_path):
            try:
                os.remove(file_path)
            except OSError:
                pass

    db.session.delete(b)
    db.session.commit()
    flash("徽章已删除")
    return redirect(url_for("manage_badges"))

@app.route("/user/profile")
@login_required
def user_profile():
    return redirect(url_for('public_profile', user_id=current_user.id))

@app.route("/user/<int:user_id>")
@login_required
def public_profile(user_id):
    user = User.query.get_or_404(user_id)
    
    # Stats
    stats = user.calculate_stats()
    
    # Total Users for Global Rarity
    total_users = User.query.count() or 1
    
    # Badges
    user_badges = UserBadge.query.filter_by(user_id=user.id).order_by(UserBadge.earned_at.desc()).all()
    earned_badge_ids = [ub.badge_id for ub in user_badges]
    all_badges = Badge.query.all()
    unearned_badges = [b for b in all_badges if b.id not in earned_badge_ids]
    
    # 💡 新增：获取用户已收集的书签
    user_bookmarks = UserBookmark.query.filter_by(user_id=user.id).order_by(UserBookmark.earned_at.desc()).all()
    
    # Recent Activity (Notes)
    recent_notes = Note.query.filter_by(user_id=user.id).order_by(Note.created_at.desc()).limit(5).all()

    # Check follow status
    is_following = current_user.is_following(user)
    is_study_partner = current_user.is_following(user) and user.is_following(current_user)

    return render_template("user/profile.html", 
                           user=user, 
                           stats=stats, 
                           user_badges=user_badges, 
                           unearned_badges=unearned_badges, 
                           user_bookmarks=user_bookmarks, # 💡 传入书签数据
                           recent_notes=recent_notes, 
                           is_following=is_following, 
                           is_study_partner=is_study_partner)

# 💡 新增：书签图鉴全屏页面路由
@app.route("/user/<int:user_id>/bookmark_book")
@login_required
def bookmark_book(user_id):
    user = User.query.get_or_404(user_id)
    
    # 获取已获得的书签
    user_bookmarks = UserBookmark.query.filter_by(user_id=user.id).order_by(UserBookmark.earned_at.desc()).all()
    earned_type_ids = [ub.bookmark_type_id for ub in user_bookmarks]
    
    # 获取未获得的书签
    all_bookmarks = BookmarkType.query.all()
    unearned_bookmarks = [b for b in all_bookmarks if b.id not in earned_type_ids]
    
    return render_template("user/bookmark_book.html", 
                           user=user,
                           user_bookmarks=user_bookmarks,
                           unearned_bookmarks=unearned_bookmarks)

@app.route("/user/<int:user_id>/achievement_book")
@login_required
def achievement_book(user_id):
    user = User.query.get_or_404(user_id)
    
    # 获取用户已获得的徽章（按获得时间倒序）
    user_badges = UserBadge.query.filter_by(user_id=user.id).order_by(UserBadge.earned_at.desc()).all()
    earned_badge_ids = [ub.badge_id for ub in user_badges]
    
    # 获取未获得的徽章
    all_badges = Badge.query.order_by(Badge.category.asc(), Badge.id.asc()).all()
    unearned_badges = [b for b in all_badges if b.id not in earned_badge_ids]
    
    # 获取所有分类（去重）
    categories = db.session.query(Badge.category).filter(Badge.category.isnot(None)).distinct().all()
    # categories is a list of tuples: [('一般',), ('学习',), ...]
    category_list = sorted(
        [c[0] for c in categories if c[0]],
        key=lambda name: (name != '一般', name)
    )
    
    # Total Users for Global Rarity
    total_users = User.query.count() or 1

    def serialize_badge_item(badge, earned_record=None):
        is_image = is_image_icon(badge.icon)
        earned_at = earned_record.earned_at if earned_record else None
        return {
            "id": badge.id,
            "name": badge.name,
            "description": badge.description or "",
            "category": badge.category or "一般",
            "rarity": badge.rarity or "common",
            "isSecret": bool(badge.is_secret),
            "isHidden": bool(badge.is_hidden),
            "isUnlocked": earned_record is not None,
            "unlockDate": earned_at.strftime('%Y.%m.%d') if earned_at else "",
            "unlockTimestamp": earned_at.isoformat() if earned_at else "",
            "isImage": is_image,
            "iconUrl": badge_icon_url(badge.icon) if is_image else "",
            "iconText": "" if is_image else badge.icon,
            "issuer": badge.issuer or "WikiBook",
            "serialNumber": earned_record.serial_number if earned_record else "",
            "conditionText": format_badge_condition(badge),
        }

    achievement_payload = {
        "user": {
            "id": user.id,
            "username": user.username,
            "avatarInitial": (user.username[:1] if user.username else "?").upper(),
        },
        "categories": category_list,
        "achievements": (
            [serialize_badge_item(ub.badge, ub) for ub in user_badges] +
            [serialize_badge_item(badge) for badge in unearned_badges]
        ),
    }
    achievement_payload_json = json.dumps(achievement_payload, ensure_ascii=False)
    
    return render_template("user/achievement_book.html", 
                         user=user,
                         user_badges=user_badges,
                         unearned_badges=unearned_badges,
                         categories=category_list,
                         total_users=total_users,
                         achievement_payload=achievement_payload,
                         achievement_payload_json=achievement_payload_json)

@app.route("/my/achievement_book")
@login_required
def my_achievement_book():
    return redirect(url_for("achievement_book", user_id=current_user.id))

@app.route("/user/<int:user_id>/follow", methods=["POST"])
@login_required
def follow_user(user_id):
    user = User.query.get_or_404(user_id)
    if user.id == current_user.id:
        return jsonify({"error": "不能关注自己"}), 400
        
    if not current_user.is_following(user):
        current_user.follow(user)
        db.session.commit()
        
        # Check for study partner (mutual follow)
        if user.is_following(current_user):
            # Notify both
            n1 = Notification(
                user_id=current_user.id,
                message=f"你和 {user.username} 成为了学友！",
                type='system',
                link=url_for('public_profile', user_id=user.id)
            )
            n2 = Notification(
                user_id=user.id,
                message=f"你和 {current_user.username} 成为了学友！",
                type='system',
                link=url_for('public_profile', user_id=current_user.id)
            )
            db.session.add_all([n1, n2])
            db.session.commit()
            flash(f"关注成功！你和 {user.username} 现在是学友了！")
        else:
            flash(f"已关注 {user.username}")
            
    return redirect(request.referrer or url_for('public_profile', user_id=user.id))

@app.route("/user/<int:user_id>/unfollow", methods=["POST"])
@login_required
def unfollow_user(user_id):
    user = User.query.get_or_404(user_id)
    if current_user.is_following(user):
        current_user.unfollow(user)
        db.session.commit()
        flash(f"已取消关注 {user.username}")
        
    return redirect(request.referrer or url_for('public_profile', user_id=user.id))

@app.route("/user/following")
@login_required
def my_following():
    # List of users I follow
    # We can also separate "Study Partners" vs "Just Following"
    
    # Study Partners (Mutual)
    partners = current_user.study_partners.all()
    partner_ids = [u.id for u in partners]
    
    # Following (excluding partners)
    following_all = current_user.followed.all()
    following_only = [u for u in following_all if u.id not in partner_ids]
    
    # Followers (who I don't follow back) - Optional, user didn't explicitly ask to see list, but good to have
    # Requirement: "自己的关注列表是可以点击查看的" -> My Following List
    
    return render_template("user/following.html", partners=partners, following=following_only)

@app.route("/api/notifications/read/all", methods=["POST"])
@login_required
def read_all_notifications():
    current_user.notifications.filter_by(is_read=False).update({"is_read": True})
    db.session.commit()
    return jsonify({"success": True})

@app.route("/user/badges")
@login_required
def my_badges():
    # Earned badges
    user_badges = UserBadge.query.filter_by(user_id=current_user.id).order_by(UserBadge.earned_at.desc()).all()
    earned_badge_ids = [ub.badge_id for ub in user_badges]
    
    # Unearned badges
    all_badges = Badge.query.all()
    unearned_badges = [b for b in all_badges if b.id not in earned_badge_ids]
    
    return render_template("user/my_badges.html", user_badges=user_badges, unearned_badges=unearned_badges)

@app.route("/user/badges/equip", methods=["POST"])
@login_required
def equip_badge():
    badge_id = request.json.get("badge_id")
    
    if badge_id is None:
        # Unequip
        current_user.selected_badge_id = None
        db.session.commit()
        return jsonify({"success": True, "message": "已卸下徽章"})
        
    # Check if user owns this badge
    ub = UserBadge.query.filter_by(user_id=current_user.id, badge_id=badge_id).first()
    if not ub:
        return jsonify({"success": False, "error": "您尚未获得该徽章"}), 403
        
    current_user.selected_badge_id = badge_id
    db.session.commit()
    return jsonify({"success": True, "message": "徽章佩戴成功"})

# Book Module Routes
@app.route("/book")
@login_required
def book_index():
    q = request.args.get("q", "").strip()
    
    # 精选笔记
    featured_notes = Note.query.filter_by(is_featured=True).order_by(Note.updated_at.desc()).all()
    
    # Last Edited Note (for Hero Section)
    last_note = Note.query.filter_by(user_id=current_user.id).order_by(Note.updated_at.desc()).first()
    
    if q:
        # 1. Title Matches
        my_title = Note.query.filter(
            Note.user_id == current_user.id,
            Note.title.contains(q)
        ).all()
        for n in my_title: n.source_type = 'my'
        
        shared_title = Note.query.join(NoteShare).filter(
            NoteShare.user_id == current_user.id,
            Note.title.contains(q)
        ).all()
        for n in shared_title: n.source_type = 'shared'
        
        title_matches = sorted(my_title + shared_title, key=lambda x: x.updated_at, reverse=True)
        
        # 2. Content Matches (exclude title matches)
        my_content = Note.query.filter(
            Note.user_id == current_user.id,
            Note.content_md.contains(q),
            ~Note.title.contains(q)
        ).all()
        for n in my_content: n.source_type = 'my'
        
        shared_content = Note.query.join(NoteShare).filter(
            NoteShare.user_id == current_user.id,
            Note.content_md.contains(q),
            ~Note.title.contains(q)
        ).all()
        for n in shared_content: n.source_type = 'shared'
        
        content_matches = sorted(my_content + shared_content, key=lambda x: x.updated_at, reverse=True)
        
        return render_template("book/index.html", 
                             featured_notes=featured_notes, 
                             q=q, 
                             title_matches=title_matches, 
                             content_matches=content_matches)
    else:
        my_notes = Note.query.filter_by(user_id=current_user.id).order_by(Note.updated_at.desc()).all()
        shared_notes = Note.query.join(NoteShare).filter(NoteShare.user_id == current_user.id).order_by(NoteShare.created_at.desc()).all()
    
        return render_template("book/index.html", my_notes=my_notes, shared_notes=shared_notes, featured_notes=featured_notes, q=q, last_note=last_note)

@app.route("/api/notes/search")
@login_required
def api_search_notes():
    q = request.args.get("q", "").strip()
    if not q:
        return jsonify({"results": []})
        
    # Search my notes
    my_notes = Note.query.filter(
        Note.user_id == current_user.id,
        (Note.title.contains(q)) | (Note.content_md.contains(q))
    ).limit(5).all()
    
    # Search shared notes
    shared_notes = Note.query.join(NoteShare).filter(
        NoteShare.user_id == current_user.id,
        (Note.title.contains(q)) | (Note.content_md.contains(q))
    ).limit(5).all()
    
    results = []
    for n in my_notes:
        results.append({
            "id": n.id,
            "title": n.title,
            "type": "my",
            "icon": n.icon
        })
    for n in shared_notes:
        results.append({
            "id": n.id,
            "title": n.title,
            "type": "shared",
            "user": n.user.username,
            "icon": n.icon
        })
        
    return jsonify({"results": results})


@app.route("/api/wikis/search")
def api_search_wikis():
    q = request.args.get("q", "").strip()
    if not q:
        return jsonify({"results": []})
        
    results = []
    
    # 1. Search Wikis (Top 3)
    wikis = Wiki.query.filter(Wiki.title.contains(q)).limit(3).all()
    for w in wikis:
        results.append({
            "type": "wiki",
            "id": w.id,
            "title": w.title,
            "desc": w.description
        })
        
    # 2. Search Pages (Top 5)
    # Join with Wiki to get wiki title
    # Search both title and content
    pages = WikiPage.query.join(Wiki).filter(
        (WikiPage.title.contains(q)) | (WikiPage.content_md.contains(q))
    ).limit(5).all()
    
    for p in pages:
        # Extract snippet if match is in content
        snippet = ""
        if q.lower() in (p.content_md or "").lower():
            # Find the position
            idx = (p.content_md or "").lower().find(q.lower())
            start = max(0, idx - 20)
            end = min(len(p.content_md), idx + 60)
            snippet = p.content_md[start:end] + "..."
            
        results.append({
            "type": "page",
            "wiki_id": p.wiki_id,
            "wiki_title": p.wiki.title,
            "slug": p.slug,
            "title": p.title,
            "snippet": snippet
        })
        
    return jsonify({"results": results})


@app.route("/book/received")
@login_required
def book_received():
    q = request.args.get("q", "").strip()
    
    if q:
        note_shares = NoteShare.query.join(Note).filter(
            NoteShare.user_id == current_user.id,
            (Note.title.contains(q)) | (Note.content_md.contains(q))
        ).order_by(NoteShare.created_at.desc()).all()
    else:
        note_shares = NoteShare.query.filter_by(user_id=current_user.id).order_by(NoteShare.created_at.desc()).all()
        
    return render_template("book/received.html", note_shares=note_shares, q=q)

@app.route("/online_users")
@login_required
def online_users():
    online_threshold = now_utc8() - timedelta(minutes=5)
    users = UserActiveStatus.query.filter(UserActiveStatus.last_active_at > online_threshold).all()
    return render_template("online_users.html", users=users)


@app.route("/book/notes")
@login_required
def my_notes():
    q = request.args.get("q", "").strip()
    
    # Base query for my notes
    query = Note.query.filter_by(user_id=current_user.id)
    
    if q:
        query = query.filter((Note.title.contains(q)) | (Note.content_md.contains(q)))
    
    notes = query.order_by(Note.updated_at.desc()).all()
    
    # Statistics
    stats = {
        "total_notes": Note.query.filter_by(user_id=current_user.id).count(),
        # 我累计阅读的笔记次数（不仅是我的笔记，也包括别人的）
        "total_reads": NoteViewLog.query.filter_by(user_id=current_user.id).count(),
        # 我的笔记被阅读的总次数（影响力）
        "total_views_received": NoteViewLog.query.join(Note).filter(Note.user_id == current_user.id).count(),
        # 我收到的分享数
        "shared_received": NoteShare.query.filter_by(user_id=current_user.id).count(),
        # Study stats
        "today": 0,
        "week": 0,
        "month": 0,
        "three_months": 0,
        "six_months": 0,
        "year": 0,
        "total": 0
    }
    
    # Calculate study stats
    now = now_utc8()
    today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
    week_start = now - timedelta(days=7)
    month_start = now - timedelta(days=30)
    three_months_start = now - timedelta(days=90)
    six_months_start = now - timedelta(days=180)
    year_start = now - timedelta(days=365)
    
    sessions = StudySession.query.filter_by(user_id=current_user.id).all()
    for s in sessions:
        duration = study_session_seconds(s)
        if duration <= 0:
            continue
        start_time = to_utc8_naive(s.start_time)
        if not start_time:
            continue
        
        stats["total"] += duration
        
        if start_time >= today_start:
            stats["today"] += duration
        if start_time >= week_start:
            stats["week"] += duration
        if start_time >= month_start:
            stats["month"] += duration
        if start_time >= three_months_start:
            stats["three_months"] += duration
        if start_time >= six_months_start:
            stats["six_months"] += duration
        if start_time >= year_start:
            stats["year"] += duration
            
    # Convert seconds to hours/minutes string
    def format_duration(seconds):
        m, s = divmod(seconds, 60)
        h, m = divmod(m, 60)
        if h > 0:
            return f"{int(h)}小时 {int(m)}分钟"
        elif m > 0:
            return f"{int(m)}分钟"
        else:
            return "少于1分钟"
            
    # Update stats with formatted duration, keeping original integer stats
    stats_formatted = stats.copy()
    for k in ["today", "week", "month", "three_months", "six_months", "year", "total"]:
        stats_formatted[k + "_formatted"] = format_duration(stats[k])
    
    return render_template("book/my_notes.html", notes=notes, stats=stats_formatted, q=q)

@app.route("/api/stats/heatmap")
@login_required
def get_heatmap_data():
    # Weights for different activities
    WEIGHTS = {
        'create': 5,  # Creating Note or Wiki History (Edit)
        'comment': 3,
        'session': 2,
        'view': 1
    }
    
    activity_counts = {}

    def process_query(model, date_col, weight=1):
        # Retrieve all history, not just one year
        results = db.session.query(date_col).filter(
            model.user_id == current_user.id
        ).all()
        for (dt,) in results:
            if dt:
                # Adjust to UTC+8 for correct date bucket
                # Assuming stored dates are naive and represent UTC or already UTC+8?
                # The app uses now_utc8() for defaults. 
                # If the column is DateTime, SQLAlchemy returns Python datetime objects.
                # If they were stored as UTC+8 (which they are per now_utc8), then the date part is correct.
                d = dt.strftime("%Y-%m-%d")
                activity_counts[d] = activity_counts.get(d, 0) + weight

    # 1. Notes created (High weight)
    process_query(Note, Note.created_at, WEIGHTS['create'])
    
    # 2. Comments (Medium weight)
    process_query(Comment, Comment.created_at, WEIGHTS['comment'])
    
    # 3. Wiki Edits (High weight)
    process_query(WikiPageHistory, WikiPageHistory.created_at, WEIGHTS['create'])
    
    # 4. Study Sessions (Medium weight)
    process_query(StudySession, StudySession.start_time, WEIGHTS['session'])

    # 5. Reading (Low weight)
    process_query(WikiViewLog, WikiViewLog.timestamp, WEIGHTS['view'])
    process_query(NoteViewLog, NoteViewLog.timestamp, WEIGHTS['view'])
    
    data = [[date, count] for date, count in activity_counts.items()]
    return jsonify({"data": data})

@app.route("/book/calendar")
@login_required
def book_calendar():
    return render_template("book/calendar.html")

@app.route("/api/notes/calendar")
@login_required
def get_calendar_notes():
    notes = Note.query.filter_by(user_id=current_user.id).all()
    events = []
    for note in notes:
        # Create a snippet from markdown content (strip basic md chars)
        content_preview = note.content_md[:200] if note.content_md else ""
        
        events.append({
            "id": note.id,
            "title": note.title,
            "start": note.created_at.strftime("%Y-%m-%d"),
            "extendedProps": {
                "content": content_preview,
                "tags": [t.name for t in note.tags],
                "created_at": note.created_at.strftime("%Y-%m-%d %H:%M")
            }
        })
    return jsonify(events)

@app.route("/book/notes/new", methods=["GET", "POST"])
@login_required
def new_note():
    if request.method == "POST":
        title = request.form.get("title", "").strip()
        content_md = request.form.get("content_md", "")
        tags_str = request.form.get("tags", "")
        
        if not title:
            flash("标题不能为空")
            return redirect(url_for("new_note"))
            
        n = Note(title=title, content_md=content_md, user_id=current_user.id)
        
        # Process tags
        n.tags = process_tags(tags_str)
        
        db.session.add(n)
        db.session.commit()
        
        # Check badges
        check_and_award_badges(current_user)
        
        return redirect(url_for("view_note", note_id=n.id))
    return render_template("book/new_note.html")

@app.route("/book/notes/<int:note_id>")
@login_required
def view_note(note_id):
    n = Note.query.get_or_404(note_id)
    
    # Check permission (owner or shared or featured)
    is_owner = n.user_id == current_user.id
    is_shared = NoteShare.query.filter_by(note_id=note_id, user_id=current_user.id).first() is not None
    is_featured = n.is_featured
    
    if not (is_owner or is_shared or current_user.is_admin or is_featured):
        abort(403)
        
    # Record view log with debounce (e.g., 5 minutes)
    try:
        last_view = NoteViewLog.query.filter_by(note_id=note_id, user_id=current_user.id).order_by(NoteViewLog.timestamp.desc()).first()
        last_view_time = to_utc8_naive(last_view.timestamp) if last_view else None
        if not last_view_time or datetime_diff_seconds(now_utc8(), last_view_time) > 300:
            log = NoteViewLog(note_id=note_id, user_id=current_user.id)
            db.session.add(log)
            db.session.commit()
            
            # Check badges for NOTE OWNER (Influencer badge)
            if not is_owner:
                owner = User.query.get(n.user_id)
                if owner:
                    check_and_award_badges(owner)
                    
    except Exception as e:
        # Ignore logging errors to not affect user experience
        print(f"Error logging view: {e}")
        
    html = markdown(n.content_md or "")
    all_users = User.query.filter(User.id != current_user.id).all() if is_owner or current_user.is_admin else []
    visible_comments = get_visible_comments(note_id=n.id)

    return render_template(
        "book/view_note.html",
        note=n,
        html=html,
        is_owner=is_owner,
        all_users=all_users,
        visible_comments=visible_comments
    )

@app.route("/book/notes/<int:note_id>/edit", methods=["GET", "POST"])
@login_required
def edit_note(note_id):
    n = Note.query.get_or_404(note_id)

    # 💡 核心锁定逻辑：
    # 检查该笔记是否有关联的 Submission 且状态为 'graded'
    lock_submission = Submission.query.filter_by(source_note_id=n.id, status='graded').first()
    if lock_submission:
        flash("该文档关联的作业已批改完成，目前已锁定不可编辑", "warning")
        return redirect(url_for("view_note", note_id=n.id))

    if n.user_id != current_user.id and not current_user.is_admin:
        abort(403)
        
    if request.method == "POST":
        # ==========================================
        # 💡 新增：手账模式 (JSON 数据提交)
        # ==========================================
        if request.is_json:
            data = request.json
            n.title = data.get("title", n.title).strip()
            
            # 只要包含 scrapbook_data，就锁定为手账模式
            if "scrapbook_data" in data:
                n.is_scrapbook = True
                n.scrapbook_data = data.get("scrapbook_data")
            else:
                n.is_scrapbook = False
                n.content_md = data.get("content_md", n.content_md)
                
            db.session.commit()
            
            # 发奖判定
            owner = User.query.get(n.user_id)
            if owner:
                check_and_award_badges(owner)
                
            return jsonify({"success": True})
            
        # ==========================================
        # 💡 传统模式：普通 Markdown 表单提交
        # ==========================================
        else:
            n.title = request.form.get("title", "").strip()
            n.content_md = request.form.get("content_md", "")
            tags_str = request.form.get("tags", "")
            
            # 如果通过传统 Markdown 编辑器保存，强制切回普通模式
            n.is_scrapbook = False
            
            # Update tags
            n.tags = process_tags(tags_str)
            
            # Update settings
            n.comment_enabled = request.form.get("comment_enabled") == "on"
            
            # 仅管理员可以设置精选
            if current_user.is_admin:
                n.is_featured = request.form.get("is_featured") == "on"
                
            db.session.commit()
            
            # Check badges for note owner (if admin featured it)
            owner = User.query.get(n.user_id)
            if owner:
                check_and_award_badges(owner)
                
            return redirect(url_for("view_note", note_id=n.id))
            
    return render_template("book/edit_note.html", note=n)

@app.route("/book/notes/<int:note_id>/comment", methods=["POST"])
@login_required
def comment_note(note_id):
    n = Note.query.get_or_404(note_id)
    # Check permission (same as view)
    is_owner = n.user_id == current_user.id
    is_shared = NoteShare.query.filter_by(note_id=note_id, user_id=current_user.id).first() is not None
    is_featured = n.is_featured
    
    if not (is_owner or is_shared or current_user.is_admin or is_featured):
        abort(403)
        
    content = request.form.get("content", "").strip()
    if content:
        c = Comment(
            content=content,
            user_id=current_user.id,
            note_id=n.id,
            is_approved=False
        )
        db.session.add(c)
        db.session.commit()
        
        # Check badges
        check_and_award_badges(current_user)
        
        flash("评论已提交，等待管理员审核通过后会展示给其他用户。", "info")
    else:
        flash("评论内容不能为空", "warning")
        
    return redirect(url_for("view_note", note_id=n.id))

@app.route("/book/notes/<int:note_id>/delete", methods=["POST"])
@login_required
def delete_note(note_id):
    n = Note.query.get_or_404(note_id)
    if n.user_id != current_user.id and not current_user.is_admin:
        abort(403)
    db.session.delete(n)
    db.session.commit()
    flash("笔记已删除")
    return redirect(url_for("book_index"))

@app.route("/book/notes/<int:note_id>/share", methods=["POST"])
@login_required
def share_note(note_id):
    n = Note.query.get_or_404(note_id)
    if n.user_id != current_user.id and not current_user.is_admin:
        abort(403)
        
    user_ids = request.form.getlist("user_ids")
    
    # Simple logic: Add new shares, ignore existing ones
    count = 0
    for uid in user_ids:
        try:
            uid = int(uid)
            if uid == n.user_id: continue
            
            exists = NoteShare.query.filter_by(note_id=n.id, user_id=uid).first()
            if not exists:
                share = NoteShare(note_id=n.id, user_id=uid)
                db.session.add(share)
                count += 1
        except ValueError:
            continue
            
    db.session.commit()
    
    # Check badges for current user (Share count)
    check_and_award_badges(current_user)
    
    flash(f"已分享给 {count} 位用户")
    return redirect(url_for("view_note", note_id=n.id))

@app.route("/book/notes/quick_create", methods=["POST"])
@login_required
def quick_create_note():
    data = request.json
    content = data.get("content", "").strip()
    custom_title = data.get("title", "").strip()
    
    if not content and not custom_title:
        return jsonify({"success": False, "error": "内容或标题不能为空"}), 400
        
    try:
        # Title: YYYY-MM-DD HH:mm:ss if not provided
        title = custom_title if custom_title else now_utc8().strftime("%Y-%m-%d %H:%M:%S")
        n = Note(title=title, content_md=content, user_id=current_user.id)
        db.session.add(n)
        db.session.commit()
        
        # Check badges
        check_and_award_badges(current_user)
        
        return jsonify({"success": True, "note_id": n.id})
    except Exception as e:
        db.session.rollback()
        return jsonify({"success": False, "error": str(e)}), 500

@app.route("/book/share/<int:share_id>/delete", methods=["POST"])
@login_required
def delete_share(share_id):
    share = NoteShare.query.get_or_404(share_id)
    
    # 只能删除自己的分享记录
    if share.user_id != current_user.id:
        return jsonify({"success": False, "error": "无权删除"}), 403
    
    try:
        db.session.delete(share)
        db.session.commit()
        return jsonify({"success": True})
    except Exception as e:
        db.session.rollback()
        return jsonify({"success": False, "error": str(e)}), 500

@app.route("/book/share/<int:share_id>/convert", methods=["POST"])
@login_required
def convert_note(share_id):
    share = NoteShare.query.get_or_404(share_id)
    
    # 只能转换自己的分享记录
    if share.user_id != current_user.id:
        return jsonify({"success": False, "error": "无权转换"}), 403
    
    try:
        original_note = share.note
        
        # 创建新的笔记，标题包含发送人姓名
        new_title = f"来自{original_note.user.username}的笔记：{original_note.title}"
        new_note = Note(
            title=new_title,
            content_md=original_note.content_md,
            user_id=current_user.id,
            is_featured=False
        )
        
        db.session.add(new_note)
        
        # 删除分享记录
        db.session.delete(share)
        
        db.session.commit()
        return jsonify({"success": True})
    except Exception as e:
        db.session.rollback()
        return jsonify({"success": False, "error": str(e)}), 500

@app.route("/wiki/<int:wiki_id>/page/<string:slug>/convert", methods=["POST"])
@login_required
def convert_wiki_page(wiki_id, slug):
    wiki = Wiki.query.get_or_404(wiki_id)
    if not can_view_wiki(wiki_id):
        return jsonify({"success": False, "error": "无权访问"}), 403
        
    page = WikiPage.query.filter_by(wiki_id=wiki_id, slug=slug).first_or_404()
    
    try:
        new_title = f"[Wiki] {page.title}"
        new_note = Note(
            title=new_title,
            content_md=page.content_md,
            user_id=current_user.id,
            is_featured=False
        )
        
        db.session.add(new_note)
        db.session.commit()
        return jsonify({"success": True, "note_id": new_note.id})
    except Exception as e:
        db.session.rollback()
        return jsonify({"success": False, "error": str(e)}), 500

@app.route("/api/link-preview")
def link_preview():
    url = request.args.get("url")
    if not url:
        return {"error": "Missing URL"}, 400
    
    try:
        headers = {"User-Agent": "Mozilla/5.0 (compatible; LinkPreviewBot/1.0)"}
        resp = requests.get(url, headers=headers, timeout=5)
        resp.raise_for_status()
        
        soup = BeautifulSoup(resp.content, "html.parser")
        
        title = soup.title.string if soup.title else ""
        og_title = soup.find("meta", property="og:title")
        if og_title and og_title.get("content"): title = og_title.get("content")
            
        description = ""
        og_desc = soup.find("meta", property="og:description")
        if og_desc:
            description = og_desc.get("content", "")
        else:
            meta_desc = soup.find("meta", attrs={"name": "description"})
            if meta_desc: description = meta_desc.get("content", "")
            
        image = ""
        og_image = soup.find("meta", property="og:image")
        if og_image:
            image = og_image.get("content", "")
            if image and not image.startswith("http"):
                from urllib.parse import urljoin
                image = urljoin(url, image)
        
        if not image:
             link_icon = soup.find("link", rel="icon") or soup.find("link", rel="shortcut icon")
             if link_icon:
                 image = link_icon.get("href", "")
                 from urllib.parse import urljoin
                 image = urljoin(url, image)

        return {
            "title": title.strip() if title else url,
            "description": description.strip()[:200] + "..." if len(description) > 200 else description.strip(),
            "image": image,
            "url": url
        }
    except Exception as e:
        # Return basic info on error
        return {
            "title": url,
            "description": "No preview available",
            "image": "",
            "url": url
        }

@app.route("/api/heartbeat", methods=["POST"])
@login_required
def heartbeat():
    try:
        data = request.get_json() or {}
        path = data.get("path", "/")
        
        # 容错：即使前端传来非法数据也能安全转换
        pomo_state = data.get("pomo_state", "IDLE")
        try:
            pomo_end_time = float(data.get("pomo_end_time", 0))
        except (ValueError, TypeError):
            pomo_end_time = 0.0
        
        # 💡 状态推断引擎：按优先级从高到低匹配
        action = "正在潜水"
        if "/admin" in path: action = "后台管理中"
        elif "/settings" in path: action = "修改偏好设置"
        elif "/leaderboard" in path: action = "查看巅峰排行榜"
        elif "/sticker_walls" in path: action = "欣赏贴纸墙"
        elif "/announcements" in path: action = "阅读系统公告"
        elif "achievement_book" in path: action = "翻阅成就图鉴"
        elif "/user/following" in path: action = "查看学友列表"
        elif "/user/" in path: action = "访问个人主页"
        elif "/online_users" in path: action = "查看在线用户"
        elif "/book/calendar" in path: action = "回顾学习日历"
        elif "/book/received" in path: action = "查看收到的笔记"
        elif "/edit" in path: action = "编辑创作中..."
        elif "/new" in path: action = "新建内容中..."
        elif "/notes/" in path: action = "阅读精选笔记"
        elif "/pages/" in path: action = "阅读 Wiki 百科"
        elif "/book" in path: action = "浏览个人笔记库"
        elif "/wikis/" in path: action = "浏览 Wiki 知识库"
        elif path == "/" or path == "": action = "在广场闲逛"
        
        # 番茄钟状态拥有最高显示优先级
        if pomo_state == "WORK": action = "🍅 专注中..."
        elif pomo_state == "REST": action = "☕ 休息中..."

        now = now_utc8()
        
        status = UserActiveStatus.query.get(current_user.id)
        if not status:
            status = UserActiveStatus(user_id=current_user.id)
            db.session.add(status)
        
        was_offline = False
        online_threshold = now - timedelta(minutes=5)
        last_active_at = to_utc8_naive(status.last_active_at)
        if last_active_at and last_active_at < online_threshold:
            was_offline = True
        elif last_active_at is None:
            was_offline = True
            
        status.last_active_at = now
        status.current_path = path
        status.current_action = action
        status.pomo_state = pomo_state
        status.pomo_end_time = pomo_end_time
        
        # 💡 只在自己刚才处于离线，并且学友目前【在线】时，才互相发送问候
        if was_offline:
            partners = current_user.study_partners.all()
            for p in partners:
                if p.is_online:
                    n = Notification(
                        user_id=p.id,
                        message=f"你的学友 {current_user.username} 上线了",
                        type='friend_login',
                        link=url_for('public_profile', user_id=current_user.id)
                    )
                    db.session.add(n)
        
        session_record = StudySession.query.filter_by(user_id=current_user.id).order_by(StudySession.end_time.desc()).first()
        gap_limit = timedelta(minutes=5)
        session_end_time = to_utc8_naive(session_record.end_time) if session_record else None
        
        if session_record and session_end_time and datetime_diff_seconds(now, session_end_time) < gap_limit.total_seconds():
            session_record.end_time = now
        else:
            session_record = StudySession(user_id=current_user.id, start_time=now, end_time=now)
            db.session.add(session_record)
            
        # 暴力销毁自己没看到的历史垃圾通知 (2分钟前的统统丢弃)
        Notification.query.filter(
            Notification.user_id == current_user.id, 
            Notification.type == 'friend_login',
            Notification.created_at < now - timedelta(minutes=2)
        ).delete()
        
        db.session.commit()
        
        # 容错：防止发奖出错引发全局崩溃
        try:
            check_and_award_badges(current_user)
        except Exception:
            db.session.rollback()
        
        notifications = []
        unread_notifs = current_user.notifications.filter_by(is_read=False).order_by(Notification.created_at.desc()).all()
        for n in unread_notifs:
            notifications.append({
                "id": n.id,
                "message": n.message,
                "type": n.type,
                "link": n.link,
                "created_at": n.created_at.strftime("%H:%M")
            })

        has_new_badges = UserBadge.query.filter_by(user_id=current_user.id, is_notified=False).count() > 0
        has_new_bookmarks = UserBookmark.query.filter_by(user_id=current_user.id, is_notified=False).count() > 0

        return jsonify({
            "status": "ok", 
            "notifications": notifications, 
            "has_new_badges": has_new_badges,
            "has_new_bookmarks": has_new_bookmarks
        })
        
    except Exception as e:
        db.session.rollback()
        # 出错时返回 JSON 而不是抛出 500 HTML 页面，保护前端不崩溃
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route("/api/bookmarks/unnotified", methods=["GET"])
@login_required
def api_unnotified_bookmarks():
    # 查询所有未被通知的新书签
    ubs = UserBookmark.query.filter_by(user_id=current_user.id, is_notified=False).all()
    if not ubs:
        return jsonify([])
    
    data = []
    for ub in ubs:
        ub.is_notified = True  # 阅后即焚，标记为已通知
        b_type = ub.bookmark_type
        is_image = is_image_icon(b_type.icon)
        
        data.append({
            "name": b_type.name,
            "icon": badge_icon_url(b_type.icon) if is_image else b_type.icon,
            "is_image": is_image,
            "rarity": b_type.rarity
        })
        
    db.session.commit()
    return jsonify(data)


@app.route("/api/badges/unnotified", methods=["GET"])
@login_required
def api_unnotified_badges():
    ubs = UserBadge.query.filter_by(user_id=current_user.id, is_notified=False).all()
    if not ubs:
        return jsonify([])
    
    data = []
    for ub in ubs:
        ub.is_notified = True 
        icon_val = ub.badge.icon
        is_image = is_image_icon(icon_val)
        
        # 安全获取时间，防止旧数据没有时间导致 500 报错
        time_str = ub.earned_at.strftime("%y-%m-%d") if ub.earned_at else now_utc8().strftime("%y-%m-%d")
        
        data.append({
            "name": ub.badge.name,
            "description": ub.badge.description,
            "icon": badge_icon_url(icon_val) if is_image else icon_val,
            "is_image": is_image,
            "earned_at": time_str,
            "serial_number": ub.serial_number or f"SF-{int(time.time())}",
            "user_name": ub.user.username,
            "avatar_char": ub.user.username[0].upper() if ub.user.username else "U",
            "rarity": ub.badge.rarity  # <--- 新增：把稀有度传给前端
        })
        
    db.session.commit()
    return jsonify(data)

@app.route("/api/notes/new_view", methods=["POST"])
@login_required
def log_note_view():
    # Hook for 'total_views_received' badge of the NOTE OWNER
    # Triggered when someone views a note. 
    # But currently view log is in view_note route (GET).
    # We can check badges there.
    pass

@app.context_processor
def inject_helpers():
    # Count online users (active in last 5 mins)
    online_threshold = now_utc8() - timedelta(minutes=5)
    online_count = UserActiveStatus.query.filter(UserActiveStatus.last_active_at > online_threshold).count()
    
    def get_badge_usage(user_id, badge_id):
        return UserSticker.query.filter_by(user_id=user_id, badge_id=badge_id).count()
    
    # 💡 新增：动态计算全局贴纸上限 (基础 10 张 + 完成的委托数量)
    def get_global_sticker_limit(user):
        if not user.is_authenticated:
            return 10
        completed_tasks = Submission.query.filter_by(user_id=user.id, status='graded').count()
        return 10 + completed_tasks
    
    return dict(
        can_view_wiki=can_view_wiki, 
        can_edit_wiki=can_edit_wiki,
        online_user_count=online_count,
        get_badge_usage=get_badge_usage,
        get_global_sticker_limit=get_global_sticker_limit,
        get_phone_apps=get_visible_phone_apps,
        phone_app_url=resolve_phone_app_url,
        phone_internal_app_meta=PHONE_INTERNAL_APP_META
    )




# ==========================================
# 🔖 书签自动化发奖核心引擎 (宽容匹配版)
# ==========================================
def check_and_award_bookmarks(user):
    # 1. 查找当前用户所有已有的书签类型 ID
    earned_ubs = UserBookmark.query.filter_by(user_id=user.id).all()
    earned_type_ids = [ub.bookmark_type_id for ub in earned_ubs]
    
    # 2. 获取所有该用户还没获得的书签模板
    available_types = BookmarkType.query.filter(BookmarkType.id.notin_(earned_type_ids)).all()
    awarded_count = 0
    
    for b_type in available_types:
        if b_type.condition_type == 'manual':
            continue

        earned = False
        val = b_type.condition_value
        ctype = b_type.condition_type  # 从数据库读取出的实际条件文本
        
        try:
            # 💡 条件 1：累计番茄专注 (次)
            if ctype in ['pomo_count']:
                count = PomodoroRecord.query.filter_by(user_id=user.id).count()
                if count >= val: earned = True
                
            # 💡 条件 2：累计专注时长 (分钟)
            elif ctype in ['pomo_duration', 'study_hours', 'study_minutes']:
                records = PomodoroRecord.query.filter_by(user_id=user.id).all()
                total_mins = sum([r.duration_minutes for r in records if r.duration_minutes])
                # 注意：如果原来 condition_type 叫 study_hours，这里要换算成小时
                if ctype == 'study_hours':
                    if (total_mins / 60.0) >= val: earned = True
                else:
                    if total_mins >= val: earned = True
                
            # 💡 条件 3：连续打卡登录 (天)
            elif ctype in ['streak_days']:
                sessions = StudySession.query.filter_by(user_id=user.id).order_by(StudySession.start_time.desc()).all()
                dates = sorted(list(set([s.start_time.date() for s in sessions])), reverse=True)
                streak = 0
                if dates:
                    today = now_utc8().date()
                    if dates[0] == today: 
                        streak = 1
                        current = today
                    elif dates[0] == today - timedelta(days=1): 
                        streak = 1
                        current = today - timedelta(days=1)
                    if streak > 0:
                        for i in range(1, len(dates)):
                            if dates[i] == current - timedelta(days=1): 
                                streak += 1
                                current = dates[i]
                            else: break
                if streak >= val: earned = True
                
            # 💡 条件 4：沉淀私人笔记达 (篇) - 兼容 book_note_count 变量
            elif ctype in ['note_count', 'book_note_count']:
                count = Note.query.filter_by(user_id=user.id).count()
                if count >= val: earned = True
                
            # 💡 条件 5：创建公开Wiki (个)
            elif ctype in ['wiki_create_count', 'public_wiki_count']:
                count = Wiki.query.filter_by(created_by_id=user.id, visibility='public').count()
                if count >= val: earned = True
                
            # 💡 条件 6：Wiki获得订阅数 (次)
            elif ctype in ['wiki_subscriptions', 'wiki_subscribe_count']:
                count = Subscription.query.join(Wiki).filter(Wiki.created_by_id == user.id).count()
                if count >= val: earned = True
                
            # 💡 条件 7：便签转化为笔记 (次)
            elif ctype in ['sticky_to_note']:
                # 通过匹配 quick_create 自动生成的默认标题作为依据
                count = Note.query.filter(Note.user_id == user.id, Note.title.like('%便签灵感归档%')).count()
                if count >= val: earned = True
                
            # 💡 条件 8：单篇Wiki阅读达 (分钟)
            elif ctype in ['single_wiki_read_time', 'wiki_read_time']:
                sessions = StudySession.query.filter_by(user_id=user.id).all()
                max_duration = 0
                for s in sessions:
                    dur = study_session_seconds(s) / 60.0
                    if dur > max_duration: max_duration = dur
                if max_duration >= val: earned = True
                
        except Exception as e:
            print(f"[书签引擎解析警告] 未知条件 {ctype}，报错: {e}")

        # 💡 执行发奖
        if earned:
            # 再次确认防止重复发放
            existing_ub = UserBookmark.query.filter_by(user_id=user.id, bookmark_type_id=b_type.id).first()
            if not existing_ub:
                ub = UserBookmark(user_id=user.id, bookmark_type_id=b_type.id, is_notified=False)
                db.session.add(ub)
                awarded_count += 1

    if awarded_count > 0:
        try:
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            print(f"[书签发奖出错] {e}")

# ==========================================
# 3. 自动触发发奖核心服务 (用户动作触发) - 终极游戏化升级版
# ==========================================
def check_and_award_badges(user):
    # 💡 防御 1：直接查库而不是用 user.earned_badges 缓存，避免拿不到最新数据
    earned_ubs = UserBadge.query.filter_by(user_id=user.id).all()
    earned_badge_ids = [ub.badge_id for ub in earned_ubs]
    
    available_badges = Badge.query.filter(Badge.id.notin_(earned_badge_ids)).all()
    awarded_count = 0
    
    for badge in available_badges:
        if badge.condition_type == 'manual':
            continue
            
        if badge.total_limit is not None and badge.issued_count >= badge.total_limit:
            continue

        earned = False
        val = badge.condition_value
        tid = badge.target_id  # 💡 提取目标 ID (用于新条件)
         
        # --- 原始条件判断逻辑 ---
        if badge.condition_type == 'all_users':
            earned = True
        elif badge.condition_type == 'note_count':
            count = Note.query.filter_by(user_id=user.id).count()
            if count >= val: earned = True
        elif badge.condition_type == 'featured_count':
            count = Note.query.filter_by(user_id=user.id, is_featured=True).count()
            if count >= val: earned = True
        elif badge.condition_type == 'wiki_edit_count':
            count = WikiPageHistory.query.filter_by(user_id=user.id).count()
            if count >= val: earned = True
        elif badge.condition_type == 'comment_count':
            count = Comment.query.filter_by(user_id=user.id).count()
            if count >= val: earned = True
        elif badge.condition_type == 'night_owl_sessions':
            sessions = StudySession.query.filter_by(user_id=user.id).all()
            night_days = set()
            for s in sessions:
                start_time = to_utc8_naive(s.start_time)
                if not start_time:
                    continue
                h = start_time.hour
                if h >= 23 or h < 4:
                    logical_date = (start_time - timedelta(hours=4)).date()
                    night_days.add(logical_date)
            if len(night_days) >= val: earned = True
        elif badge.condition_type == 'early_bird':
            sessions = StudySession.query.filter_by(user_id=user.id).all()
            early_days = set()
            for s in sessions:
                start_time = to_utc8_naive(s.start_time)
                if not start_time:
                    continue
                h = start_time.hour
                if 5 <= h < 8:
                    early_days.add(start_time.date())
            if len(early_days) >= val: earned = True
        elif badge.condition_type == 'weekend_warrior':
            sessions = StudySession.query.filter_by(user_id=user.id).all()
            total_seconds = 0
            for s in sessions:
                start_time = to_utc8_naive(s.start_time)
                if start_time and start_time.weekday() in [5, 6]:
                    duration = study_session_seconds(s)
                    if duration > 0:
                        total_seconds += duration
            total_hours = total_seconds / 3600
            if total_hours >= val: earned = True
        elif badge.condition_type == 'long_session_count':
            sessions = StudySession.query.filter_by(user_id=user.id).all()
            long_count = 0
            for s in sessions:
                duration = study_session_seconds(s)
                if duration >= 7200:
                    long_count += 1
            if long_count >= val: earned = True
        elif badge.condition_type == 'share_count':
            count = NoteShare.query.join(Note).filter(Note.user_id == user.id).count()
            if count >= val: earned = True
        elif badge.condition_type == 'total_views_received':
            count = NoteViewLog.query.join(Note).filter(Note.user_id == user.id).count()
            if count >= val: earned = True
        elif badge.condition_type == 'wiki_create_count':
            from sqlalchemy import func
            subquery = db.session.query(func.min(WikiPageHistory.id)).group_by(WikiPageHistory.page_id)
            count = WikiPageHistory.query.filter(WikiPageHistory.id.in_(subquery), WikiPageHistory.user_id == user.id).count()
            if count >= val: earned = True
        elif badge.condition_type == 'study_hours':
            sessions = StudySession.query.filter_by(user_id=user.id).all()
            total_seconds = sum(study_session_seconds(s) for s in sessions)
            total_hours = total_seconds / 3600
            if total_hours >= val: earned = True
        elif badge.condition_type == 'pomo_count':
            count = PomodoroRecord.query.filter_by(user_id=user.id).count()
            if count >= val: earned = True
        elif badge.condition_type == 'streak_days':
            sessions = StudySession.query.filter_by(user_id=user.id).order_by(StudySession.start_time.desc()).all()
            dates = sorted(list(set([s.start_time.date() for s in sessions])), reverse=True)
            streak = 0
            if dates:
                today = now_utc8().date()
                if dates[0] == today:
                    streak = 1
                    current = today
                elif dates[0] == today - timedelta(days=1):
                    streak = 1
                    current = today - timedelta(days=1)
                else:
                    streak = 0 
                if streak > 0:
                    for i in range(1, len(dates)):
                        if dates[i] == current - timedelta(days=1):
                            streak += 1
                            current = dates[i]
                        else:
                            break
            if streak >= val: earned = True
        elif badge.condition_type == 'login_days_in_range':
            if badge.start_time and badge.end_time:
                sessions = StudySession.query.filter(
                    StudySession.user_id == user.id,
                    StudySession.start_time >= badge.start_time,
                    StudySession.start_time <= badge.end_time
                ).all()
                login_days = set([s.start_time.date() for s in sessions])
                if len(login_days) >= val: earned = True

        # ==========================================
        # 🚀 进阶逻辑：全新的 5 大游戏化触发条件
        # ==========================================
        
        # 💡 条件 1：用户等级达标 (Lv.X)
        elif badge.condition_type == 'user_level':
            # 直接调用之前在 User 模型中写好的 level 属性
            if user.level >= val: earned = True
            
        # 💡 条件 2：指定 Wiki 学习时长达标 (分钟)
        elif badge.condition_type == 'wiki_specific_duration':
            if tid:
                # 统计该用户在该指定 wiki_id 下的日志数量 (粗略估算：1个log约代表10分钟停留)
                # 也可以根据你的业务逻辑调整为更精确的 session 时长计算
                view_logs_count = WikiViewLog.query.filter_by(user_id=user.id, wiki_id=tid).count()
                estimated_minutes = view_logs_count * 10
                if estimated_minutes >= val: earned = True

        # 💡 条件 3：完成指定的任务
        elif badge.condition_type == 'assignment_complete':
            if tid:
                # 查找该作业的提交记录，且状态必须为已批改 (graded)
                sub = Submission.query.filter_by(user_id=user.id, assignment_id=tid, status='graded').first()
                if sub: earned = True
                
        # 💡 条件 4：在指定任务中获得星星数量达标
        elif badge.condition_type == 'assignment_stars':
            if tid:
                sub = Submission.query.filter_by(user_id=user.id, assignment_id=tid, status='graded').first()
                if sub and (sub.stars_earned or 0) >= val: earned = True
                
        # 💡 条件 5：在指定任务中等第达标
        elif badge.condition_type == 'assignment_grade':
            if tid:
                sub = Submission.query.filter_by(user_id=user.id, assignment_id=tid, status='graded').first()
                if sub and sub.grade:
                    # 设定等第的数字权重，6 代表最高级
                    grade_map = {'S+': 6, 'S': 5, 'A': 4, 'B': 3, 'C': 2, 'D': 1}
                    # 获取当前提交的等第权重
                    current_grade_val = grade_map.get(sub.grade.upper(), 0)
                    # 只要大于等于管理员设定的权重值 (val) 就发奖
                    if current_grade_val >= val: earned = True

        elif badge.condition_type in OJ_BADGE_CONDITION_LABELS:
            earned = meets_oj_badge_condition(user, badge)

        # -------------------------------------------------------------
        
        if earned:
            # 💡 防御 2：确认确实没发过才发
            existing_ub = UserBadge.query.filter_by(user_id=user.id, badge_id=badge.id).first()
            if not existing_ub:
                ub = UserBadge(user_id=user.id, badge_id=badge.id, is_notified=False)
                
                # 💡 防御 3：加入短 UUID 哈希，彻底解决高并发下 serial_number 重复导致的崩溃回滚！
                current_count = UserBadge.query.filter_by(badge_id=badge.id).count()
                ub.serial_number = f"{badge.serial_prefix}-{str(current_count + 1).zfill(3)}-{uuid.uuid4().hex[:4].upper()}"
                
                db.session.add(ub)
                badge.issued_count = current_count + 1
                awarded_count += 1
                
                partners = user.study_partners.all()
                for p in partners:
                    n = Notification(
                        user_id=p.id,
                        message=f"你的学友 {user.username} 获得了 {badge.name} 徽章！",
                        type='friend_achievement',
                        link=url_for('public_profile', user_id=user.id)
                    )
                    db.session.add(n)
            
    if awarded_count > 0:
        try:
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            print(f"[发奖出错] {e}")

    check_and_award_bookmarks(user)

@app.route("/api/stickers/<page_type>/publish", methods=["POST"])
@login_required
def api_publish_sticker_snapshot(page_type):
    if page_type not in ['wiki', 'book', 'profile']:
        return jsonify({"error": "Invalid page type"}), 400
        
    data = request.json
    if not data:
         return jsonify({"error": "无效的请求数据，请确保Content-Type为application/json"}), 400

    title = data.get("title", f"{current_user.username}的{page_type}板")
    caption = data.get("caption", "")
    visibility = data.get("visibility", "followers") # Default to followers
    
    # Get current live stickers
    current_stickers = UserSticker.query.filter_by(user_id=current_user.id, page_type=page_type).all()
    
    if not current_stickers:
        return jsonify({"error": "Cannot publish empty board"}), 400
        
    stickers_data = [{
        "badge_id": s.badge_id,
        "badge_icon": badge_icon_url(s.badge.icon) if s.badge else "",
        "x": s.x,
        "y": s.y,
        "rotation": s.rotation,
        "scale": s.scale,
        "z_index": s.z_index
    } for s in current_stickers]
    
    snapshot = StickerBoardSnapshot(
        user_id=current_user.id,
        page_type=page_type,
        title=title,
        caption=caption,
        visibility=visibility,
        data_json=stickers_data
    )
    
    try:
        db.session.add(snapshot)
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        return jsonify({"error": f"数据库错误: {str(e)}"}), 500
    
    return jsonify({"success": True, "snapshot_id": snapshot.id})

@app.route("/api/stickers/snapshots", methods=["GET"])
@login_required
def api_sticker_snapshots():
    user_id_param = request.args.get('user_id')
    target_user_id = current_user.id
    if user_id_param:
        try:
            target_user_id = int(user_id_param)
        except ValueError:
            pass
            
    query = StickerBoardSnapshot.query.filter_by(user_id=target_user_id).order_by(StickerBoardSnapshot.created_at.desc())
    
    # Permission Check
    if target_user_id != current_user.id:
        # Filter based on visibility
        # If I am a follower, I can see 'followers' and 'public'
        # If I am just a stranger, I can only see 'public'
        target_user = User.query.get(target_user_id)
        if not target_user:
            return jsonify({"error": "User not found"}), 404
            
        if current_user.is_following(target_user):
            query = query.filter(StickerBoardSnapshot.visibility.in_(['followers', 'public']))
        else:
            query = query.filter(StickerBoardSnapshot.visibility == 'public')
            
    snapshots = query.all()
    
    return jsonify({
        "snapshots": [{
            "id": s.id,
            "page_type": s.page_type,
            "title": s.title,
            "caption": s.caption,
            "created_at": s.created_at.isoformat(),
            "likes_count": s.likes_count,
            "preview_count": len(s.data_json)
        } for s in snapshots]
    })

@app.route("/api/stickers/snapshots/<int:snapshot_id>", methods=["GET"])
@login_required
def api_sticker_snapshot_detail(snapshot_id):
    snapshot = StickerBoardSnapshot.query.get_or_404(snapshot_id)
    
    # Permission Check
    if snapshot.user_id != current_user.id:
        if snapshot.visibility == 'private':
            return jsonify({"error": "Permission denied"}), 403
        if snapshot.visibility == 'followers':
            target_user = User.query.get(snapshot.user_id)
            if not current_user.is_following(target_user):
                 return jsonify({"error": "Permission denied"}), 403
                 
    return jsonify({
        "id": snapshot.id,
        "user_id": snapshot.user_id,
        "user_name": snapshot.user.username if snapshot.user else "Unknown",
        "page_type": snapshot.page_type,
        "title": snapshot.title,
        "caption": snapshot.caption,
        "created_at": snapshot.created_at.isoformat(),
        "stickers": snapshot.data_json
    })

# System Announcement Routes

@app.route("/admin/announcements")
@login_required
def manage_announcements():
    if not current_user.is_admin:
        abort(403)
    announcements = SystemAnnouncement.query.order_by(SystemAnnouncement.created_at.desc()).all()
    return render_template("admin/manage_announcements.html", announcements=announcements)

@app.route("/admin/announcements/new", methods=["GET", "POST"])
@login_required
def new_announcement():
    if not current_user.is_admin:
        abort(403)
    if request.method == "POST":
        title = request.form.get("title", "").strip()
        content = request.form.get("content", "")
        if not title:
            flash("标题不能为空")
            return redirect(url_for("new_announcement"))
        
        sa = SystemAnnouncement(
            title=title, 
            content=content, 
            created_by_id=current_user.id
        )
        db.session.add(sa)
        db.session.commit()
        flash("通知已发布")
        return redirect(url_for("manage_announcements"))
    return render_template("admin/edit_announcement.html", announcement=None)

@app.route("/admin/announcements/<int:id>/edit", methods=["GET", "POST"])
@login_required
def edit_announcement_route(id):
    if not current_user.is_admin:
        abort(403)
    sa = SystemAnnouncement.query.get_or_404(id)
    if request.method == "POST":
        sa.title = request.form.get("title", "").strip()
        sa.content = request.form.get("content", "")
        sa.is_active = request.form.get("is_active") == "on"
        db.session.commit()
        flash("通知已更新")
        return redirect(url_for("manage_announcements"))
    return render_template("admin/edit_announcement.html", announcement=sa)

@app.route("/admin/announcements/<int:id>/delete", methods=["POST"])
@login_required
def delete_announcement(id):
    if not current_user.is_admin:
        abort(403)
    sa = SystemAnnouncement.query.get_or_404(id)
    # Cascade delete confirmations
    UserAnnouncementConfirmation.query.filter_by(announcement_id=id).delete()
    db.session.delete(sa)
    db.session.commit()
    flash("通知已删除")
    return redirect(url_for("manage_announcements"))

@app.route("/admin/announcements/<int:id>/stats")
@login_required
def announcement_stats(id):
    if not current_user.is_admin:
        abort(403)
    sa = SystemAnnouncement.query.get_or_404(id)
    confirmations = UserAnnouncementConfirmation.query.filter_by(announcement_id=id).all()
    confirmed_user_ids = [c.user_id for c in confirmations]
    
    all_users = User.query.all()
    confirmed_users = []
    unconfirmed_users = []
    
    for u in all_users:
        if u.id in confirmed_user_ids:
            # Attach confirmed time
            c = next((x for x in confirmations if x.user_id == u.id), None)
            u.confirmed_at = c.confirmed_at if c else None
            confirmed_users.append(u)
        else:
            unconfirmed_users.append(u)
            
    return render_template("admin/announcement_stats.html", announcement=sa, confirmed_users=confirmed_users, unconfirmed_users=unconfirmed_users)

@app.route("/api/announcements/check")
@login_required
def check_announcements():
    # 1. Find all active announcements
    active_anns = SystemAnnouncement.query.filter_by(is_active=True).order_by(SystemAnnouncement.created_at.desc()).all()
    
    if not active_anns:
        return jsonify({"has_unconfirmed": False, "has_history": False})
        
    # 2. Check confirmations
    confirmed_ids = [c.announcement_id for c in current_user.announcement_confirmations]
    
    # 3. Find first unconfirmed
    unconfirmed = None
    for ann in active_anns:
        if ann.id not in confirmed_ids:
            unconfirmed = ann
            break # Just need the latest one or the first one? usually show latest first.
            
    has_history = len(active_anns) > 0 # Basically if any exist, history is valid.
    
    if unconfirmed:
        html = markdown(unconfirmed.content or "")
        return jsonify({
            "has_unconfirmed": True,
            "has_history": has_history,
            "announcement": {
                "id": unconfirmed.id,
                "title": unconfirmed.title,
                "html": html,
                "created_at": unconfirmed.created_at.strftime("%Y-%m-%d")
            }
        })
    else:
        return jsonify({
            "has_unconfirmed": False,
            "has_history": has_history
        })

@app.route("/api/announcements/<int:id>/confirm", methods=["POST"])
@login_required
def confirm_announcement(id):
    sa = SystemAnnouncement.query.get_or_404(id)
    if not sa.is_active:
        return jsonify({"error": "Announcement is not active"}), 400
        
    exists = UserAnnouncementConfirmation.query.filter_by(user_id=current_user.id, announcement_id=id).first()
    if not exists:
        conf = UserAnnouncementConfirmation(user_id=current_user.id, announcement_id=id)
        db.session.add(conf)
        db.session.commit()
        
    return jsonify({"success": True})

@app.route("/api/announcements/history")
@login_required
def announcement_history():
    # Return list of all active announcements
    anns = SystemAnnouncement.query.filter_by(is_active=True).order_by(SystemAnnouncement.created_at.desc()).all()
    
    data = []
    for ann in anns:
        # Check if confirmed
        is_confirmed = UserAnnouncementConfirmation.query.filter_by(user_id=current_user.id, announcement_id=ann.id).first() is not None
        html = markdown(ann.content or "")
        data.append({
            "id": ann.id,
            "title": ann.title,
            "html": html,
            "created_at": ann.created_at.strftime("%Y-%m-%d"),
            "is_confirmed": is_confirmed
        })
        
    return jsonify({"announcements": data})

@app.route("/admin/announcements/files", methods=["GET"])
@login_required
def manage_announcement_files():
    if not current_user.is_admin:
        abort(403)
    files = AnnouncementFile.query.order_by(AnnouncementFile.uploaded_at.desc()).all()
    return render_template("admin/announcement_files.html", files=files)

@app.route("/admin/announcements/files/upload", methods=["POST"])
@login_required
def upload_announcement_file():
    if not current_user.is_admin:
        return {"error": "Permission denied"}, 403
        
    if "image" not in request.files:
        return {"error": "No file part"}, 400
    file = request.files["image"]
    if file.filename == "":
        return {"error": "No selected file"}, 400
    if file:
        filename = secure_filename(file.filename)
        ext = os.path.splitext(filename)[1].lower()
        if ext not in [".jpg", ".jpeg", ".png", ".gif", ".webp"]:
             return {"error": f"File type '{ext}' not allowed. Allowed: jpg, png, gif, webp"}, 400
        unique_filename = str(uuid.uuid4()) + ext
        
        # Ensure uploads directory exists
        upload_dir = os.path.join(app.config["UPLOAD_FOLDER"], "announcements")
        if not os.path.exists(upload_dir):
            os.makedirs(upload_dir)
            
        file.save(os.path.join(upload_dir, unique_filename))
        
        file_path = url_for("static", filename=f"uploads/announcements/{unique_filename}")
        
        # Record in DB
        af = AnnouncementFile(
            filename=filename, 
            file_path=file_path, 
            uploaded_by_id=current_user.id
        )
        db.session.add(af)
        db.session.commit()
        
        return {"data": {"filePath": file_path, "filename": filename}}

@app.route("/admin/announcements/files/<int:file_id>/delete", methods=["POST"])
@login_required
def delete_announcement_file(file_id):
    if not current_user.is_admin:
        abort(403)
        
    af = AnnouncementFile.query.get_or_404(file_id)
    
    # Optional: Delete actual file from disk
    # filename = af.file_path.split("/")[-1]
    # file_path = os.path.join(app.config["UPLOAD_FOLDER"], "announcements", filename)
    # if os.path.exists(file_path):
    #     os.remove(file_path)
    
    db.session.delete(af)
    db.session.commit()
    flash("文件已删除")
    return redirect(url_for("manage_announcement_files"))

@app.route("/announcements")
@login_required
def announcements_view():
    # 获取所有处于激活状态的系统公告
    anns = SystemAnnouncement.query.filter_by(is_active=True).order_by(SystemAnnouncement.created_at.desc()).all()
    
    # 获取 URL 中选中的公告 ID
    selected_id = request.args.get('id', type=int)
    
    announcements_data = []
    selected_ann = None
    
    for a in anns:
        data = {
            'id': a.id,
            'title': a.title,
            'html': markdown(a.content or ""),
            'created_at': a.created_at
        }
        announcements_data.append(data)
        
        # 寻找匹配的公告
        if selected_id and a.id == selected_id:
            selected_ann = data
            
    # 如果没有指定 ID 或指定的 ID 不存在，默认选中最新的一条
    if not selected_ann and announcements_data:
        selected_ann = announcements_data[0]
        
    return render_template("announcements.html", announcements=announcements_data, selected_ann=selected_ann)


def upgrade_db():
    """Upgrade database schema manually"""
    with app.app_context():
        # 💡 1. 确保所有新表 (比如 PomodoroRecord) 都存在
        db.create_all()
        
        inspector = db.inspect(db.engine)
        
            # 💡 1. 升级 Assignment 表，增加星级设定
        if 'assignment' in inspector.get_table_names():
            assign_columns = [c['name'] for c in inspector.get_columns('assignment')]
            with db.engine.connect() as conn:
                needs_commit = False
                # ... (之前的 content_md, target_type 等检查代码保留，这里略过) ...
                
                # 👇 重点：补上 star_level 字段
                if 'star_level' not in assign_columns:
                    conn.execute(db.text("ALTER TABLE assignment ADD COLUMN star_level INTEGER DEFAULT 1"))
                    needs_commit = True
                
                if needs_commit:
                    conn.commit()

        # 💡 2. 升级 Submission 表，增加实得星级
        if 'submission' in inspector.get_table_names():
            sub_columns = [c['name'] for c in inspector.get_columns('submission')]
            with db.engine.connect() as conn:
                needs_commit = False
                
                # 👇 重点：补上 stars_earned 字段
                if 'stars_earned' not in sub_columns:
                    conn.execute(db.text("ALTER TABLE submission ADD COLUMN stars_earned INTEGER DEFAULT 0"))
                    needs_commit = True
                    
                if needs_commit:
                    conn.commit()
        # 升级 Badge 表
        if 'badge' in inspector.get_table_names():
            badge_columns = [c['name'] for c in inspector.get_columns('badge')]
            with db.engine.connect() as conn:
                if 'target_id' not in badge_columns:
                    conn.execute(db.text("ALTER TABLE badge ADD COLUMN target_id INTEGER"))
                    conn.commit()
        
        if 'note' in inspector.get_table_names():
            note_columns = [c['name'] for c in inspector.get_columns('note')]
            with db.engine.connect() as conn:
                needs_commit = False
                if 'is_scrapbook' not in note_columns:
                    conn.execute(db.text("ALTER TABLE note ADD COLUMN is_scrapbook BOOLEAN DEFAULT 0"))
                    needs_commit = True
                if 'scrapbook_data' not in note_columns:
                    conn.execute(db.text("ALTER TABLE note ADD COLUMN scrapbook_data JSON"))
                    needs_commit = True
                if needs_commit:
                    conn.commit()

        if 'comment' in inspector.get_table_names():
            comment_columns = [c['name'] for c in inspector.get_columns('comment')]
            with db.engine.connect() as conn:
                needs_commit = False
                added_is_approved = False

                if 'is_approved' not in comment_columns:
                    conn.execute(db.text("ALTER TABLE comment ADD COLUMN is_approved BOOLEAN DEFAULT 0"))
                    added_is_approved = True
                    needs_commit = True

                if 'approved_at' not in comment_columns:
                    conn.execute(db.text("ALTER TABLE comment ADD COLUMN approved_at DATETIME"))
                    needs_commit = True

                if 'approved_by_id' not in comment_columns:
                    conn.execute(db.text("ALTER TABLE comment ADD COLUMN approved_by_id INTEGER"))
                    needs_commit = True

                # 新增审核功能时，历史评论默认全部进入待审核
                if added_is_approved:
                    conn.execute(db.text("UPDATE comment SET is_approved = 0 WHERE is_approved IS NULL"))
                    conn.execute(db.text("UPDATE comment SET approved_at = NULL"))
                    conn.execute(db.text("UPDATE comment SET approved_by_id = NULL"))

                if needs_commit:
                    conn.commit()

        if 'wiki' in inspector.get_table_names():
            wiki_columns = [c['name'] for c in inspector.get_columns('wiki')]
            with db.engine.connect() as conn:
                needs_commit = False
                if 'icon_url' not in wiki_columns:
                    conn.execute(db.text("ALTER TABLE wiki ADD COLUMN icon_url VARCHAR(255)"))
                    needs_commit = True
                if 'icon_scale' not in wiki_columns:
                    conn.execute(db.text("ALTER TABLE wiki ADD COLUMN icon_scale FLOAT DEFAULT 1.0"))
                    needs_commit = True
                if 'icon_position_x' not in wiki_columns:
                    conn.execute(db.text("ALTER TABLE wiki ADD COLUMN icon_position_x FLOAT DEFAULT 50.0"))
                    needs_commit = True
                if 'icon_position_y' not in wiki_columns:
                    conn.execute(db.text("ALTER TABLE wiki ADD COLUMN icon_position_y FLOAT DEFAULT 50.0"))
                    needs_commit = True
                if needs_commit:
                    conn.commit()
                

@app.route('/service/wiki/pages')
@login_required
def service_wiki_pages():
    wiki_id = request.args.get('wiki_id', type=int)
    ns = request.args.get('ns', default=None)
    
    if not wiki_id:
        return jsonify({'error': 'wiki_id is required'}), 400
        
    wiki = Wiki.query.get_or_404(wiki_id)
    
    if not can_view_wiki(wiki.id):
        return jsonify({'error': 'Permission denied'}), 403

    query = WikiPage.query.filter_by(wiki_id=wiki_id)
    
    if ns:
        query = query.filter(WikiPage.title.like(f"{ns}:%"))
        
    pages = query.order_by(WikiPage.updated_at.desc()).all()
    
    pages_data = []
    for p in pages:
        parts = p.title.split(':', 1)
        namespace = parts[0] if len(parts) > 1 else 'Main'
        
        pages_data.append({
            'id': p.id,
            'title': p.title,
            'slug': p.slug,
            'namespace': namespace,
            'updated_at': p.updated_at.isoformat() if p.updated_at else None
        })
        
    return jsonify({'pages': pages_data})

@app.route("/wikis/<int:wiki_id>/contributors")
@login_required
def wiki_contributors(wiki_id):
    w = Wiki.query.get_or_404(wiki_id)
    
    # Get all pages
    page_ids = db.session.query(WikiPage.id).filter_by(wiki_id=wiki_id).all()
    page_ids = [p[0] for p in page_ids]
    
    contributors = []
    try:
        if page_ids:
            contributors = db.session.query(
                User, 
                db.func.count(WikiPageHistory.id).label('edit_count')
            ).join(
                WikiPageHistory, WikiPageHistory.user_id == User.id
            ).filter(
                WikiPageHistory.page_id.in_(page_ids)
            ).group_by(
                User.id
            ).order_by(
                db.text('edit_count DESC')
            ).all()
    except Exception as e:
        print(f"Error fetching contributors for list: {e}")
        contributors = []
        
    return render_template("wiki/contributors.html", wiki=w, contributors=contributors)

@app.route("/wikis/<int:wiki_id>/pages/<slug>")
@login_required
def view_page(wiki_id, slug):
    if not can_view_wiki(wiki_id):
        abort(403)
    w = Wiki.query.get_or_404(wiki_id)
    p = WikiPage.query.filter_by(wiki_id=wiki_id, slug=slug).first_or_404()
    
    now = now_utc8()
    
    last_log = WikiViewLog.query.filter_by(wiki_id=wiki_id, user_id=current_user.id)\
        .order_by(WikiViewLog.timestamp.desc()).first()
    
    should_commit = False
    last_log_time = to_utc8_naive(last_log.timestamp) if last_log else None
    
    if not last_log_time or datetime_diff_seconds(now, last_log_time) > 600:
        log = WikiViewLog(wiki_id=wiki_id, user_id=current_user.id, timestamp=now)
        db.session.add(log)
        should_commit = True

    view_key = f"viewed_page_{p.id}"
    last_view_ts = session.get(view_key)
    current_ts = time.time()
    
    if not last_view_ts or (current_ts - last_view_ts) > 300:
        p.view_count += 1
        session[view_key] = current_ts
        should_commit = True
    
    if should_commit:
        db.session.commit()

    # 1. 获取所有文件夹和页面
    all_folders = WikiFolder.query.filter_by(wiki_id=wiki_id).all()
    all_pages = WikiPage.query.filter_by(wiki_id=wiki_id).all()
    
    # 2. 构建混合目录树 (支持文件夹和独立页面混排)
    root_items = []
    for f in all_folders:
        f.item_type = 'folder'
        f.children = sorted([page for page in all_pages if page.folder_id == f.id], key=lambda x: x.order_weight)
        root_items.append(f)
        
    for page in all_pages:
        if page.folder_id is None:
            page.item_type = 'page'
            root_items.append(page)
            
    root_items = sorted(root_items, key=lambda x: x.order_weight)
    
    html = markdown(p.content_md or "")
    html = parse_assignment_tags(html, p.id) # 💡 调用解析器
    is_subscribed = Subscription.query.filter_by(wiki_id=wiki_id, user_id=current_user.id).first() is not None
    
    contributors = []
    try:
        if all_pages:
            page_ids_list = [pg.id for pg in all_pages]
            contributors = db.session.query(
                User, 
                db.func.count(WikiPageHistory.id).label('edit_count')
            ).join(
                WikiPageHistory, WikiPageHistory.user_id == User.id
            ).filter(
                WikiPageHistory.page_id.in_(page_ids_list)
            ).group_by(
                User.id
            ).order_by(
                db.text('edit_count DESC')
            ).all()
    except Exception as e:
        print(f"Error fetching contributors: {e}")
        contributors = []

    my_recent_notes = []
    if current_user.is_authenticated:
        my_recent_notes = Note.query.filter_by(user_id=current_user.id).order_by(Note.updated_at.desc()).limit(15).all()
    visible_comments = get_visible_comments(wiki_page_id=p.id)

    return render_template("wiki/view_page.html", 
                           wiki=w, 
                           page=p, 
                           root_items=root_items, # 💡 传入树形结构
                           pages=all_pages,       # 兼容旧逻辑保留
                           html=html, 
                           can_edit=can_edit_wiki(wiki_id), 
                           is_subscribed=is_subscribed, 
                           contributors=contributors, 
                           my_recent_notes=my_recent_notes,
                           visible_comments=visible_comments)

# ==========================================
# ✨ 全员贴纸墙 (Global Sticker Wall) 模型与路由
# ==========================================

# 贴纸墙权限关联表
wall_classes = db.Table('wall_classes',
    db.Column('wall_id', db.Integer, db.ForeignKey('sticker_wall.id'), primary_key=True),
    db.Column('class_id', db.Integer, db.ForeignKey('class.id'), primary_key=True)
)
wall_groups = db.Table('wall_groups',
    db.Column('wall_id', db.Integer, db.ForeignKey('sticker_wall.id'), primary_key=True),
    db.Column('group_id', db.Integer, db.ForeignKey('group.id'), primary_key=True)
)
wall_users = db.Table('wall_users',
    db.Column('wall_id', db.Integer, db.ForeignKey('sticker_wall.id'), primary_key=True),
    db.Column('user_id', db.Integer, db.ForeignKey('user.id'), primary_key=True)
)

class StickerWall(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    description = db.Column(db.String(255))
    created_by_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=False)
    created_at = db.Column(db.DateTime, default=now_utc8)
    visibility = db.Column(db.String(20), default="public") # public, restricted
    bg_style = db.Column(db.String(50), default="dots") # dots, grid, lines, empty
    
    allowed_classes = db.relationship('Class', secondary=wall_classes, lazy='subquery')
    allowed_groups = db.relationship('Group', secondary=wall_groups, lazy='subquery')
    allowed_users = db.relationship('User', secondary=wall_users, lazy='subquery')
    
    stickers = db.relationship("WallSticker", backref="wall", cascade="all, delete-orphan")

    def is_visible_to(self, user):
        return True

    def can_arrange_stickers(self, user):
        if not user.is_authenticated:
            return False
        if self.visibility == 'public':
            return True
        if user.is_admin: return True
        if user in self.allowed_users: return True
        if user.student_class and user.student_class in self.allowed_classes: return True
        if not set(user.groups).isdisjoint(set(self.allowed_groups)): return True
        return False

class WallSticker(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    wall_id = db.Column(db.Integer, db.ForeignKey("sticker_wall.id"), nullable=False)
    user_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=False)
    badge_id = db.Column(db.Integer, db.ForeignKey("badge.id"), nullable=False)
    x = db.Column(db.Float, nullable=False) # X坐标百分比
    y = db.Column(db.Float, nullable=False) # Y坐标百分比
    rotation = db.Column(db.Float, default=0.0) # 旋转角度
    created_at = db.Column(db.DateTime, default=now_utc8)
    
    user = db.relationship("User")
    badge = db.relationship("Badge")


class WikiFolder(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    wiki_id = db.Column(db.Integer, db.ForeignKey('wiki.id'), nullable=False)
    name = db.Column(db.String(100), nullable=False)
    order_weight = db.Column(db.Integer, default=0) # 用于排序

    # 关联
    wiki = db.relationship('Wiki', backref=db.backref('folders', cascade='all, delete-orphan'))

class StickyNote(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    content = db.Column(db.Text, default="")
    # 便签颜色 (默认黄色，还支持 pink, blue, green)
    color = db.Column(db.String(20), default="yellow") 
    # 记录在屏幕上的坐标 (百分比或像素)
    pos_x = db.Column(db.Integer, default=100)
    pos_y = db.Column(db.Integer, default=100)
    z_index = db.Column(db.Integer, default=1000) # 层级，点击时置顶
    
    created_at = db.Column(db.DateTime, default=now_utc8)
    updated_at = db.Column(db.DateTime, default=now_utc8, onupdate=now_utc8)

    # 关联 User (方便通过 user.sticky_notes 获取)
    user = db.relationship('User', backref=db.backref('sticky_notes', cascade='all, delete-orphan', lazy=True))

class PomodoroRecord(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    duration_minutes = db.Column(db.Integer, default=45) # 默认一次 45 分钟
    completed_at = db.Column(db.DateTime, default=now_utc8)

    # 关联 User
    user = db.relationship('User', backref=db.backref('pomodoros', cascade='all, delete-orphan', lazy='dynamic'))


# --- 后台管理员路由：管理贴纸墙 ---
@app.route("/admin/sticker_walls", methods=["GET", "POST"])
@login_required
def manage_sticker_walls():
    if not current_user.is_admin:
        abort(403)
        
    if request.method == "POST":
        name = request.form.get("name", "").strip()
        description = request.form.get("description", "").strip()
        visibility = request.form.get("visibility", "public")
        bg_style = request.form.get("bg_style", "dots")
        if name:
            wall = StickerWall(name=name, description=description, visibility=visibility, bg_style=bg_style, created_by_id=current_user.id)
            db.session.add(wall)
            db.session.commit()
            flash("贴纸墙创建成功！")
        return redirect(url_for("manage_sticker_walls"))
        
    walls = StickerWall.query.order_by(StickerWall.created_at.desc()).all()
    return render_template("admin/manage_sticker_walls.html", walls=walls)

@app.route("/admin/sticker_walls/<int:wall_id>/edit", methods=["GET", "POST"])
@login_required
def edit_sticker_wall(wall_id):
    if not current_user.is_admin:
        abort(403)
        
    wall = StickerWall.query.get_or_404(wall_id)
    
    if request.method == "POST":
        wall.name = request.form.get("name", "").strip()
        wall.description = request.form.get("description", "").strip()
        wall.visibility = request.form.get("visibility", "public")
        wall.bg_style = request.form.get("bg_style", "dots")
        
        # 处理权限
        if wall.visibility == 'restricted':
            # 更新班级权限
            class_ids = request.form.getlist("allowed_classes")
            wall.allowed_classes = Class.query.filter(Class.id.in_(class_ids)).all()
            
            # 更新小组权限
            group_ids = request.form.getlist("allowed_groups")
            wall.allowed_groups = Group.query.filter(Group.id.in_(group_ids)).all()
            
            # 更新用户权限
            user_ids = request.form.getlist("allowed_users")
            wall.allowed_users = User.query.filter(User.id.in_(user_ids)).all()
        else:
            # 如果是公开的，清空所有限制
            wall.allowed_classes = []
            wall.allowed_groups = []
            wall.allowed_users = []
            
        db.session.commit()
        flash("贴纸墙设置已更新！")
        return redirect(url_for("manage_sticker_walls"))
        
    classes = Class.query.order_by(Class.name).all()
    groups = Group.query.order_by(Group.name).all()
    all_users = User.query.order_by(User.username).all()
    
    return render_template("admin/edit_sticker_wall.html", 
                         wall=wall, 
                         classes=classes, 
                         groups=groups, 
                         all_users=all_users)

@app.route("/admin/sticker_walls/<int:wall_id>/delete", methods=["POST"])
@login_required
def delete_sticker_wall(wall_id):
    if not current_user.is_admin:
        abort(403)
    wall = StickerWall.query.get_or_404(wall_id)
    db.session.delete(wall)
    db.session.commit()
    flash("贴纸墙已删除")
    return redirect(url_for("manage_sticker_walls"))

@app.route("/sticker_walls")
@login_required
def sticker_walls_view():
    all_walls = StickerWall.query.order_by(StickerWall.created_at.desc()).all()
    visible_walls = all_walls
    
    # 1. 获取用户拥有的所有徽章数量 (分组统计)
    user_badges = UserBadge.query.filter_by(user_id=current_user.id).all()
    badge_counts = {}
    for ub in user_badges:
        if ub.badge_id not in badge_counts:
            badge_counts[ub.badge_id] = {'badge': ub.badge, 'owned': 0}
        badge_counts[ub.badge_id]['owned'] += 1
        
    # 2. 获取用户在【所有贴纸墙】上已经贴出的徽章数量
    used_stickers = WallSticker.query.filter_by(user_id=current_user.id).all()
    used_counts = {}
    for s in used_stickers:
        used_counts[s.badge_id] = used_counts.get(s.badge_id, 0) + 1
        
    # 3. 构造前端抽屉所需的库存数据
    my_inventory = []
    for bid, data in badge_counts.items():
        used = used_counts.get(bid, 0)
        my_inventory.append({
            'badge': data['badge'],
            'owned': data['owned'],
            'used': used,
            'available': data['owned'] - used
        })
    
    # 按剩余可用数量排序
    my_inventory.sort(key=lambda x: (-x['available'], x['badge'].id))
    my_total_stickers = len(used_stickers) # 全局已用总数
    
    # 💡 核心新增：动态计算贴纸上限
    completed_tasks_count = Submission.query.filter_by(user_id=current_user.id, status='graded').count()
    max_stickers_limit = 10 + completed_tasks_count
    
    return render_template("sticker_walls.html", 
                           walls=visible_walls, 
                           my_inventory=my_inventory, 
                           my_total_stickers=my_total_stickers,
                           max_stickers_limit=max_stickers_limit) # 传入模板

# ==========================================
# 2. API路由：接管 sticker.js 的保存和加载
# ==========================================
# ==========================================
# API路由：接管 sticker.js 的保存和加载
# ==========================================
@app.route("/api/stickers/<page_type>", methods=["GET", "POST"])
@login_required
def api_stickers(page_type):
    # 验证页面类型，新增了 'wall' 支持
    if page_type not in ['wiki', 'book', 'profile', 'wall']:
        return jsonify({"error": "Invalid page type"}), 400
        
    if request.method == "GET":
        # --------------------------------------------------
        # 1. 贴纸墙 (wall) 获取逻辑
        # --------------------------------------------------
        if page_type == 'wall':
            # sticker.js 默认通过 user_id 参数传目标 ID，对于 wall 来说，它就是 wall_id
            wall_id = request.args.get('user_id', type=int) 
            wall = StickerWall.query.get_or_404(wall_id)
            can_arrange = wall.can_arrange_stickers(current_user)
                
            # 拉取该墙面上所有的贴纸（包括别人的）
            stickers = WallSticker.query.filter_by(wall_id=wall_id).all()
            
            # 计算当前用户在所有墙上的贴纸使用总量，用于前端实时更新库存状态
            global_used = db.session.query(WallSticker.badge_id, db.func.count(WallSticker.id))\
                .filter_by(user_id=current_user.id)\
                .group_by(WallSticker.badge_id).all()
            inventory_usage = {bid: count for bid, count in global_used}

            return jsonify({
                "stickers": [{
                    "id": s.id,
                    "badge_id": s.badge_id,
                    # 兼容不同类型的图标渲染
                    "badge_icon": badge_icon_url(s.badge.icon) if is_image_icon(s.badge.icon) else s.badge.icon,
                    "badge_name": s.badge.name,
                    "x": s.x,
                    "y": s.y,
                    "rotation": s.rotation,
                    "scale": 1.0, # 贴纸墙固定大小，但需要传给前端保持数据结构
                    "z_index": s.id,
                    "user_name": s.user.username,
                    "created_at": s.created_at.strftime("%Y-%m-%d %H:%M"),
                    # 关键控制权限：只有当前贴纸的主人才能选中/拖拽/删除
                    "is_editable": can_arrange and (s.user_id == current_user.id)
                } for s in stickers],
                "inventory_usage": inventory_usage
            })
            
        # --------------------------------------------------
        # 2. 原有的 Profile / Wiki / Book 获取逻辑
        # --------------------------------------------------
        target_user_id = request.args.get('user_id', current_user.id, type=int)
        
        # 如果是看别人的贴纸板，校验学友权限
        if target_user_id != current_user.id:
            target_user = User.query.get(target_user_id)
            if not target_user or not (current_user.is_following(target_user) and target_user.is_following(current_user)):
                 return jsonify({"error": "Permission denied: Not study partners"}), 403

        stickers = UserSticker.query.filter_by(user_id=target_user_id, page_type=page_type).all()
        return jsonify({
            "stickers": [{
                "id": s.id, 
                "badge_id": s.badge_id,
                "badge_name": s.badge.name, # <--- 新增这一行，解决悬浮窗报错
                "badge_icon": badge_icon_url(s.badge.icon) if s.badge else "",
                "x": s.x, 
                "y": s.y, 
                "rotation": s.rotation, 
                "scale": s.scale, 
                "z_index": s.z_index,
                # 个人主页/笔记类：如果是当前用户自己，就有编辑权限
                "is_editable": target_user_id == current_user.id 
            } for s in stickers]
        })
        
    elif request.method == "POST":
        data = request.json
        stickers_data = data.get("stickers", [])
        
        # --------------------------------------------------
        # 1. 贴纸墙 (wall) 保存逻辑 (协作模式增量覆盖)
        # --------------------------------------------------
        if page_type == 'wall':
            wall_id = request.args.get('user_id', type=int)
            wall = StickerWall.query.get_or_404(wall_id)
            if not wall.can_arrange_stickers(current_user):
                return jsonify({"error": "Permission denied: You cannot arrange stickers on this wall"}), 403
            
            # [全局额度强校验]
            # 计算除了当前墙以外，该用户在其他所有的墙上贴了多少张
            other_walls_count = WallSticker.query.filter(
                WallSticker.user_id == current_user.id, 
                WallSticker.wall_id != wall_id
            ).count()
            
            # 本次要保存的贴纸数量
            incoming_count = len(stickers_data)
            
            # 严格限制最多只能贴 10 张
            # 💡 动态获取当前用户的贴纸上限
            completed_tasks_count = Submission.query.filter_by(user_id=current_user.id, status='graded').count()
            max_limit = 10 + completed_tasks_count

            # 严格限制最多只能贴 max_limit 张
            if other_walls_count + incoming_count > max_limit:
                return jsonify({"error": f"OVERLOAD // 保存失败：您的全局贴纸总数不可超过 {max_limit} 张！多去完成任务可以提升上限。"}), 400
            # [安全校验：防止篡改他人贴纸]
            # 后端只信任当前登录用户 (current_user.id)，无视前端传来的 user_id 或贴纸归属信息
            # 我们采取“先清空我的，再插入新的”策略，这样无论前端传什么，操作范围永远局限于“我的贴纸”
            
            # 1. 仅删掉当前用户在这面墙上的旧贴纸
            WallSticker.query.filter_by(user_id=current_user.id, wall_id=wall_id).delete()
            
            # 2. 插入前端传过来的最新贴纸数组
            for s in stickers_data:
                # 忽略前端传来的 id, user_id 等敏感字段，全部强制归属为当前用户
                new_sticker = WallSticker(
                    wall_id=wall_id, 
                    user_id=current_user.id, 
                    badge_id=s.get("badge_id"),
                    x=s.get("x", 50), 
                    y=s.get("y", 50), 
                    rotation=s.get("rotation", 0)
                )
                db.session.add(new_sticker)
                    
            db.session.commit()
            return jsonify({"success": True})
            
        # --------------------------------------------------
        # 2. 原有的 Profile / Book / Wiki 保存逻辑
        # --------------------------------------------------
        # 完全清空该用户在这个场景下的所有贴纸
        UserSticker.query.filter_by(user_id=current_user.id, page_type=page_type).delete()
        
        # 重新插入
        for s in stickers_data:
            new_sticker = UserSticker(
                user_id=current_user.id, 
                badge_id=s.get("badge_id"), 
                page_type=page_type,
                x=s.get("x", 50), 
                y=s.get("y", 50), 
                rotation=s.get("rotation", 0),
                scale=s.get("scale", 1), 
                z_index=s.get("z_index", 1)
            )
            db.session.add(new_sticker)
            
        db.session.commit()
        return jsonify({"success": True})

@app.route("/api/stickers/profile/clear", methods=["POST"])
@login_required
def api_clear_profile_stickers():
    """一键回收个人主页贴纸墙上的所有贴纸"""
    try:
        # 只删除当前用户在个人主页(profile)场景下的贴纸
        UserSticker.query.filter_by(
            user_id=current_user.id, 
            page_type='profile'
        ).delete()
        
        db.session.commit()
        return jsonify({"success": True, "message": "所有贴纸已收回库中"})
    except Exception as e:
        db.session.rollback()
        return jsonify({"success": False, "error": str(e)}), 500

@app.route("/api/stickers/all/nuclear_reset", methods=["POST"])
@login_required
def api_nuclear_reset_stickers():
    """核弹级重置：清空该用户在全站（主页、Wiki、笔记）的所有贴纸记录"""
    try:
        # 不看 page_type，直接根据 user_id 全删
        UserSticker.query.filter_by(user_id=current_user.id).delete()
        db.session.commit()
        return jsonify({"success": True, "message": "全站贴纸已彻底重置"})
    except Exception as e:
        db.session.rollback()
        return jsonify({"success": False, "error": str(e)}), 500

# ==========================================
# 🏆 巅峰竞技场 (Leaderboard) 路由
# ==========================================
LEADERBOARD_METRICS = {
    "duration": {
        "label": "肝帝榜",
        "icon": "fa-stopwatch",
        "unit": "小时",
        "description": "累计专注时长",
    },
    "pomo": {
        "label": "番茄榜",
        "icon": "fa-clock",
        "unit": "个",
        "description": "完成番茄钟数量",
    },
    "contribution": {
        "label": "贡献榜",
        "icon": "fa-pen-nib",
        "unit": "源力",
        "description": "笔记与 Wiki 编辑综合得分",
    },
    "streak": {
        "label": "毅力榜",
        "icon": "fa-fire",
        "unit": "天",
        "description": "连续打卡天数",
    },
    "oj_ac": {
        "label": "OJ AC榜",
        "icon": "fa-check-circle",
        "unit": "次",
        "description": "OJ 通过题目数量",
    },
    "oj_attempts": {
        "label": "OJ 尝试榜",
        "icon": "fa-code",
        "unit": "次",
        "description": "OJ 提交尝试数量",
    },
}

LEADERBOARD_PERIODS = {
    "weekly": "本周",
    "all": "总榜",
}


def build_leaderboard_snapshot(users, metric, period, weekly_start):
    ranking_data = []
    start_date = weekly_start if period == "weekly" else datetime.min
    metric_meta = LEADERBOARD_METRICS[metric]

    for u in users:
        score = 0

        if metric == "duration":
            sessions = StudySession.query.filter(
                StudySession.user_id == u.id,
                StudySession.start_time >= start_date,
            ).all()
            total_sec = sum(study_session_seconds(s) for s in sessions)
            score = round(total_sec / 3600, 1)

        elif metric == "contribution":
            notes_query = Note.query.filter_by(user_id=u.id)
            edits_query = WikiPageHistory.query.filter_by(user_id=u.id)
            if period == "weekly":
                notes_query = notes_query.filter(Note.created_at >= start_date)
                edits_query = edits_query.filter(WikiPageHistory.created_at >= start_date)

            note_count = notes_query.count()
            edit_count = edits_query.count()
            score = note_count * 10 + edit_count * 2

        elif metric == "streak":
            stats = u.calculate_stats()
            score = stats.get("streak_days", 0)

        elif metric == "pomo":
            pomo_query = PomodoroRecord.query.filter_by(user_id=u.id)
            if period == "weekly":
                pomo_query = pomo_query.filter(PomodoroRecord.completed_at >= start_date)
            score = pomo_query.count()

        elif metric == "oj_ac":
            judge_query = JudgeTask.query.filter_by(user_id=u.id, status="accepted")
            judge_query = judge_query.filter(JudgeTask.problem_id.isnot(None))
            if period == "weekly":
                judge_query = judge_query.filter(JudgeTask.created_at >= start_date)
            score = judge_query.with_entities(JudgeTask.problem_id).distinct().count()

        elif metric == "oj_attempts":
            judge_query = JudgeTask.query.filter_by(user_id=u.id)
            judge_query = judge_query.filter(JudgeTask.problem_id.isnot(None))
            if period == "weekly":
                judge_query = judge_query.filter(JudgeTask.created_at >= start_date)
            score = judge_query.count()

        if score > 0 or u.id == current_user.id:
            ranking_data.append({
                "user": u,
                "score": score,
                "badge_count": len(u.earned_badges[:3]),
            })

    ranking_data.sort(key=lambda x: x["score"], reverse=True)

    rows = []
    my_rank = None
    top_three = []

    for index, data in enumerate(ranking_data, start=1):
        user = data["user"]
        row = {
            "rank": index,
            "user_id": user.id,
            "name": user.real_name or user.username,
            "username": user.username,
            "initial": (user.real_name or user.username or "?")[:1].upper(),
            "class_id": user.class_id,
            "class_name": user.student_class.name if user.student_class else "",
            "group_ids": [group.id for group in user.groups],
            "score": data["score"],
            "unit": metric_meta["unit"],
            "badge_count": data["badge_count"],
            "is_current_user": user.id == current_user.id,
            "profile_url": url_for("public_profile", user_id=user.id),
        }
        rows.append(row)
        if index <= 3:
            top_three.append(row)
        if row["is_current_user"]:
            my_rank = row

    gap_to_next = 0
    if my_rank and my_rank["rank"] > 1:
        gap_to_next = round(rows[my_rank["rank"] - 2]["score"] - my_rank["score"], 1)

    return {
        "metric": metric,
        "period": period,
        "label": metric_meta["label"],
        "icon": metric_meta["icon"],
        "description": metric_meta["description"],
        "unit": metric_meta["unit"],
        "period_label": LEADERBOARD_PERIODS[period],
        "rows": rows,
        "top_three": top_three,
        "top_score": rows[0]["score"] if rows else 0,
        "total_records": len(rows),
        "my_rank": my_rank,
        "gap_to_next": gap_to_next,
    }


@app.route("/leaderboard")
@login_required
def leaderboard():
    metric = request.args.get("metric", "duration")
    period = request.args.get("period", "all")
    scope = request.args.get("scope", "all")
    if metric not in LEADERBOARD_METRICS:
        metric = "duration"
    if period not in LEADERBOARD_PERIODS:
        period = "all"

    users = User.query.all()
    groups = Group.query.order_by(Group.name.asc()).all()
    leaderboard_scopes = {
        "all": {
            "label": "全部",
            "icon": "fa-globe",
            "description": "显示所有学生的排行",
        }
    }
    if current_user.class_id:
        leaderboard_scopes["class"] = {
            "label": "本班",
            "icon": "fa-users",
            "class_id": current_user.class_id,
            "description": current_user.student_class.name if current_user.student_class else "当前班级",
        }
    for group in groups:
        leaderboard_scopes[f"group:{group.id}"] = {
            "label": group.name,
            "icon": "fa-user-friends",
            "group_id": group.id,
            "description": group.description or "自定义学生组",
        }
    if scope not in leaderboard_scopes:
        scope = "all"

    now = now_utc8()
    weekly_start = now - timedelta(days=now.weekday())
    weekly_start = weekly_start.replace(hour=0, minute=0, second=0, microsecond=0)

    leaderboard_payload = {}
    for period_key in LEADERBOARD_PERIODS:
        leaderboard_payload[period_key] = {}
        for metric_key in LEADERBOARD_METRICS:
            leaderboard_payload[period_key][metric_key] = build_leaderboard_snapshot(
                users,
                metric_key,
                period_key,
                weekly_start,
            )

    return render_template(
        "leaderboard.html",
        leaderboard_payload=leaderboard_payload,
        leaderboard_metrics=LEADERBOARD_METRICS,
        leaderboard_periods=LEADERBOARD_PERIODS,
        initial_metric=metric,
        initial_period=period,
        initial_scope=scope,
        leaderboard_scopes=leaderboard_scopes,
    )

# ==========================================
# ✨ 全局便签 (Sticky Notes) 接口
# ==========================================

@app.route('/api/sticky_notes', methods=['GET'])
@login_required
def get_sticky_notes():
    """获取当前用户的所有便签"""
    notes = StickyNote.query.filter_by(user_id=current_user.id).all()
    return jsonify([{
        'id': n.id, 
        'content': n.content, 
        'color': n.color,
        'pos_x': n.pos_x, 
        'pos_y': n.pos_y, 
        'z_index': n.z_index
    } for n in notes])

@app.route('/api/sticky_notes', methods=['POST'])
@login_required
def create_sticky_note():
    """新建一个便签"""
    # 强制限制最多 5 个便签
    count = StickyNote.query.filter_by(user_id=current_user.id).count()
    if count >= 5:
        return jsonify({'success': False, 'error': '为了保持屏幕整洁，最多只能贴 5 个便签哦！'}), 400
    
    data = request.json or {}
    # 新便签错开位置显示
    offset = count * 30
    
    note = StickyNote(
        user_id=current_user.id,
        content="",
        color=data.get('color', 'yellow'),
        pos_x=100 + offset, # 默认 x 坐标
        pos_y=100 + offset, # 默认 y 坐标
        z_index=1000 + count
    )
    db.session.add(note)
    db.session.commit()
    
    return jsonify({
        'success': True, 
        'note': {
            'id': note.id, 'content': note.content, 'color': note.color,
            'pos_x': note.pos_x, 'pos_y': note.pos_y, 'z_index': note.z_index
        }
    })

@app.route('/api/sticky_notes/<int:note_id>', methods=['PUT'])
@login_required
def update_sticky_note(note_id):
    """更新便签 (内容、位置、颜色、层级)"""
    note = StickyNote.query.get_or_404(note_id)
    if note.user_id != current_user.id:
        return jsonify({'success': False, 'error': '无权限'}), 403
    
    data = request.json
    if 'content' in data: note.content = data['content']
    if 'color' in data: note.color = data['color']
    if 'pos_x' in data: note.pos_x = data['pos_x']
    if 'pos_y' in data: note.pos_y = data['pos_y']
    if 'z_index' in data: note.z_index = data['z_index']
    
    db.session.commit()
    return jsonify({'success': True})

@app.route('/api/sticky_notes/<int:note_id>', methods=['DELETE'])
@login_required
def delete_sticky_note(note_id):
    """撕掉(删除)便签"""
    note = StickyNote.query.get_or_404(note_id)
    if note.user_id == current_user.id:
        db.session.delete(note)
        db.session.commit()
        return jsonify({'success': True})
    return jsonify({'success': False}), 403


# ==========================================
# 🍅 番茄钟 (Pomodoro) 接口
# ==========================================
@app.route('/api/pomodoro/today', methods=['GET'])
@login_required
def get_today_pomodoro():
    try:
        today_start = now_utc8().replace(hour=0, minute=0, second=0, microsecond=0)
        count = PomodoroRecord.query.filter(
            PomodoroRecord.user_id == current_user.id,
            PomodoroRecord.completed_at >= today_start
        ).count()
        return jsonify({'success': True, 'count': count})
    except Exception as e:
        return jsonify({'success': False, 'count': 0, 'error': str(e)})

@app.route('/api/pomodoro/complete', methods=['POST'])
@login_required
def complete_pomodoro():
    try:
        data = request.get_json() or {}
        duration = data.get('duration', 45)
        
        record = PomodoroRecord(user_id=current_user.id, duration_minutes=duration)
        db.session.add(record)
        db.session.commit()
        
        today_start = now_utc8().replace(hour=0, minute=0, second=0, microsecond=0)
        count = PomodoroRecord.query.filter(
            PomodoroRecord.user_id == current_user.id,
            PomodoroRecord.completed_at >= today_start
        ).count()
        
        return jsonify({'success': True, 'count': count})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'count': 0, 'error': str(e)})

# 作业类定义

class Assignment(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    title = db.Column(db.String(200), nullable=False)
    content_md = db.Column(db.Text, nullable=True)  
    target_type = db.Column(db.String(50), default='all') 
    target_value = db.Column(db.String(200), nullable=True) 
    created_at = db.Column(db.DateTime, default=now_utc8)
    deadline = db.Column(db.DateTime, nullable=True)
    is_active = db.Column(db.Boolean, default=True)
    created_by_id = db.Column(db.Integer, db.ForeignKey('user.id'))
    created_by = db.relationship('User', backref='created_assignments', foreign_keys=[created_by_id])
    
    # 👇 加上这一行！用来应付旧数据库的 NOT NULL 检查 (0代表独立作业)
    wiki_page_id = db.Column(db.Integer, default=0) 

    star_level = db.Column(db.Integer, default=1) # 💡 任务难度星级 (1-5)
    
    submissions = db.relationship('Submission', backref='assignment', lazy='dynamic')

class Submission(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    assignment_id = db.Column(db.Integer, db.ForeignKey('assignment.id'), nullable=False)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    is_viewed = db.Column(db.Boolean, default=False) # 💡 新增：是否已查看批改反馈
    
    # 快照核心：保存提交时的笔记 ID 和内容
    source_note_id = db.Column(db.Integer, db.ForeignKey('note.id'), nullable=True)
    content_snapshot_md = db.Column(db.Text, nullable=False)
    content = db.Column(db.Text, nullable=True)
    
    # 批阅字段
    grade = db.Column(db.String(20)) # 优/良/中/差
    feedback = db.Column(db.Text)    # 评语
    status = db.Column(db.String(20), default='pending') # pending, graded, redo
    
    submitted_at = db.Column(db.DateTime, default=now_utc8)
    graded_at = db.Column(db.DateTime, nullable=True)

    user = db.relationship('User', backref=db.backref('submissions', lazy='dynamic'))

    stars_earned = db.Column(db.Integer, default=0) # 💡 学生获得的星级


class JudgeTask(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    submission_id = db.Column(db.Integer, db.ForeignKey('submission.id'), nullable=True)
    problem_id = db.Column(db.Integer, db.ForeignKey('problem.id'), nullable=True, index=True)
    assignment_id = db.Column(db.Integer, db.ForeignKey('oj_assignment.id'), nullable=True, index=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True)
    language = db.Column(db.String(32), nullable=False, default='python')
    source_code = db.Column(db.Text, nullable=False)
    test_cases = db.Column(db.JSON, nullable=False, default=list)
    status = db.Column(db.String(32), nullable=False, default='queued')  # queued, running, accepted, failed, system_error
    queue_job_id = db.Column(db.String(64), nullable=True, index=True)
    runtime_image = db.Column(db.String(255), nullable=True)
    time_limit_ms = db.Column(db.Integer, nullable=False, default=2000)
    memory_limit_mb = db.Column(db.Integer, nullable=False, default=256)
    total_score = db.Column(db.Integer, nullable=False, default=0)
    passed_count = db.Column(db.Integer, nullable=False, default=0)
    total_count = db.Column(db.Integer, nullable=False, default=0)
    result_summary = db.Column(db.JSON, nullable=True)
    error_message = db.Column(db.Text, nullable=True)
    created_at = db.Column(db.DateTime, default=now_utc8)
    started_at = db.Column(db.DateTime, nullable=True)
    finished_at = db.Column(db.DateTime, nullable=True)

    submission = db.relationship('Submission', backref=db.backref('judge_tasks', lazy='dynamic'))
    problem = db.relationship('Problem', backref=db.backref('judge_tasks', lazy='dynamic'))
    assignment = db.relationship('OJAssignment', backref=db.backref('judge_tasks', lazy='dynamic'))
    user = db.relationship('User', backref=db.backref('judge_tasks', lazy='dynamic'))


class JudgeTaskResult(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    task_id = db.Column(db.Integer, db.ForeignKey('judge_task.id'), nullable=False)
    case_index = db.Column(db.Integer, nullable=False)
    status = db.Column(db.String(32), nullable=False)
    score = db.Column(db.Integer, nullable=False, default=0)
    time_ms = db.Column(db.Integer, nullable=False, default=0)
    memory_kb = db.Column(db.Integer, nullable=False, default=0)
    input_data = db.Column(db.Text, nullable=True)
    expected_output = db.Column(db.Text, nullable=True)
    actual_output = db.Column(db.Text, nullable=True)
    stderr_text = db.Column(db.Text, nullable=True)
    created_at = db.Column(db.DateTime, default=now_utc8)

    task = db.relationship('JudgeTask', backref=db.backref('results', cascade='all, delete-orphan', lazy='dynamic'))


class Problem(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    problem_code = db.Column(db.String(40), nullable=False, unique=True, index=True)
    title = db.Column(db.String(200), nullable=False)
    slug = db.Column(db.String(200), unique=True, nullable=False)
    statement_md = db.Column(db.Text, nullable=False, default="")
    input_spec_md = db.Column(db.Text, nullable=False, default="")
    output_spec_md = db.Column(db.Text, nullable=False, default="")
    hint_md = db.Column(db.Text, nullable=False, default="")
    source = db.Column(db.String(200), nullable=True)
    difficulty = db.Column(db.String(20), nullable=False, default="medium")
    time_limit_ms = db.Column(db.Integer, nullable=False, default=2000)
    memory_limit_mb = db.Column(db.Integer, nullable=False, default=256)
    allowed_languages = db.Column(db.JSON, nullable=False, default=list)
    is_visible = db.Column(db.Boolean, nullable=False, default=True)
    created_by_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    created_at = db.Column(db.DateTime, default=now_utc8)
    updated_at = db.Column(db.DateTime, default=now_utc8, onupdate=now_utc8)

    created_by = db.relationship('User', backref='created_problems', foreign_keys=[created_by_id])
    test_cases = db.relationship(
        'ProblemTestCase',
        backref='problem',
        cascade='all, delete-orphan',
        order_by='ProblemTestCase.sort_order.asc()'
    )

    @property
    def uid_text(self):
        return f"{self.id:04d}"

    @property
    def testcase_count(self):
        return len(self.test_cases)

    @property
    def sample_cases(self):
        return [case for case in self.test_cases if case.case_type == 'sample']

    @property
    def hidden_cases(self):
        return [case for case in self.test_cases if case.case_type == 'hidden']

    @property
    def allowed_languages_text(self):
        return ", ".join(self.allowed_languages or [])


class ProblemTestCase(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    problem_id = db.Column(db.Integer, db.ForeignKey('problem.id'), nullable=False)
    case_type = db.Column(db.String(20), nullable=False, default='hidden')  # sample, hidden
    input_data = db.Column(db.Text, nullable=False, default="")
    expected_output = db.Column(db.Text, nullable=False, default="")
    score = db.Column(db.Integer, nullable=False, default=1)
    sort_order = db.Column(db.Integer, nullable=False, default=0)
    created_at = db.Column(db.DateTime, default=now_utc8)


class ProblemFile(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    problem_id = db.Column(db.Integer, db.ForeignKey('problem.id'), nullable=False)
    filename = db.Column(db.String(255), nullable=False)
    stored_filename = db.Column(db.String(255), nullable=False)
    file_path = db.Column(db.String(500), nullable=False)
    content_type = db.Column(db.String(120), nullable=True)
    file_size = db.Column(db.Integer, nullable=False, default=0)
    uploaded_by_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    uploaded_at = db.Column(db.DateTime, default=now_utc8)

    problem = db.relationship('Problem', backref=db.backref('files', cascade='all, delete-orphan', lazy='dynamic'))
    uploader = db.relationship('User', backref='uploaded_problem_files', foreign_keys=[uploaded_by_id])

    @property
    def extension(self):
        return os.path.splitext(self.filename or "")[1].lower()

    @property
    def markdown_embed(self):
        if self.extension in ALLOWED_IMAGE_EXTENSIONS:
            return f"![{self.filename}]({self.file_path})"
        return f"[{self.filename}]({self.file_path})"

    @property
    def is_text_editable(self):
        return self.extension in TEXT_EDITABLE_PROBLEM_FILE_EXTENSIONS

    @property
    def is_testdata_file(self):
        return self.extension in TESTDATA_FILE_EXTENSIONS


oj_assignment_classes = db.Table(
    'oj_assignment_classes',
    db.Column('assignment_id', db.Integer, db.ForeignKey('oj_assignment.id'), primary_key=True),
    db.Column('class_id', db.Integer, db.ForeignKey('class.id'), primary_key=True),
)

oj_assignment_groups = db.Table(
    'oj_assignment_groups',
    db.Column('assignment_id', db.Integer, db.ForeignKey('oj_assignment.id'), primary_key=True),
    db.Column('group_id', db.Integer, db.ForeignKey('group.id'), primary_key=True),
)

oj_assignment_users = db.Table(
    'oj_assignment_users',
    db.Column('assignment_id', db.Integer, db.ForeignKey('oj_assignment.id'), primary_key=True),
    db.Column('user_id', db.Integer, db.ForeignKey('user.id'), primary_key=True),
)

oj_assignment_managers = db.Table(
    'oj_assignment_managers',
    db.Column('assignment_id', db.Integer, db.ForeignKey('oj_assignment.id'), primary_key=True),
    db.Column('user_id', db.Integer, db.ForeignKey('user.id'), primary_key=True),
)


class OJAssignment(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    title = db.Column(db.String(200), nullable=False)
    description_md = db.Column(db.Text, nullable=True)
    start_at = db.Column(db.DateTime, nullable=True)
    end_at = db.Column(db.DateTime, nullable=True)
    extension_days = db.Column(db.Integer, nullable=False, default=0)
    visibility = db.Column(db.String(20), nullable=False, default='all')  # all, restricted
    is_active = db.Column(db.Boolean, nullable=False, default=True)
    created_by_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    created_at = db.Column(db.DateTime, default=now_utc8)
    updated_at = db.Column(db.DateTime, default=now_utc8, onupdate=now_utc8)

    created_by = db.relationship('User', backref='created_oj_assignments', foreign_keys=[created_by_id])
    allowed_classes = db.relationship('Class', secondary=oj_assignment_classes, lazy='subquery')
    allowed_groups = db.relationship('Group', secondary=oj_assignment_groups, lazy='subquery')
    allowed_users = db.relationship('User', secondary=oj_assignment_users, lazy='subquery')
    managers = db.relationship('User', secondary=oj_assignment_managers, lazy='subquery')
    problem_links = db.relationship(
        'OJAssignmentProblem',
        backref='assignment',
        cascade='all, delete-orphan',
        order_by='OJAssignmentProblem.sort_order.asc()',
    )
    files = db.relationship('OJAssignmentFile', backref='assignment', cascade='all, delete-orphan', lazy='dynamic')

    @property
    def problem_count(self):
        return len(self.problem_links)

    @property
    def deadline_with_extension(self):
        if not self.end_at:
            return None
        return self.end_at + timedelta(days=self.extension_days or 0)

    @property
    def target_text(self):
        if self.visibility == 'all':
            return '全部学生'
        parts = []
        parts.extend(cls.name for cls in self.allowed_classes)
        parts.extend(group.name for group in self.allowed_groups)
        parts.extend((user.real_name or user.username) for user in self.allowed_users)
        return '、'.join(parts) or '未设置'

    def is_manager(self, user):
        if not user.is_authenticated:
            return False
        return bool(user.is_admin or user.id == self.created_by_id or user in self.managers)

    def is_visible_to(self, user):
        if not user.is_authenticated:
            return False
        if self.is_manager(user):
            return True
        if not self.is_active:
            return False
        if self.visibility == 'all':
            return True
        if user in self.allowed_users:
            return True
        if user.student_class and user.student_class in self.allowed_classes:
            return True
        return not set(user.groups).isdisjoint(set(self.allowed_groups))


class OJAssignmentProblem(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    assignment_id = db.Column(db.Integer, db.ForeignKey('oj_assignment.id'), nullable=False)
    problem_id = db.Column(db.Integer, db.ForeignKey('problem.id'), nullable=False)
    sort_order = db.Column(db.Integer, nullable=False, default=0)

    problem = db.relationship('Problem')


class OJAssignmentFile(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    assignment_id = db.Column(db.Integer, db.ForeignKey('oj_assignment.id'), nullable=False)
    filename = db.Column(db.String(255), nullable=False)
    stored_filename = db.Column(db.String(255), nullable=False)
    file_path = db.Column(db.String(500), nullable=False)
    content_type = db.Column(db.String(120), nullable=True)
    file_size = db.Column(db.Integer, nullable=False, default=0)
    uploaded_by_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    uploaded_at = db.Column(db.DateTime, default=now_utc8)

    uploader = db.relationship('User', backref='uploaded_oj_assignment_files', foreign_keys=[uploaded_by_id])

    @property
    def extension(self):
        return os.path.splitext(self.filename or "")[1].lower()

    @property
    def markdown_embed(self):
        if self.extension in ALLOWED_IMAGE_EXTENSIONS:
            return f"![{self.filename}]({self.file_path})"
        return f"[{self.filename}]({self.file_path})"


def can_view_problem(problem, user):
    if problem.is_visible:
        return True
    return bool(user.is_authenticated and user.is_admin)


def get_problem_or_404_for_current_user(slug):
    problem = Problem.query.filter_by(slug=slug).first_or_404()
    if not can_view_problem(problem, current_user):
        abort(404)
    return problem


def normalize_problem_code(raw_code):
    normalized = (raw_code or "").strip().upper()
    normalized = re.sub(r"\s+", "", normalized)
    return normalized


def next_problem_code():
    existing_codes = {
        code for (code,) in db.session.query(Problem.problem_code).all()
        if code
    }
    next_number = 1000
    while True:
        candidate = f"P{next_number}"
        if candidate not in existing_codes:
            return candidate
        next_number += 1


def validate_problem_code(raw_code, problem_id=None):
    problem_code = normalize_problem_code(raw_code)
    if not problem_code:
        raise ValueError("题目编号不能为空")
    if not re.match(r"^[A-Z][A-Z0-9_-]{1,39}$", problem_code):
        raise ValueError("题目编号需以字母开头，只能包含字母、数字、下划线或短横线")

    query = Problem.query.filter_by(problem_code=problem_code)
    if problem_id is not None:
        query = query.filter(Problem.id != problem_id)
    if query.first():
        raise ValueError(f"题目编号 {problem_code} 已存在")
    return problem_code


OJ_STATUS_META = {
    'queued': {'label': '排队中', 'badge': 'badge-info', 'tone': 'info'},
    'running': {'label': '评测中', 'badge': 'badge-warning', 'tone': 'warning'},
    'accepted': {'label': '通过', 'badge': 'badge-success', 'tone': 'success'},
    'failed': {'label': '未通过', 'badge': 'badge-error', 'tone': 'danger'},
    'system_error': {'label': '系统错误', 'badge': 'badge-error', 'tone': 'danger'},
}


def oj_status_meta(status):
    return OJ_STATUS_META.get(status or '', {'label': status or '未知', 'badge': 'badge-ghost', 'tone': 'neutral'})


OJ_CASE_STATUS_META = {
    'queued': {'label': 'Queued', 'tone': 'info'},
    'running': {'label': 'Running', 'tone': 'warning'},
    'accepted': {'label': 'AC', 'tone': 'success'},
    'wrong_answer': {'label': 'Wrong Answer', 'tone': 'danger'},
    'runtime_error': {'label': 'Runtime Error', 'tone': 'danger'},
    'time_limit_exceeded': {'label': 'Time Limit Exceeded', 'tone': 'warning'},
    'compile_error': {'label': 'Compile Error', 'tone': 'danger'},
    'system_error': {'label': 'System Error', 'tone': 'danger'},
    'failed': {'label': 'Failed', 'tone': 'danger'},
}


def oj_case_status_meta(status):
    return OJ_CASE_STATUS_META.get(status or '', {'label': status or 'Unknown', 'tone': 'neutral'})


def oj_submission_status_meta(task, results=None):
    if task.status in {'queued', 'running'}:
        return oj_case_status_meta(task.status)
    if task.status == 'accepted':
        return oj_case_status_meta('accepted')
    if task.result_summary and task.result_summary.get('failure_reason') == 'compile_error':
        return oj_case_status_meta('compile_error')

    ordered_results = results
    if ordered_results is None:
        ordered_results = task.results.order_by(JudgeTaskResult.case_index.asc()).all()
    first_failed = next((result for result in ordered_results if result.status != 'accepted'), None)
    if first_failed:
        return oj_case_status_meta(first_failed.status)
    return oj_case_status_meta(task.status)


def _jedi_unavailable_response():
    return jsonify({
        'ok': False,
        'message': 'Python 智能提示服务不可用：缺少 jedi 依赖。',
        'items': [],
    }), 503


def _jedi_position(payload):
    line = max(int(payload.get('line') or 1), 1)
    column = max(int(payload.get('column') or 1) - 1, 0)
    return line, column


def _jedi_script(source_code):
    return jedi.Script(source_code or '', path='solution.py')


def _jedi_doc(name):
    try:
        doc = name.docstring(raw=False)
    except Exception:
        doc = ''
    return doc.strip()


OJ_PYTHON_DOC_HINTS = {
    'print': '打印输出：最常用的输出函数，把内容显示到屏幕。sep 控制多个值之间的分隔符，end 控制输出结尾。',
    'input': '读取一行输入：返回字符串，常用于读取 OJ 标准输入中的单行数据。',
    'range': '生成整数序列：常用于 for 循环。range(n) 表示 0 到 n-1。',
    'len': '获取长度：返回字符串、列表、元组、字典等对象中的元素数量。',
    'map': '批量转换：常和 input().split() 配合，把一行输入转换成整数列表。',
    'int': '转成整数：常用于把 input() 读入的字符串转换成数字。',
    'float': '转成浮点数：把字符串或数字转换为小数。',
    'bool': '转成布尔值：返回 True 或 False，常用于条件判断。',
    'str': '转成字符串：常用于拼接输出或格式化结果。',
    'list': '转成列表：常用于 list(map(int, input().split()))。',
    'tuple': '转成元组：返回不可变序列，适合表示不会再修改的一组值。',
    'dict': '创建字典：保存键值对，适合计数、映射和快速查询。',
    'set': '创建集合：自动去重，支持交集、并集等集合运算。',
    'sum': '求和：返回可迭代对象中所有数字的总和。',
    'max': '最大值：返回多个值或序列中的最大元素。',
    'min': '最小值：返回多个值或序列中的最小元素。',
    'sorted': '排序：返回一个新的已排序列表，不修改原对象。',
    'enumerate': '枚举序列：同时得到下标和值，常写作 for i, x in enumerate(a)。',
    'zip': '打包序列：把多个序列按位置组合，常用于同时遍历多个列表。',
    'abs': '绝对值：返回数字的非负大小。',
    'all': '全部为真：可迭代对象中所有元素都为真时返回 True。',
    'any': '任一为真：可迭代对象中至少一个元素为真时返回 True。',
    'reversed': '反向迭代：返回一个从后往前遍历的迭代器。',
    'split': '分割字符串：按分隔符把字符串拆成列表。OJ 中常用 input().split() 拆分一行输入。',
    'append': '追加元素：把一个元素添加到列表末尾。',
    'type': '查看类型：返回对象所属的类型。',
    'try': '异常处理语句：尝试执行一段代码，并用 except 捕获可能出现的错误。',
}

OJ_PYTHON_DOC_DETAILS = {
    'print': '打印输出',
    'input': '读取输入',
    'range': '整数序列',
    'len': '获取长度',
    'map': '批量转换',
    'int': '转成整数',
    'float': '转成浮点数',
    'bool': '转成布尔值',
    'str': '转成字符串',
    'list': '转成列表',
    'tuple': '元组',
    'dict': '字典',
    'set': '集合',
    'sum': '求和',
    'max': '最大值',
    'min': '最小值',
    'sorted': '排序',
    'enumerate': '下标和值',
    'zip': '组合遍历',
    'abs': '绝对值',
    'all': '全部为真',
    'any': '任一为真',
    'reversed': '反向迭代',
    'split': '分割字符串',
    'append': '列表追加',
    'type': '查看类型',
    'try': '异常处理',
}

OJ_PYTHON_SIGNATURE_HINTS = {
    'print': {
        'label': "print(*values, sep=' ', end='\\n', file=None, flush=False)",
        'documentation': '打印输出：把一个或多个值输出到屏幕。sep 用来设置多个值之间的分隔符，end 用来设置输出结尾，默认会换行。',
        'parameters': {
            '*values: object': '要输出的一个或多个对象，会先被转换成字符串。',
            'sep: Optional[str]=...': '多个输出值之间的分隔符，默认是一个空格。',
            'end: Optional[str]=...': '输出结尾，默认是换行符。设为 "" 可以不换行。',
            'file: Optional[SupportsWrite[str]]=...': '输出目标，OJ 中通常保持默认即可。',
            'flush: bool=...': '是否立即刷新输出缓冲，OJ 中通常保持默认即可。',
        },
    },
    'input': {
        'label': 'input(prompt=None)',
        'documentation': '读取一行输入并返回字符串。OJ 中常用 input() 读取一行，再配合 split()、int()、map() 处理数据。',
    },
    'int': {
        'label': 'int(x=0, base=10)',
        'documentation': '转换为整数。常见写法：int(input()) 读取一个整数，list(map(int, input().split())) 读取一行多个整数。base 表示字符串数字的进制。',
    },
    'float': {
        'label': 'float(x=0.0)',
        'documentation': '转换为浮点数，也就是小数。适合题目输入包含小数时使用。',
    },
    'str': {
        'label': 'str(object="")',
        'documentation': '转换为字符串。常用于拼接输出或把数字转换成文本。',
    },
    'len': {
        'label': 'len(s)',
        'documentation': '返回对象长度，例如字符串长度、列表元素个数、字典键值对数量。',
    },
    'range': {
        'label': 'range(start, stop[, step])',
        'documentation': '生成整数序列，常用于 for 循环。range(n) 生成 0 到 n-1；range(a, b) 生成 a 到 b-1。',
    },
    'map': {
        'label': 'map(function, iterable, ...)',
        'documentation': '把函数应用到序列的每个元素。OJ 常用：list(map(int, input().split()))，表示把一行输入拆分后全部转成整数。',
    },
    'list': {
        'label': 'list(iterable=())',
        'documentation': '创建列表。列表可以修改、追加和排序，适合保存一组数据。',
    },
    'tuple': {
        'label': 'tuple(iterable=())',
        'documentation': '创建元组。元组不可修改，适合表示固定的一组值。',
    },
    'dict': {
        'label': 'dict(...)',
        'documentation': '创建字典，用键值对保存数据。常用于计数、映射和快速查询。',
    },
    'set': {
        'label': 'set(iterable=())',
        'documentation': '创建集合。集合会自动去重，适合判断元素是否出现、求交集或并集。',
    },
    'sum': {
        'label': 'sum(iterable, start=0)',
        'documentation': '求和。返回序列中数字的总和，可选 start 作为初始值。',
    },
    'max': {
        'label': 'max(iterable) 或 max(a, b, ...)',
        'documentation': '返回最大值。可以传入一个序列，也可以传入多个值。',
    },
    'min': {
        'label': 'min(iterable) 或 min(a, b, ...)',
        'documentation': '返回最小值。可以传入一个序列，也可以传入多个值。',
    },
    'sorted': {
        'label': 'sorted(iterable, key=None, reverse=False)',
        'documentation': '返回新的已排序列表，不会修改原对象。reverse=True 表示降序。',
    },
    'enumerate': {
        'label': 'enumerate(iterable, start=0)',
        'documentation': '遍历时同时得到下标和值，例如：for i, x in enumerate(a):',
    },
    'zip': {
        'label': 'zip(*iterables)',
        'documentation': '把多个序列按位置配对。常用于同时遍历多个列表。',
    },
    'split': {
        'label': 'split(sep=None, maxsplit=-1)',
        'documentation': '把字符串按分隔符拆成列表。默认按空白字符拆分，OJ 常用 input().split() 读取一行多个数据。',
    },
    'append': {
        'label': 'append(x)',
        'documentation': '把元素 x 添加到列表末尾。这个方法会修改原列表，返回 None。',
    },
}


def _oj_python_doc(name, fallback=''):
    hint = OJ_PYTHON_DOC_HINTS.get(name)
    if hint and fallback:
        return f"{hint}\n\n---\n\n{fallback}"
    return hint or fallback


def _oj_python_detail(name, fallback=''):
    return OJ_PYTHON_DOC_DETAILS.get(name) or fallback


def _oj_python_signature_hint(name):
    return OJ_PYTHON_SIGNATURE_HINTS.get(name) or {}


def _oj_python_signature_label(signature):
    hint = _oj_python_signature_hint(signature.name)
    return hint.get('label') or signature.to_string()


def _oj_python_signature_doc(signature):
    hint = _oj_python_signature_hint(signature.name)
    fallback = _jedi_doc(signature)
    return hint.get('documentation') or _oj_python_doc(signature.name, fallback)


def _oj_python_param_doc(signature, param):
    hint = _oj_python_signature_hint(signature.name)
    param_docs = hint.get('parameters') or {}
    return param_docs.get(param.to_string()) or _jedi_doc(param)


app.jinja_env.globals["oj_status_meta"] = oj_status_meta
app.jinja_env.globals["oj_case_status_meta"] = oj_case_status_meta


def build_oj_judge_cases(problem):
    return [
        {
            'input': testcase.input_data,
            'expected_output': testcase.expected_output,
            'score': testcase.score,
            'case_type': testcase.case_type,
        }
        for testcase in problem.test_cases
    ]


def can_view_oj_judge_task(task, user):
    return bool(user.is_authenticated and (user.is_admin or task.user_id == user.id))


def get_oj_judge_task_or_404(task_id):
    task = JudgeTask.query.filter(JudgeTask.problem_id.isnot(None), JudgeTask.id == task_id).first_or_404()
    if not can_view_oj_judge_task(task, current_user):
        abort(404)
    return task


def translate_python_error(stderr_text):
    error_text = (stderr_text or "").strip()
    if not error_text:
        return ""

    line_match = re.search(r'line\s+(\d+)', error_text)
    line_hint = f"第 {line_match.group(1)} 行：" if line_match else ""

    never_closed_match = re.search(r"SyntaxError:\s*'([^']+)'\s+was never closed", error_text)
    if never_closed_match:
        opener = never_closed_match.group(1)
        closer_map = {"(": ")", "[": "]", "{": "}"}
        closer = closer_map.get(opener, "")
        if closer:
            return f"语法错误：{line_hint}左括号 {opener} 没有闭合，请补上对应的右括号 {closer}。"
        return f"语法错误：{line_hint}{opener} 没有闭合，请检查成对符号。"

    if "unterminated string literal" in error_text or "EOL while scanning string literal" in error_text:
        return f"语法错误：{line_hint}字符串没有闭合，请检查引号是否成对出现。"

    if "expected ':'" in error_text:
        return f"语法错误：{line_hint}这里缺少冒号 :，请检查 if、for、while、def、class 等语句结尾。"

    if "Perhaps you forgot a comma" in error_text:
        return f"语法错误：{line_hint}这里可能缺少逗号，请检查列表、函数参数或多个值之间的分隔。"

    if "invalid syntax" in error_text:
        return f"语法错误：{line_hint}这一行附近的写法不符合 Python 语法，请重点检查括号、冒号、逗号、引号和运算符。"

    translations = [
        ("SyntaxError", "语法错误：代码写法不符合 Python 语法，请检查括号、冒号、引号或语句结构。"),
        ("IndentationError", "缩进错误：Python 对缩进很敏感，请检查空格层级是否一致。"),
        ("NameError", "名称错误：使用了未定义的变量或函数，请检查变量名是否拼写正确、是否已经赋值。"),
        ("TypeError", "类型错误：操作的数据类型不匹配，例如把字符串当数字计算，或函数参数类型不对。"),
        ("ValueError", "值错误：数据格式不符合预期，例如把无法转换的内容转成整数。"),
        ("IndexError", "下标越界：访问了列表、字符串等不存在的位置。"),
        ("KeyError", "键不存在：访问了字典里不存在的键。"),
        ("EOFError", "输入不足：程序还在等待输入，但测试数据已经读完。"),
        ("ZeroDivisionError", "除零错误：程序中出现了除以 0 的运算。"),
        ("ModuleNotFoundError", "模块不存在：判题环境中没有安装你导入的模块。"),
        ("ImportError", "导入错误：模块或模块中的对象导入失败。"),
        ("RecursionError", "递归过深：递归调用次数超过限制，可能需要改成循环或优化递归出口。"),
    ]
    for marker, message in translations:
        if marker in error_text:
            return message
    return "Python 运行错误：程序运行时发生异常，请根据下方错误信息定位问题。"


def build_oj_failure_feedback(task, results):
    if task.status == 'accepted':
        return None
    if task.status in {'queued', 'running'}:
        return None

    first_failed = next((result for result in results if result.status != 'accepted'), None)
    if first_failed and first_failed.status == 'wrong_answer':
        return {
            'title': 'Wrong Answer',
            'message': '程序正常运行了，但输出结果和标准答案不一致。请重点检查输出格式、边界情况和算法逻辑。',
            'detail': '',
            'tone': 'warning',
        }

    if first_failed and first_failed.status == 'runtime_error':
        detail = (first_failed.stderr_text or "").strip()
        return {
            'title': '运行错误',
            'message': translate_python_error(detail) if task.language == 'python' else '程序运行时发生错误，请查看错误输出定位问题。',
            'detail': detail,
            'tone': 'error',
        }

    if first_failed and first_failed.status == 'time_limit_exceeded':
        return {
            'title': 'Time Limit Exceeded',
            'message': '程序运行超时了。请检查是否存在死循环，或尝试优化算法复杂度。',
            'detail': '',
            'tone': 'warning',
        }

    if task.result_summary and task.result_summary.get('failure_reason') == 'compile_error':
        detail = (task.error_message or "").strip()
        return {
            'title': '编译错误',
            'message': '代码没有通过编译，请根据编译器输出检查语法、头文件或类型声明。',
            'detail': detail,
            'tone': 'error',
        }

    if task.status == 'system_error':
        return {
            'title': '系统错误',
            'message': '判题服务遇到系统级错误，这通常不是代码答案本身的问题。',
            'detail': (task.error_message or "").strip(),
            'tone': 'error',
        }

    return {
        'title': '未通过',
        'message': '提交未通过，请查看左侧测试点状态和输出差异。',
        'detail': (task.error_message or "").strip(),
        'tone': 'warning',
    }


def normalize_language_list(raw_languages):
    default_languages = ['python', 'cpp', 'c']
    if not raw_languages:
        return default_languages

    parsed = []
    seen = set()
    for item in raw_languages.split(','):
        lang = item.strip().lower()
        if not lang or lang in seen:
            continue
        parsed.append(lang)
        seen.add(lang)

    return parsed or default_languages


def build_problem_slug(title, problem_id=None):
    base_slug = secure_filename((title or "").strip()).lower().strip('-')
    if not base_slug:
        base_slug = f"problem-{uuid.uuid4().hex[:8]}"

    candidate = base_slug
    suffix = 2
    while True:
        query = Problem.query.filter_by(slug=candidate)
        if problem_id is not None:
            query = query.filter(Problem.id != problem_id)
        if query.first() is None:
            return candidate
        candidate = f"{base_slug}-{suffix}"
        suffix += 1


def collect_problem_testcases_from_form(form):
    case_types = form.getlist('case_type')
    inputs = form.getlist('case_input')
    outputs = form.getlist('case_output')
    scores = form.getlist('case_score')

    testcases = []
    row_count = max(len(case_types), len(inputs), len(outputs), len(scores))
    for index in range(row_count):
        case_type = (case_types[index] if index < len(case_types) else 'hidden').strip().lower() or 'hidden'
        input_data = inputs[index] if index < len(inputs) else ''
        expected_output = outputs[index] if index < len(outputs) else ''
        score_raw = scores[index] if index < len(scores) else '1'

        if not input_data.strip() and not expected_output.strip():
            continue

        try:
            score = max(int(score_raw), 0)
        except (TypeError, ValueError):
            score = 1

        if case_type not in {'sample', 'hidden'}:
            case_type = 'hidden'

        testcases.append(
            {
                'case_type': case_type,
                'input_data': input_data,
                'expected_output': expected_output,
                'score': score,
                'sort_order': len(testcases),
            }
        )

    return testcases


def zip_entry_name(info):
    return (info.filename or "").replace("\\", "/").strip()


def is_safe_zip_entry(info):
    name = zip_entry_name(info)
    if not name or name.endswith("/"):
        return False
    if name.startswith("/") or name.startswith("../") or "/../" in name:
        return False
    return True


def zip_basename(info):
    return os.path.basename(zip_entry_name(info))


def read_zip_text(zip_file, info):
    return decode_zip_text(zip_file.read(info))


def parse_oj_problem_zip_metadata(zip_file, safe_infos, fallback_title):
    meta = {}
    for info in safe_infos:
        if zip_basename(info).lower() in {"problem.json", "metadata.json", "meta.json"}:
            try:
                parsed = json.loads(read_zip_text(zip_file, info) or "{}")
            except json.JSONDecodeError as exc:
                raise ValueError(f"{zip_basename(info)} 不是有效的 JSON：{exc}") from exc
            if isinstance(parsed, dict):
                meta = parsed.get("problem") if isinstance(parsed.get("problem"), dict) else parsed
            break

    text_files = {}
    for info in safe_infos:
        basename = zip_basename(info).lower()
        if basename in {"statement.md", "readme.md", "description.md", "input.md", "output.md", "hint.md"}:
            text_files[basename] = read_zip_text(zip_file, info).strip()

    title = (meta.get("title") or meta.get("name") or fallback_title).strip()
    if not title:
        title = "导入题目"

    raw_languages = meta.get("allowed_languages", meta.get("languages", "python,cpp,c"))
    if isinstance(raw_languages, (list, tuple)):
        raw_languages = ",".join(str(item) for item in raw_languages)

    visible_value = meta.get("is_visible", meta.get("visible", True))
    if isinstance(visible_value, str):
        visible = visible_value.strip().lower() not in {"0", "false", "no", "off", "hidden"}
    else:
        visible = bool(visible_value)

    statement_md = (
        meta.get("statement_md")
        or meta.get("statement")
        or text_files.get("statement.md")
        or text_files.get("readme.md")
        or text_files.get("description.md")
        or f"## 题目描述\n\n{title}\n"
    )

    return {
        "problem_code": (meta.get("problem_code") or meta.get("code") or "").strip(),
        "title": title,
        "slug": (meta.get("slug") or "").strip(),
        "difficulty": (meta.get("difficulty") or "medium").strip().lower(),
        "time_limit_ms": int(meta.get("time_limit_ms") or meta.get("time_limit") or 2000),
        "memory_limit_mb": int(meta.get("memory_limit_mb") or meta.get("memory_limit") or 256),
        "allowed_languages": str(raw_languages or "python,cpp,c"),
        "is_visible": visible,
        "statement_md": statement_md,
        "input_spec_md": meta.get("input_spec_md") or meta.get("input_spec") or text_files.get("input.md", ""),
        "output_spec_md": meta.get("output_spec_md") or meta.get("output_spec") or text_files.get("output.md", ""),
        "hint_md": meta.get("hint_md") or meta.get("hint") or text_files.get("hint.md", ""),
        "source": (meta.get("source") or "ZIP 导入").strip(),
    }


def import_oj_problem_from_zip(file_storage, user):
    if not file_storage or not file_storage.filename:
        raise ValueError("请选择 zip 文件")

    original_filename = secure_filename(file_storage.filename)
    if os.path.splitext(original_filename)[1].lower() != ".zip":
        raise ValueError("仅支持 .zip 文件")

    zip_bytes = file_storage.read()
    if not zip_bytes:
        raise ValueError("zip 文件为空")

    saved_files = []
    reserved_filenames = set()
    try:
        with zipfile.ZipFile(io.BytesIO(zip_bytes)) as zip_file:
            safe_infos = [info for info in zip_file.infolist() if is_safe_zip_entry(info)]
            if not safe_infos:
                raise ValueError("zip 内没有可导入的文件")
            if len(safe_infos) > 500:
                raise ValueError("zip 内文件过多，最多支持 500 个文件")

            total_size = sum(info.file_size for info in safe_infos)
            if total_size > 64 * 1024 * 1024:
                raise ValueError("zip 解压后文件总量超过 64MB")

            fallback_title = os.path.splitext(original_filename)[0] or "导入题目"
            meta = parse_oj_problem_zip_metadata(zip_file, safe_infos, fallback_title)

            if meta["problem_code"]:
                problem_code = validate_problem_code(meta["problem_code"])
            else:
                problem_code = next_problem_code()

            difficulty = meta["difficulty"] if meta["difficulty"] in {"easy", "medium", "hard"} else "medium"
            problem = Problem(
                problem_code=problem_code,
                title=meta["title"],
                slug=build_problem_slug(meta["slug"] or meta["title"]),
                difficulty=difficulty,
                time_limit_ms=max(meta["time_limit_ms"], 100),
                memory_limit_mb=max(meta["memory_limit_mb"], 16),
                allowed_languages=normalize_language_list(meta["allowed_languages"]),
                is_visible=meta["is_visible"],
                created_by_id=user.id,
                statement_md=meta["statement_md"],
                input_spec_md=meta["input_spec_md"],
                output_spec_md=meta["output_spec_md"],
                hint_md=meta["hint_md"],
                source=meta["source"] or None,
            )
            db.session.add(problem)
            db.session.flush()

            imported_files = {}
            skipped_files = 0
            for info in safe_infos:
                basename = zip_basename(info)
                if not basename:
                    continue
                if not is_allowed_problem_file(basename):
                    skipped_files += 1
                    continue

                filename = make_unique_problem_filename(problem.id, basename, reserved_filenames)
                content_bytes = zip_file.read(info)
                problem_file = create_problem_asset_from_bytes(
                    problem.id,
                    filename,
                    content_bytes,
                    "application/octet-stream",
                    user.id,
                )
                db.session.add(problem_file)
                saved_files.append(problem_file.file_path)
                imported_files[filename] = {
                    "file": problem_file,
                    "text": decode_zip_text(content_bytes) if problem_file.is_testdata_file else "",
                }

            grouped_files = {}
            for filename, payload in imported_files.items():
                stem, ext = os.path.splitext(filename)
                ext = ext.lower()
                if ext not in {".in", ".out", ".ans"}:
                    continue
                grouped_files.setdefault(stem, {})[ext] = payload

            imported_cases = 0
            for stem, pair in sorted(grouped_files.items(), key=lambda item: item[0]):
                output_payload = pair.get(".out") or pair.get(".ans")
                if ".in" not in pair or not output_payload:
                    continue

                case_type = "sample" if stem.lower().startswith(("sample", "example")) else "hidden"
                problem.test_cases.append(
                    ProblemTestCase(
                        case_type=case_type,
                        input_data=pair[".in"]["text"],
                        expected_output=output_payload["text"],
                        score=1,
                        sort_order=imported_cases,
                    )
                )
                imported_cases += 1

            db.session.commit()
            return problem, imported_cases, len(imported_files), skipped_files
    except zipfile.BadZipFile as exc:
        db.session.rollback()
        raise ValueError("zip 文件无法读取或已损坏") from exc
    except Exception:
        db.session.rollback()
        for file_path in saved_files:
            remove_problem_asset_file(file_path)
        raise


def ensure_oj_authoring_schema():
    required_tables = {
        'problem': Problem.__table__,
        'problem_test_case': ProblemTestCase.__table__,
        'problem_file': ProblemFile.__table__,
        'oj_assignment': OJAssignment.__table__,
        'oj_assignment_problem': OJAssignmentProblem.__table__,
        'oj_assignment_file': OJAssignmentFile.__table__,
        'oj_assignment_classes': oj_assignment_classes,
        'oj_assignment_groups': oj_assignment_groups,
        'oj_assignment_users': oj_assignment_users,
        'oj_assignment_managers': oj_assignment_managers,
    }
    created_tables = []

    with db.engine.begin() as connection:
        existing_tables = set(inspect(connection).get_table_names())
        for table_name, table in required_tables.items():
            if table_name in existing_tables:
                continue
            table.create(bind=connection, checkfirst=True)
            created_tables.append(table_name)

        if 'problem' in existing_tables:
            inspector = inspect(connection)
            problem_columns = {column['name'] for column in inspector.get_columns('problem')}
            if 'problem_code' not in problem_columns:
                connection.execute(db.text("ALTER TABLE problem ADD COLUMN problem_code VARCHAR(40)"))
            problem_rows = connection.execute(db.text("SELECT id, problem_code FROM problem ORDER BY id ASC")).fetchall()
            for row in problem_rows:
                if row.problem_code:
                    continue
                connection.execute(
                    db.text("UPDATE problem SET problem_code = :problem_code WHERE id = :problem_id"),
                    {"problem_code": f"P{999 + row.id}", "problem_id": row.id},
                )
            problem_indexes = {index['name'] for index in inspector.get_indexes('problem')}
            if 'ix_problem_problem_code' not in problem_indexes:
                connection.execute(db.text("CREATE UNIQUE INDEX IF NOT EXISTS ix_problem_problem_code ON problem (problem_code)"))

        if 'judge_task' in existing_tables:
            inspector = inspect(connection)
            judge_task_columns = {column['name'] for column in inspector.get_columns('judge_task')}
            if 'problem_id' not in judge_task_columns:
                connection.execute(db.text("ALTER TABLE judge_task ADD COLUMN problem_id INTEGER"))
            if 'assignment_id' not in judge_task_columns:
                connection.execute(db.text("ALTER TABLE judge_task ADD COLUMN assignment_id INTEGER"))
            judge_task_indexes = {index['name'] for index in inspector.get_indexes('judge_task')}
            if 'ix_judge_task_problem_id' not in judge_task_indexes:
                connection.execute(db.text("CREATE INDEX IF NOT EXISTS ix_judge_task_problem_id ON judge_task (problem_id)"))
            if 'ix_judge_task_assignment_id' not in judge_task_indexes:
                connection.execute(db.text("CREATE INDEX IF NOT EXISTS ix_judge_task_assignment_id ON judge_task (assignment_id)"))

    return created_tables


def get_visible_oj_assignments_for_user(user):
    base_query = OJAssignment.query.order_by(OJAssignment.start_at.desc(), OJAssignment.created_at.desc(), OJAssignment.id.desc())
    if not user.is_authenticated:
        return []
    if user.is_admin:
        return base_query.all()
    return [assignment for assignment in base_query.all() if assignment.is_visible_to(user)]


def get_oj_assignment_or_404_visible(assignment_id):
    assignment = OJAssignment.query.get_or_404(assignment_id)
    if not assignment.is_visible_to(current_user):
        abort(404)
    return assignment


def resolve_oj_assignment_targets(assignment, class_ids, group_ids, user_ids, manager_ids):
    class_ids = {int(value) for value in class_ids if str(value).isdigit()}
    group_ids = {int(value) for value in group_ids if str(value).isdigit()}
    user_ids = {int(value) for value in user_ids if str(value).isdigit()}
    manager_ids = {int(value) for value in manager_ids if str(value).isdigit()}

    assignment.allowed_classes = Class.query.filter(Class.id.in_(class_ids)).order_by(Class.name.asc()).all() if class_ids else []
    assignment.allowed_groups = Group.query.filter(Group.id.in_(group_ids)).order_by(Group.name.asc()).all() if group_ids else []
    assignment.allowed_users = User.query.filter(User.id.in_(user_ids)).order_by(User.username.asc()).all() if user_ids else []
    assignment.managers = User.query.filter(User.id.in_(manager_ids)).order_by(User.username.asc()).all() if manager_ids else []


def resolve_oj_assignment_problem_links(assignment, raw_problem_ids):
    parsed_ids = []
    seen_ids = set()
    for raw_value in raw_problem_ids:
        if not str(raw_value).isdigit():
            continue
        problem_id = int(raw_value)
        if problem_id in seen_ids:
            continue
        seen_ids.add(problem_id)
        parsed_ids.append(problem_id)

    problems = Problem.query.filter(Problem.id.in_(parsed_ids)).all() if parsed_ids else []
    problem_map = {problem.id: problem for problem in problems}
    assignment.problem_links.clear()
    for sort_order, problem_id in enumerate(parsed_ids):
        problem = problem_map.get(problem_id)
        if not problem:
            continue
        assignment.problem_links.append(
            OJAssignmentProblem(problem_id=problem.id, sort_order=sort_order)
        )


def build_oj_assignment_form_state(assignment=None):
    return {
        'title': assignment.title if assignment else '',
        'description_md': assignment.description_md if assignment else '',
        'start_at': assignment.start_at.strftime('%Y-%m-%dT%H:%M') if assignment and assignment.start_at else '',
        'end_at': assignment.end_at.strftime('%Y-%m-%dT%H:%M') if assignment and assignment.end_at else '',
        'extension_days': assignment.extension_days if assignment else 0,
        'visibility': assignment.visibility if assignment else 'all',
        'is_active': '1' if (assignment.is_active if assignment else True) else '0',
        'selected_class_ids': [item.id for item in assignment.allowed_classes] if assignment else [],
        'selected_group_ids': [item.id for item in assignment.allowed_groups] if assignment else [],
        'selected_user_ids': [item.id for item in assignment.allowed_users] if assignment else [],
        'selected_manager_ids': [item.id for item in assignment.managers] if assignment else [],
        'selected_problem_ids': [link.problem_id for link in assignment.problem_links] if assignment else [],
    }


def get_assignment_for_problem_request(problem=None):
    assignment_id = request.values.get('assignment_id', type=int)
    if not assignment_id:
        return None
    assignment = get_oj_assignment_or_404_visible(assignment_id)
    if problem is not None:
        assignment_problem_ids = {link.problem_id for link in assignment.problem_links}
        if problem.id not in assignment_problem_ids:
            abort(404)
    return assignment


def assignment_problem_latest_status(assignment, user):
    problem_ids = [link.problem_id for link in assignment.problem_links]
    latest_by_problem = {}
    accepted_problem_ids = set()
    if not user.is_authenticated or not problem_ids:
        return latest_by_problem, accepted_problem_ids

    tasks = (
        JudgeTask.query
        .filter(
            JudgeTask.user_id == user.id,
            JudgeTask.problem_id.in_(problem_ids),
            JudgeTask.assignment_id == assignment.id,
        )
        .order_by(JudgeTask.problem_id.asc(), JudgeTask.created_at.desc(), JudgeTask.id.desc())
        .all()
    )
    for task in tasks:
        if task.status == 'accepted':
            accepted_problem_ids.add(task.problem_id)
        latest_by_problem.setdefault(task.problem_id, task)
    return latest_by_problem, accepted_problem_ids


def oj_task_display_meta(task):
    if not task:
        return {'label': '未提交', 'badge': 'badge-ghost'}
    if task.status == 'accepted':
        return {'label': 'AC', 'badge': 'badge-success'}
    if task.result_summary and task.result_summary.get('failure_reason') == 'compile_error':
        return {'label': '编译错误', 'badge': 'badge-error'}

    first_failed = task.results.order_by(JudgeTaskResult.case_index.asc()).filter(JudgeTaskResult.status != 'accepted').first()
    if first_failed:
        mapping = {
            'wrong_answer': {'label': 'WA', 'badge': 'badge-error'},
            'runtime_error': {'label': 'RE', 'badge': 'badge-error'},
            'time_limit_exceeded': {'label': 'TLE', 'badge': 'badge-warning'},
        }
        return mapping.get(first_failed.status, {'label': first_failed.status, 'badge': 'badge-error'})

    meta = oj_status_meta(task.status)
    return {'label': meta['label'], 'badge': meta['badge']}


def get_oj_assignment_participants(assignment):
    if assignment.visibility == 'all':
        return User.query.filter_by(is_admin=False).order_by(User.id.asc()).all()

    users = {}
    for user in assignment.allowed_users:
        if not user.is_admin:
            users[user.id] = user

    for class_item in assignment.allowed_classes:
        for user in class_item.users:
            if not user.is_admin:
                users[user.id] = user

    for group_item in assignment.allowed_groups:
        for user in group_item.users:
            if not user.is_admin:
                users[user.id] = user

    return sorted(
        users.values(),
        key=lambda item: (
            (item.student_class.name if item.student_class else 'zzzz'),
            item.real_name or item.username,
            item.username,
        ),
    )


def parse_assignment_tags(html_content, wiki_page_id):
    """
    支持“ID钢印”机制的终极解析器
    """
    # 💡 匹配 [作业#15:标题] 或是以前的 [作业:标题]
    # match.group(1) 捕获 "#15" (如果不存在则是 None)
    # match.group(2) 捕获标题内容
    pattern = r'\[作业(#\d+)?[:：]\s*(.*?)\]'
    
    def replace_func(match):
        id_part = match.group(1)
        title = match.group(2).strip()
        
        assign = None
        if id_part:
            # 如果有钢印，绝对精准地通过 ID 查找
            assign_id = int(id_part.replace('#', ''))
            assign = Assignment.query.get(assign_id)
            # 以防万一数据库没更新，我们强制将标题显示为 Markdown 里的新标题
            if assign: title = assign.title 
        else:
            # 兼容老数据：按标题查找
            assign = Assignment.query.filter_by(wiki_page_id=wiki_page_id, title=title).first()
        
        if not assign:
            if current_user.is_authenticated and current_user.is_admin:
                return f'<div class="badge badge-error gap-2 my-2 py-3 px-4 rounded-xl font-bold">⚠️ 请点击“保存”以激活作业：{title}</div>'
            return f'<span class="text-stone-400 italic text-sm">[作业 "{title}" 准备中...]</span>'

        # ============ 管理员视角 ============
        if current_user.is_authenticated and current_user.is_admin:
            return f'''
            <div class="my-6 p-6 border-2 border-dashed border-primary/30 rounded-[2rem] bg-primary/5 flex flex-col md:flex-row items-center justify-between gap-4 shadow-inner">
                <div class="flex items-center gap-4">
                    <div class="w-12 h-12 rounded-2xl bg-primary/10 flex items-center justify-center text-primary text-xl">
                        <i class="fas fa-tasks"></i>
                    </div>
                    <div>
                        <p class="text-[10px] uppercase font-black opacity-40 tracking-widest leading-none mb-1">Assignment ID:{assign.id}</p>
                        <h4 class="font-black text-stone-800 dark:text-white text-lg">{title}</h4>
                    </div>
                </div>
                <button onclick="app.assignments.openConsole({assign.id}, '{title}')" class="btn btn-primary rounded-2xl px-8 shadow-lg shadow-primary/20 hover:scale-105 transition-transform">
                    进入批阅矩阵
                </button>
            </div>
            '''
        
        # ============ 学生视角 ============
        elif current_user.is_authenticated:
            sub = Submission.query.filter_by(assignment_id=assign.id, user_id=current_user.id).first()
            
            status_dot = '<div class="w-2.5 h-2.5 rounded-full bg-stone-300"></div>'
            status_text = "未提交"
            btn_text = "📤 提交我的笔记"
            btn_class = "btn-accent shadow-accent/20"
            
            if sub:
                if sub.status == 'graded':
                    status_dot = '<div class="w-2.5 h-2.5 rounded-full bg-emerald-500 shadow-[0_0_8px_rgba(16,185,129,0.5)]"></div>'
                    status_text = f"已批改 ({sub.grade})"
                    btn_text = "查看反馈"
                    btn_class = "btn-ghost bg-stone-100 dark:bg-base-200"
                else:
                    status_dot = '<div class="w-2.5 h-2.5 rounded-full bg-amber-400 animate-pulse shadow-[0_0_8px_rgba(251,191,36,0.5)]"></div>'
                    status_text = "待批阅"
                    btn_text = "更新快照"
                    btn_class = "btn-primary shadow-primary/20"

            return f'''
            <div class="my-6 p-6 border border-stone-200 dark:border-white/5 rounded-[2rem] bg-white dark:bg-base-300/50 flex flex-col sm:flex-row items-center justify-between gap-4 shadow-sm">
                <div class="flex items-center gap-4">
                    <div class="w-12 h-12 rounded-2xl bg-stone-50 dark:bg-base-200 flex items-center justify-center text-2xl">📝</div>
                    <div>
                        <h4 class="font-black text-stone-800 dark:text-stone-100 text-lg leading-tight">{title}</h4>
                        <div class="flex items-center gap-2 mt-1.5">
                            {status_dot}
                            <span class="text-[10px] font-black uppercase tracking-tighter opacity-50">{status_text}</span>
                        </div>
                    </div>
                </div>
                <button onclick="app.assignments.openSubmitModal({assign.id}, '{title}')" class="btn {btn_class} rounded-2xl px-8 transition-all active:scale-95">
                    {btn_text}
                </button>
            </div>
            '''
        return ""
            
    return re.sub(pattern, replace_func, html_content)

@app.route("/api/assignments/submit", methods=["POST"])
@login_required
def submit_assignment():
    data = request.json
    assign_id = data.get("assignment_id")
    note_id = data.get("note_id")
    
    note = Note.query.get_or_404(note_id)
    if note.user_id != current_user.id:
        return jsonify({"success": False, "error": "只能提交自己的笔记"}), 403
        
    # 核心：创建提交记录或覆盖现有记录
    sub = Submission.query.filter_by(assignment_id=assign_id, user_id=current_user.id).first()
    
    if sub and sub.status == 'graded':
        return jsonify({"success": False, "error": "作业已批改，无法修改"}), 403
        
    if not sub:
        sub = Submission(assignment_id=assign_id, user_id=current_user.id)
        db.session.add(sub)
        
    # 💡 关键：执行快照
    sub.source_note_id = note.id
    sub.content_snapshot_md = note.content_md # 复制当前内容
    sub.submitted_at = now_utc8()
    sub.status = 'pending'
    
    db.session.commit()
    return jsonify({"success": True})

@app.route("/api/assignments/get_submission", methods=["POST"])
@login_required
def get_submission_content():
    if not current_user.is_admin: abort(403)
    data = request.json
    assign_id = data.get("assignment_id")
    user_id = data.get("user_id")
    
    sub = Submission.query.filter_by(assignment_id=assign_id, user_id=user_id).first()
    if not sub:
        return jsonify({"success": False, "error": "该学生尚未提交"})
    
    # 💡 使用 markdown 转换快照内容给前端看
    content_html = markdown(sub.content_snapshot_md or "*内容为空*")
    return jsonify({
        "success": True, 
        "html": content_html, 
        "feedback": sub.feedback or "", 
        "grade": sub.grade or "A"
    })

@app.route("/api/assignments/<int:assign_id>/submissions")
@login_required
def get_assignment_submissions(assign_id):
    """
    教师端核心：获取所有学生（包含已交和未交）的作业提交矩阵
    """
    if not current_user.is_admin: 
        abort(403)
    
    # 1. 查找作业配置
    assign = Assignment.query.get_or_404(assign_id)
    
    # 2. 获取所有学生（非管理员用户），按班级和学号排序
    all_students = User.query.filter_by(is_admin=False).order_by(User.class_id, User.student_id).all()
    
    # 3. 获取该作业目前已有的提交记录
    subs = Submission.query.filter_by(assignment_id=assign_id).all()
    # 转换为字典提高查询效率：{user_id: submission_object}
    sub_map = {s.user_id: s for s in subs}
    
    students_data = []
    for stu in all_students:
        s = sub_map.get(stu.id)
        students_data.append({
            "user_id": stu.id,
            "username": stu.username,
            "real_name": stu.real_name or stu.username,
            "class_name": stu.class_name or "未分配",
            "status": s.status if s else "unsubmitted", # 状态：unsubmitted(灰), pending(黄), graded(绿)
            "grade": s.grade if s else "",
            "submitted_at": s.submitted_at.strftime("%m-%d %H:%M") if s else ""
        })
        
    return jsonify({
        "title": assign.title, 
        "students": students_data
    })


# 1. 渲染精选 Wiki 管理页面
@app.route("/admin/manage_featured_wikis")
@login_required
def manage_featured_wikis():
    if not current_user.is_admin:
        abort(403)
        
    # 查询已精选的 Wiki (按设置的顺序排序)
    featured_wikis = Wiki.query.filter_by(is_featured=True).order_by(Wiki.featured_order.asc()).all()
    # 查询未精选的 Wiki (按创建时间倒序)
    unfeatured_wikis = Wiki.query.filter_by(is_featured=False).order_by(Wiki.created_at.desc()).all()
    
    return render_template("admin/manage_featured_wikis.html", 
                           featured=featured_wikis, 
                           unfeatured=unfeatured_wikis)

# 2. 接收 AJAX 请求，保存新的顺序
@app.route("/api/admin/update_featured_wikis", methods=["POST"])
@login_required
def update_featured_wikis():
    if not current_user.is_admin:
        return jsonify({"success": False, "error": "权限不足"}), 403
        
    data = request.json
    featured_ids = data.get("featured_ids", []) # 这是一个有序的 Wiki ID 列表
    
    try:
        # 先将所有 Wiki 的精选状态重置
        Wiki.query.update({Wiki.is_featured: False, Wiki.featured_order: 0})
        
        # 根据前端传来的 ID 列表，重新设置精选状态和顺序
        for index, w_id in enumerate(featured_ids):
            wiki = Wiki.query.get(w_id)
            if wiki:
                wiki.is_featured = True
                wiki.featured_order = index
                
        db.session.commit()
        return jsonify({"success": True})
    except Exception as e:
        db.session.rollback()
        return jsonify({"success": False, "error": str(e)})


# 书签管理路由

# ==========================================
# 管理员：书签种类管理后台
# ==========================================

@app.route('/admin/bookmarks')
@login_required
def manage_bookmarks():
    if not current_user.is_admin:
        flash('权限不足，仅管理员可访问此页面', 'error')
        return redirect(url_for('index'))
    
    # 按照稀有度和创建时间排序
    bookmarks = BookmarkType.query.order_by(
        db.case(
            {'legendary': 1, 'epic': 2, 'rare': 3, 'common': 4}, 
            value=BookmarkType.rarity
        ),
        BookmarkType.created_at.desc()
    ).all()
    
    return render_template('admin/manage_bookmarks.html', bookmarks=bookmarks)

@app.route('/admin/bookmarks/save', methods=['POST'])
@login_required
def save_bookmark_type():
    if not current_user.is_admin:
        flash('权限不足', 'error')
        return redirect(url_for('manage_bookmarks'))
        
    b_id = request.form.get('id')
    name = request.form.get('name', '').strip()
    description = request.form.get('description', '').strip()
    rarity = request.form.get('rarity', 'common')
    condition_type = request.form.get('condition_type', 'manual')
    condition_value = int(request.form.get('condition_value', 0))
    
    if not name:
        flash('书签名称不能为空', 'error')
        return redirect(url_for('manage_bookmarks'))

    # 💡 核心修改：处理图片上传
    icon_file = request.files.get('icon_file')
    icon_url = None

    if icon_file and icon_file.filename:
        # 确保保存目录存在
        upload_dir = os.path.join(app.root_path, 'static', 'uploads', 'bookmarks')
        os.makedirs(upload_dir, exist_ok=True)
        
        # 生成安全且唯一的文件名 (防冲突)
        ext = icon_file.filename.rsplit('.', 1)[1].lower() if '.' in icon_file.filename else 'png'
        filename = secure_filename(f"bookmark_{int(time.time())}.{ext}")
        save_path = os.path.join(upload_dir, filename)
        
        icon_file.save(save_path)
        icon_url = f"/static/uploads/bookmarks/{filename}" # 存入数据库的相对路径

    # 数据库操作
    if b_id:
        # 编辑已有书签
        b_type = BookmarkType.query.get_or_404(int(b_id))
        if icon_url:
            b_type.icon = icon_url # 如果上传了新图就替换，否则保留老图
        msg = f'书签 "{name}" 已成功更新！'
    else:
        # 新建书签
        if not icon_url:
            flash('新建书签必须上传图片！', 'error')
            return redirect(url_for('manage_bookmarks'))
        b_type = BookmarkType()
        b_type.icon = icon_url
        db.session.add(b_type)
        msg = f'新书签 "{name}" 已成功创建！'
        
    b_type.name = name
    b_type.description = description
    b_type.rarity = rarity
    b_type.condition_type = condition_type
    b_type.condition_value = condition_value
    
    try:
        db.session.commit()
        flash(msg, 'success')
    except Exception as e:
        db.session.rollback()
        flash('保存失败，请检查数据库', 'error')
        
    return redirect(url_for('manage_bookmarks'))

@app.route('/admin/bookmarks/<int:b_id>/delete', methods=['POST'])
@login_required
def delete_bookmark_type(b_id):
    if not current_user.is_admin:
        return jsonify({'success': False, 'error': '权限不足'}), 403
        
    b_type = BookmarkType.query.get_or_404(b_id)
    try:
        db.session.delete(b_type)
        db.session.commit()
        flash(f'书签 "{b_type.name}" 已被删除', 'success')
    except Exception as e:
        db.session.rollback()
        flash('删除失败，可能已有用户获得了该书签', 'error')
        
    return redirect(url_for('manage_bookmarks'))


# ==========================================
# 用户端：获取我的书签图鉴与坐标 API
# ==========================================
@app.route('/api/bookmarks/my')
@login_required
def get_my_bookmarks():
    # 按照获得时间倒序，最新的在最前面
    user_bms = UserBookmark.query.filter_by(user_id=current_user.id).order_by(UserBookmark.earned_at.desc()).all()
    
    results = []
    for ub in user_bms:
        b_type = ub.bookmark_type
        results.append({
            'id': ub.id,
            'name': b_type.name,
            'icon': b_type.icon,
            'rarity': b_type.rarity,
            'is_placed': ub.is_placed,
            'target_title': ub.target_title,
            'target_snippet': ub.target_snippet,
            'target_url': ub.target_url,
            'target_block_id': ub.target_block_id,
            'earned_at': ub.earned_at.strftime('%Y-%m-%d') if ub.earned_at else ''
        })
        
        # 如果是新获得的，顺便把提醒状态消除
        if not ub.is_notified:
            ub.is_notified = True

    # 顺带把状态保存
    db.session.commit()
            
    return jsonify({'success': True, 'bookmarks': results})


# ==========================================
# 书签动作：夹入书签与渲染当前页书签
# ==========================================

@app.route('/api/bookmarks/place', methods=['POST'])
@login_required
def place_bookmark():
    data = request.json
    ub_id = data.get('user_bookmark_id')
    ub = UserBookmark.query.get_or_404(ub_id)
    
    # 安全校验：只能用自己的书签
    if ub.user_id != current_user.id:
        return jsonify({'success': False, 'error': '无权操作'}), 403
        
    ub.is_placed = True
    ub.target_url = data.get('target_url')
    ub.target_block_id = data.get('target_block_id')
    ub.target_title = data.get('target_title', '未命名页面')
    
    # 处理摘要文字，超过30个字加省略号
    snippet = data.get('target_snippet', '')
    ub.target_snippet = snippet[:30] + '...' if len(snippet) > 30 else snippet
    
    ub.placed_at = datetime.utcnow()
    
    try:
        db.session.commit()
        return jsonify({'success': True, 'icon': ub.bookmark_type.icon})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)})

# ==========================================
# 书签动作：取下已夹入的书签
# ==========================================
@app.route('/api/bookmarks/unplace', methods=['POST'])
@login_required
def unplace_bookmark():
    data = request.json
    ub_id = data.get('user_bookmark_id')
    ub = UserBookmark.query.get_or_404(ub_id)
    
    # 安全校验：只能操作自己的书签
    if ub.user_id != current_user.id:
        return jsonify({'success': False, 'error': '无权操作'}), 403
        
    # 清空夹入位置的数据，恢复为闲置状态
    ub.is_placed = False
    ub.target_url = None
    ub.target_block_id = None
    ub.target_title = None
    ub.target_snippet = None
    ub.placed_at = None
    
    try:
        db.session.commit()
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)})


# ==========================================
# 书签动作：渲染当前页书签 (互关学友版)
# ==========================================
@app.route('/api/bookmarks/page')
def get_page_bookmarks():
    target_url = request.args.get('url')
    if not target_url:
        return jsonify({'success': False, 'bookmarks': []})
        
    # 查询夹在这个 URL 上的所有书签
    placed_bms = UserBookmark.query.filter_by(target_url=target_url, is_placed=True).all()
    
    results = []
    
    # 必须是登录状态才能看书签
    if current_user.is_authenticated:
        # 💡 性能优化关键：
        # 利用你写好的 study_partners 属性，一次性查出所有学友的 ID 集合
        # 这样在下面的循环中就是 O(1) 的内存匹配，极大减轻数据库压力
        partner_ids = {u.id for u in current_user.study_partners.all()}
        
        for bm in placed_bms:
            is_mine = (bm.user_id == current_user.id)
            is_partner = (bm.user_id in partner_ids)
            
            # 🛡️ 严格拦截：只有当书签是【自己的】或者【互相关注的学友】时，才下发数据
            if is_mine or is_partner:
                results.append({
                    'block_id': bm.target_block_id,
                    'icon': bm.bookmark_type.icon,
                    'user_name': bm.user.username,
                    'is_mine': is_mine
                })
                
    return jsonify({'success': True, 'bookmarks': results})

# 测试使用
@app.route('/admin/bookmarks/<int:b_id>/test_grant', methods=['POST'])
@login_required
def test_grant_bookmark(b_id):
    if not current_user.is_admin:
        return jsonify({'success': False})
    
    # 检查是否已经拥有一个未使用的同款书签
    existing = UserBookmark.query.filter_by(user_id=current_user.id, bookmark_type_id=b_id, is_placed=False).first()
    if existing:
        flash('你的抽屉里已经有一枚闲置的该书签了，先去把它夹在书里吧！', 'warning')
    else:
        # 强制发给自己
        new_bm = UserBookmark(user_id=current_user.id, bookmark_type_id=b_id)
        db.session.add(new_bm)
        db.session.commit()
        flash('🎉 测试发放成功！快去右上角【书签抽屉】或【个人主页】看看吧！', 'success')
        
    return redirect(url_for('manage_bookmarks'))

# ==========================================
# 📊 星露谷式结算系统 (高级沉浸版)
# ==========================================
@app.route('/api/reports/check')
@login_required
def check_learning_reports():
    status = UserActiveStatus.query.get(current_user.id)
    if not status:
        return jsonify({"has_report": False})
        
    now = now_utc8()
    today = now.date()
    
    last_d = status.last_daily_report
    last_w = status.last_weekly_report
    last_m = status.last_monthly_report
    
    report_type = None
    start_dt = None
    end_dt = None
    title = ""
    
    # 获取自然周一和自然月一号
    this_month_first = today.replace(day=1)
    this_monday = today - timedelta(days=today.weekday())
    
    # 优先级 1: 月报 (如果没看过上个月的)
    if last_m is None or last_m < this_month_first:
        report_type = 'monthly'
        end_dt = datetime.combine(this_month_first, datetime.min.time())
        start_dt = (this_month_first - timedelta(days=1)).replace(day=1)
        start_dt = datetime.combine(start_dt, datetime.min.time())
        title = f"🌕 {start_dt.strftime('%Y年%m月')} 学习月报"
        
    # 优先级 2: 周报 (如果没看过上周的)
    elif last_w is None or last_w < this_monday:
        report_type = 'weekly'
        end_dt = datetime.combine(this_monday, datetime.min.time())
        start_dt = end_dt - timedelta(days=7)
        title = f"📅 上周结算 ({start_dt.strftime('%m.%d')} - {(end_dt-timedelta(days=1)).strftime('%m.%d')})"
        
    # 优先级 3: 日报 (如果没看过昨天的)
    elif last_d is None or last_d < today:
        report_type = 'daily'
        end_dt = datetime.combine(today, datetime.min.time())
        start_dt = end_dt - timedelta(days=1)
        title = f"📝 昨日出货单 ({start_dt.strftime('%m.%d')})"
        
    if not report_type:
        return jsonify({"has_report": False})
        
    # 1. 基础数据查询
    sessions = StudySession.query.filter(StudySession.user_id == current_user.id, StudySession.start_time >= start_dt, StudySession.start_time < end_dt).all()
    duration = sum(study_session_seconds(s) for s in sessions)
    duration_mins = int(duration / 60)
    
    notes = Note.query.filter(Note.user_id == current_user.id, Note.created_at >= start_dt, Note.created_at < end_dt).count()
    pomos = PomodoroRecord.query.filter(PomodoroRecord.user_id == current_user.id, PomodoroRecord.completed_at >= start_dt, PomodoroRecord.completed_at < end_dt).count()
    
    # 💡 2. 新增进阶数据：Wiki贡献与社区互动
    wiki_edits = WikiPageHistory.query.filter(WikiPageHistory.user_id == current_user.id, WikiPageHistory.created_at >= start_dt, WikiPageHistory.created_at < end_dt).count()
    comments = Comment.query.filter(Comment.user_id == current_user.id, Comment.created_at >= start_dt, Comment.created_at < end_dt).count()
    
    # 💡 3. 计算“综合源力值” (EXP) 与 评级
    # 公式：每分钟1分 + 每番茄5分 + 每笔记15分 + 每次Wiki贡献10分 + 每次评论2分
    exp = duration_mins + (pomos * 5) + (notes * 15) + (wiki_edits * 10) + (comments * 2)
    
    rating = 'C'
    if report_type == 'daily':
        if exp >= 150: rating = 'S'
        elif exp >= 60: rating = 'A'
        elif exp >= 30: rating = 'B'
    elif report_type == 'weekly':
        if exp >= 700: rating = 'S'
        elif exp >= 300: rating = 'A'
        elif exp >= 100: rating = 'B'
    elif report_type == 'monthly':
        if exp >= 2500: rating = 'S'
        elif exp >= 1000: rating = 'A'
        elif exp >= 400: rating = 'B'
    
    # 4. 获取徽章
    badges = UserBadge.query.filter(UserBadge.user_id == current_user.id, UserBadge.earned_at >= start_dt, UserBadge.earned_at < end_dt).all()
    badge_list = []
    for b in badges:
        icon_val = b.badge.icon
        is_img = is_image_icon(icon_val)
        badge_list.append({
            "name": b.badge.name,
            "icon": badge_icon_url(icon_val) if is_img else icon_val,
            "is_image": is_img
        })
        
    # 💡 智能静默逻辑（新增的维度也纳入判定，全为0才不弹窗）
    if duration_mins == 0 and notes == 0 and pomos == 0 and wiki_edits == 0 and comments == 0 and len(badge_list) == 0:
        if report_type == 'monthly':
            status.last_monthly_report = today
            status.last_weekly_report = today
            status.last_daily_report = today
        elif report_type == 'weekly':
            status.last_weekly_report = today
            status.last_daily_report = today
        elif report_type == 'daily':
            status.last_daily_report = today
        db.session.commit()
        return jsonify({"has_report": False})
        
    return jsonify({
        "has_report": True,
        "type": report_type,
        "title": title,
        "stats": {
            "duration": duration_mins,
            "pomos": pomos,
            "notes": notes,
            "wiki_edits": wiki_edits,
            "comments": comments,
            "exp": exp,
            "rating": rating,
            "badges": badge_list
        }
    })

@app.route('/api/reports/mark_read', methods=['POST'])
@login_required
def mark_report_read():
    data = request.json
    report_type = data.get('type')
    status = UserActiveStatus.query.get(current_user.id)
    
    if status:
        today = now_utc8().date()
        # 只要看了高级别报告，低级别的也视为已读
        if report_type == 'monthly':
            status.last_monthly_report = today
            status.last_weekly_report = today
            status.last_daily_report = today
        elif report_type == 'weekly':
            status.last_weekly_report = today
            status.last_daily_report = today
        elif report_type == 'daily':
            status.last_daily_report = today
        db.session.commit()
        
    return jsonify({"success": True})


# ==========================================
# 🛠️ 开发者专用：重置报告状态 API
# ==========================================
@app.route('/api/reports/test_reset/<report_type>', methods=['POST'])
@login_required
def test_reset_report(report_type):
    if not current_user.is_admin:
        abort(403)

    if report_type not in {'daily', 'weekly', 'monthly'}:
        return jsonify({"success": False, "error": "invalid_report_type"}), 400
    
    status = UserActiveStatus.query.get(current_user.id)
    if not status:
        return jsonify({"success": False})
        
    today = now_utc8().date()
    
    # 💡 核心逻辑：为了精准触发某一种报告，必须把比它优先级高的报告标记为“今天已看”，把目标报告设为 None
    if report_type == 'daily':
        status.last_monthly_report = today
        status.last_weekly_report = today
        status.last_daily_report = None
    elif report_type == 'weekly':
        status.last_monthly_report = today
        status.last_weekly_report = None  # <--- 设为空，下次必弹周报
        status.last_daily_report = today
    elif report_type == 'monthly':
        status.last_monthly_report = None
        
    db.session.commit()
    return jsonify({"success": True})


# ==========================================
# 📌 独立告示栏 (学生端)
# ==========================================
@app.route("/bulletin")
@login_required
def bulletin_view():
    return render_template("bulletin.html")


# ==========================================
# 🧩 OJ 题库：学生 / 教师可见视图
# ==========================================
@app.route('/oj/problems')
@login_required
def oj_problem_list():
    ensure_oj_authoring_schema()

    q = request.args.get('q', '').strip()
    difficulty = request.args.get('difficulty', '').strip().lower()
    visibility = request.args.get('visibility', 'visible').strip().lower()
    payload = build_oj_problem_list_payload(q, difficulty, visibility)
    return render_template(
        'oj/oj_shell.html',
        page_title='OJ 题库',
        shell_payload={
            "initialView": "problems",
            "problemList": payload,
            "urls": {
                "problemList": url_for("oj_problem_list"),
                "problemListJson": url_for("oj_problem_list_json"),
                "submissionList": url_for("oj_submission_list"),
                "submissionListJson": url_for("oj_submission_list_json"),
                "assignmentList": url_for("oj_assignment_list"),
                "assignmentListJson": url_for("oj_assignment_list_json"),
                "adminProblems": url_for("manage_oj_problems") if current_user.is_admin else None,
                "createProblem": url_for("create_oj_problem") if current_user.is_admin else None,
                "createAssignment": url_for("create_oj_assignment") if current_user.is_admin else None,
            },
            "currentUser": {"isAdmin": bool(current_user.is_admin)},
        },
    )


@app.route('/oj/problems.json')
@login_required
def oj_problem_list_json():
    ensure_oj_authoring_schema()
    q = request.args.get('q', '').strip()
    difficulty = request.args.get('difficulty', '').strip().lower()
    visibility = request.args.get('visibility', 'visible').strip().lower()
    return jsonify({"ok": True, "problemList": build_oj_problem_list_payload(q, difficulty, visibility)})


@app.route('/oj/problems/<slug>')
@login_required
def oj_problem_detail(slug):
    ensure_oj_authoring_schema()
    problem = Problem.query.filter_by(slug=slug).first_or_404()
    active_assignment = get_assignment_for_problem_request(problem)
    if not can_view_problem(problem, current_user) and not active_assignment:
        abort(404)

    return render_template(
        'oj/oj_shell.html',
        page_title=f'{problem.title} - OJ 题库',
        shell_payload={
            "initialView": "problemDetail",
            "problemDetail": serialize_oj_problem_detail(problem, active_assignment),
            "urls": {
                "problemList": url_for("oj_problem_list"),
                "problemListJson": url_for("oj_problem_list_json"),
                "submissionList": url_for("oj_submission_list"),
                "submissionListJson": url_for("oj_submission_list_json"),
                "assignmentList": url_for("oj_assignment_list"),
                "assignmentListJson": url_for("oj_assignment_list_json"),
                "adminProblems": url_for("manage_oj_problems") if current_user.is_admin else None,
            },
            "currentUser": {"isAdmin": bool(current_user.is_admin)},
        },
    )


@app.route('/oj/problems/<slug>.json')
@login_required
def oj_problem_detail_json(slug):
    ensure_oj_authoring_schema()
    problem = Problem.query.filter_by(slug=slug).first_or_404()
    active_assignment = get_assignment_for_problem_request(problem)
    if not can_view_problem(problem, current_user) and not active_assignment:
        abort(404)
    return jsonify({"ok": True, "problemDetail": serialize_oj_problem_detail(problem, active_assignment)})


@app.route('/oj/problems/<slug>/code')
@login_required
def oj_problem_code(slug):
    ensure_oj_authoring_schema()
    problem = Problem.query.filter_by(slug=slug).first_or_404()
    active_assignment = get_assignment_for_problem_request(problem)
    if not can_view_problem(problem, current_user) and not active_assignment:
        abort(404)

    statement_html, statement_has_sample_pairs = render_problem_statement(problem.statement_md or "")
    sample_cases = problem.sample_cases
    my_latest_task = (
        JudgeTask.query
        .filter_by(
            problem_id=problem.id,
            user_id=current_user.id,
            assignment_id=active_assignment.id if active_assignment else None,
        )
        .order_by(JudgeTask.created_at.desc(), JudgeTask.id.desc())
        .first()
    )

    return render_template(
        'oj/problem_code.html',
        problem=problem,
        sample_cases=sample_cases,
        statement_html=statement_html,
        statement_has_sample_pairs=statement_has_sample_pairs,
        my_latest_task=my_latest_task,
        active_assignment=active_assignment,
    )


@app.route('/oj/problems/<slug>/syntax-check', methods=['POST'])
@login_required
def oj_problem_syntax_check(slug):
    ensure_oj_authoring_schema()
    problem = Problem.query.filter_by(slug=slug).first_or_404()
    active_assignment = get_assignment_for_problem_request(problem)
    if not can_view_problem(problem, current_user) and not active_assignment:
        abort(404)

    payload = request.get_json(silent=True) or {}
    language = (payload.get('language') or '').strip().lower()
    source_code = payload.get('source_code') or ''

    if language != 'python':
        return jsonify({
            'ok': True,
            'diagnostics': [],
            'message': '当前仅对 Python 代码提供语法检查。',
        })

    if not source_code.strip():
        return jsonify({
            'ok': True,
            'diagnostics': [],
            'message': '输入代码后会自动检查 Python 语法。',
        })

    try:
        ast.parse(source_code)
    except SyntaxError as exc:
        line = exc.lineno or 1
        column = max((exc.offset or 1) - 1, 0)
        return jsonify({
            'ok': False,
            'diagnostics': [{
                'row': max(line - 1, 0),
                'column': column,
                'text': exc.msg or 'Python 语法错误',
                'type': 'error',
            }],
            'message': f"第 {line} 行：{exc.msg or 'Python 语法错误'}",
        })

    return jsonify({
        'ok': True,
        'diagnostics': [],
        'message': 'Python 语法检查通过。',
    })


@app.route('/oj/problems/<slug>/python-completions', methods=['POST'])
@login_required
def oj_problem_python_completions(slug):
    ensure_oj_authoring_schema()
    problem = Problem.query.filter_by(slug=slug).first_or_404()
    active_assignment = get_assignment_for_problem_request(problem)
    if not can_view_problem(problem, current_user) and not active_assignment:
        abort(404)

    if jedi is None:
        return _jedi_unavailable_response()

    payload = request.get_json(silent=True) or {}
    source_code = payload.get('source_code') or ''
    line, column = _jedi_position(payload)

    try:
        completions = _jedi_script(source_code).complete(line, column)
    except Exception as exc:
        return jsonify({'ok': False, 'message': str(exc), 'items': []}), 200

    items = []
    for completion in completions[:80]:
        doc = _oj_python_doc(completion.name, _jedi_doc(completion))
        items.append({
            'label': completion.name,
            'insert_text': completion.name,
            'detail': _oj_python_detail(completion.name, completion.description),
            'kind': completion.type,
            'documentation': doc[:2400],
        })

    return jsonify({'ok': True, 'items': items})


@app.route('/oj/problems/<slug>/python-hover', methods=['POST'])
@login_required
def oj_problem_python_hover(slug):
    ensure_oj_authoring_schema()
    problem = Problem.query.filter_by(slug=slug).first_or_404()
    active_assignment = get_assignment_for_problem_request(problem)
    if not can_view_problem(problem, current_user) and not active_assignment:
        abort(404)

    if jedi is None:
        return _jedi_unavailable_response()

    payload = request.get_json(silent=True) or {}
    source_code = payload.get('source_code') or ''
    line, column = _jedi_position(payload)

    try:
        names = _jedi_script(source_code).help(line, column)
    except Exception as exc:
        return jsonify({'ok': False, 'message': str(exc), 'items': []}), 200

    items = []
    for name in names[:4]:
        doc = _oj_python_doc(name.name, _jedi_doc(name))
        if not doc:
            continue
        items.append({
            'name': name.name,
            'detail': name.description,
            'documentation': doc[:3200],
        })

    return jsonify({'ok': True, 'items': items})


@app.route('/oj/problems/<slug>/python-signatures', methods=['POST'])
@login_required
def oj_problem_python_signatures(slug):
    ensure_oj_authoring_schema()
    problem = Problem.query.filter_by(slug=slug).first_or_404()
    active_assignment = get_assignment_for_problem_request(problem)
    if not can_view_problem(problem, current_user) and not active_assignment:
        abort(404)

    if jedi is None:
        return _jedi_unavailable_response()

    payload = request.get_json(silent=True) or {}
    source_code = payload.get('source_code') or ''
    line, column = _jedi_position(payload)

    try:
        signatures = _jedi_script(source_code).get_signatures(line, column)
    except Exception as exc:
        return jsonify({'ok': False, 'message': str(exc), 'signatures': []}), 200

    signature_items = []
    active_signature = 0
    active_parameter = 0
    for signature in signatures[:5]:
        parameters = []
        for param in signature.params:
            parameters.append({
                'label': param.to_string(),
                'documentation': _oj_python_param_doc(signature, param)[:1200],
            })
        signature_items.append({
            'label': _oj_python_signature_label(signature),
            'documentation': _oj_python_signature_doc(signature)[:2400],
            'parameters': parameters,
        })
        active_parameter = max(getattr(signature, 'index', 0) or 0, 0)

    return jsonify({
        'ok': True,
        'activeSignature': active_signature,
        'activeParameter': active_parameter,
        'signatures': signature_items,
    })


@app.route('/oj/problems/<slug>/submit', methods=['GET', 'POST'])
@login_required
def submit_oj_problem(slug):
    ensure_oj_authoring_schema()
    problem = Problem.query.filter_by(slug=slug).first_or_404()
    active_assignment = get_assignment_for_problem_request(problem)
    if not can_view_problem(problem, current_user) and not active_assignment:
        abort(404)

    if request.method == 'GET':
        return render_template('oj/problem_submit.html', problem=problem, active_assignment=active_assignment)

    language = request.form.get('language', '').strip().lower()
    source_code = request.form.get('source_code', '')
    allowed_languages = problem.allowed_languages or ['python', 'cpp', 'c']

    if language not in allowed_languages:
        flash('这道题暂不支持所选语言。', 'error')
        return redirect(url_for('oj_problem_detail', slug=problem.slug, assignment_id=active_assignment.id if active_assignment else None))

    if not source_code.strip():
        flash('请先填写代码再提交。', 'error')
        return redirect(url_for('oj_problem_detail', slug=problem.slug, assignment_id=active_assignment.id if active_assignment else None))

    test_cases = build_oj_judge_cases(problem)
    if not test_cases:
        flash('这道题还没有配置测试数据，暂时不能提交。', 'error')
        return redirect(url_for('oj_problem_detail', slug=problem.slug, assignment_id=active_assignment.id if active_assignment else None))

    task = JudgeTask(
        problem_id=problem.id,
        assignment_id=active_assignment.id if active_assignment else None,
        user_id=current_user.id,
        language=language,
        source_code=source_code,
        test_cases=test_cases,
        status='queued',
        runtime_image=app.config.get('JUDGE_RUNTIME_IMAGE'),
        time_limit_ms=problem.time_limit_ms,
        memory_limit_mb=problem.memory_limit_mb,
        total_count=len(test_cases),
    )
    db.session.add(task)
    db.session.commit()

    try:
        job = enqueue_judge_task(task.id)
        task.queue_job_id = job.id
        task.status = 'queued'
        db.session.commit()
        flash('代码已提交，正在进入评测队列。', 'success')
    except Exception as exc:
        task.status = 'system_error'
        task.finished_at = now_utc8()
        task.error_message = f"提交已保存，但无法连接评测队列：{exc}"
        task.result_summary = {
            'status': 'system_error',
            'error_message': str(exc),
        }
        db.session.commit()
        flash('提交已保存，但评测队列暂不可用。请检查 Redis / Judge Worker。', 'warning')

    return redirect(url_for('oj_submission_detail', task_id=task.id))


@app.route('/oj/submissions')
@login_required
def oj_submission_list():
    ensure_oj_authoring_schema()
    status_filter = request.args.get('status', '').strip().lower()
    language_filter = request.args.get('language', '').strip()
    assignment_id_filter = request.args.get('assignment_id', type=int)
    user_filter = request.args.get('user', '').strip() if current_user.is_admin else ''
    problem_filter = request.args.get('problem', '').strip()
    problem_id_filter = request.args.get('problem_id', type=int)
    payload = build_oj_submission_list_payload(status_filter, language_filter, user_filter, problem_filter, problem_id_filter, assignment_id_filter)
    return render_template(
        'oj/oj_shell.html',
        page_title='OJ 提交记录',
        shell_payload={
            "initialView": "submissions",
            "submissionList": payload,
            "urls": {
                "problemList": url_for("oj_problem_list"),
                "problemListJson": url_for("oj_problem_list_json"),
                "submissionList": url_for("oj_submission_list"),
                "submissionListJson": url_for("oj_submission_list_json"),
                "assignmentList": url_for("oj_assignment_list"),
                "assignmentListJson": url_for("oj_assignment_list_json"),
            },
            "currentUser": {"isAdmin": bool(current_user.is_admin)},
        },
    )


@app.route('/oj/submissions.json')
@login_required
def oj_submission_list_json():
    ensure_oj_authoring_schema()
    return jsonify({
        "ok": True,
        "submissionList": build_oj_submission_list_payload(
            request.args.get('status', '').strip().lower(),
            request.args.get('language', '').strip(),
            request.args.get('user', '').strip() if current_user.is_admin else '',
            request.args.get('problem', '').strip(),
            request.args.get('problem_id', type=int),
            request.args.get('assignment_id', type=int),
        ),
    })


@app.route('/oj/submissions/<int:task_id>')
@login_required
def oj_submission_detail(task_id):
    ensure_oj_authoring_schema()
    task = get_oj_judge_task_or_404(task_id)
    if task.user_id == current_user.id and task.status in OJ_FINAL_STATUSES:
        check_and_award_badges(current_user)
    results = task.results.order_by(JudgeTaskResult.case_index.asc()).all()
    return render_template(
        'oj/oj_shell.html',
        page_title=f'提交 #{task.id} - OJ',
        shell_payload={
            "initialView": "submissionDetail",
            "submissionDetail": serialize_submission_detail(task, results),
            "urls": {
                "problemList": url_for("oj_problem_list"),
                "problemListJson": url_for("oj_problem_list_json"),
                "submissionList": url_for("oj_submission_list"),
                "submissionListJson": url_for("oj_submission_list_json"),
                "assignmentList": url_for("oj_assignment_list"),
                "assignmentListJson": url_for("oj_assignment_list_json"),
            },
            "currentUser": {"isAdmin": bool(current_user.is_admin)},
        },
    )


@app.route('/oj/submissions/<int:task_id>.json')
@login_required
def oj_submission_detail_json(task_id):
    ensure_oj_authoring_schema()
    task = get_oj_judge_task_or_404(task_id)
    results = task.results.order_by(JudgeTaskResult.case_index.asc()).all()
    return jsonify({"ok": True, "submissionDetail": serialize_submission_detail(task, results)})


@app.route('/api/oj/submissions/<int:task_id>')
@login_required
def oj_submission_status_api(task_id):
    ensure_oj_authoring_schema()
    task = get_oj_judge_task_or_404(task_id)
    if task.user_id == current_user.id and task.status in OJ_FINAL_STATUSES:
        check_and_award_badges(current_user)
    meta = oj_submission_status_meta(task)
    return jsonify({
        'id': task.id,
        'status': task.status,
        'status_label': meta['label'],
        'status_tone': meta['tone'],
        'passed_count': task.passed_count,
        'total_count': task.total_count,
        'total_score': task.total_score,
        'finished_at': task.finished_at.isoformat() if task.finished_at else None,
        'error_message': task.error_message,
    })


# ==========================================
# 🧠 管理员：OJ 题库管理
# ==========================================
@app.route('/admin/oj/problems', methods=['GET'])
@login_required
def manage_oj_problems():
    if not current_user.is_admin:
        abort(403)

    created_tables = ensure_oj_authoring_schema()
    if created_tables:
        flash('检测到旧数据库，已自动补建 OJ 题库数据表。', 'success')

    try:
        problems = Problem.query.order_by(Problem.problem_code.asc(), Problem.id.asc()).all()
    except OperationalError:
        db.session.rollback()
        created_tables = ensure_oj_authoring_schema()
        if created_tables:
            flash('OJ 题库数据表已自动修复，请继续操作。', 'success')
        problems = Problem.query.order_by(Problem.problem_code.asc(), Problem.id.asc()).all()
    return render_template(
        'admin/manage_oj_problems.html',
        problems=problems,
        problems_json=[serialize_admin_problem(problem) for problem in problems],
    )


@app.route('/admin/oj/problems.json', methods=['GET'])
@login_required
def manage_oj_problems_json():
    if not current_user.is_admin:
        abort(403)

    ensure_oj_authoring_schema()
    problems = Problem.query.order_by(Problem.problem_code.asc(), Problem.id.asc()).all()
    return jsonify({
        "ok": True,
        "problems": [serialize_admin_problem(problem) for problem in problems],
        "count": len(problems),
    })


@app.route('/admin/oj/problems/<int:problem_id>.json', methods=['GET'])
@login_required
def admin_oj_problem_json(problem_id):
    if not current_user.is_admin:
        abort(403)

    ensure_oj_authoring_schema()
    problem = Problem.query.get_or_404(problem_id)
    return jsonify({
        "ok": True,
        "problem": serialize_problem_edit_workspace(problem),
    })


@app.route('/admin/oj/problems/new', methods=['GET', 'POST'])
@login_required
def create_oj_problem():
    if not current_user.is_admin:
        abort(403)

    created_tables = ensure_oj_authoring_schema()
    if created_tables:
        flash('检测到旧数据库，已自动补建 OJ 题库数据表。', 'success')

    draft = {
        'problem_code': next_problem_code(),
        'title': '',
        'slug': '',
        'difficulty': 'medium',
        'time_limit_ms': 2000,
        'memory_limit_mb': 256,
        'allowed_languages': 'python,cpp,c',
        'is_visible': '1',
        'statement_md': """## 题目描述

在这里写题目的核心描述。

## 输入格式

说明输入内容。

## 输出格式

说明输出内容。

## 样例

```input1
1
```

```output1
1
```

## 数据范围与提示

补充数据规模、边界条件或提示信息。
""",
        'input_spec_md': '',
        'output_spec_md': '',
        'hint_md': '',
        'source': '',
    }

    if request.method == 'POST':
        draft = {
            'problem_code': request.form.get('problem_code', '').strip(),
            'title': request.form.get('title', '').strip(),
            'slug': request.form.get('slug', '').strip(),
            'difficulty': request.form.get('difficulty', 'medium').strip().lower() or 'medium',
            'time_limit_ms': request.form.get('time_limit_ms', 2000, type=int) or 2000,
            'memory_limit_mb': request.form.get('memory_limit_mb', 256, type=int) or 256,
            'allowed_languages': request.form.get('allowed_languages', 'python,cpp,c').strip(),
            'is_visible': request.form.get('is_visible', '1'),
            'statement_md': request.form.get('statement_md', '').strip(),
            'input_spec_md': request.form.get('input_spec_md', '').strip(),
            'output_spec_md': request.form.get('output_spec_md', '').strip(),
            'hint_md': request.form.get('hint_md', '').strip(),
            'source': request.form.get('source', '').strip(),
        }

        if not draft['title']:
            flash('题目标题不能为空', 'error')
            return render_template('admin/create_oj_problem.html', draft=draft)

        try:
            problem_code = validate_problem_code(draft['problem_code'])
        except ValueError as exc:
            flash(str(exc), 'error')
            return render_template('admin/create_oj_problem.html', draft=draft)

        final_slug = build_problem_slug(draft['slug'] or draft['title'])
        new_problem = Problem(
            problem_code=problem_code,
            title=draft['title'],
            slug=final_slug,
            difficulty=draft['difficulty'] if draft['difficulty'] in {'easy', 'medium', 'hard'} else 'medium',
            time_limit_ms=max(draft['time_limit_ms'], 100),
            memory_limit_mb=max(draft['memory_limit_mb'], 16),
            allowed_languages=normalize_language_list(draft['allowed_languages']),
            is_visible=draft['is_visible'] == '1',
            created_by_id=current_user.id,
            statement_md=draft['statement_md'],
            input_spec_md=draft['input_spec_md'],
            output_spec_md=draft['output_spec_md'],
            hint_md=draft['hint_md'],
            source=draft['source'] or None,
        )
        db.session.add(new_problem)
        db.session.commit()
        flash(f'OJ 题目《{new_problem.title}》已创建。接下来可以配置文件空间和测试数据。', 'success')
        return redirect(url_for('edit_oj_problem', problem_id=new_problem.id))

    return render_template('admin/create_oj_problem.html', draft=draft)


@app.route('/admin/oj/problems/import_zip', methods=['POST'])
@login_required
def import_oj_problem_zip():
    if not current_user.is_admin:
        abort(403)

    created_tables = ensure_oj_authoring_schema()
    if created_tables:
        flash('检测到旧数据库，已自动补建 OJ 题库数据表。', 'success')

    try:
        problem, imported_cases, imported_files, skipped_files = import_oj_problem_from_zip(
            request.files.get('problem_zip'),
            current_user,
        )
    except ValueError as exc:
        flash(f'导入 zip 失败：{exc}', 'error')
        return redirect(url_for('create_oj_problem'))
    except Exception as exc:
        app.logger.exception("Import OJ problem zip failed")
        flash(f'导入 zip 失败：{exc}', 'error')
        return redirect(url_for('create_oj_problem'))

    message = f'已从 zip 创建题目《{problem.title}》，导入 {imported_files} 个文件、{imported_cases} 组测试点。'
    if skipped_files:
        message += f' 跳过 {skipped_files} 个不支持的文件。'
    flash(message, 'success')
    return redirect(url_for('edit_oj_problem', problem_id=problem.id))


@app.route('/admin/oj/problems/<int:problem_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_oj_problem(problem_id):
    if not current_user.is_admin:
        abort(403)

    created_tables = ensure_oj_authoring_schema()
    if created_tables:
        flash('检测到旧数据库，已自动补建 OJ 题库数据表。', 'success')

    problem = Problem.query.get_or_404(problem_id)

    if request.method == 'POST':
        raw_problem_code = request.form.get('problem_code', '').strip()
        title = request.form.get('title', '').strip()
        slug = request.form.get('slug', '').strip()

        if not title:
            if is_ajax_request():
                return jsonify({"ok": False, "message": "题目标题不能为空"}), 400
            flash('题目标题不能为空', 'error')
            return redirect(url_for('edit_oj_problem', problem_id=problem.id))

        try:
            problem.problem_code = validate_problem_code(raw_problem_code, problem_id=problem.id)
        except ValueError as exc:
            if is_ajax_request():
                return jsonify({"ok": False, "message": str(exc)}), 400
            flash(str(exc), 'error')
            return redirect(url_for('edit_oj_problem', problem_id=problem.id))

        problem.title = title
        problem.slug = build_problem_slug(slug or title, problem_id=problem.id)
        problem.statement_md = request.form.get('statement_md', '').strip()
        if 'input_spec_md' in request.form:
            problem.input_spec_md = request.form.get('input_spec_md', '').strip()
        if 'output_spec_md' in request.form:
            problem.output_spec_md = request.form.get('output_spec_md', '').strip()
        if 'hint_md' in request.form:
            problem.hint_md = request.form.get('hint_md', '').strip()
        problem.source = request.form.get('source', '').strip() or None

        difficulty = request.form.get('difficulty', 'medium').strip().lower()
        problem.difficulty = difficulty if difficulty in {'easy', 'medium', 'hard'} else 'medium'
        problem.time_limit_ms = max(request.form.get('time_limit_ms', 2000, type=int) or 2000, 100)
        problem.memory_limit_mb = max(request.form.get('memory_limit_mb', 256, type=int) or 256, 16)
        problem.allowed_languages = normalize_language_list(request.form.get('allowed_languages', 'python,cpp,c'))
        problem.is_visible = request.form.get('is_visible') == '1'

        db.session.commit()
        if is_ajax_request():
            return jsonify({
                "ok": True,
                "message": f'题目《{problem.title}》已更新。',
                "problem": serialize_problem_edit_workspace(problem),
                "summary": serialize_admin_problem(problem),
            })
        flash(f'题目《{problem.title}》已更新。', 'success')
        return redirect(url_for('edit_oj_problem', problem_id=problem.id))

    problem_files = ProblemFile.query.filter_by(problem_id=problem.id).all()
    data_file_count = sum(1 for problem_file in problem_files if problem_file.is_testdata_file)
    asset_file_count = sum(1 for problem_file in problem_files if not problem_file.is_testdata_file)
    return render_template(
        'admin/oj_admin_shell.html',
        page_title=f'编辑 OJ 题目 - {problem.title}',
        shell_payload={
            "initialView": "edit",
            "currentProblemId": problem.id,
            "problemsUrl": url_for('manage_oj_problems_json'),
            "problem": serialize_problem_edit_workspace(problem),
            "problemUrl": url_for('admin_oj_problem_json', problem_id=problem.id),
            "urls": {
                "adminList": url_for('manage_oj_problems'),
                "publicList": url_for('oj_problem_list'),
                "createProblem": url_for('create_oj_problem'),
            },
        },
    )


@app.route('/admin/oj/problems/<int:problem_id>/files')
@login_required
def manage_oj_problem_files(problem_id):
    if not current_user.is_admin:
        abort(403)

    created_tables = ensure_oj_authoring_schema()
    if created_tables:
        flash('检测到旧数据库，已自动补建 OJ 题库数据表。', 'success')

    problem = Problem.query.get_or_404(problem_id)
    return render_template(
        'admin/manage_oj_problem_files.html',
        problem=problem,
        workspace_payload=serialize_problem_file_workspace(problem),
        **build_problem_file_workspace_context(problem),
    )


@app.route('/admin/oj/problems/<int:problem_id>/files.json')
@login_required
def manage_oj_problem_files_json(problem_id):
    if not current_user.is_admin:
        abort(403)

    ensure_oj_authoring_schema()
    problem = Problem.query.get_or_404(problem_id)
    return jsonify({
        "ok": True,
        "workspace": serialize_problem_file_workspace(problem),
    })


@app.route('/admin/oj/problems/<int:problem_id>/files/text/create', methods=['POST'])
@login_required
def create_text_oj_problem_file(problem_id):
    if not current_user.is_admin:
        abort(403)

    ensure_oj_authoring_schema()
    problem = Problem.query.get_or_404(problem_id)
    filename = request.form.get('filename', '').strip()
    content = request.form.get('content', '')

    try:
        sanitized_name = secure_filename(filename)
        if ProblemFile.query.filter_by(problem_id=problem.id, filename=sanitized_name).first():
            raise ValueError("已存在同名文件")
        problem_file = create_text_problem_asset(problem.id, filename, content, current_user.id)
        db.session.add(problem_file)
        db.session.flush()
        rebuild_hidden_problem_testcases_from_files(problem)
        db.session.commit()
        return problem_file_workspace_response(problem, f'已创建文件：{problem_file.filename}', 'success')
    except ValueError as exc:
        db.session.rollback()
        return problem_file_workspace_response(problem, f'创建文件失败：{exc}', 'error', 400)


@app.route('/admin/oj/problems/<int:problem_id>/files/text/create_pair', methods=['POST'])
@login_required
def create_text_oj_problem_file_pair(problem_id):
    if not current_user.is_admin:
        abort(403)

    ensure_oj_authoring_schema()
    problem = Problem.query.get_or_404(problem_id)
    case_index = request.form.get('case_index', '').strip()
    input_content = request.form.get('input_content', '')
    output_content = request.form.get('output_content', '')
    overwrite = request.form.get('overwrite') == '1'

    if not case_index.isdigit() or int(case_index) <= 0:
        return problem_file_workspace_response(problem, '组号必须是大于 0 的整数。', 'error', 400)

    input_filename = f'{int(case_index)}.in'
    output_filename = f'{int(case_index)}.out'
    existing_input = ProblemFile.query.filter_by(problem_id=problem.id, filename=input_filename).first()
    existing_output = ProblemFile.query.filter_by(problem_id=problem.id, filename=output_filename).first()

    if (existing_input or existing_output) and not overwrite:
        return problem_file_workspace_response(problem, f'文件 {input_filename} 或 {output_filename} 已存在。勾选覆盖后可直接更新。', 'error', 400)

    try:
        if existing_input:
            update_text_problem_asset(existing_input, input_filename, input_content)
        else:
            db.session.add(create_text_problem_asset(problem.id, input_filename, input_content, current_user.id))

        if existing_output:
            update_text_problem_asset(existing_output, output_filename, output_content)
        else:
            db.session.add(create_text_problem_asset(problem.id, output_filename, output_content, current_user.id))

        db.session.flush()
        rebuild_hidden_problem_testcases_from_files(problem)
        db.session.commit()
        action_text = '已覆盖更新' if overwrite and (existing_input or existing_output) else '已创建'
        return problem_file_workspace_response(problem, f'{action_text}文件对：{input_filename} / {output_filename}，并已自动更新测试点。', 'success')
    except ValueError as exc:
        db.session.rollback()
        return problem_file_workspace_response(problem, f'创建文件对失败：{exc}', 'error', 400)


@app.route('/admin/oj/problems/<int:problem_id>/files/<int:file_id>/text', methods=['POST'])
@login_required
def update_text_oj_problem_file(problem_id, file_id):
    if not current_user.is_admin:
        abort(403)

    ensure_oj_authoring_schema()
    problem = Problem.query.get_or_404(problem_id)
    problem_file = ProblemFile.query.filter_by(id=file_id, problem_id=problem.id).first_or_404()
    filename = request.form.get('filename', '').strip()
    content = request.form.get('content', '')

    try:
        sanitized_name = secure_filename(filename)
        exists = ProblemFile.query.filter_by(problem_id=problem.id, filename=sanitized_name).filter(ProblemFile.id != problem_file.id).first()
        if exists:
            raise ValueError("已存在同名文件")
        update_text_problem_asset(problem_file, filename, content)
        db.session.flush()
        rebuild_hidden_problem_testcases_from_files(problem)
        db.session.commit()
        return problem_file_workspace_response(problem, f'已更新文件：{problem_file.filename}，并已自动更新测试点。', 'success')
    except ValueError as exc:
        db.session.rollback()
        return problem_file_workspace_response(problem, f'更新文件失败：{exc}', 'error', 400)


@app.route('/admin/oj/problems/<int:problem_id>/files/upload', methods=['POST'])
@login_required
def upload_oj_problem_files(problem_id):
    if not current_user.is_admin:
        abort(403)

    ensure_oj_authoring_schema()
    problem = Problem.query.get_or_404(problem_id)

    files = request.files.getlist('files')
    if not files:
        return problem_file_workspace_response(problem, '没有选择要上传的文件。', 'error', 400)

    uploaded_count = 0
    for file_storage in files:
        if not file_storage or not file_storage.filename:
            continue
        original_filename = secure_filename(file_storage.filename)
        if ProblemFile.query.filter_by(problem_id=problem.id, filename=original_filename).first():
            if not is_ajax_request():
                flash(f'已存在同名文件：{original_filename}', 'error')
            continue
        try:
            saved_file = save_problem_asset(file_storage, problem.id)
        except ValueError as exc:
            if not is_ajax_request():
                flash(f'文件 {file_storage.filename} 上传失败：{exc}', 'error')
            continue

        db.session.add(
            ProblemFile(
                problem_id=problem.id,
                filename=saved_file['filename'],
                stored_filename=saved_file['stored_filename'],
                file_path=saved_file['file_path'],
                content_type=saved_file['content_type'],
                file_size=saved_file['file_size'],
                uploaded_by_id=current_user.id,
            )
        )
        uploaded_count += 1

    db.session.flush()
    rebuild_hidden_problem_testcases_from_files(problem)
    db.session.commit()

    if uploaded_count:
        return problem_file_workspace_response(problem, f'已上传 {uploaded_count} 个题目文件。', 'success')
    return problem_file_workspace_response(problem, '没有成功上传任何文件。', 'error', 400)


@app.route('/admin/oj/problems/<int:problem_id>/files/<int:file_id>/delete', methods=['POST'])
@login_required
def delete_oj_problem_file(problem_id, file_id):
    if not current_user.is_admin:
        abort(403)

    ensure_oj_authoring_schema()
    problem = Problem.query.get_or_404(problem_id)
    problem_file = ProblemFile.query.filter_by(id=file_id, problem_id=problem.id).first_or_404()

    remove_problem_asset_file(problem_file.file_path)
    db.session.delete(problem_file)
    db.session.flush()
    rebuild_hidden_problem_testcases_from_files(problem)
    db.session.commit()
    return problem_file_workspace_response(problem, f'已删除文件：{problem_file.filename}，并已自动更新测试点。', 'success')


@app.route('/admin/oj/problems/<int:problem_id>/files/import_testcases', methods=['POST'])
@login_required
def import_oj_problem_testcases_from_files(problem_id):
    if not current_user.is_admin:
        abort(403)

    ensure_oj_authoring_schema()
    problem = Problem.query.get_or_404(problem_id)
    mode = request.form.get('import_mode', 'append').strip().lower()

    files = ProblemFile.query.filter_by(problem_id=problem.id).order_by(ProblemFile.filename.asc()).all()
    grouped_files = {}
    for problem_file in files:
        stem, ext = os.path.splitext(problem_file.filename)
        ext = ext.lower()
        if ext not in {'.in', '.out'}:
            continue
        grouped_files.setdefault(stem, {})[ext] = problem_file

    paired_entries = []
    for stem, pair in grouped_files.items():
        if '.in' in pair and '.out' in pair:
            paired_entries.append((stem, pair['.in'], pair['.out']))

    if not paired_entries:
        return problem_file_workspace_response(problem, '没有找到可配对的 .in / .out 文件。', 'error', 400)

    paired_entries.sort(key=lambda item: item[0])

    if mode == 'replace':
        problem.test_cases.clear()
        db.session.flush()

    sort_order_start = len(problem.test_cases)
    imported_count = 0

    for _, input_file, output_file in paired_entries:
        input_absolute_path = os.path.join(app.root_path, input_file.file_path.lstrip('/'))
        output_absolute_path = os.path.join(app.root_path, output_file.file_path.lstrip('/'))

        if not os.path.exists(input_absolute_path) or not os.path.exists(output_absolute_path):
            continue

        with open(input_absolute_path, 'r', encoding='utf-8') as input_handle:
            input_data = input_handle.read()
        with open(output_absolute_path, 'r', encoding='utf-8') as output_handle:
            expected_output = output_handle.read()

        problem.test_cases.append(
            ProblemTestCase(
                case_type='hidden',
                input_data=input_data,
                expected_output=expected_output,
                score=1,
                sort_order=sort_order_start + imported_count,
            )
        )
        imported_count += 1

    db.session.commit()

    if imported_count:
        action_text = '覆盖并导入' if mode == 'replace' else '追加导入'
        return problem_file_workspace_response(problem, f'已{action_text} {imported_count} 组测试点。', 'success')
    return problem_file_workspace_response(problem, '找到文件对了，但没有成功导入任何测试点。', 'error', 400)


@app.route('/admin/oj/problems/<int:problem_id>/delete', methods=['POST'])
@login_required
def delete_oj_problem(problem_id):
    if not current_user.is_admin:
        abort(403)

    created_tables = ensure_oj_authoring_schema()
    if created_tables:
        flash('检测到旧数据库，已自动补建 OJ 题库数据表。', 'success')

    problem = Problem.query.get_or_404(problem_id)
    title = problem.title
    for problem_file in problem.files.all():
        remove_problem_asset_file(problem_file.file_path)
    db.session.delete(problem)
    db.session.commit()
    if is_ajax_request():
        return jsonify({
            "ok": True,
            "message": f'OJ 题目《{title}》已删除。',
        })
    flash(f'OJ 题目《{title}》已删除。', 'success')
    return redirect(url_for('manage_oj_problems'))


@app.route('/oj/assignments')
@login_required
def oj_assignment_list():
    ensure_oj_authoring_schema()
    return render_template(
        'oj/oj_shell.html',
        page_title='OJ 作业',
        shell_payload={
            "initialView": "assignments",
            "assignmentList": build_oj_assignment_list_payload(),
            "urls": {
                "problemList": url_for("oj_problem_list"),
                "problemListJson": url_for("oj_problem_list_json"),
                "submissionList": url_for("oj_submission_list"),
                "submissionListJson": url_for("oj_submission_list_json"),
                "assignmentList": url_for("oj_assignment_list"),
                "assignmentListJson": url_for("oj_assignment_list_json"),
                "createAssignment": url_for("create_oj_assignment") if current_user.is_admin else None,
                "adminProblems": url_for("manage_oj_problems") if current_user.is_admin else None,
            },
            "currentUser": {"isAdmin": bool(current_user.is_admin)},
        },
    )


@app.route('/oj/assignments.json')
@login_required
def oj_assignment_list_json():
    ensure_oj_authoring_schema()
    return jsonify({"ok": True, "assignmentList": build_oj_assignment_list_payload()})


@app.route('/oj/assignments/<int:assignment_id>')
@login_required
def oj_assignment_detail(assignment_id):
    ensure_oj_authoring_schema()
    assignment = get_oj_assignment_or_404_visible(assignment_id)
    return render_template(
        'oj/oj_shell.html',
        page_title=f'{assignment.title} - OJ 作业',
        shell_payload={
            "initialView": "assignmentDetail",
            "assignmentDetail": build_oj_assignment_detail_payload(assignment),
            "urls": {
                "problemList": url_for("oj_problem_list"),
                "problemListJson": url_for("oj_problem_list_json"),
                "submissionList": url_for("oj_submission_list"),
                "submissionListJson": url_for("oj_submission_list_json"),
                "assignmentList": url_for("oj_assignment_list"),
                "assignmentListJson": url_for("oj_assignment_list_json"),
                "adminProblems": url_for("manage_oj_problems") if current_user.is_admin else None,
            },
            "currentUser": {"isAdmin": bool(current_user.is_admin)},
        },
    )


@app.route('/oj/assignments/<int:assignment_id>.json')
@login_required
def oj_assignment_detail_json(assignment_id):
    ensure_oj_authoring_schema()
    assignment = get_oj_assignment_or_404_visible(assignment_id)
    return jsonify({"ok": True, "assignmentDetail": build_oj_assignment_detail_payload(assignment)})


@app.route('/oj/assignments/<int:assignment_id>/scoreboard')
@login_required
def oj_assignment_scoreboard(assignment_id):
    ensure_oj_authoring_schema()
    assignment = get_oj_assignment_or_404_visible(assignment_id)
    return render_template(
        'oj/oj_shell.html',
        page_title=f'{assignment.title} - 作业成绩表',
        shell_payload={
            "initialView": "assignmentScoreboard",
            "assignmentScoreboard": build_oj_assignment_scoreboard_payload(assignment),
            "urls": {
                "problemList": url_for("oj_problem_list"),
                "problemListJson": url_for("oj_problem_list_json"),
                "submissionList": url_for("oj_submission_list"),
                "submissionListJson": url_for("oj_submission_list_json"),
                "assignmentList": url_for("oj_assignment_list"),
                "assignmentListJson": url_for("oj_assignment_list_json"),
                "adminProblems": url_for("manage_oj_problems") if current_user.is_admin else None,
            },
            "currentUser": {"isAdmin": bool(current_user.is_admin)},
        },
    )


@app.route('/oj/assignments/<int:assignment_id>/scoreboard.json')
@login_required
def oj_assignment_scoreboard_json(assignment_id):
    ensure_oj_authoring_schema()
    assignment = get_oj_assignment_or_404_visible(assignment_id)
    return jsonify({"ok": True, "assignmentScoreboard": build_oj_assignment_scoreboard_payload(assignment)})


@app.route('/admin/oj/assignments/new', methods=['GET', 'POST'])
@login_required
def create_oj_assignment():
    if not current_user.is_admin:
        abort(403)

    ensure_oj_authoring_schema()
    form_state = build_oj_assignment_form_state()

    if request.method == 'POST':
        form_state = {
            'title': request.form.get('title', '').strip(),
            'description_md': request.form.get('description_md', '').strip(),
            'start_at': request.form.get('start_at', '').strip(),
            'end_at': request.form.get('end_at', '').strip(),
            'extension_days': request.form.get('extension_days', 0, type=int) or 0,
            'visibility': request.form.get('visibility', 'all').strip() or 'all',
            'is_active': request.form.get('is_active', '1'),
            'selected_class_ids': [int(value) for value in request.form.getlist('class_ids') if value.isdigit()],
            'selected_group_ids': [int(value) for value in request.form.getlist('group_ids') if value.isdigit()],
            'selected_user_ids': [int(value) for value in request.form.getlist('user_ids') if value.isdigit()],
            'selected_manager_ids': [int(value) for value in request.form.getlist('manager_ids') if value.isdigit()],
            'selected_problem_ids': [int(value) for value in request.form.getlist('problem_ids') if value.isdigit()],
        }

        if not form_state['title']:
            flash('作业标题不能为空。', 'error')
        elif not form_state['selected_problem_ids']:
            flash('至少要选择一道 OJ 题目。', 'error')
        else:
            assignment = OJAssignment(
                title=form_state['title'],
                description_md=form_state['description_md'],
                start_at=datetime.strptime(form_state['start_at'], '%Y-%m-%dT%H:%M') if form_state['start_at'] else None,
                end_at=datetime.strptime(form_state['end_at'], '%Y-%m-%dT%H:%M') if form_state['end_at'] else None,
                extension_days=max(int(form_state['extension_days'] or 0), 0),
                visibility=form_state['visibility'] if form_state['visibility'] in {'all', 'restricted'} else 'all',
                is_active=form_state['is_active'] == '1',
                created_by_id=current_user.id,
            )
            resolve_oj_assignment_targets(
                assignment,
                form_state['selected_class_ids'],
                form_state['selected_group_ids'],
                form_state['selected_user_ids'],
                form_state['selected_manager_ids'],
            )
            resolve_oj_assignment_problem_links(assignment, form_state['selected_problem_ids'])
            db.session.add(assignment)
            db.session.commit()
            flash(f'作业《{assignment.title}》已创建。', 'success')
            return redirect(url_for('edit_oj_assignment', assignment_id=assignment.id))

    return render_template(
        'admin/create_oj_assignment.html',
        form_state=form_state,
        classes=Class.query.order_by(Class.name.asc()).all(),
        groups=Group.query.order_by(Group.name.asc()).all(),
        users=User.query.order_by(User.username.asc()).all(),
        problems=Problem.query.order_by(Problem.problem_code.asc(), Problem.id.asc()).all(),
        assignment=None,
    )


@app.route('/admin/oj/assignments/<int:assignment_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_oj_assignment(assignment_id):
    if not current_user.is_admin:
        abort(403)

    ensure_oj_authoring_schema()
    assignment = OJAssignment.query.get_or_404(assignment_id)
    form_state = build_oj_assignment_form_state(assignment)

    if request.method == 'POST':
        form_state = {
            'title': request.form.get('title', '').strip(),
            'description_md': request.form.get('description_md', '').strip(),
            'start_at': request.form.get('start_at', '').strip(),
            'end_at': request.form.get('end_at', '').strip(),
            'extension_days': request.form.get('extension_days', 0, type=int) or 0,
            'visibility': request.form.get('visibility', 'all').strip() or 'all',
            'is_active': request.form.get('is_active', '1'),
            'selected_class_ids': [int(value) for value in request.form.getlist('class_ids') if value.isdigit()],
            'selected_group_ids': [int(value) for value in request.form.getlist('group_ids') if value.isdigit()],
            'selected_user_ids': [int(value) for value in request.form.getlist('user_ids') if value.isdigit()],
            'selected_manager_ids': [int(value) for value in request.form.getlist('manager_ids') if value.isdigit()],
            'selected_problem_ids': [int(value) for value in request.form.getlist('problem_ids') if value.isdigit()],
        }

        if not form_state['title']:
            flash('作业标题不能为空。', 'error')
        elif not form_state['selected_problem_ids']:
            flash('至少要选择一道 OJ 题目。', 'error')
        else:
            assignment.title = form_state['title']
            assignment.description_md = form_state['description_md']
            assignment.start_at = datetime.strptime(form_state['start_at'], '%Y-%m-%dT%H:%M') if form_state['start_at'] else None
            assignment.end_at = datetime.strptime(form_state['end_at'], '%Y-%m-%dT%H:%M') if form_state['end_at'] else None
            assignment.extension_days = max(int(form_state['extension_days'] or 0), 0)
            assignment.visibility = form_state['visibility'] if form_state['visibility'] in {'all', 'restricted'} else 'all'
            assignment.is_active = form_state['is_active'] == '1'
            resolve_oj_assignment_targets(
                assignment,
                form_state['selected_class_ids'],
                form_state['selected_group_ids'],
                form_state['selected_user_ids'],
                form_state['selected_manager_ids'],
            )
            resolve_oj_assignment_problem_links(assignment, form_state['selected_problem_ids'])
            db.session.commit()
            flash(f'作业《{assignment.title}》已更新。', 'success')
            return redirect(url_for('edit_oj_assignment', assignment_id=assignment.id))

    return render_template(
        'admin/create_oj_assignment.html',
        form_state=form_state,
        classes=Class.query.order_by(Class.name.asc()).all(),
        groups=Group.query.order_by(Group.name.asc()).all(),
        users=User.query.order_by(User.username.asc()).all(),
        problems=Problem.query.order_by(Problem.problem_code.asc(), Problem.id.asc()).all(),
        assignment=assignment,
        assignment_files=assignment.files.order_by(OJAssignmentFile.uploaded_at.desc(), OJAssignmentFile.id.desc()).all(),
    )


@app.route('/admin/oj/assignments/<int:assignment_id>/files/upload', methods=['POST'])
@login_required
def upload_oj_assignment_files(assignment_id):
    if not current_user.is_admin:
        abort(403)

    ensure_oj_authoring_schema()
    assignment = OJAssignment.query.get_or_404(assignment_id)
    files = request.files.getlist('files')
    uploaded_count = 0

    for file_storage in files:
        if not file_storage or not file_storage.filename:
            continue
        try:
            saved_file = save_oj_assignment_asset(file_storage, assignment.id)
        except ValueError as exc:
            flash(f'文件 {file_storage.filename} 上传失败：{exc}', 'error')
            continue

        db.session.add(
            OJAssignmentFile(
                assignment_id=assignment.id,
                filename=saved_file['filename'],
                stored_filename=saved_file['stored_filename'],
                file_path=saved_file['file_path'],
                content_type=saved_file['content_type'],
                file_size=saved_file['file_size'],
                uploaded_by_id=current_user.id,
            )
        )
        uploaded_count += 1

    db.session.commit()
    if uploaded_count:
        flash(f'已上传 {uploaded_count} 个作业文件。', 'success')
    return redirect(url_for('edit_oj_assignment', assignment_id=assignment.id) + '#assignment-files')


@app.route('/admin/oj/assignments/<int:assignment_id>/files/<int:file_id>/delete', methods=['POST'])
@login_required
def delete_oj_assignment_file(assignment_id, file_id):
    if not current_user.is_admin:
        abort(403)

    ensure_oj_authoring_schema()
    assignment = OJAssignment.query.get_or_404(assignment_id)
    assignment_file = OJAssignmentFile.query.filter_by(id=file_id, assignment_id=assignment.id).first_or_404()
    remove_oj_assignment_asset_file(assignment_file.file_path)
    db.session.delete(assignment_file)
    db.session.commit()
    flash(f'已删除文件：{assignment_file.filename}', 'success')
    return redirect(url_for('edit_oj_assignment', assignment_id=assignment.id) + '#assignment-files')


# ==========================================
# 🛠️ 管理员：作业发布控制台
# ==========================================
@app.route("/manage_assignments", methods=["GET", "POST"])
@login_required
def manage_assignments():
    if not current_user.is_admin:
        abort(403)
        
    if request.method == "POST":
        title = request.form.get("title")
        content_md = request.form.get("content_md")
        target_type = request.form.get("target_type")
        target_value = request.form.get("target_value")
        deadline_str = request.form.get("deadline")
        
        star_level = request.form.get("star_level", 1, type=int)
        deadline = datetime.strptime(deadline_str, "%Y-%m-%dT%H:%M") if deadline_str else None
        
        new_assign = Assignment(
            title=title, content_md=content_md, target_type=target_type, 
            target_value=target_value, deadline=deadline, 
            star_level=star_level,
            created_by_id=current_user.id, wiki_page_id=0
        )
        db.session.add(new_assign)
        db.session.commit()
        
        flash("🎉 委托发布成功！", "success")
        return redirect(url_for('manage_assignments'))
        
    # 💡 获取所有用户
    all_users = User.query.all()
    
    # 💡 动态获取现有的班级和小组 (做安全判断，以防你的User表还没建这两个字段)
    classes = []
    groups = []
    if hasattr(User, 'class_name'):
        classes = [r[0] for r in db.session.query(User.class_name).filter(User.class_name.isnot(None)).distinct().all() if r[0]]
    if hasattr(User, 'group_name'):
        groups = [r[0] for r in db.session.query(User.group_name).filter(User.group_name.isnot(None)).distinct().all() if r[0]]

    assignments = Assignment.query.order_by(Assignment.created_at.desc()).all()
    return render_template(
        "admin/manage_assignments.html", 
        assignments=assignments, 
        all_users=all_users, 
        classes=classes, 
        groups=groups
    )

# ==========================================
# 🛠️ 管理员：编辑具体任务 API
# ==========================================
@app.route('/admin/assignments/<int:assign_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_assignment(assign_id):
    if not current_user.is_admin:
        abort(403)
        
    assign = Assignment.query.get_or_404(assign_id)
    
    if request.method == 'POST':
        assign.title = request.form.get("title")
        assign.content_md = request.form.get("content_md")
        assign.target_type = request.form.get("target_type")
        assign.target_value = request.form.get("target_value")
        deadline_str = request.form.get("deadline")
        assign.star_level = request.form.get("star_level", 1, type=int)
        
        if deadline_str:
            assign.deadline = datetime.strptime(deadline_str, "%Y-%m-%dT%H:%M")
        else:
            assign.deadline = None
            
        db.session.commit()
        flash("SYSTEM UPDATED // 任务节点已成功覆写", "success")
        return redirect(url_for('global_grading_hub'))

    # 动态获取现有的班级和小组供下拉菜单使用
    classes = []
    groups = []
    if hasattr(User, 'class_name'):
        classes = [r[0] for r in db.session.query(User.class_name).filter(User.class_name.isnot(None)).distinct().all() if r[0]]
    if hasattr(User, 'group_name'):
        groups = [r[0] for r in db.session.query(User.group_name).filter(User.group_name.isnot(None)).distinct().all() if r[0]]

    return render_template("admin/edit_assignment.html", assign=assign, classes=classes, groups=groups)

# ==========================================
# 📡 刷新我的任务 API (带权限过滤 & 👑管理员增强)
# ==========================================
@app.route('/api/assignments/my_quests')
@login_required
def get_my_quests():
    # 1. 基础查询：获取所有任务
    query = Assignment.query
    
    # 2. 👑 管理员特权：如果是管理员，直接获取全部任务，不做任何过滤
    if current_user.is_admin:
        my_assignments = query.order_by(Assignment.created_at.desc()).all()
    else:
        # 3. 普通村民逻辑：只显示针对 全员、自己班级、自己小组 或 自己的任务
        from sqlalchemy import or_
        
        filters = [Assignment.target_type == 'all']
        
        if hasattr(current_user, 'class_name') and current_user.class_name:
            filters.append(db.and_(Assignment.target_type == 'class', Assignment.target_value == current_user.class_name))
            
        if hasattr(current_user, 'group_name') and current_user.group_name:
            filters.append(db.and_(Assignment.target_type == 'group', Assignment.target_value == current_user.group_name))
            
        filters.append(db.and_(Assignment.target_type == 'personal', Assignment.target_value == str(current_user.id)))
        
        my_assignments = query.filter(or_(*filters)).order_by(Assignment.created_at.desc()).all()

    # 4. 组装数据
    quests = []
    for am in my_assignments:
        sub = Submission.query.filter_by(assignment_id=am.id, user_id=current_user.id).first()
        
        # 💡 核心新增：如果当前是管理员，实时计算该任务下有多少【待批改 (pending)】的作业
        pending_count = 0
        if current_user.is_admin:
            pending_count = Submission.query.filter_by(assignment_id=am.id, status='pending').count()

        quests.append({
            "id": am.id,
            "title": am.title,
            "issuer": am.created_by.username if am.created_by else "系统",
            "status": sub.status if sub else "active",
            "is_viewed": sub.is_viewed if sub else False,
            "grade": sub.grade if sub else None,
            "stars_earned": sub.stars_earned if sub else 0,
            "star_level": am.star_level,
            "page_url": url_for('view_assignment', assign_id=am.id),
            
            # 💡 新增管理员专属字段：传给前端 JS 用于渲染快捷按钮和红点角标
            "is_admin": current_user.is_admin,
            "grade_url": url_for('grade_assignment', assign_id=am.id),
            "pending_count": pending_count
        })
    
    return jsonify({"quests": quests})

# ==========================================
# 👑 管理员专属：全局批改大厅 (总览所有任务进度)
# ==========================================
@app.route('/admin/grading_hub')
@login_required
def global_grading_hub():
    if not current_user.is_admin:
        abort(403)
        
    # 获取所有作业（按创建时间倒序）
    all_assignments = Assignment.query.order_by(Assignment.created_at.desc()).all()
    
    assignments_data = []
    for assign in all_assignments:
        # 计算该作业下的提交总数和待批改数
        total_subs = Submission.query.filter_by(assignment_id=assign.id).count()
        pending_subs = Submission.query.filter_by(assignment_id=assign.id, status='pending').count()
        
        # 只在列表里展示有人提交过的作业
        if total_subs > 0:
            assignments_data.append({
                'assignment': assign,
                'total_subs': total_subs,
                'pending_subs': pending_subs
            })
            
    # 排序：有待批改的排在前面，其次按发布时间倒序
    assignments_data.sort(key=lambda x: (x['pending_subs'] > 0, x['assignment'].created_at), reverse=True)
    
    return render_template('admin/grading_hub.html', assignments_data=assignments_data)


# ==========================================
# 👑 管理员：具体任务的分屏批阅控制台
# ==========================================
@app.route('/assignment/<int:assign_id>/grade', methods=['GET', 'POST'])
@login_required
def grade_assignment(assign_id):
    if not current_user.is_admin:
        abort(403)
    
    assignment = Assignment.query.get_or_404(assign_id)
    
    # 1. 处理批改提交请求
    if request.method == 'POST':
        sub_id = request.form.get('submission_id')
        grade = request.form.get('grade')
        feedback = request.form.get('feedback')
        stars_earned = request.form.get('stars_earned', assignment.star_level, type=int)
        
        sub = Submission.query.get(sub_id)
        if sub and sub.assignment_id == assignment.id:
            sub.stars_earned = stars_earned
            sub.grade = grade
            sub.feedback = feedback
            sub.status = 'graded'
            sub.is_viewed = False # 设为未看，触发学生端提醒
            db.session.commit()
            flash(f"已成功为 {sub.user.real_name or sub.user.username} 批改作业！", "success")
            
        return redirect(url_for('grade_assignment', assign_id=assignment.id))
        
    # 2. GET 请求：准备分屏界面的所有数据
    # 按提交时间先后排序
    submissions = Submission.query.filter_by(assignment_id=assignment.id).order_by(Submission.submitted_at.desc()).all()
    
    classes_set = set()
    subs_data = []
    
    for sub in submissions:
        c_name = sub.user.class_name or "未分配班级"
        classes_set.add(c_name)
        
        # 将 Markdown 转换为 HTML，方便前端直接渲染
        html_content = markdown(sub.content_snapshot_md or "*内容为空*")
        
        subs_data.append({
            'id': sub.id,
            'user_id': sub.user.id,
            'username': sub.user.username,
            'real_name': sub.user.real_name or sub.user.username,
            'class_name': c_name,
            'status': sub.status,
            'grade': sub.grade or 'A',
            'stars_earned': sub.stars_earned or assignment.star_level,
            'feedback': sub.feedback or '',
            'submitted_at': sub.submitted_at.strftime("%m-%d %H:%M"),
            'content_html': html_content
        })
        
    classes_list = sorted(list(classes_set))
    
    return render_template('admin/grade_assignment.html', 
                           assignment=assignment, 
                           submissions_json=subs_data, 
                           classes=classes_list)

# ==========================================
# 📜 学生端：查看委托与提交作业详情页
# ==========================================
@app.route('/assignment/<int:assign_id>', methods=['GET', 'POST'])
@login_required
def view_assignment(assign_id):
    assignment = Assignment.query.get_or_404(assign_id)
    
    # 获取当前用户的提交记录
    submission = Submission.query.filter_by(assignment_id=assignment.id, user_id=current_user.id).first()
    
    # 💡 如果是刚批改完进来查看的，自动把“已查看”标记为 True，去掉广场上的绿色感叹号
    if submission and submission.status == 'graded' and not submission.is_viewed:
        submission.is_viewed = True
        db.session.commit()
    
    # POST 请求：处理学生提交的作业
    if request.method == 'POST':
        content = request.form.get('content')
        
        if not submission:
            # 第一次提交
            submission = Submission(
                assignment_id=assignment.id, 
                user_id=current_user.id, 
                content=content, 
                content_snapshot_md=content,
                status='pending'
            )
            db.session.add(submission)
        else:
            # 重新提交修改
            submission.content = content
            submission.status = 'pending'
            submission.content_snapshot_md = content
            # submission.submitted_at = now_utc8() # 如果有提交时间字段可以解开这行
            
        db.session.commit()
        flash("✨ 委托提交成功！请等待批阅。", "success")
        return redirect(url_for('view_assignment', assign_id=assignment.id))
        
    return render_template('view_assignment.html', assignment=assignment, submission=submission, now_utc8=now_utc8)

# ==========================================
# 🗑️ 管理员：撕下(删除)委托 API
# ==========================================
@app.route('/api/assignments/<int:assign_id>', methods=['DELETE'])
@login_required
def delete_assignment_api(assign_id):
    # 安全拦截：只有管理员能删
    if not current_user.is_admin:
        return jsonify({"success": False, "error": "权限不足"}), 403
        
    assign = Assignment.query.get_or_404(assign_id)
    
    # 💡 必须先删除关联的提交记录，否则数据库会因为外键约束报错
    Submission.query.filter_by(assignment_id=assign.id).delete()
    
    # 彻底删除该任务
    db.session.delete(assign)
    db.session.commit()
    
    return jsonify({"success": True})


# ==========================================
# 🎛️ 管理员：白噪声唱片管理
# ==========================================
@app.route('/admin/vinyls', methods=['GET', 'POST'])
@login_required
def manage_vinyls():
    if not current_user.is_admin:
        abort(403)
        
    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        condition_type = request.form.get('condition_type', 'all_users')
        condition_value = int(request.form.get('condition_value', 0))
        
        cover_file = request.files.get('cover_file')
        audio_file = request.files.get('audio_file')
        
        if not name or not cover_file or not audio_file:
            flash("名称、封面和音频文件缺一不可", "error")
            return redirect(url_for('manage_vinyls'))
            
        # 保存目录
        cover_dir = os.path.join(app.root_path, 'static', 'uploads', 'vinyls', 'covers')
        audio_dir = os.path.join(app.root_path, 'static', 'uploads', 'vinyls', 'audio')
        os.makedirs(cover_dir, exist_ok=True)
        os.makedirs(audio_dir, exist_ok=True)
        
        # 处理封面图
        cover_ext = cover_file.filename.rsplit('.', 1)[-1].lower()
        cover_filename = f"cover_{uuid.uuid4().hex[:8]}.{cover_ext}"
        cover_file.save(os.path.join(cover_dir, cover_filename))
        cover_url = f"/static/uploads/vinyls/covers/{cover_filename}"
        
        # 处理音频流 (支持 mp3, wav, ogg)
        audio_ext = audio_file.filename.rsplit('.', 1)[-1].lower()
        audio_filename = f"audio_{uuid.uuid4().hex[:8]}.{audio_ext}"
        audio_file.save(os.path.join(audio_dir, audio_filename))
        audio_url = f"/static/uploads/vinyls/audio/{audio_filename}"
        
        # 存入数据库
        v = VinylRecord(name=name, cover_url=cover_url, audio_url=audio_url, 
                        condition_type=condition_type, condition_value=condition_value)
        db.session.add(v)
        db.session.commit()
        
        # 💡 如果是全员发放，直接分发
        if condition_type == 'all_users':
            users = User.query.all()
            for u in users:
                if not UserVinyl.query.filter_by(user_id=u.id, vinyl_id=v.id).first():
                    db.session.add(UserVinyl(user_id=u.id, vinyl_id=v.id))
            db.session.commit()
            
        flash(f"唱片 {name} 压制成功并已发布！", "success")
        return redirect(url_for('manage_vinyls'))
        
    vinyls = VinylRecord.query.order_by(VinylRecord.created_at.desc()).all()
    # 借用徽章的条件字典展示
    condition_map = {
        'all_users': '全员派发', 'pomo_count': '番茄达人(次)', 'study_hours': '专注时长(小时)',
        'streak_days': '连续打卡(天)', 'note_count': '产出笔记(篇)'
    }
    return render_template('admin/manage_vinyls.html', vinyls=vinyls, condition_map=condition_map)


@app.route('/admin/vinyls/<int:v_id>/delete', methods=['POST'])
@login_required
def delete_vinyl(v_id):
    if not current_user.is_admin: abort(403)
    v = VinylRecord.query.get_or_404(v_id)
    db.session.delete(v)
    db.session.commit()
    flash("唱片已销毁", "success")
    return redirect(url_for('manage_vinyls'))

# ==========================================
# 🎧 学生端：赛博频段 (Cyber Radio)
# ==========================================
@app.route('/radio')
@login_required
def cyber_radio():
    # 1. 查出用户已获得的唱片 ID
    user_vinyls = UserVinyl.query.filter_by(user_id=current_user.id).all()
    earned_v_ids = [uv.vinyl_id for uv in user_vinyls]
    
    # 2. 取出所有唱片
    all_vinyls = VinylRecord.query.order_by(VinylRecord.created_at.asc()).all()
    
    # 借用条件中英文字典
    condition_map = {
        'all_users': 'DEFAULT_INIT // 默认配发', 'pomo_count': 'POMO_REQ // 专注次数需达', 
        'study_hours': 'TIME_REQ // 累计时长需达', 'streak_days': 'STREAK_REQ // 连续打卡需达', 
        'note_count': 'DATA_REQ // 产出笔记需达'
    }
    
    return render_template('cyber_radio.html', all_vinyls=all_vinyls, earned_v_ids=earned_v_ids, condition_map=condition_map)
    

def ensure_platform_seed_data():
    os.makedirs(app.config["UPLOAD_FOLDER"], exist_ok=True)

    if MarkdownStyle.query.first() is None:
        db.session.add(MarkdownStyle(css_text=""))
        db.session.commit()

    seed_default_phone_apps()
    ensure_builtin_phone_apps()


def initialize_platform(run_schema=False):
    if run_schema:
        db.create_all()
        upgrade_db()
        try:
            migrate_stamp(revision="head")
        except Exception as exc:
            app.logger.warning(f"Unable to stamp migration head automatically: {exc}")
        repair_postgres_integer_pk_sequences()
    ensure_platform_seed_data()


def repair_postgres_integer_pk_sequences():
    if db.engine.dialect.name != "postgresql":
        return []

    repaired = []
    preparer = db.engine.dialect.identifier_preparer

    with db.engine.begin() as connection:
        inspector = inspect(connection)
        existing_tables = set(inspector.get_table_names())
        schema = connection.execute(db.text("SELECT current_schema()")).scalar() or "public"

        for table in db.metadata.sorted_tables:
            if table.name not in existing_tables:
                continue

            pk_columns = list(table.primary_key.columns)
            if len(pk_columns) != 1:
                continue

            column = pk_columns[0]
            if not isinstance(column.type, Integer):
                continue

            default_value = connection.execute(
                db.text(
                    """
                    SELECT column_default
                    FROM information_schema.columns
                    WHERE table_schema = :schema
                      AND table_name = :table_name
                      AND column_name = :column_name
                    """
                ),
                {"schema": schema, "table_name": table.name, "column_name": column.name},
            ).scalar()

            sequence_name = connection.execute(
                db.text("SELECT pg_get_serial_sequence(:table_name, :column_name)"),
                {"table_name": f"{schema}.{table.name}", "column_name": column.name},
            ).scalar()

            quoted_table = preparer.quote(table.name)
            quoted_column = preparer.quote(column.name)

            if not sequence_name:
                raw_sequence_name = f"{table.name}_{column.name}_seq"
                quoted_sequence = preparer.quote(raw_sequence_name)
                sequence_literal = raw_sequence_name.replace("'", "''")
                connection.execute(db.text(f"CREATE SEQUENCE IF NOT EXISTS {quoted_sequence}"))
                connection.execute(db.text(f"ALTER SEQUENCE {quoted_sequence} OWNED BY {quoted_table}.{quoted_column}"))
                connection.execute(
                    db.text(
                        f"ALTER TABLE {quoted_table} ALTER COLUMN {quoted_column} "
                        f"SET DEFAULT nextval('{sequence_literal}'::regclass)"
                    )
                )
                sequence_name = raw_sequence_name
                repaired.append(f"{table.name}.{column.name}: added sequence default")
            elif not default_value:
                sequence_literal = sequence_name.replace("'", "''")
                connection.execute(
                    db.text(
                        f"ALTER TABLE {quoted_table} ALTER COLUMN {quoted_column} "
                        f"SET DEFAULT nextval('{sequence_literal}'::regclass)"
                    )
                )
                repaired.append(f"{table.name}.{column.name}: restored sequence default")

            if sequence_name:
                sequence_literal = sequence_name.replace("'", "''")
                connection.execute(
                    db.text(
                        f"""
                        SELECT setval(
                            '{sequence_literal}'::regclass,
                            GREATEST(COALESCE((SELECT MAX({quoted_column}) FROM {quoted_table}), 0), 1),
                            COALESCE((SELECT MAX({quoted_column}) FROM {quoted_table}), 0) > 0
                        )
                        """
                    )
                )

    return repaired


@app.cli.group("wikibook")
def wikibook_cli():
    """Operational commands for the Wikibook platform."""


@wikibook_cli.command("init-platform")
@click.option("--with-schema/--no-schema", default=False, help="Create/upgrade schema before seeding platform data.")
def init_platform_command(with_schema):
    """Prepare DB schema and baseline seed data."""
    initialize_platform(run_schema=with_schema)
    click.echo("Wikibook platform initialization complete.")


@wikibook_cli.command("repair-postgres-sequences")
def repair_postgres_sequences_command():
    """Repair PostgreSQL integer primary key sequence defaults after imports/migrations."""
    repaired = repair_postgres_integer_pk_sequences()
    if repaired:
        for item in repaired:
            click.echo(item)
        click.echo(f"Repaired {len(repaired)} primary key sequence default(s).")
    else:
        click.echo("No PostgreSQL primary key sequence defaults needed repair.")


@wikibook_cli.command("enqueue-judge-task")
@click.argument("task_id", type=int)
def enqueue_judge_task_command(task_id):
    """Push an existing judge task into the Redis queue."""
    task = JudgeTask.query.get(task_id)
    if task is None:
        raise click.ClickException(f"JudgeTask {task_id} does not exist.")

    job = enqueue_judge_task(task_id)
    task.queue_job_id = job.id
    task.status = 'queued'
    db.session.commit()
    click.echo(f"Queued JudgeTask {task_id} as job {job.id}.")




if __name__ == "__main__":
    with app.app_context():
        if app.config.get("AUTO_INIT_DB"):
            initialize_platform(run_schema=True)
        
    app.run(
        host=os.environ.get("HOST", "0.0.0.0"),
        port=int(os.environ.get("PORT", "5009")),
        debug=os.environ.get("FLASK_DEBUG", "false").lower() == "true",
    )
